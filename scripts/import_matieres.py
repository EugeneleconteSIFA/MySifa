"""
SIFA — Import matières premières depuis Excel
Usage : python scripts/import_matieres.py [chemin_vers_fichier.xlsx]

Ce script peuple directement la base de données SQLite à partir du classeur
"Prix matière première support adhésif SIFA".

Pré-requis : le serveur MySifa doit avoir tourné au moins une fois pour que
la migration 16 ait créé les tables matiere_params / matiere_base / matiere_config.
Si ce n'est pas le cas, lancez d'abord "python main.py" quelques secondes.

Idempotent : relancer le script met à jour les lignes existantes.
"""

import sys
import os
import sqlite3
import openpyxl
from datetime import datetime

# ── Chemin DB ─────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH  = os.getenv("DB_PATH", os.path.join(BASE_DIR, "data", "production.db"))

# ── Chemin Excel ──────────────────────────────────────────────────
if len(sys.argv) > 1:
    EXCEL_PATH = sys.argv[1]
else:
    # Cherche dans data/ un fichier .xlsx contenant "prix" ou "matiere"
    data_dir = os.path.join(BASE_DIR, "data")
    candidates = []
    if os.path.exists(data_dir):
        candidates = sorted([
            os.path.join(data_dir, f)
            for f in os.listdir(data_dir)
            if f.lower().endswith(".xlsx") and any(k in f.lower() for k in ("prix", "matiere", "matiè"))
        ])
    if candidates:
        EXCEL_PATH = candidates[-1]  # Le plus récent (alphabétiquement)
        print(f"Fichier détecté automatiquement : {os.path.basename(EXCEL_PATH)}")
    else:
        print("Erreur : aucun fichier Excel trouvé dans data/")
        print("Usage : python scripts/import_matieres.py /chemin/vers/prix.xlsx")
        sys.exit(1)

NOW = datetime.now().isoformat()

# ──────────────────────────────────────────────────────────────────
# Schéma des tables attendu par app/routers/matiere_prix.py
#
# matiere_params :
#   id, categorie, code, designation, fournisseur, poids_m2,
#   prix_eur_m2, prix_usd_kg, taux_change, incidence_dollar,
#   transport_total, appellation, grammage, notes, updated_at
#
# matiere_base :
#   id, ref_interne, designation, frontal, type_adhesion, adhesif,
#   silicone, glassine, marqueur, prix_cohesio, prix_rotoflex,
#   rotoflex_supplement_eur_m2, updated_at
#
# matiere_config : cle, valeur, updated_at
# ──────────────────────────────────────────────────────────────────

# Mapping colonnes Parametres (index 0-based, d'après le classeur SIFA)
# Col 0  : code / catégorie (S, GLS, E, P, VB, CO, TP…)
# Col 1  : poids / référence numérique
# Col 2  : désignation
# Col 3  : notes / date MAJ
# Col 5  : prix ancien
# Col 6  : fournisseur principal
# Col 7  : prix €/m² (PRIX PRINCIPAL)
# Col 8  : taxes ou gaz (coefficient additionnel)
# Col 10 : incidence dollar/taxes import
# Col 13 : prix au kg en USD
# Col 19 : transport au kg (€/kg)
# Col 20 : transport au m² (€/m²) — utilisé en transport_total si disponible
# Col 22 : taux de conversion USD→EUR
# Col 23 : poids au m² (kg/m²)
# Col 24 : fournisseur alternatif
# Col 25 : appellation (code court)
# Col 26 : grammage

# En-têtes de section dans la feuille Parametres
SECTION_PATTERNS = [
    ("1- silicone",          "Silicone"),
    ("2- glassine",          "Glassine"),
    ("3- adhésif enlevable", "Adhésif enlevable"),
    ("3- adhesif enlevable", "Adhésif enlevable"),
    ("4- adhésif",           "Adhésif"),
    ("4- adhesif",           "Adhésif"),
    ("5- adhésif congélation","Adhésif congélation"),
    ("5- adhesif congelation","Adhésif congélation"),
    ("5- adhésif permanent", "Adhésif permanent"),
    ("5- adhesif permanent", "Adhésif permanent"),
    ("5- adhésif pneumatique","Adhésif pneumatique"),
    ("5- adhesif pneumatique","Adhésif pneumatique"),
    ("6- velin",             "Velin"),
    ("6- couché",            "Couché"),
    ("6- couche",            "Couché"),
    ("7- thermique eco",     "Thermique Eco"),
    ("8- thermique pro",     "Thermique Pro"),
    ("9-",                   "PP / Polyester"),
]


def _safe_float(v):
    if v is None:
        return None
    try:
        f = float(v)
        if f != f:  # NaN
            return None
        return f
    except (ValueError, TypeError):
        return None


def _safe_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s.lower() not in ("none", "nan") else None


def _safe_int(v):
    f = _safe_float(v)
    if f is None:
        return None
    try:
        return int(round(f))
    except (ValueError, TypeError):
        return None


def _is_section_header(designation: str):
    """Retourne (categorie, True) si la désignation est un en-tête de section, sinon (None, False)."""
    dl = designation.lower().strip()
    for pattern, cat in SECTION_PATTERNS:
        if dl.startswith(pattern) or pattern in dl:
            return cat, True
    return None, False


def ensure_tables(conn):
    """
    Crée les tables si absentes, en cohérence avec app/core/database.py migration 16.
    Ne touche pas aux tables si elles existent déjà avec le bon schéma.
    """
    conn.execute("""
        CREATE TABLE IF NOT EXISTS schema_migrations (
            version INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            applied_at TEXT NOT NULL
        )
    """)

    existing = [r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table'"
    ).fetchall()]

    if "matiere_config" not in existing:
        conn.execute("""
            CREATE TABLE matiere_config (
                cle        TEXT PRIMARY KEY,
                valeur     TEXT NOT NULL,
                updated_at TEXT
            )
        """)
        conn.execute("INSERT OR IGNORE INTO matiere_config VALUES ('marge_erreur','5',?)", (NOW,))
        conn.execute("INSERT OR IGNORE INTO matiere_config VALUES ('taux_change_usd','0.85',?)", (NOW,))
        conn.execute("INSERT OR IGNORE INTO matiere_config VALUES ('supplement_rotoflex_eur_m2','0.06',?)", (NOW,))
        print("  Table matiere_config créée.")

    if "matiere_params" not in existing:
        conn.execute("""
            CREATE TABLE matiere_params (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                categorie        TEXT NOT NULL DEFAULT '',
                code             TEXT,
                designation      TEXT NOT NULL,
                fournisseur      TEXT,
                poids_m2         REAL,
                prix_eur_m2      REAL,
                prix_usd_kg      REAL,
                taux_change      REAL DEFAULT 1.0,
                incidence_dollar REAL DEFAULT 1.0,
                transport_total  REAL DEFAULT 0,
                appellation      TEXT,
                grammage         INTEGER,
                notes            TEXT,
                updated_at       TEXT
            )
        """)
        print("  Table matiere_params créée.")
    else:
        # Ajouter les colonnes manquantes si la table existait avec un ancien schéma
        existing_cols = [r[1] for r in conn.execute("PRAGMA table_info(matiere_params)").fetchall()]
        for col, defn in [
            ("incidence_dollar", "REAL DEFAULT 1.0"),
            ("transport_total",  "REAL DEFAULT 0"),
        ]:
            if col not in existing_cols:
                conn.execute(f"ALTER TABLE matiere_params ADD COLUMN {col} {defn}")
                print(f"  Colonne matiere_params.{col} ajoutée.")

    if "matiere_base" not in existing:
        conn.execute("""
            CREATE TABLE matiere_base (
                id                         INTEGER PRIMARY KEY AUTOINCREMENT,
                ref_interne                INTEGER,
                designation                TEXT NOT NULL,
                frontal                    TEXT,
                type_adhesion              TEXT,
                adhesif                    TEXT,
                silicone                   TEXT,
                glassine                   TEXT,
                marqueur                   TEXT,
                prix_cohesio               REAL,
                prix_rotoflex              REAL,
                rotoflex_supplement_eur_m2 REAL,
                updated_at                 TEXT
            )
        """)
        print("  Table matiere_base créée.")
    else:
        existing_cols = [r[1] for r in conn.execute("PRAGMA table_info(matiere_base)").fetchall()]
        if "rotoflex_supplement_eur_m2" not in existing_cols:
            conn.execute("ALTER TABLE matiere_base ADD COLUMN rotoflex_supplement_eur_m2 REAL")
            print("  Colonne matiere_base.rotoflex_supplement_eur_m2 ajoutée.")

    conn.execute(
        "INSERT OR IGNORE INTO schema_migrations (version, name, applied_at) VALUES (?,?,?)",
        (16, "matiere_params_base_config", NOW)
    )
    conn.commit()


def import_parametres(ws, conn):
    """
    Parse la feuille 'Parametres' du classeur SIFA et insère dans matiere_params.

    Mapping colonnes (index 0-based) :
      0=code, 1=poids_ref, 2=désignation, 3=notes, 5=prix_ancien,
      6=fournisseur, 7=prix_eur_m2, 8=taxes_gaz, 10=incidence_dollar,
      13=prix_usd_kg, 20=transport_m2, 22=taux_change, 23=poids_m2,
      24=fournisseur_alt, 25=appellation, 26=grammage
    """
    current_category = "Autre"
    imported = 0
    updated  = 0

    # Lire le taux de change global (ligne 1, col G = index 6)
    row1_vals = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    taux_global = _safe_float(row1_vals[6]) if len(row1_vals) > 6 else None
    if taux_global and taux_global > 0:
        conn.execute(
            "INSERT OR REPLACE INTO matiere_config (cle, valeur, updated_at) VALUES (?,?,?)",
            ("taux_change_usd", str(taux_global), NOW)
        )

    for row_values in ws.iter_rows(min_row=3, values_only=True):
        r = list(row_values) + [None] * 28
        r = r[:28]

        designation = _safe_str(r[2])
        if not designation:
            continue

        # Détection en-tête de section
        cat, is_header = _is_section_header(designation)
        if is_header:
            current_category = cat
            continue

        # Glassine et sans-silicone ont le code GLS/SGLN dans col 0
        code_raw = _safe_str(r[0]) or ""
        if code_raw.upper() in ("GLS",) and current_category == "Silicone":
            current_category = "Glassine"

        # Ignorer les lignes sans aucune donnée de prix
        prix_eur = _safe_float(r[7])
        prix_usd = _safe_float(r[13])
        if prix_eur is None and prix_usd is None:
            continue

        # Notes : date MAJ + infos supplémentaires
        notes_parts = []
        if _safe_str(r[3]):
            notes_parts.append(_safe_str(r[3]))
        if _safe_str(r[5]):
            notes_parts.append(f"Prix ancien : {_safe_str(r[5])}")
        if _safe_float(r[8]) is not None:
            notes_parts.append(f"Taxes/gaz : {r[8]}")
        if _safe_float(r[19]) is not None:
            notes_parts.append(f"Transport/kg : {r[19]:.4f} €/kg")
        if _safe_float(r[18]) is not None:
            notes_parts.append(f"Transport/lot 26t : {r[18]:.0f} €")

        # Transport au m² : utiliser la valeur calculée col 20 si disponible,
        # sinon calculer à partir de transport/kg * poids_m2
        transport_m2 = _safe_float(r[20])
        if transport_m2 is None:
            t_kg = _safe_float(r[19])
            p_m2 = _safe_float(r[23])
            if t_kg is not None and p_m2 is not None:
                transport_m2 = t_kg * p_m2

        code        = _safe_str(r[0])
        fournisseur = _safe_str(r[6]) or _safe_str(r[24])
        appellation = _safe_str(r[25])

        grammage = None
        for col in (r[26], r[1]):
            g = _safe_float(col)
            if g is not None and 10 <= g <= 1000:
                grammage = int(round(g))
                break

        poids_m2         = _safe_float(r[23])
        incidence_dollar = _safe_float(r[10]) or 1.0
        taux_change      = _safe_float(r[22]) or taux_global or 1.0

        notes_str = " | ".join(notes_parts) if notes_parts else None

        existing = conn.execute(
            "SELECT id FROM matiere_params WHERE designation=? AND categorie=?",
            (designation, current_category)
        ).fetchone()

        if existing:
            conn.execute("""
                UPDATE matiere_params SET
                    code=?, fournisseur=?, grammage=?, poids_m2=?,
                    prix_eur_m2=?, prix_usd_kg=?, taux_change=?,
                    incidence_dollar=?, transport_total=?,
                    appellation=?, notes=?, updated_at=?
                WHERE id=?
            """, (
                code, fournisseur, grammage, poids_m2,
                prix_eur, prix_usd, taux_change,
                incidence_dollar, transport_m2 or 0,
                appellation, notes_str, NOW,
                existing[0]
            ))
            updated += 1
        else:
            conn.execute("""
                INSERT INTO matiere_params (
                    categorie, code, designation, fournisseur, grammage,
                    poids_m2, prix_eur_m2, prix_usd_kg, taux_change,
                    incidence_dollar, transport_total,
                    appellation, notes, updated_at
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                current_category, code, designation, fournisseur, grammage,
                poids_m2, prix_eur, prix_usd, taux_change,
                incidence_dollar, transport_m2 or 0,
                appellation, notes_str, NOW
            ))
            imported += 1

    conn.commit()
    return imported, updated


def import_base_matieres(ws, conn):
    """
    Parse la feuille 'Base_matières' du classeur SIFA.
    Colonnes : A=ref, B=désignation, C=frontal, D=type, E=adhésif,
               F=silicone, G=glassine, H=marqueur, I=prix_cohesio, J=prix_rotoflex
    """
    imported = 0
    updated  = 0
    ROTOFLEX_DEFAULT_SUPPLEMENT = 0.06  # €/m² par défaut

    for row_values in ws.iter_rows(min_row=2, values_only=True):
        r = list(row_values) + [None] * 13
        r = r[:13]

        # Ligne de groupe (ex : 'VELIN', 'COUCHE'...) : col A texte, col B vide
        col_a = _safe_str(r[0])
        col_b = _safe_str(r[1])
        designation = col_b

        if not designation or designation.strip() in ("", "            "):
            continue

        # Vérifier qu'on a au moins les colonnes descriptives
        frontal  = _safe_str(r[2])
        type_adh = _safe_str(r[3])
        adhesif  = _safe_str(r[4])
        if not frontal and not type_adh:
            continue

        ref_int = None
        v = r[0]
        if v is not None:
            f = _safe_float(v)
            if f is not None and 100 <= f <= 9999:
                ref_int = int(round(f))

        silicone      = _safe_str(r[5])
        glassine      = _safe_str(r[6])
        marqueur      = _safe_str(r[7])
        prix_cohesio  = _safe_float(r[8])
        prix_rotoflex = _safe_float(r[9])

        # Calculer le supplément Rotoflex
        sup = None
        if prix_cohesio is not None and prix_rotoflex is not None:
            sup = round(prix_rotoflex - prix_cohesio, 6)
        elif prix_cohesio is not None:
            prix_rotoflex = prix_cohesio + ROTOFLEX_DEFAULT_SUPPLEMENT
            sup = ROTOFLEX_DEFAULT_SUPPLEMENT

        existing = conn.execute(
            "SELECT id FROM matiere_base WHERE designation=?",
            (designation,)
        ).fetchone()

        if existing:
            conn.execute("""
                UPDATE matiere_base SET
                    ref_interne=?, frontal=?, type_adhesion=?,
                    adhesif=?, silicone=?, glassine=?, marqueur=?,
                    prix_cohesio=?, prix_rotoflex=?,
                    rotoflex_supplement_eur_m2=?, updated_at=?
                WHERE id=?
            """, (
                ref_int, frontal, type_adh,
                adhesif, silicone, glassine, marqueur,
                prix_cohesio, prix_rotoflex,
                sup, NOW,
                existing[0]
            ))
            updated += 1
        else:
            conn.execute("""
                INSERT INTO matiere_base (
                    ref_interne, designation, frontal, type_adhesion,
                    adhesif, silicone, glassine, marqueur,
                    prix_cohesio, prix_rotoflex,
                    rotoflex_supplement_eur_m2, updated_at
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                ref_int, designation, frontal, type_adh,
                adhesif, silicone, glassine, marqueur,
                prix_cohesio, prix_rotoflex,
                sup, NOW
            ))
            imported += 1

    conn.commit()
    return imported, updated


def main():
    print("=" * 60)
    print("Import matières premières SIFA")
    print(f"  Excel : {EXCEL_PATH}")
    print(f"  DB    : {DB_PATH}")
    print("=" * 60)

    if not os.path.exists(EXCEL_PATH):
        print(f"\nErreur : fichier Excel introuvable : {EXCEL_PATH}")
        sys.exit(1)

    if not os.path.exists(DB_PATH):
        print(f"\nErreur : base de données introuvable : {DB_PATH}")
        print("Lancez l'application une première fois pour initialiser la DB, puis relancez ce script.")
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    except Exception as e:
        print(f"\nErreur ouverture Excel : {e}")
        sys.exit(1)

    print(f"\nFeuilles détectées : {wb.sheetnames}")

    conn = sqlite3.connect(DB_PATH, timeout=15)
    conn.row_factory = sqlite3.Row

    # Vérifier/créer les tables
    print("\n[1/3] Vérification du schéma DB...")
    ensure_tables(conn)
    print("      OK")

    # Import feuille Parametres
    sheet_params = next((s for s in wb.sheetnames if "param" in s.lower()), None)
    if sheet_params:
        print(f"\n[2/3] Import feuille '{sheet_params}'...")
        n_imp, n_upd = import_parametres(wb[sheet_params], conn)
        print(f"      {n_imp} lignes insérées, {n_upd} mises à jour")
    else:
        print("\n[2/3] Feuille Parametres introuvable — ignorée.")
        n_imp = 0

    # Import feuille Base_matières
    sheet_base = next(
        (s for s in wb.sheetnames if "base" in s.lower() or ("mat" in s.lower() and "param" not in s.lower())),
        None
    )
    if sheet_base:
        print(f"\n[3/3] Import feuille '{sheet_base}'...")
        n_imp2, n_upd2 = import_base_matieres(wb[sheet_base], conn)
        print(f"      {n_imp2} lignes insérées, {n_upd2} mises à jour")
    else:
        print("\n[3/3] Feuille Base_matières introuvable — ignorée.")
        n_imp2 = 0

    # Résumé
    total_params = conn.execute("SELECT COUNT(*) FROM matiere_params").fetchone()[0]
    total_base   = conn.execute("SELECT COUNT(*) FROM matiere_base").fetchone()[0]
    cats = conn.execute(
        "SELECT categorie, COUNT(*) as n FROM matiere_params GROUP BY categorie ORDER BY categorie"
    ).fetchall()

    print("\n" + "=" * 60)
    print("Résumé")
    print(f"  matiere_params : {total_params} lignes totales")
    for c in cats:
        print(f"    [{c[0]}] : {c[1]} lignes")
    print(f"  matiere_base   : {total_base} lignes totales")
    print()
    print("Pour déployer sur le VPS :")
    print("  ./deploy.sh --db")
    print("=" * 60)

    conn.close()


if __name__ == "__main__":
    main()
