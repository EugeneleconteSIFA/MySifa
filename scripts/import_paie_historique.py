"""
Import payroll historical data from xlsx into MySifa database.

Usage (on the VPS, from the repo root):
    python3 scripts/import_paie_historique.py

The script expects the xlsx file at:
    data/uploads/paie_historique.xlsx

(copy your 'pour export data.xlsx' there before running)

Creates paie_employes + paie_variables tables if missing.
Creates inactive placeholder users for employees without accounts.
Uses fuzzy name matching to avoid duplicates when spelling differs slightly.
"""
from __future__ import annotations

import json
import os
import re
import sqlite3
import unicodedata
from datetime import datetime
from difflib import SequenceMatcher

import openpyxl

# ─── Paths ───────────────────────────────────────────────────────────────────

BASE_DIR  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH   = os.environ.get("DB_PATH", os.path.join(BASE_DIR, "data", "production.db"))
XLSX_PATH = os.path.join(BASE_DIR, "data", "uploads", "paie_historique.xlsx")

# ─── Normalize helpers ───────────────────────────────────────────────────────

def normalize(s: str) -> str:
    """Lowercase, strip accents, remove punctuation, collapse spaces."""
    if not s:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.lower()
    s = re.sub(r"['\"‘’“”]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def name_key(s: str) -> str:
    """Order-independent normalized name key (sort words)."""
    return " ".join(sorted(normalize(s).split()))


def name_similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, normalize(a), normalize(b)).ratio()


FUZZY_THRESHOLD = 0.82   # similarity score to consider a match

# ─── Label → field key mapping ───────────────────────────────────────────────

LABEL_MAP: dict[str, str] = {
    normalize("Nom"):                                    "nom",
    normalize("Prénom"):                                 "prenom",
    normalize("Matricule"):                              "matricule",
    normalize("Compteur HS M-1"):                        "compteur_hs_m1",
    normalize("Contrat (TYPE)"):                         "contrat_type",
    normalize("Date début"):                             "date_debut",
    normalize("Date fin"):                               "date_fin",
    normalize("Nb d’heures de base"):               "nb_heures_base",
    normalize("Nb d'heures de base"):                    "nb_heures_base",
    normalize("Nb d’heures à payer"):          "nb_heures_payer",
    normalize("Nb d'heures à payer"):                    "nb_heures_payer",
    normalize("Taux horaire"):                           "taux_horaire",
    normalize("Salaire mensuel"):                        "salaire_mensuel",
    normalize("Augmentation de Salaire"):                "augmentation_salaire",
    normalize("Commissions sur ventes"):                 "commissions_ventes",
    normalize("MUTUELLE"):                               "mutuelle",
    normalize("Avantages en natures voiture"):           "avantage_voiture",
    normalize("Heure de nuit"):                          "heures_nuit",
    normalize("dont Heures DE NUIT férié"):              "heures_nuit_ferie",
    normalize("dont Heures DE NUIT dimanche"):           "heures_nuit_dimanche",
    normalize("dont Heure de nuit dimanche férié"):      "heures_nuit_dimanche_ferie",
    normalize("Heures sup 25 %"):                        "heures_sup_25",
    normalize("Heures sup 50 %"):                        "heures_sup_50",
    normalize("Heures Supplémentaires DE NUIT"):         "heures_sup_nuit",
    normalize("Panier (6,47€ par jours)"):               "panier",
    normalize("Nb d’heures jour férié ( +150%)"):  "heures_ferie",
    normalize("Nb d'heures jour férié ( +150%)"):        "heures_ferie",
    normalize("Prime ancienneté"):                       "prime_anciennete",
    normalize("Prime d’objectifs"):                 "prime_objectifs",
    normalize("Prime d'objectifs"):                      "prime_objectifs",
    normalize("Prime inflation"):                        "prime_inflation",
    normalize("Prime exceptionnelle"):                   "prime_exceptionnelle",
    normalize("Solde tout compte (oui non)"):            "solde_tout_compte",
    normalize("Prime équipe"):                           "prime_equipe",
    normalize("Absence en heures"):                      "absence_heures",
    normalize("Absence maladie en heures"):              "absence_maladie_heures",
    normalize("Absence maladie en jours"):               "absence_maladie_jours",
    normalize("Absence Deces familial - Mariage"):       "absence_deces_mariage",
    normalize("Absence congés payés en heures"):         "absence_cp_heures",
    normalize("Absence congés payés en jours"):          "absence_cp_jours",
    normalize("Absence RTT"):                            "absence_rtt",
    normalize("Absence Congés sans solde heures"):       "absence_css_heures",
    normalize("Absence Congés sans solde jours"):        "absence_css_jours",
    normalize("heures d’absence non justifiés"): "absence_non_justifie_h",
    normalize("heures d'absence non justifiés"):         "absence_non_justifie_h",
    normalize("Jours d’absence non justifiés"): "absence_non_justifie_j",
    normalize("Jours d'absence non justifiés"):          "absence_non_justifie_j",
    normalize("Absence justifiée non payée Heures"):     "absence_justifiee_np_h",
    normalize("Absence justifiée non payée Jours"):      "absence_justifiee_np_j",
    normalize("Absence AT en heures"):                   "absence_at_heures",
    normalize("Absence AT en jours"):                    "absence_at_jours",
    normalize("Mi-temps thérapeutique"):                 "mi_temps_therapeutique",
    normalize("Absence Chomage partiel"):                "absence_chomage_partiel",
    normalize("Absence congés parentale"):               "absence_conge_parentale",
    normalize("DATE CONGES PAYES"):                      "date_conges_payes",
    normalize("Frais professionnels"):                   "frais_pro",
    normalize("Frais remboursement transport"):          "frais_transport",
    normalize("Prêt SIFA"):                              "pret_sifa",
    normalize("ATD"):                                    "atd",
    normalize("Acompte exceptionnel"):                   "acompte_exceptionnel",
    normalize("information"):                            "information",
}

FIXED_FIELDS = {
    "matricule", "contrat_type", "date_debut", "date_fin",
    "nb_heures_base", "taux_horaire", "salaire_mensuel",
    "prime_anciennete", "mutuelle", "avantage_voiture",
}


# ─── DB setup ────────────────────────────────────────────────────────────────

def setup_db(conn: sqlite3.Connection) -> None:
    conn.execute("PRAGMA foreign_keys = OFF")
    conn.execute("""
    CREATE TABLE IF NOT EXISTS paie_employes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER UNIQUE,
        matricule TEXT, contrat_type TEXT DEFAULT 'CDI',
        date_debut TEXT, date_fin TEXT,
        nb_heures_base REAL, taux_horaire REAL,
        salaire_mensuel REAL, prime_anciennete REAL,
        mutuelle INTEGER DEFAULT 0, avantage_voiture REAL,
        updated_at TEXT, updated_by TEXT,
        FOREIGN KEY (user_id) REFERENCES users(id)
    )""")
    conn.execute("""
    CREATE TABLE IF NOT EXISTS paie_variables (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        annee INTEGER NOT NULL, mois INTEGER NOT NULL,
        data TEXT NOT NULL DEFAULT '{}',
        updated_at TEXT NOT NULL, updated_by TEXT,
        UNIQUE(user_id, annee, mois),
        FOREIGN KEY (user_id) REFERENCES users(id)
    )""")
    # Record migration 7 if not already present
    try:
        existing = conn.execute(
            "SELECT 1 FROM schema_migrations WHERE version=7"
        ).fetchone()
        if not existing:
            conn.execute(
                "INSERT INTO schema_migrations(version, applied_at) VALUES(7,?)",
                (datetime.utcnow().isoformat(),)
            )
    except Exception:
        pass  # schema_migrations might not exist yet
    conn.commit()


# ─── Name matching ────────────────────────────────────────────────────────────

def find_user_id(
    full_name: str,
    user_by_key: dict[str, int],
    all_user_names: list[tuple[str, int]],
) -> int | None:
    # 1. Exact key match (order-independent)
    key = name_key(full_name)
    if key in user_by_key:
        return user_by_key[key]

    # 2. Fuzzy match against all user names
    best_score = 0.0
    best_id = None
    for (uname, uid) in all_user_names:
        score = name_similarity(full_name, uname)
        if score > best_score:
            best_score = score
            best_id = uid

    if best_score >= FUZZY_THRESHOLD:
        return best_id

    return None


# ─── Main import ─────────────────────────────────────────────────────────────

def run_import(db_path: str, xlsx_path: str) -> None:
    if not os.path.exists(xlsx_path):
        print(f"[error] xlsx not found: {xlsx_path}")
        print("  → Copy your payroll xlsx to data/uploads/paie_historique.xlsx")
        return

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    setup_db(conn)

    now = datetime.utcnow().isoformat()

    # Load all users (including inactive ones created by a previous run)
    db_users = conn.execute("SELECT id, nom FROM users").fetchall()
    user_by_key: dict[str, int] = {name_key(u["nom"]): u["id"] for u in db_users}
    all_user_names: list[tuple[str, int]] = [(u["nom"], u["id"]) for u in db_users]
    print(f"[db] {len(db_users)} users loaded")

    # Cache to avoid re-matching the same xlsx name across sheets
    name_to_user_id: dict[str, int] = {}

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    stats = {"matched": 0, "created": 0, "vars": 0}

    for sheet_name in ["01-2026", "02-2026", "03-2026"]:
        if sheet_name not in wb.sheetnames:
            print(f"[warn] Sheet {sheet_name} not found — skipping")
            continue

        mois_str, annee_str = sheet_name.split("-")
        mois  = int(mois_str)
        annee = int(annee_str)
        ws = wb[sheet_name]
        print(f"\n=== {sheet_name}  (annee={annee}, mois={mois})  "
              f"rows={ws.max_row} cols={ws.max_column} ===")

        # Build row_index → field_key map
        row_map: dict[int, str] = {}
        for ri in range(1, ws.max_row + 1):
            cell_val = ws.cell(ri, 1).value
            if cell_val is None:
                continue
            norm = normalize(str(cell_val))
            if norm in LABEL_MAP:
                row_map[ri] = LABEL_MAP[norm]
            else:
                print(f"  [unmapped] row {ri}: '{cell_val}'")

        nom_row    = next((r for r, k in row_map.items() if k == "nom"),    None)
        prenom_row = next((r for r, k in row_map.items() if k == "prenom"), None)
        if nom_row is None:
            print("  [error] No 'Nom' row — skipping sheet")
            continue
        print(f"  [ok] {len(row_map)} labels mapped")

        for col in range(2, ws.max_column + 1):
            nom_val    = ws.cell(nom_row, col).value
            prenom_val = ws.cell(prenom_row, col).value if prenom_row else None

            if not nom_val:
                continue

            nom_val    = str(nom_val).strip()
            prenom_val = str(prenom_val).strip() if prenom_val else ""
            full_name  = f"{prenom_val} {nom_val}".strip()

            # Use cached result if same name seen on a previous sheet
            if full_name in name_to_user_id:
                user_id = name_to_user_id[full_name]
            else:
                user_id = find_user_id(full_name, user_by_key, all_user_names)

                if user_id is None:
                    # Create inactive placeholder user
                    prenom_norm = normalize(prenom_val).split()
                    nom_norm    = normalize(nom_val).split()
                    first_part  = prenom_norm[0] if prenom_norm else "x"
                    last_part   = nom_norm[0]    if nom_norm    else "y"
                    identifiant = f"{first_part}.{last_part}"

                    # Make identifiant unique if needed
                    base_ident = identifiant
                    suffix = 2
                    while conn.execute(
                        "SELECT 1 FROM users WHERE identifiant=?", (identifiant,)
                    ).fetchone():
                        identifiant = f"{base_ident}{suffix}"
                        suffix += 1

                    display_nom = f"{prenom_val} {nom_val}".title()
                    conn.execute(
                        "INSERT INTO users"
                        "(nom, email, password_hash, role, actif, identifiant, created_at) "
                        "VALUES (?,?,?,?,?,?,?)",
                        (display_nom, f"paie_{identifiant}@sifa.local", "",
                         "fabrication", 0, identifiant, now),
                    )
                    user_id = conn.execute(
                        "SELECT last_insert_rowid()"
                    ).fetchone()[0]
                    user_by_key[name_key(display_nom)] = user_id
                    all_user_names.append((display_nom, user_id))
                    stats["created"] += 1
                    print(f"  [created]  {full_name!r} → id={user_id}"
                          f"  ({identifiant})")
                else:
                    stats["matched"] += 1
                    matched_name = next(
                        n for n, i in all_user_names if i == user_id
                    )
                    if normalize(matched_name) != normalize(full_name):
                        print(f"  [fuzzy]    {full_name!r} → '{matched_name}'"
                              f"  (id={user_id})")
                    else:
                        print(f"  [matched]  {full_name!r} → id={user_id}")

                name_to_user_id[full_name] = user_id

            # Extract field values
            field_data: dict[str, object] = {}
            for ri, fk in row_map.items():
                if fk in ("nom", "prenom"):
                    continue
                val = ws.cell(ri, col).value
                if val is not None:
                    if hasattr(val, "isoformat"):
                        val = val.strftime("%Y-%m-%d")
                    field_data[fk] = val

            fixed_data = {k: v for k, v in field_data.items() if k in FIXED_FIELDS}
            var_data   = {k: v for k, v in field_data.items() if k not in FIXED_FIELDS}

            # Upsert fixed employee data (only if no record yet — don't overwrite)
            if fixed_data:
                if not conn.execute(
                    "SELECT 1 FROM paie_employes WHERE user_id=?", (user_id,)
                ).fetchone():
                    fcols = ", ".join(["user_id"] + list(fixed_data.keys()))
                    fvals = [user_id] + list(fixed_data.values())
                    conn.execute(
                        f"INSERT INTO paie_employes ({fcols}) "
                        f"VALUES ({', '.join(['?']*len(fvals))})",
                        fvals,
                    )

            # Upsert monthly variable data
            conn.execute(
                "INSERT INTO paie_variables"
                "(user_id, annee, mois, data, updated_at, updated_by) "
                "VALUES (?,?,?,?,?,?) "
                "ON CONFLICT(user_id, annee, mois) DO UPDATE SET "
                "data=excluded.data, updated_at=excluded.updated_at",
                (user_id, annee, mois, json.dumps(var_data, ensure_ascii=False),
                 now, "import_historique"),
            )
            stats["vars"] += 1

    conn.commit()
    conn.close()

    print(f"\n{'='*50}")
    print(f"Import terminé :")
    print(f"  Employés trouvés dans la DB  : {stats['matched']}")
    print(f"  Employés créés (inactifs)    : {stats['created']}")
    print(f"  Fiches paie variables créées : {stats['vars']}")
    print(f"{'='*50}")


if __name__ == "__main__":
    print(f"DB   : {DB_PATH}")
    print(f"XLSX : {XLSX_PATH}")
    run_import(DB_PATH, XLSX_PATH)
