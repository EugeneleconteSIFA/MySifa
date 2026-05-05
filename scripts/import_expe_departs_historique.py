"""
MyExpé — Import Excel vers historique des départs (expe_departs.statut='valide').

Usage:
  python scripts/import_expe_departs_historique.py "/chemin/vers/Suivi départ 1.xlsx"

Notes:
- Insère des départs en statut 'valide' (historique) avec validated_at = now.
- Dédoublonne au mieux sur (date_enlevement, ref_sifa, arc, no_bl, no_cde_transport, client, transporteur).
- DB_PATH peut être surchargé via env var (comme le runtime).
"""

import os
import re
import sys
import sqlite3
import unicodedata
from datetime import datetime
from typing import Any, Optional

import openpyxl


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.getenv("DB_PATH", os.path.join(BASE_DIR, "data", "production.db"))


def _norm_header(v: Any) -> str:
    t = str(v or "").strip().lower()
    t = unicodedata.normalize("NFD", t)
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    t = re.sub(r"[^a-z0-9]+", "_", t)
    return t.strip("_")


def _norm_sheet_name(v: Any) -> str:
    t = str(v or "").strip().lower()
    t = unicodedata.normalize("NFD", t)
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _pick_worksheet(wb: openpyxl.Workbook, desired_name: str):
    want = _norm_sheet_name(desired_name)
    for sname in wb.sheetnames:
        if _norm_sheet_name(sname) == want:
            return wb[sname]
    return wb.active


def _as_date_yyyy_mm_dd(v: Any) -> Optional[str]:
    if v is None or v == "":
        return None
    # openpyxl peut donner un datetime/date natif
    if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
        try:
            return f"{int(v.year):04d}-{int(v.month):02d}-{int(v.day):02d}"
        except Exception:
            pass
    s = str(v).strip()
    if not s:
        return None
    # dd/mm/yyyy
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:04d}-{mo:02d}-{d:02d}"
    # yyyy-mm-dd
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s[:10]):
        return s[:10]
    return None


def _as_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _as_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", ".").replace("\u202f", "").replace(" ", ""))
    except ValueError:
        return None


def _pick(row: dict, *keys: str) -> Any:
    for k in keys:
        if k in row and row[k] not in (None, ""):
            return row[k]
    return None


def _dedupe_key(d: dict) -> tuple:
    return (
        d.get("date_enlevement") or "",
        (d.get("ref_sifa") or "").strip().lower(),
        (d.get("arc") or "").strip().lower(),
        (d.get("no_bl") or "").strip().lower(),
        (d.get("no_cde_transport") or "").strip().lower(),
        (d.get("client") or "").strip().lower(),
        (d.get("transporteur") or "").strip().lower(),
    )


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage : python scripts/import_expe_departs_historique.py /chemin/vers/fichier.xlsx")
        return 2

    excel_path = sys.argv[1]
    # Copie depuis Finder / notes : parfois préfixé "terminé@"
    if excel_path.startswith("terminé@") or excel_path.startswith("termine@"):
        excel_path = excel_path.split("@", 1)[1]

    if not os.path.exists(excel_path):
        print(f"Erreur : fichier introuvable: {excel_path}")
        return 2

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = _pick_worksheet(wb, "terminé")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Erreur : feuille vide.")
        return 1

    headers = [_norm_header(x) for x in rows[0]]
    idx = {h: i for i, h in enumerate(headers) if h}

    # Mapping tolérant (plusieurs variantes possibles)
    def get_cell(r: tuple, *h: str) -> Any:
        for k in h:
            nk = _norm_header(k)
            if nk in idx and idx[nk] < len(r):
                return r[idx[nk]]
        return None

    parsed = []
    for r in rows[1:]:
        if not r or all(v is None or str(v).strip() == "" for v in r):
            continue

        date_enl = _as_date_yyyy_mm_dd(
            get_cell(r, "date_enlevement", "date d'enlevement", "date enlèvement", "date enl")
        )
        # fallback : si une colonne "date" unique existe
        if not date_enl:
            date_enl = _as_date_yyyy_mm_dd(get_cell(r, "date"))

        d = {
            "date_enlevement": date_enl,
            "affreteurs": _as_str(get_cell(r, "affreteurs", "affreteur", "affretement", "affr")),
            "transporteur": _as_str(get_cell(r, "transporteur", "transp")),
            "client": _as_str(get_cell(r, "client")),
            "code_postal_destination": _as_str(
                get_cell(r, "code_postal_destination", "code postal", "destination", "cp", "code_postal")
            ),
            "ref_sifa": _as_str(get_cell(r, "ref_sifa", "ref sifa", "reference sifa", "ref")),
            "arc": _as_str(get_cell(r, "arc")),
            "no_cde_transport": _as_str(
                get_cell(r, "no_cde_transport", "n commande transport", "commande transport", "no commande transporteur")
            ),
            "no_bl": _as_str(get_cell(r, "no_bl", "bl", "n bl", "no bl")),
            "nb_palette": _as_float(get_cell(r, "nb_palette", "nombre de palettes", "palettes", "pal")),
            "poids_total_kg": _as_float(get_cell(r, "poids_total_kg", "poids total", "poids", "kg")),
            "date_livraison": _as_date_yyyy_mm_dd(
                get_cell(r, "date_livraison", "date livraison", "livraison", "date liv")
            ),
        }

        # Exiger au moins une date d'enlèvement + un champ d'identification
        if not d["date_enlevement"]:
            # si vraiment absent : ignore la ligne
            continue

        parsed.append(d)

    if not parsed:
        print("Aucune ligne importable (vérifie les en-têtes et la colonne date d'enlèvement).")
        return 1

    now = datetime.now().isoformat(timespec="seconds")
    validated_by = "import_excel"

    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    try:
        inserted = 0
        skipped = 0
        for d in parsed:
            k = _dedupe_key(d)
            ex = conn.execute(
                """
                SELECT id FROM expe_departs
                WHERE statut='valide'
                  AND COALESCE(substr(date_enlevement,1,10),'')=?
                  AND LOWER(COALESCE(ref_sifa,''))=?
                  AND LOWER(COALESCE(arc,''))=?
                  AND LOWER(COALESCE(no_bl,''))=?
                  AND LOWER(COALESCE(no_cde_transport,''))=?
                  AND LOWER(COALESCE(client,''))=?
                  AND LOWER(COALESCE(transporteur,''))=?
                LIMIT 1
                """,
                k,
            ).fetchone()
            if ex:
                skipped += 1
                continue

            conn.execute(
                """
                INSERT INTO expe_departs (
                  date_enlevement, affreteurs, transporteur, client, code_postal_destination,
                  ref_sifa, arc, no_cde_transport, no_bl, nb_palette, poids_total_kg, date_livraison,
                  statut, created_at, created_by_email, validated_at, validated_by_email
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?, 'valide', ?, ?, ?, ?)
                """,
                (
                    d["date_enlevement"],
                    d["affreteurs"],
                    d["transporteur"],
                    d["client"],
                    d["code_postal_destination"],
                    d["ref_sifa"],
                    d["arc"],
                    d["no_cde_transport"],
                    d["no_bl"],
                    d["nb_palette"],
                    d["poids_total_kg"],
                    d["date_livraison"],
                    now,
                    validated_by,
                    now,
                    validated_by,
                ),
            )
            inserted += 1

        conn.commit()
        print(f"Import terminé. Insérés: {inserted} | Déjà présents (skip): {skipped} | Total lus: {len(parsed)}")
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())

