#!/usr/bin/env python3
"""
Import / enrichissement de l'annuaire fournisseurs depuis un export ERP Excel.

Ce que fait ce script
---------------------
Rapproche chaque ligne du fichier Excel d'une fiche `fournisseurs_fsc`
existante, complète les champs vides de celles qu'il retrouve, et crée celles
qui manquent. Il ne remplace jamais une valeur déjà saisie dans MySifa : la
fiche tenue à la main fait foi, l'export ERP ne fait que combler les trous.
`--ecraser` inverse cette règle, champ par champ, et se dit à voix haute dans
le rapport.

Trois temps, comme l'import du catalogue produits
------------------------------------------------
    --inventaire   Ne touche à rien. Montre le rapprochement proposé, les
                   colonnes du fichier qui ne mènent nulle part, et la
                   distribution des valeurs. C'est ce qu'on lit AVANT de
                   décider quoi que ce soit.
    --simulation   Rejoue tout le travail d'écriture et affiche le diff
                   champ par champ, sans committer. C'est le défaut : lancé
                   sans option, le script ne modifie rien.
    --appliquer    Écrit.

Relançable sans doublon : le rapprochement se fait sur le nom normalisé
(accents, ponctuation et formes juridiques ignorés), puis sur le SIRET, puis
sur le numéro de TVA intracommunautaire. Un deuxième passage ne trouve donc
plus rien à créer.

Le fichier Excel de l'ERP
------------------------
Son `xl/styles.xml` porte l'attribut `biltinId` au lieu de `builtinId` — une
faute de frappe de l'outil qui l'exporte. openpyxl refuse le fichier avec un
TypeError obscur. On répare la copie en mémoire plutôt que de demander à
l'utilisateur de réenregistrer son fichier dans Excel.

Ce que le script ne devine pas
------------------------------
- **La certification FSC.** L'export n'en dit rien. Une fiche créée ici part
  donc à `has_fsc = 0`, sans licence ni certificat. Marquer l'inverse ferait
  apparaître des fournisseurs comme certifiés sans qu'aucun document ne le
  soutienne — exactement ce qu'un audit de chaîne de contrôle cherche.
- **Les catégories.** Le fichier n'en renseigne que 5 sur 199. Les fiches
  créées arrivent donc sans catégorie, à compléter dans Paramètres. C'est
  aussi ce qui alimente les favoris de la recherche fournisseur : une fiche
  sans catégorie reste trouvable, elle ne remonte simplement pas en tête.

Exemples
--------
    python3 scripts/import_fournisseurs_excel.py --fichier "Table Fournisseurs.xlsx" --inventaire
    python3 scripts/import_fournisseurs_excel.py --fichier "Table Fournisseurs.xlsx"
    python3 scripts/import_fournisseurs_excel.py --fichier "Table Fournisseurs.xlsx" --appliquer
"""

from __future__ import annotations

import argparse
import io
import json
import os
import re
import sqlite3
import sys
import unicodedata
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

RACINE = Path(__file__).resolve().parent.parent
if str(RACINE) not in sys.path:
    sys.path.insert(0, str(RACINE))


# ═══════════════════════════════════════════════════════════════════
#  Lecture du fichier Excel
# ═══════════════════════════════════════════════════════════════════

def ouvrir_classeur(chemin: Path):
    """Charge le .xlsx, en réparant au besoin le styles.xml de l'ERP.

    La réparation se fait sur une copie en mémoire : le fichier de
    l'utilisateur n'est jamais réécrit.
    """
    try:
        import openpyxl
    except ImportError:
        raise SystemExit(
            "openpyxl est requis pour lire le fichier Excel.\n"
            "  pip install openpyxl"
        )

    try:
        return openpyxl.load_workbook(chemin, data_only=True, read_only=False)
    except TypeError as e:
        if "biltinId" not in str(e) and "builtinId" not in str(e):
            raise
        print("  ! styles.xml malformé (biltinId) — réparation en mémoire.")

    tampon = io.BytesIO()
    with zipfile.ZipFile(chemin) as zin, zipfile.ZipFile(tampon, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/styles.xml":
                data = data.replace(b"biltinId", b"builtinId")
            zout.writestr(item, data)
    tampon.seek(0)
    return openpyxl.load_workbook(tampon, data_only=True)


def cle_entete(s) -> str:
    """Normalise un intitulé de colonne : « N.TVA  » et « n tva » se rejoignent."""
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def lire_lignes(chemin: Path) -> tuple[list[dict], list[str]]:
    """Renvoie (lignes en dict indexé par clé d'entête, intitulés d'origine)."""
    wb = ouvrir_classeur(chemin)
    ws = wb[wb.sheetnames[0]]
    brut = list(ws.iter_rows(values_only=True))
    if not brut:
        raise SystemExit("Feuille vide.")
    entetes = list(brut[0])
    cles = [cle_entete(h) for h in entetes]
    lignes = []
    for r in brut[1:]:
        if not any(c is not None and str(c).strip() for c in r):
            continue
        d = {}
        for i, k in enumerate(cles):
            if not k:
                continue
            v = r[i] if i < len(r) else None
            if isinstance(v, str):
                v = v.strip()
            d[k] = v if v not in ("", None) else None
        d["_ligne_excel"] = brut.index(r) + 1
        lignes.append(d)
    return lignes, [str(h) if h is not None else "" for h in entetes]


# ═══════════════════════════════════════════════════════════════════
#  Normalisation des valeurs
# ═══════════════════════════════════════════════════════════════════

# Même logique que app/core/migrations/2026_08_03_annuaire_fournisseurs.py :
# les deux doivent rapprocher les mêmes fiches, sinon l'import crée des
# doublons que la migration avait su éviter.
_FORMES_JURIDIQUES = (
    "sa", "sas", "sarl", "sasu", "gmbh", "ltd", "bv", "nv", "spa", "srl", "inc",
    "snc", "eurl", "scop", "scp", "gie", "ag", "plc", "oy", "ab", "as",
)


def norm_nom(s) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c)).lower()
    s = "".join(ch if ch.isalnum() else " " for ch in s)
    mots = [m for m in s.split() if m not in _FORMES_JURIDIQUES and len(m) > 1]
    return " ".join(mots) or " ".join(s.split())


def squash(s) -> str:
    """Nom réduit à ses seuls caractères alphanumériques.

    « 2 D M S.A.S. » → « 2dmsas », « 2DM » → « 2dm ». Sert uniquement à
    DÉTECTER une ressemblance, jamais à décider d'un rapprochement : deux
    noms tassés proches peuvent être deux sociétés différentes (« ABIX » et
    « ABI »). Le rapprochement automatique reste sur le nom normalisé, le
    SIRET et la TVA — trois clés qui ne se trompent pas.
    """
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c)).lower()
    s = re.sub(r"[^a-z0-9]+", "", s)
    for forme in sorted(_FORMES_JURIDIQUES, key=len, reverse=True):
        if s.endswith(forme) and len(s) > len(forme) + 2:
            s = s[: -len(forme)]
            break
    return s


def proches(nouveaux: list[dict], fiches: list, seuil: float = 0.86) -> list[tuple]:
    """Quasi-doublons : fiche existante ressemblant à un nom qu'on va créer.

    On les CRÉE quand même — refuser de créer sur une ressemblance ferait
    perdre des fournisseurs bien réels — mais on les signale nommément pour
    que la fusion se fasse en connaissance de cause, dans Paramètres →
    Fournisseurs → Doublons.
    """
    from difflib import SequenceMatcher
    index = [(squash(dict(f).get("nom")), dict(f)) for f in fiches]
    index = [(s, f) for s, f in index if s]
    trouves = []
    for c in nouveaux:
        sc = squash(c.get("nom"))
        if not sc:
            continue
        meilleur, ratio = None, 0.0
        for s, f in index:
            # Un préfixe strict est le cas typique (« 2dm » dans « 2dmsas ») :
            # SequenceMatcher le note bas quand les longueurs divergent.
            if sc == s or (len(sc) >= 3 and len(s) >= 3 and (sc.startswith(s) or s.startswith(sc))):
                r = 0.95
            else:
                r = SequenceMatcher(None, sc, s).ratio()
            if r > ratio:
                meilleur, ratio = f, r
        if meilleur and ratio >= seuil:
            trouves.append((c, meilleur, round(ratio, 2)))
    return trouves


def texte(v, maxlen: int = 0) -> str | None:
    if v is None:
        return None
    s = re.sub(r"\s+", " ", str(v)).strip()
    if s in ("", "-", "--", "/", "n/a", "N/A", "."):
        return None
    if maxlen and len(s) > maxlen:
        s = s[:maxlen]
    return s or None


def tel(v) -> str | None:
    """Uniformise un numéro : « 03.20.26.45.26 » et « +333 20 56 25 22 » gardent
    leur forme lisible, mais les séparateurs hétérogènes sont ramenés à
    l'espace pour qu'une recherche sur le numéro fonctionne."""
    s = texte(v)
    if not s:
        return None
    plus = s.lstrip().startswith("+")
    chiffres = re.sub(r"\D", "", s)
    if len(chiffres) < 6:
        return None            # « 0 », « - », reste d'une saisie abandonnée
    if plus:
        return "+" + chiffres
    if len(chiffres) == 10 and chiffres.startswith("0"):
        return " ".join(chiffres[i:i + 2] for i in range(0, 10, 2))
    return chiffres


def email(v) -> str | None:
    s = texte(v)
    if not s:
        return None
    s = s.split(";")[0].split(",")[0].strip().lower()
    return s if re.match(r"^[^@\s]+@[^@\s]+\.[a-z]{2,}$", s) else None


def ville(v) -> str | None:
    """Nettoie un nom de ville de deux artefacts de l'export ERP.

    1. Le suffixe « -> 59200 ». 24 lignes sur 199 le portent, et le code qui
       suit la flèche est CHAQUE FOIS identique à la colonne C.P. — vérifié sur
       les 24. C'est de la redondance d'affichage, pas une donnée : la laisser
       donnerait « CHAMP/DRAC -> 38560 » comme nom de ville dans la fiche, et
       une recherche sur « Champ » suivie du code postal ne trouverait rien.

    2. Le nom écrit deux fois de suite, sans séparateur : « DEVILLE LES
       ROUENDEVILLE LES ROUEN ». On ne défait ce doublement que s'il est
       EXACT — la première moitié égale la seconde. Un rapprochement plus
       souple risquerait de tronquer une ville dont le nom répète un mot.
    """
    s = texte(v, 120)
    if not s:
        return None
    s = re.split(r"\s*-+>\s*", s)[0].strip()
    if not s:
        return None
    moitie = len(s) // 2
    if len(s) % 2 == 0 and s[:moitie] == s[moitie:] and moitie >= 4:
        s = s[:moitie]
    return s.strip() or None


def siret(v) -> str | None:
    s = texte(v)
    if not s:
        return None
    chiffres = re.sub(r"\D", "", s)
    # 14 chiffres = SIRET, 9 = SIREN. Tout le reste est une saisie douteuse
    # qu'on préfère laisser dehors plutôt que d'indexer.
    return chiffres if len(chiffres) in (9, 14) else None


def tva(v) -> str | None:
    s = texte(v)
    if not s:
        return None
    s = re.sub(r"[^A-Za-z0-9]", "", s).upper()
    return s if re.match(r"^[A-Z]{2}[0-9A-Z]{6,13}$", s) else None


def entier(v) -> int | None:
    """Délai d'expédition en jours.

    0 n'est pas retenu : 157 lignes sur 199 le portent, c'est le défaut du
    champ dans l'ERP, pas un engagement d'expédition le jour même. L'écrire
    afficherait « expédié J+0 » sur les trois quarts de l'annuaire — un
    engagement que personne n'a pris.
    """
    if v is None:
        return None
    try:
        n = int(float(str(v).replace(",", ".")))
    except (TypeError, ValueError):
        return None
    if n <= 0 or n > 365:
        return None
    return n


_DEVISES = {
    "E": "EUR", "EUR": "EUR", "€": "EUR",
    "DOL": "USD", "USD": "USD", "$": "USD",
    "L": "GBP", "GBP": "GBP", "£": "GBP",
    "CHF": "CHF", "SEK": "SEK", "PLN": "PLN", "DKK": "DKK", "CNY": "CNY",
}


def devise(v) -> str | None:
    s = texte(v)
    return _DEVISES.get(s.upper()) if s else None


def code_liste(v) -> str | None:
    """Extrait le libellé d'une valeur ERP « 12 - Carte VISA » → « carte visa »."""
    s = texte(v)
    if not s:
        return None
    s = re.sub(r"^\s*\d+\s*-\s*", "", s).strip()
    return s or None


# Le libellé ERP → le code du référentiel config.py. Une entrée absente d'ici
# est signalée dans le rapport plutôt que silencieusement perdue : c'est ainsi
# qu'on découvre qu'un mode de règlement manque au référentiel.
_MAP_REGLEMENT = {
    "virement": "virement",
    "virement sur proforma": "virement_proforma",
    "virement 8j avant expedition": "virement_pre_expedition",
    "virement avant expedition": "virement_pre_expedition",
    "comptant": "comptant",
    "carte visa": "carte",
    "carte bancaire": "carte",
    "cheque": "cheque",
    "prelevement": "prelevement",
    "lcr acceptee": "lcr_acceptee",
    "lcr non acceptee": "lcr_non_acceptee",
    "bor": "bor",
}

_MAP_LIVRAISON = {
    "par nos soins": "par_nos_soins",
    "par vos soins": "par_vos_soins",
    "enlevement": "enlevement",
    "enlevement sur place": "enlevement",
}

_MAP_TVA = {
    "soumis a tva": "soumis_tva",
    "exonere": "exonere",
    "cee": "cee",
    "export": "export",
}

_MAP_CATEGORIE = {
    "fournisseur adhesif": "adhesif",
    "adhesif": "adhesif",
    "fournisseur frontal": "frontal",
    "frontal": "frontal",
    "fournisseur glassine": "glassine",
    "glassine": "glassine",
    "fournisseur carton": "carton",
    "carton": "carton",
    "fournisseur mandrin": "mandrin",
    "mandrin": "mandrin",
    "fournisseur palette": "palette",
    "palette": "palette",
    "complexe": "complexe",
    "negoce": "negoce",
    "sous traitant": "sous_traitant",
    "sous-traitant": "sous_traitant",
}


def cle_map(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c)).lower()
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", s)).strip()


def via_map(libelle: str | None, table: dict, inconnus: Counter, quoi: str):
    if not libelle:
        return None
    code = table.get(cle_map(libelle))
    if not code:
        inconnus[f"{quoi}: {libelle}"] += 1
    return code


# ═══════════════════════════════════════════════════════════════════
#  Extraction d'une ligne Excel → champs MySifa
# ═══════════════════════════════════════════════════════════════════

# Colonnes lues, par clé d'entête normalisée.
COL = {
    "code":        "code",
    "nom":         "raisonsociale",
    "adresse1":    "adresse1",
    "adresse2":    "adresse2",
    "bp":          "bp",
    "cp":          "cp",
    "ville":       "ville",
    "cpays":       "cpays",
    "pays":        "pays",
    "siret":       "siret",
    "rcs":         "rcs",
    "tva":         "ntva",
    "tel":         "telephone",
    "fax":         "telecopie",
    "email":       "email",
    "cat1":        "categorie1",
    "cat2":        "categorie2",
    "cat3":        "categorie3",
    "livraison":   "modedelivraison",
    "delai":       "nbejexp",
    "devise":      "devise",
    "reglement":   "modedereglement",
    "fiscale":     "posfiscale",
    "comptable":   "codecomptable",
    "etat":        "etat",
}

# Colonnes volontairement ignorées : soit vides dans l'export, soit sans
# équivalent métier dans MySifa. Listées ici pour que le rapport
# d'inventaire distingue « ignorée sciemment » de « oubliée ».
IGNOREES = {
    "no": "numéro de ligne ERP",
    "ean": "vide dans l'export",
    "nif": "3 valeurs, hétérogènes (code NAF / TVA espagnole)",
    "ftp": "constante « 1 - Non »",
    "remise": "toujours 0",
    "escompte": "quasi toujours 0",
    "encoursautorise": "quasi toujours 0",
    "montantfranco": "toujours 0",
    "notrenoclt": "10 valeurs sur 199",
    "cc": "3 valeurs, sémantique inconnue",
    "ir": "vide", "is": "vide", "if": "vide",
    "datecreation": "métadonnée ERP",
    "datemodification": "métadonnée ERP",
    "codecomptable": "reporté en note de fiche",
    "code": "reporté en note de fiche",
}


def extraire(ligne: dict, inconnus: Counter) -> dict:
    g = ligne.get

    nom = texte(g(COL["nom"])) or texte(g(COL["code"]))

    # Adresse : trois colonnes ERP pour une seule ligne d'adresse en base.
    # On les recolle dans l'ordre de lecture d'une enveloppe.
    morceaux = [texte(g(COL["adresse1"])), texte(g(COL["adresse2"])), texte(g(COL["bp"]))]
    adresse = ", ".join(m for m in morceaux if m) or None

    cats = []
    for k in ("cat1", "cat2", "cat3"):
        c = via_map(code_liste(g(COL[k])), _MAP_CATEGORIE, inconnus, "catégorie")
        if c and c not in cats:
            cats.append(c)

    # Note ERP : on ne reporte que ce qui n'est pas déjà lisible ailleurs.
    # « Code ERP : ACCESSOIRES ASUS » sur une fiche nommée « ACCESSOIRES
    # ASUS » n'informe personne, et le code comptable vaut le plus souvent
    # « F » + le code. Sur 199 lignes, cette garde évite d'écrire une note
    # inutile dans la grande majorité des fiches.
    note_erp = []
    code_erp = texte(g(COL["code"]))
    comptable = texte(g(COL["comptable"]))
    if code_erp and norm_nom(code_erp) != norm_nom(nom):
        note_erp.append("Code ERP : " + code_erp)
    if comptable and code_erp and comptable.upper() != ("F" + code_erp).upper():
        note_erp.append("Code comptable : " + comptable)
    elif comptable and not code_erp:
        note_erp.append("Code comptable : " + comptable)

    return {
        "nom": nom,
        "adresse": adresse,
        "code_postal": texte(g(COL["cp"]), 20),
        "ville": ville(g(COL["ville"])),
        "pays": (texte(g(COL["cpays"])) or "FR").upper()[:2],
        "siret": siret(g(COL["siret"])),
        "tva_intracom": tva(g(COL["tva"])),
        "rcs": texte(g(COL["rcs"]), 60),
        "telephone": tel(g(COL["tel"])),
        "fax": tel(g(COL["fax"])),
        "email": email(g(COL["email"])),
        "price_currency": devise(g(COL["devise"])),
        "mode_reglement": via_map(code_liste(g(COL["reglement"])), _MAP_REGLEMENT,
                                  inconnus, "mode de règlement"),
        "mode_livraison": via_map(code_liste(g(COL["livraison"])), _MAP_LIVRAISON,
                                  inconnus, "mode de livraison"),
        "delai_expedition_jours": entier(g(COL["delai"])),
        "regime_tva": via_map(code_liste(g(COL["fiscale"])), _MAP_TVA, inconnus, "régime TVA"),
        "categories": cats,
        "_note_erp": " · ".join(note_erp) or None,
        "_ligne_excel": ligne.get("_ligne_excel"),
    }


# Champs simplement recopiés dans la colonne de même nom.
CHAMPS_SIMPLES = (
    "adresse", "code_postal", "ville", "pays", "siret", "tva_intracom", "rcs",
    "telephone", "fax", "email", "price_currency", "mode_reglement",
    "mode_livraison", "delai_expedition_jours", "regime_tva",
)


# ═══════════════════════════════════════════════════════════════════
#  Base de données
# ═══════════════════════════════════════════════════════════════════

# Sentinelle de la cible « - » dans le fichier d'alias : ligne à ne pas importer.
IGNORER = object()


def lire_alias(chemin: Path) -> dict:
    """Table de correspondance « nom de l'export » → fiche existante.

    Pourquoi ce fichier existe
    --------------------------
    Les fiches historiques de l'annuaire portent un nom d'usage court — « UPM »,
    « Ricoh », « Avery » — et rien d'autre que leur licence FSC. L'export ERP,
    lui, nomme les mêmes sociétés en raison sociale complète : « UPM RAFLATAC »,
    « RICOH Industrie France SAS », « AVERY DENNISON MATERIALS SALES FRANCE SAS ».

    Aucune règle automatique ne peut trancher : « UPM » est un préfixe de
    « UPM RAFLATAC », mais « ABI » est aussi un préfixe de « ABIX » sans être la
    même société. Rapprocher sur la ressemblance ferait fusionner des
    fournisseurs distincts ; ne rien rapprocher créerait deux fiches par
    société — l'une avec la licence FSC, l'autre avec l'adresse et le SIRET.

    Cette seconde issue est la pire : dans l'écran de réception, les deux
    apparaissent, une seule porte le badge FSC, et rien n'empêche d'attraper
    l'autre. Un claim de chaîne de contrôle tombe là-dessus.

    D'où ce fichier, écrit une fois, relu à chaque import : deux colonnes
    séparées par `;`, `nom dans l'export` puis `id de fiche` (ou nom exact de
    fiche). Les lignes vides et celles commençant par `#` sont ignorées.
    Générez-en un pré-rempli avec --generer-alias, relisez-le, corrigez-le.
    """
    if not chemin.exists():
        raise SystemExit(f"Fichier d'alias introuvable : {chemin}")
    table = {}
    for num, ligne in enumerate(chemin.read_text(encoding="utf-8-sig").splitlines(), 1):
        ligne = ligne.strip()
        if not ligne or ligne.startswith("#"):
            continue
        bouts = [b.strip() for b in ligne.split(";")]
        if len(bouts) < 2 or not bouts[0]:
            print(f"  ! alias ligne {num} ignorée (deux colonnes attendues) : {ligne}")
            continue
        # Commentaire en fin de ligne : le fichier généré en met un (le score et
        # le nom de la fiche visée) pour que la relecture soit possible sans
        # ouvrir la base à côté. Il ne doit pas se retrouver dans la valeur.
        cible = bouts[1].split("#", 1)[0].strip()
        if not cible:
            print(f"  ! alias ligne {num} ignorée (cible vide) : {ligne}")
            continue
        # Cible « - » : ne pas importer cette ligne du tout. Le cas d'usage est
        # l'export qui liste deux fois la même relation commerciale sous deux
        # entités juridiques (siège et filiale de vente) alors qu'on ne les
        # distingue pas à l'achat. Même fichier, même syntaxe : une seule chose
        # à comprendre pour décider du sort d'une ligne.
        table[norm_nom(bouts[0])] = IGNORER if cible in ("-", "ignorer", "ignore") else cible
    return table


def resoudre_alias(table: dict, fiches: list) -> dict:
    """Convertit la cible de chaque alias (id ou nom) en fiche réelle."""
    par_id = {int(dict(f)["id"]): dict(f) for f in fiches}
    par_nom = {norm_nom(dict(f)["nom"]): dict(f) for f in fiches}
    sortie, inconnus = {}, []
    for cle, cible in table.items():
        if cible is IGNORER:
            sortie[cle] = IGNORER
            continue
        fiche = None
        if str(cible).isdigit():
            fiche = par_id.get(int(cible))
        if fiche is None:
            fiche = par_nom.get(norm_nom(cible))
        if fiche is None:
            inconnus.append(cible)
            continue
        sortie[cle] = fiche
    if inconnus:
        # Un alias qui ne pointe sur rien est une faute de frappe, pas une
        # absence : le signaler évite de croire le rapprochement fait.
        print(f"  ! {len(inconnus)} alias sans fiche correspondante : "
              + ", ".join(str(x) for x in inconnus[:8]))
    return sortie


def ecrire_alias_propose(chemin: Path, quasi: list) -> None:
    """Écrit un fichier d'alias pré-rempli à partir des quasi-doublons."""
    lignes = [
        "# Alias d'import fournisseurs — À RELIRE AVANT USAGE",
        "#",
        "# Une ligne = « le nom dans l'export ERP » ; « la fiche à enrichir ».",
        "# Chaque ligne ci-dessous est une PROPOSITION fondée sur une simple",
        "# ressemblance de nom. Deux sociétés différentes peuvent se ressembler :",
        "# vérifiez, puis commentez (#) ou supprimez les lignes fausses.",
        "#",
        "# Les fiches citées ici ne portent souvent que leur licence FSC. Les",
        "# aliaser, c'est leur donner l'adresse, le SIRET et le téléphone de",
        "# l'export SANS créer de seconde fiche — donc sans qu'un opérateur",
        "# puisse choisir, en réception, celle qui n'a pas la licence.",
        "#",
        "# Cible « - » au lieu d'un id : la ligne n'est pas importée du tout",
        "# (export qui liste deux fois la même relation sous deux entités).",
        "#",
        "# nom dans l'export ; id ou nom de la fiche  (ou « - » pour ignorer)",
    ]
    for c, f, r in sorted(quasi, key=lambda x: -x[2]):
        lignes.append(f"{c['nom']};{f['id']}    # ~{r} → « {f['nom']} »"
                      + (f" · licence {f['licence']}" if f.get("licence") else ""))
    chemin.write_text("\n".join(lignes) + "\n", encoding="utf-8")
    print(f"\nAlias proposé écrit dans {chemin} ({len(quasi)} ligne(s)).")
    print("  Relisez-le, corrigez-le, puis relancez avec --alias " + str(chemin))


def ouvrir_base(chemin: str | None) -> sqlite3.Connection:
    if not chemin:
        try:
            from config import DB_PATH
            chemin = DB_PATH
        except Exception:
            raise SystemExit("Impossible de lire DB_PATH depuis config.py — "
                             "passez --db explicitement.")
    p = Path(chemin)
    if not p.exists():
        raise SystemExit(f"Base introuvable : {p}")
    conn = sqlite3.connect(str(p))
    conn.row_factory = sqlite3.Row
    return conn


def colonnes(conn: sqlite3.Connection, table: str) -> set[str]:
    return {r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()}


def charger_annuaire(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute("SELECT * FROM fournisseurs_fsc").fetchall()


def index_rapprochement(fiches: list[sqlite3.Row], cols: set[str]) -> dict:
    """Trois index, du plus fiable au plus large. Le SIRET tranche avant le nom :
    deux raisons sociales peuvent différer pour un même établissement."""
    par_siret, par_tva, par_nom = {}, {}, {}
    for f in fiches:
        d = dict(f)
        if "siret" in cols and d.get("siret"):
            par_siret.setdefault(re.sub(r"\D", "", str(d["siret"])), d)
        if "tva_intracom" in cols and d.get("tva_intracom"):
            par_tva.setdefault(re.sub(r"[^A-Z0-9]", "", str(d["tva_intracom"]).upper()), d)
        par_nom.setdefault(norm_nom(d.get("nom")), d)
    return {"siret": par_siret, "tva": par_tva, "nom": par_nom}


def rapprocher(cand: dict, idx: dict) -> tuple[dict | None, str]:
    # L'alias passe avant le SIRET : c'est une décision humaine explicite, et
    # les fiches concernées n'ont justement pas de SIRET à comparer.
    alias = idx.get("alias") or {}
    n_alias = norm_nom(cand.get("nom"))
    if n_alias in alias and alias[n_alias] is not IGNORER:
        return alias[n_alias], "alias"
    if cand.get("siret") and cand["siret"] in idx["siret"]:
        return idx["siret"][cand["siret"]], "siret"
    if cand.get("tva_intracom") and cand["tva_intracom"] in idx["tva"]:
        return idx["tva"][cand["tva_intracom"]], "tva"
    n = norm_nom(cand.get("nom"))
    if n and n in idx["nom"]:
        return idx["nom"][n], "nom"
    return None, ""


def vide(v) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def planifier(cand: dict, fiche: dict | None, cols: set[str], ecraser: bool) -> dict:
    """Calcule ce qui serait écrit. Renvoie {champ: (avant, après)}."""
    diff = {}
    for champ in CHAMPS_SIMPLES:
        if champ not in cols:
            continue
        neuf = cand.get(champ)
        if neuf is None:
            continue
        actuel = fiche.get(champ) if fiche else None
        # « pays » et « price_currency » ont un défaut en base ('FR', 'EUR') :
        # une valeur par défaut n'est pas une valeur saisie, elle se laisse
        # compléter. Sans cette exception, aucune fiche existante ne
        # recevrait jamais son pays réel.
        defaut = (champ == "pays" and actuel == "FR") or \
                 (champ == "price_currency" and actuel == "EUR")
        if not vide(actuel) and not defaut and not ecraser:
            continue
        if str(actuel or "") == str(neuf):
            continue
        diff[champ] = (actuel, neuf)

    # Catégories : on ajoute, on ne retire jamais. Le fichier n'en renseigne
    # presque aucune ; les remplacer effacerait le travail fait dans
    # Paramètres.
    if "categories" in cols and cand.get("categories"):
        actuelles = []
        if fiche and fiche.get("categories"):
            try:
                p = json.loads(fiche["categories"])
                actuelles = p if isinstance(p, list) else []
            except (ValueError, TypeError):
                actuelles = []
        fusion = list(actuelles)
        for c in cand["categories"]:
            if c not in fusion:
                fusion.append(c)
        if fusion != actuelles:
            diff["categories"] = (json.dumps(actuelles, ensure_ascii=False),
                                  json.dumps(fusion, ensure_ascii=False))

    # Note ERP : ajoutée une seule fois, jamais dupliquée à la relance.
    if "notes" in cols and cand.get("_note_erp"):
        actuel = (fiche or {}).get("notes") or ""
        if cand["_note_erp"] not in actuel:
            nouveau = (actuel.rstrip() + "\n" + cand["_note_erp"]).strip() \
                if actuel.strip() else cand["_note_erp"]
            diff["notes"] = (actuel or None, nouveau)

    return diff


# ═══════════════════════════════════════════════════════════════════
#  Rapports
# ═══════════════════════════════════════════════════════════════════

def titre(s: str) -> None:
    print("\n" + s)
    print("─" * min(len(s), 76))


def inventaire(lignes, entetes, candidats, apparies, nouveaux,
                inconnus, cols, quasi) -> None:
    titre(f"Fichier : {len(lignes)} ligne(s)")
    manquants = [c for c in CHAMPS_SIMPLES if c not in cols]
    if manquants:
        print("  ! Colonnes absentes de la base (migration non jouée ?) : "
              + ", ".join(manquants))

    titre("Colonnes du fichier")
    vus = {cle_entete(h) for h in entetes}
    utilisees = set(COL.values())
    for h in entetes:
        k = cle_entete(h)
        if not k:
            continue
        remplies = sum(1 for l in lignes if l.get(k) not in (None, ""))
        if k in utilisees:
            cible = [nom for nom, ck in COL.items() if ck == k][0]
            etat = f"→ {cible}"
        elif k in IGNOREES:
            etat = f"· ignorée ({IGNOREES[k]})"
        else:
            etat = "?? NON TRAITÉE — à examiner"
        print(f"  {h[:26]:<28} {remplies:>4}/{len(lignes)}  {etat}")

    titre(f"Rapprochement : {len(apparies)} fiche(s) retrouvée(s), "
          f"{len(nouveaux)} à créer")
    par_motif = Counter(m for _, _, m in apparies)
    for motif, n in par_motif.most_common():
        print(f"  {n:>4} par {motif}")

    if nouveaux:
        titre(f"À créer ({len(nouveaux)})")
        for c in nouveaux[:200]:
            lieu = " · ".join(x for x in [c.get("code_postal"), c.get("ville"),
                                          c.get("pays")] if x)
            print(f"  + {(c.get('nom') or '?')[:38]:<40} {lieu}")

    if quasi:
        titre(f"QUASI-DOUBLONS à vérifier ({len(quasi)})")
        print("  Ces fiches SERONT créées : la ressemblance ne suffit pas à décider.")
        print("  Après import : Paramètres → Fournisseurs → Doublons pour fusionner.")
        for c, f, r in quasi:
            print(f"  ~{r}  créer « {c['nom']} »   ≈   #{f['id']} « {f['nom']} »")

    if inconnus:
        titre("Valeurs non reconnues (référentiel à compléter ?)")
        for k, n in inconnus.most_common():
            print(f"  {n:>4} × {k}")


def rapport_ecritures(plans, nouveaux_plans, ecraser: bool) -> None:
    titre("Enrichissement des fiches existantes")
    if not plans:
        print("  Rien à compléter — les fiches retrouvées sont déjà à jour.")
    else:
        par_champ = Counter()
        for _, diff in plans:
            for champ in diff:
                par_champ[champ] += 1
        for champ, n in par_champ.most_common():
            print(f"  {n:>4} × {champ}")
        print(f"\n  {len(plans)} fiche(s) modifiée(s)"
              + ("  [--ecraser ACTIF : les valeurs existantes sont remplacées]"
                 if ecraser else "  (valeurs existantes préservées)"))
        titre("Détail (30 premières)")
        for fiche, diff in plans[:30]:
            print(f"  #{fiche['id']} {fiche['nom']}")
            for champ, (av, ap) in diff.items():
                a = "∅" if vide(av) else str(av)[:34]
                b = str(ap)[:34]
                print(f"       {champ:<24} {a:<36} → {b}")

    titre(f"Créations ({len(nouveaux_plans)})")
    print("  has_fsc = 0 et aucune catégorie : l'export n'en dit rien, ces deux")
    print("  informations restent à saisir dans Paramètres.")


# ═══════════════════════════════════════════════════════════════════
#  Écriture
# ═══════════════════════════════════════════════════════════════════

def appliquer(conn, plans, nouveaux_plans, cols, auteur: str) -> tuple[int, int]:
    now = datetime.now().isoformat(timespec="seconds")
    modifiees = 0
    for fiche, diff in plans:
        sets, vals = [], []
        for champ, (_, ap) in diff.items():
            sets.append(f"{champ}=?")
            vals.append(ap)
        if "updated_at" in cols:
            sets.append("updated_at=?")
            vals.append(now)
        vals.append(fiche["id"])
        conn.execute(f"UPDATE fournisseurs_fsc SET {', '.join(sets)} WHERE id=?", vals)
        modifiees += 1

    creees = 0
    for cand, diff in nouveaux_plans:
        champs = ["nom"]
        vals = [cand["nom"]]
        for champ, (_, ap) in diff.items():
            champs.append(champ)
            vals.append(ap)
        # Explicites, pour ne pas dépendre du DEFAULT de la colonne : une
        # fiche créée par import n'est pas certifiée FSC tant que personne ne
        # l'a dit.
        for champ, v in (("has_fsc", 0), ("actif", 1), ("updated_at", now)):
            if champ in cols and champ not in champs:
                champs.append(champ)
                vals.append(v)
        try:
            conn.execute(
                f"INSERT INTO fournisseurs_fsc ({', '.join(champs)}) "
                f"VALUES ({', '.join('?' * len(champs))})", vals
            )
            creees += 1
        except sqlite3.IntegrityError as e:
            # Le nom est UNIQUE : une collision ici veut dire que le
            # rapprochement a laissé passer un cas. On le signale au lieu de
            # faire échouer tout l'import.
            print(f"  ! « {cand['nom']} » non créé : {e}")

    conn.commit()
    return modifiees, creees


# ═══════════════════════════════════════════════════════════════════

def main() -> int:
    ap = argparse.ArgumentParser(
        description="Import / enrichissement de l'annuaire fournisseurs depuis un export ERP.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--fichier", required=True, help="Export Excel (.xlsx)")
    ap.add_argument("--db", default=None, help="Base SQLite (défaut : DB_PATH de config.py)")
    ap.add_argument("--inventaire", action="store_true",
                    help="N'écrit rien : montre le rapprochement et les colonnes")
    ap.add_argument("--simulation", action="store_true",
                    help="Rejoue l'écriture sans committer (comportement par défaut)")
    ap.add_argument("--appliquer", action="store_true", help="Écrit en base")
    ap.add_argument("--alias", default=None,
                    help="Fichier de correspondances « nom export ; fiche » (voir --generer-alias)")
    ap.add_argument("--generer-alias", default=None, metavar="CHEMIN",
                    help="Écrit un fichier d'alias pré-rempli depuis les quasi-doublons, puis s'arrête")
    ap.add_argument("--ecraser", action="store_true",
                    help="Remplace aussi les valeurs déjà renseignées dans MySifa")
    ap.add_argument("--auteur", default=os.getenv("USER") or "import",
                    help="Nom porté par la trace d'import")
    args = ap.parse_args()

    chemin = Path(args.fichier).expanduser()
    if not chemin.exists():
        raise SystemExit(f"Fichier introuvable : {chemin}")

    print(f"Lecture de {chemin.name}…")
    lignes, entetes = lire_lignes(chemin)

    inconnus: Counter = Counter()
    candidats = [extraire(l, inconnus) for l in lignes]
    sans_nom = [c for c in candidats if not c.get("nom")]
    candidats = [c for c in candidats if c.get("nom")]
    if sans_nom:
        print(f"  ! {len(sans_nom)} ligne(s) sans raison sociale ni code — ignorée(s).")

    # Doublons internes au fichier : deux lignes pour un même fournisseur.
    # La première gagne, la seconde est signalée — fusionner deux lignes
    # d'export au petit bonheur produirait une fiche que personne n'a voulue.
    vus, propres, doublons_fichier = {}, [], []
    for c in candidats:
        n = norm_nom(c["nom"])
        if n in vus:
            doublons_fichier.append((c["nom"], vus[n]))
            continue
        vus[n] = c["nom"]
        propres.append(c)
    candidats = propres

    conn = ouvrir_base(args.db)
    cols = colonnes(conn, "fournisseurs_fsc")
    if not cols:
        raise SystemExit("Table fournisseurs_fsc absente de cette base.")
    fiches = charger_annuaire(conn)
    idx = index_rapprochement(fiches, cols)
    if args.alias:
        idx["alias"] = resoudre_alias(lire_alias(Path(args.alias).expanduser()), fiches)
        print(f"  Alias : {len(idx['alias'])} correspondance(s) chargée(s).")

    ignorees = []
    if idx.get("alias"):
        gardees = []
        for c in candidats:
            if idx["alias"].get(norm_nom(c["nom"])) is IGNORER:
                ignorees.append(c)
            else:
                gardees.append(c)
        candidats = gardees
        if ignorees:
            titre(f"Lignes ignorées sur demande du fichier d'alias ({len(ignorees)})")
            for c in ignorees:
                lieu = " · ".join(x for x in (c.get("ville"), c.get("pays")) if x)
                print(f"  - « {c['nom']} »" + (f"   {lieu}" if lieu else ""))

    apparies, nouveaux = [], []
    for c in candidats:
        fiche, motif = rapprocher(c, idx)
        if fiche:
            apparies.append((c, fiche, motif))
        else:
            nouveaux.append(c)

    # ── Collisions : deux lignes de l'export visant LA MÊME fiche ────────────
    #
    # Le cas réel qui a motivé cette garde : « UPM RAFLATAC » (Pompey, TVA
    # FR77…) était aliasé vers la fiche « UPM », et la ligne « UPM » (Tampere,
    # TVA FI10…) visait la même fiche par son nom. Les deux enrichissements
    # s'appliquaient l'un après l'autre sur la même fiche, et le second écrasait
    # les champs du premier : la fiche finissait avec l'adresse finlandaise et
    # la note « Code ERP : RAFLATAC ». Deux sociétés en une, sans un mot dans le
    # rapport — et l'entité française sans fiche du tout.
    #
    # Règle retenue : la clé la plus forte garde la fiche, les autres lignes
    # partent en création. Deux numéros de TVA différents, ce sont deux
    # personnes morales ; les empiler sur une fiche est toujours faux.
    FORCE = {"alias": 0, "siret": 1, "tva": 2, "nom": 3}
    par_cible = defaultdict(list)
    for item in apparies:
        par_cible[item[1]["id"]].append(item)

    collisions = []
    apparies = []
    for fid, lot in par_cible.items():
        if len(lot) == 1:
            apparies.append(lot[0])
            continue
        lot.sort(key=lambda x: FORCE.get(x[2], 9))
        gagnant = lot[0]
        apparies.append(gagnant)
        for perdant in lot[1:]:
            collisions.append((perdant[0], gagnant[0], gagnant[1], perdant[2]))
            nouveaux.append(perdant[0])

    if collisions:
        titre(f"COLLISIONS ÉCARTÉES ({len(collisions)})")
        print("  Deux lignes de l'export visaient la même fiche. La ligne rapprochée")
        print("  par la clé la plus forte l'enrichit ; l'autre part en création.")
        for perdue, gardee, fiche, motif in collisions:
            print(f"  fiche #{fiche['id']} « {fiche['nom']} » enrichie par « {gardee['nom']} »")
            print(f"      « {perdue['nom']} » (rapprochée par {motif}) sera créée à part.")
        # Le nom est UNIQUE en base : si la ligne écartée porte exactement le nom
        # d'une fiche existante, sa création échouera. Le dire ici, pas au
        # moment de l'INSERT, laisse le temps de trancher avant d'écrire.
        noms_pris = {norm_nom(dict(f)["nom"]) for f in fiches}
        bloquees = [c for c, _, _, _ in collisions if norm_nom(c["nom"]) in noms_pris]
        if bloquees:
            print()
            print("  ATTENTION — ces lignes portent le nom d'une fiche existante et ne")
            print("  pourront pas être créées telles quelles (le nom est unique) :")
            for c in bloquees:
                print(f"      « {c['nom']} »")
            print("  Tranchez d'abord : renommez la fiche existante pour distinguer les")
            print("  deux entités, ou ajoutez un alias pour dire laquelle est laquelle.")

    print(f"  Annuaire MySifa : {len(fiches)} fiche(s).")
    if doublons_fichier:
        titre(f"Doublons dans le fichier ({len(doublons_fichier)}) — 2e ligne ignorée")
        for a, b in doublons_fichier[:20]:
            print(f"  « {a} » ≈ « {b} »")

    quasi = proches(nouveaux, fiches)

    if args.generer_alias:
        if not quasi:
            print("\nAucun quasi-doublon : pas d'alias à proposer.")
            return 0
        ecrire_alias_propose(Path(args.generer_alias).expanduser(), quasi)
        return 0

    inventaire(lignes, entetes, candidats, apparies, nouveaux, inconnus, cols, quasi)

    if args.inventaire:
        print("\nMode inventaire : rien n'a été écrit.")
        return 0

    plans = []
    for c, fiche, _ in apparies:
        diff = planifier(c, fiche, cols, args.ecraser)
        if diff:
            plans.append((fiche, diff))
    nouveaux_plans = [(c, planifier(c, None, cols, args.ecraser)) for c in nouveaux]

    rapport_ecritures(plans, nouveaux_plans, args.ecraser)

    if not args.appliquer:
        print("\nSimulation : rien n'a été écrit. Relancez avec --appliquer pour valider.")
        return 0

    modifiees, creees = appliquer(conn, plans, nouveaux_plans, cols, args.auteur)
    titre("Appliqué")
    print(f"  {modifiees} fiche(s) enrichie(s), {creees} fiche(s) créée(s).")
    print("  À faire dans Paramètres → Fournisseurs : cocher la certification FSC")
    print("  des fiches concernées, et renseigner les catégories (elles pilotent")
    print("  les favoris de la recherche fournisseur).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
