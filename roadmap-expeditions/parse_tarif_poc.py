"""
Preuve de concept — normalisation d'une grille tarifaire transporteur
vers le schema expe_tarifs (cf. ROADMAP-EXPEDITIONS.md sec.5.1).

Cible : grille "SIFA 010126 - P U" (compte 100346) — 2 feuilles : POIDS + PALETTE.
Sortie : poc_tarifs_100346.csv (1 ligne = 1 cellule de grille normalisee).

C'est l'embryon du futur skill `import-tarif-transporteur` / endpoint de parsing.
"""
import openpyxl, csv, re, os, math
from collections import Counter

SRC = os.environ.get("SRC", "/sessions/eloquent-bold-euler/mnt/uploads/SIFA 010126 - P U-1.xlsx")
OUTDIR = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(OUTDIR, "poc_tarifs_100346.csv")
TRANSPORTEUR = "Compte 100346"

def floor_from(v):
    try: return int(math.floor(float(v)))
    except Exception: return 0

def dept_code(label):
    m = re.search(r"\((\w{1,3})\)", str(label or ""))
    return m.group(1).zfill(2) if m else None

def unite_norm(u):
    s = str(u or "").strip().lower()
    if "100" in s: return "au_100kg"
    if "forfait" in s: return "forfait"
    if "kg" in s: return "au_kg"
    return s or "forfait"

rows = []
wb = openpyxl.load_workbook(SRC, data_only=True)
for ws in wb.worksheets:
    a8 = str(ws["A8"].value or "").upper()
    if "POIDS" in a8:   base = "poids"
    elif "PALETTE" in a8: base = "palette"
    else: continue
    # En-tetes : ligne 10 = borne basse (DE), 11 = borne haute (A), 12 = unite ; data des la ligne 13
    cols = []
    for c in range(3, ws.max_column + 1):
        to_v = ws.cell(row=11, column=c).value
        if to_v is None: continue
        try: tmax = float(to_v)
        except Exception: continue
        tmin = floor_from(ws.cell(row=10, column=c).value)
        unite = unite_norm(ws.cell(row=12, column=c).value)
        cols.append((c, tmin, tmax, unite))
    for r in range(13, ws.max_row + 1):
        lbl = ws.cell(row=r, column=1).value
        code = dept_code(lbl)
        if not code: continue
        for (c, tmin, tmax, unite) in cols:
            price = ws.cell(row=r, column=c).value
            if price is None or str(price).strip() == "": continue
            try: price = float(price)
            except Exception: continue
            rows.append({
                "transporteur": TRANSPORTEUR,
                "type_envoi": "messagerie",
                "base_calcul": base,
                "zone_type": "departement",
                "zone_valeur": code,
                "dept_label": re.sub(r"^\(\w+\)\s*", "", str(lbl)).strip(),
                "tranche_min": tmin,
                "tranche_max": (int(tmax) if base == "palette" else tmax),
                "prix": round(price, 2),
                "unite": unite,
                "mini_perception": "",
                "source_filename": os.path.basename(SRC),
            })

fields = ["transporteur","type_envoi","base_calcul","zone_type","zone_valeur","dept_label",
          "tranche_min","tranche_max","prix","unite","mini_perception","source_filename"]
with open(OUT, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=fields); w.writeheader(); w.writerows(rows)

print("CSV ecrit :", OUT)
print("Lignes generees :", len(rows))
print("Par base_calcul :", dict(Counter(r["base_calcul"] for r in rows)))
print("Par unite       :", dict(Counter(r["unite"] for r in rows)))
print("Departements distincts :", len(set(r["zone_valeur"] for r in rows)))
print("\nApercu (10 premieres lignes) :")
for r in rows[:10]:
    print(f"  {r['base_calcul']:8} dpt{r['zone_valeur']} {r['dept_label'][:18]:18} [{r['tranche_min']}-{r['tranche_max']}] {r['prix']:>7} {r['unite']}")
