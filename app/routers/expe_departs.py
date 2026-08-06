"""
MyExpé — suivi des départs (exportations).
Accès : utilisateurs avec droit application « expe ».
"""
import csv
import io
import json
import os
import re
import shutil
import sqlite3
import unicodedata
import uuid
from io import BytesIO
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Body, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse

from app.services.audit_service import log_action
from app.services.email_service import (
    email_expe_devis_confirmation,
    email_expe_rfq_transport,
    send_email,
)
from config import public_base_url
from app.services import expe_evenements as expe_ev
from app.services.expe_transporteurs_seed import seed_expe_transporteurs_if_empty
from database import get_db
from services.auth_service import get_current_user, user_can_write_expe, user_has_app_access

router = APIRouter()

_PARIS = ZoneInfo("Europe/Paris")

TARIF_UPLOAD_DIR = "data/uploads/transporteurs"
os.makedirs(TARIF_UPLOAD_DIR, exist_ok=True)

_ALLOWED_TARIF_EXT = {".pdf", ".xlsx", ".xls", ".jpg", ".jpeg", ".png", ".webp", ".gif"}


def _require_expe(request: Request) -> dict:
    user = get_current_user(request)
    if not user_has_app_access(user, "expe"):
        raise HTTPException(status_code=403, detail="Accès MyExpé requis")
    return user


def _require_expe_write(request: Request) -> dict:
    user = _require_expe(request)
    if not user_can_write_expe(user):
        raise HTTPException(status_code=403, detail="Accès MyExpé en lecture seule")
    return user


def _today_paris_iso() -> str:
    return datetime.now(_PARIS).date().isoformat()


def _norm_search(s: str) -> str:
    t = unicodedata.normalize("NFD", (s or "").lower())
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _row_blob(d: dict) -> str:
    parts = [
        d.get("date_enlevement"),
        d.get("affreteurs"),
        d.get("transporteur"),
        d.get("client"),
        d.get("code_postal_destination"),
        d.get("ref_sifa"),
        d.get("arc"),
        d.get("no_cde_transport"),
        d.get("no_bl"),
        d.get("type_palette_label"),
        d.get("type_palette_reference"),
        d.get("type_colis"),
        d.get("nb_palette"),
        d.get("poids_total_kg"),
        d.get("date_livraison"),
        d.get("created_by_email"),
        d.get("validated_by_email"),
        d.get("validated_at"),
    ]
    return _norm_search(" ".join(str(p) for p in parts if p is not None and str(p) != ""))


_HIST_SEARCH_COLS = (
    "d.date_enlevement",
    "d.affreteurs",
    "d.transporteur",
    "d.client",
    "d.code_postal_destination",
    "d.ref_sifa",
    "d.arc",
    "d.no_cde_transport",
    "d.no_bl",
    "mp.reference",
    "mp.designation",
    "d.type_colis",
    "d.nb_palette",
    "d.poids_total_kg",
    "d.date_livraison",
    "d.created_by_email",
    "d.validated_by_email",
    "d.validated_at",
)


def _historique_search_clause(q: str) -> tuple[str, list[Any]]:
    """Clause SQL AND … pour la recherche multi-mots (tous les tokens requis)."""
    qt = _norm_search(q)
    if not qt:
        return "", []
    tokens = [t for t in qt.split(" ") if t]
    if not tokens:
        return "", []
    parts: list[str] = []
    params: list[Any] = []
    ncols = len(_HIST_SEARCH_COLS)
    for tok in tokens:
        likes = " OR ".join(
            f"LOWER(COALESCE(CAST({c} AS TEXT), '')) LIKE ?" for c in _HIST_SEARCH_COLS
        )
        parts.append(f"({likes})")
        params.extend([f"%{tok}%"] * ncols)
    return " AND ".join(parts), params


_DEPARTS_SELECT = """
    SELECT d.*,
           mp.reference AS type_palette_reference,
           mp.designation AS type_palette_designation,
           COALESCE(mp.is_europe, 0) AS palette_ref_is_europe,
           t.couleur AS transporteur_couleur,
           pe.reference AS planning_dossier_ref,
           pe.numero_of AS planning_numero_of
    FROM expe_departs d
    LEFT JOIN matieres_premieres mp ON mp.id = d.type_palette_matiere_id
    LEFT JOIN expe_transporteurs t ON t.id = d.transporteur_id
    LEFT JOIN planning_entries pe ON pe.id = d.planning_entry_id
"""


def _depart_dict(row) -> dict:
    d = dict(row)
    if (d.get("type_colis") or "").strip().lower() == "vrac":
        d["type_palette_label"] = "Vrac"
    else:
        ref = (d.get("type_palette_reference") or "").strip()
        des = (d.get("type_palette_designation") or "").strip()
        d["type_palette_label"] = (f"{ref} — {des}" if des else ref) if ref else None
    return d


def _validate_type_palette_matiere_id(conn, matiere_id: Any) -> Optional[int]:
    if matiere_id is None or matiere_id == "":
        return None
    try:
        mid = int(matiere_id)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Type de palette invalide.")
    row = conn.execute(
        """SELECT id FROM matieres_premieres
           WHERE id=? AND actif=1 AND categorie='palette'""",
        (mid,),
    ).fetchone()
    if not row:
        raise HTTPException(
            status_code=400,
            detail="Type de palette introuvable ou inactif (réf. MyStock).",
        )
    return mid


def _date_prefix(raw: str) -> str:
    """Extrait YYYY-MM-DD depuis une saisie date ou datetime."""
    s = (raw or "").strip()
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    return s


def _now_paris_iso() -> str:
    return datetime.now(_PARIS).replace(tzinfo=None).isoformat(timespec="seconds")


def _f(body: dict, key: str) -> Any:
    v = body.get(key)
    if v is None or v == "":
        return None
    return v


def _float_opt(body: dict, key: str) -> Any:
    v = body.get(key)
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", ".").replace("\u202f", "").replace(" ", ""))
    except ValueError:
        return None


def _int_flag(body: dict, key: str, default: Optional[int] = None) -> Optional[int]:
    if key not in body:
        return default
    v = body.get(key)
    if v is None or v == "":
        return 0
    if v in (1, True, "1", "true", "True"):
        return 1
    return 0


def _int_opt(body: dict, key: str) -> Any:
    v = body.get(key)
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def _safe_tarif_filename(name: str) -> str:
    base = Path(name or "tarif").name
    base = re.sub(r"[^\w.\- ]", "_", base, flags=re.UNICODE).strip("._ ") or "tarif"
    return base[:120]


_EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")


def _normalize_emails(value: Any) -> list[str]:
    """Accepte list[str], str JSON, str séparée par , / ; / saut de ligne. Renvoie une liste dédupliquée."""
    if value is None:
        return []
    items: list[str] = []
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return []
        # Tente JSON d'abord
        if s.startswith("["):
            try:
                arr = json.loads(s)
                if isinstance(arr, list):
                    items = [str(x).strip() for x in arr if x]
            except (json.JSONDecodeError, ValueError):
                items = []
        if not items:
            for chunk in re.split(r"[,;\n\r\t]+", s):
                v = chunk.strip()
                if v:
                    items.append(v)
    elif isinstance(value, (list, tuple)):
        for v in value:
            if v is None:
                continue
            items.append(str(v).strip())
    # filtre + dédup en conservant l'ordre
    seen: set[str] = set()
    out: list[str] = []
    for v in items:
        if not v:
            continue
        if not _EMAIL_RE.match(v):
            continue
        low = v.lower()
        if low in seen:
            continue
        seen.add(low)
        out.append(v)
    return out


def _normalize_tels(value: Any) -> list[dict]:
    """Accepte list[{numero, service}], list[str], str JSON, str libre.
    Renvoie une liste [{numero:str, service:str}, ...] dédupliquée sur (numero, service)."""
    if value is None:
        return []
    raw_items: list[Any] = []
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return []
        if s.startswith("["):
            try:
                arr = json.loads(s)
                if isinstance(arr, list):
                    raw_items = arr
            except (json.JSONDecodeError, ValueError):
                raw_items = []
        if not raw_items:
            for chunk in re.split(r"[,;\n\r\t]+", s):
                v = chunk.strip()
                if v:
                    raw_items.append(v)
    elif isinstance(value, (list, tuple)):
        raw_items = list(value)
    out: list[dict] = []
    seen: set[tuple] = set()
    for item in raw_items:
        numero = ""
        service = ""
        if isinstance(item, dict):
            numero = str(item.get("numero") or item.get("tel") or "").strip()
            service = str(item.get("service") or item.get("label") or "").strip()
        else:
            numero = str(item or "").strip()
        if not numero:
            continue
        # Limite raisonnable
        numero = numero[:40]
        service = service[:80]
        key = (numero, service.lower())
        if key in seen:
            continue
        seen.add(key)
        out.append({"numero": numero, "service": service})
    return out


def _normalize_portail(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    low = s.lower()
    # Tolère qu'on saute le scheme
    if not (low.startswith("http://") or low.startswith("https://")):
        s = "https://" + s
    return s[:512]


def _serialize_transporteur_row(row: Any) -> dict:
    d = dict(row)
    raw_emails = d.get("contact_emails")
    if isinstance(raw_emails, str) and raw_emails.strip():
        try:
            arr = json.loads(raw_emails)
            d["contact_emails"] = arr if isinstance(arr, list) else []
        except (json.JSONDecodeError, ValueError):
            d["contact_emails"] = _normalize_emails(raw_emails)
    else:
        d["contact_emails"] = []
    # Téléphones : liste normalisée [{numero, service}]
    raw_tels = d.get("contact_tels")
    if isinstance(raw_tels, str) and raw_tels.strip():
        d["contact_tels"] = _normalize_tels(raw_tels)
    elif isinstance(raw_tels, list):
        d["contact_tels"] = _normalize_tels(raw_tels)
    else:
        # Fallback : reconstruit depuis contact_tel legacy
        legacy = d.get("contact_tel")
        d["contact_tels"] = _normalize_tels(legacy) if legacy else []
    return d


def _tarif_abs_root() -> str:
    from config import BASE_DIR

    return os.path.abspath(os.path.join(BASE_DIR, TARIF_UPLOAD_DIR))


def _resolve_tarif_path(tarif_url: Optional[str]) -> Optional[str]:
    if not tarif_url:
        return None
    from config import BASE_DIR

    p = tarif_url.strip()
    if not p:
        return None
    if not os.path.isabs(p):
        p = os.path.join(BASE_DIR, p)
    abs_p = os.path.abspath(p)
    root = _tarif_abs_root()
    if abs_p != root and not abs_p.startswith(root + os.sep):
        return None
    return abs_p if os.path.isfile(abs_p) else None


def _unlink_tarif(tarif_url: Optional[str]) -> None:
    p = _resolve_tarif_path(tarif_url)
    if p:
        try:
            os.unlink(p)
        except OSError:
            pass


@router.get("/matieres-palettes")
def list_matieres_palettes_expe(request: Request):
    """Références palettes actives (catégorie palette, MyStock matières premières)."""
    _require_expe(request)
    with get_db() as conn:
        rows = conn.execute(
            """SELECT id, reference, designation, palettes_par_pile,
                      COALESCE(is_europe, 0) AS is_europe
               FROM matieres_premieres
               WHERE actif=1 AND categorie='palette'
               ORDER BY reference COLLATE NOCASE"""
        ).fetchall()
    return [dict(r) for r in rows]


def _resolve_palette_europe_flag(conn, type_palette_matiere_id, body) -> int:
    """Détermine si un départ correspond à une expédition palette Europe.

    Règle :
      1. Si l'utilisateur a forcé `palette_europe` à 0 ou 1 dans le body, on respecte.
      2. Sinon auto-détection via le flag `is_europe` sur la référence palette MyStock.
    """
    raw = body.get("palette_europe")
    if raw in (1, True, "1", "true", "True"):
        return 1
    if raw in (0, False, "0", "false", "False"):
        return 0
    if type_palette_matiere_id:
        row = conn.execute(
            "SELECT COALESCE(is_europe, 0) AS is_europe FROM matieres_premieres WHERE id=?",
            (type_palette_matiere_id,),
        ).fetchone()
        if row and int(row["is_europe"]) == 1:
            return 1
    return 0


_PAL_EUROPE_STATUTS = ("en_attente", "retournee", "perdue")


def _validate_palette_europe_statut(value) -> str:
    s = (str(value or "").strip().lower() or "en_attente")
    if s not in _PAL_EUROPE_STATUTS:
        raise HTTPException(
            status_code=400,
            detail=f"Statut palette Europe invalide ({', '.join(_PAL_EUROPE_STATUTS)}).",
        )
    return s


def _validate_planning_entry_id(conn, value) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        pid = int(value)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="ID dossier planning invalide.")
    row = conn.execute(
        "SELECT id FROM planning_entries WHERE id=?", (pid,)
    ).fetchone()
    if not row:
        raise HTTPException(status_code=400, detail="Dossier planning introuvable.")
    return pid


def _cle_no_bl(valeur: Any) -> str:
    """Forme normalisée d'un n° de BL, pour la comparaison seulement.

    « BL-1001 », « bl 1001 » et « BL1001 » désignent le même document papier.
    On ne stocke PAS cette forme : ce que l'opérateur a tapé reste ce qui est
    enregistré, c'est ce qui figure sur le document. Elle ne sert qu'à
    reconnaître un doublon.
    """
    return (
        str(valeur or "").strip().upper().replace(" ", "").replace("-", "").replace(".", "")
    )


def _check_no_bl_unique(conn, no_bl: Any, exclure_id: Optional[int] = None) -> None:
    """Refuse un n° de BL déjà porté par un autre départ.

    Premier maillon d'un audit FSC : l'auditeur arrive avec un bon de livraison
    et demande le départ correspondant. Si deux lignes répondent, la chaîne
    qu'on lui présente est peut-être la mauvaise et personne ne peut trancher.

    Le refus est un 409 structuré, pas un 400 : ce n'est pas une requête
    malformée, c'est un conflit d'état que l'opérateur peut lever en
    connaissance de cause (`no_bl_doublon_confirme: true`) — un même BL réparti
    sur deux enlèvements existe. Ce qui ne doit pas exister, c'est le doublon
    créé sans que personne ne l'ait vu.
    """
    cle = _cle_no_bl(no_bl)
    if not cle:
        return
    rows = conn.execute(
        """SELECT id, no_bl, client, date_enlevement
             FROM expe_departs
            WHERE UPPER(REPLACE(REPLACE(REPLACE(TRIM(COALESCE(no_bl,'')),' ',''),'-',''),'.','')) = ?
            LIMIT 5""",
        (cle,),
    ).fetchall()
    autres = [dict(r) for r in rows if exclure_id is None or int(r["id"]) != int(exclure_id)]
    if not autres:
        return
    raise HTTPException(
        status_code=409,
        detail={
            "code": "no_bl_doublon",
            "no_bl": str(no_bl or "").strip(),
            "departs": autres,
            "message": (
                f"Le n° de BL « {str(no_bl or '').strip()} » est déjà porté par "
                f"{len(autres)} autre(s) départ(s). Deux départs sous le même numéro "
                f"rendent la traçabilité ambiguë : un auditeur ne saurait pas lequel "
                f"correspond au document qu'il présente."
            ),
        },
    )


def _sync_no_dossier(conn, depart_id: int) -> None:
    """Recopie sur le départ la référence du dossier qu'il pointe.

    `planning_entry_id` est la source de vérité : c'est le dossier que
    l'utilisateur désigne dans le formulaire d'expédition. `no_dossier` en est
    une copie textuelle, tenue à jour ici parce que toute la chaîne FSC —
    traceur, mention à porter sur le document de vente, registre — interroge
    des références de dossier et non des identifiants de ligne de planning.

    Sans cette recopie, le champ reste vide sur tout départ créé après la
    migration 222 (son backfill depuis `ref_sifa` ne s'est joué qu'une fois),
    et un auditeur qui part d'un bon de livraison ne peut pas remonter à la
    matière : la chaîne casse au premier maillon interne.

    Deux règles :
      - la saisie prime sur la déduction. Un `no_dossier_source` hérité du
        backfill ('reconstitue', déduit de `ref_sifa`) est écrasé dès qu'un
        vrai dossier est désigné ;
      - on n'efface QUE ce que ce mécanisme a écrit. Si le dossier est retiré
        du départ, un rattachement 'reconstitue' antérieur reste en place — il
        ne vient pas d'ici, ce n'est pas à nous de le supprimer.
    """
    row = conn.execute(
        """SELECT COALESCE(
                    NULLIF(TRIM(COALESCE(pe.reference, '')), ''),
                    NULLIF(TRIM(COALESCE(pe.numero_of, '')), '')) AS ref
             FROM expe_departs d
             LEFT JOIN planning_entries pe ON pe.id = d.planning_entry_id
            WHERE d.id = ?""",
        (depart_id,),
    ).fetchone()
    ref = ((row["ref"] if row else None) or "").strip()

    if ref:
        conn.execute(
            "UPDATE expe_departs SET no_dossier=?, no_dossier_source='saisi' WHERE id=?",
            (ref, depart_id),
        )
    else:
        conn.execute(
            "UPDATE expe_departs SET no_dossier=NULL, no_dossier_source=NULL "
            " WHERE id=? AND COALESCE(no_dossier_source,'')='saisi'",
            (depart_id,),
        )


@router.get("/dossiers-disponibles")
def list_dossiers_disponibles_expe(request: Request):
    """Picker dossier pour l'écran Ajouter départ (MyExpé).

    Renvoie :
      - les 4 derniers dossiers terminés (les plus récemment mis à jour),
      - tous les dossiers en cours,
      - le prochain dossier en attente (premier par position),
      - puis le reste des dossiers (attente et terminés) pour la recherche.

    Champ `displayed_section` ('en_cours' | 'prochain' | 'termine_recent' | 'autre')
    permet de styler la liste. Champ `departs_count` indique combien de départs
    sont déjà liés à ce dossier (anti-doublon visuel).
    """
    _require_expe(request)
    with get_db() as conn:
        rows = conn.execute(
            """SELECT pe.id, pe.reference, pe.client, pe.description,
                      pe.ref_produit, pe.numero_of, pe.date_livraison,
                      pe.format_l, pe.format_h, pe.statut, pe.position,
                      pe.duree_heures, pe.updated_at,
                      m.nom AS machine_nom, m.code AS machine_code,
                      (SELECT COUNT(*) FROM expe_departs ed
                         WHERE ed.planning_entry_id = pe.id) AS departs_count,
                      COALESCE(ftm.palette_type, fta.palette_type)
                        AS ft_palette_type,
                      COALESCE(ftm.palette_nb_cartons_sol, fta.palette_nb_cartons_sol)
                        AS ft_palette_nb_cartons_sol,
                      COALESCE(ftm.palette_nb_cartons_hauteur, fta.palette_nb_cartons_hauteur)
                        AS ft_palette_nb_cartons_hauteur,
                      COALESCE(ftm.nb_au_sol, fta.nb_au_sol) AS ft_nb_au_sol,
                      COALESCE(ftm.nb_etage, fta.nb_etage) AS ft_nb_etage
               FROM planning_entries pe
               JOIN machines m ON m.id = pe.machine_id
               -- Rapprochement fiche technique par clé produit normalisée
               -- (XXX/NNNN), avec repli sur la référence textuelle. MyExpé était
               -- le dernier module à joindre sur `reference` brute : il ratait
               -- toutes les fiches dont le libellé porte une variante machine ou
               -- laize ("1315/0004 - COHESIO 1 - L570").
               --
               -- Deux tables dérivées plutôt qu'une sous-requête corrélée :
               -- SQLite n'autorise pas un ON à référencer l'alias `m`, et le
               -- MIN(id) par clé garantit une seule fiche — donc jamais de
               -- ligne de dossier dupliquée.
               --   ftm = la fiche de la machine du dossier (prioritaire)
               --   fta = n'importe quelle fiche de ce produit (repli)
               LEFT JOIN (
                   SELECT MIN(id) AS id,
                          COALESCE(NULLIF(TRIM(ref_produit_norm), ''),
                                   LOWER(TRIM(COALESCE(reference, '')))) AS k,
                          LOWER(TRIM(COALESCE(machine, ''))) AS mk
                   FROM fiches_techniques
                   GROUP BY k, mk
               ) km ON km.k = COALESCE(NULLIF(TRIM(pe.ref_produit_norm), ''),
                                       LOWER(TRIM(COALESCE(pe.ref_produit, ''))))
                   AND km.mk = LOWER(TRIM(COALESCE(m.nom, '')))
                   AND km.mk != ''
                   AND COALESCE(pe.ref_produit, '') != ''
               LEFT JOIN fiches_techniques ftm ON ftm.id = km.id
               LEFT JOIN (
                   SELECT MIN(id) AS id,
                          COALESCE(NULLIF(TRIM(ref_produit_norm), ''),
                                   LOWER(TRIM(COALESCE(reference, '')))) AS k
                   FROM fiches_techniques
                   GROUP BY k
               ) ka ON ka.k = COALESCE(NULLIF(TRIM(pe.ref_produit_norm), ''),
                                       LOWER(TRIM(COALESCE(pe.ref_produit, ''))))
                   AND COALESCE(pe.ref_produit, '') != ''
               LEFT JOIN fiches_techniques fta ON fta.id = ka.id
               ORDER BY pe.position ASC, pe.id ASC"""
        ).fetchall()

    en_cours: list = []
    attente: list = []
    termine: list = []
    for r in rows:
        d = dict(r)
        st = (d.get("statut") or "").strip().lower()
        if st == "en_cours":
            en_cours.append(d)
        elif st == "attente":
            attente.append(d)
        elif st == "termine":
            termine.append(d)

    termine.sort(key=lambda d: (d.get("updated_at") or ""), reverse=True)
    termine_recents = termine[:4]
    termine_autres = termine[4:]
    prochain = attente[0] if attente else None
    attente_autres = attente[1:]

    out: list = []
    for d in termine_recents:
        d["displayed_section"] = "termine_recent"
        out.append(d)
    for d in en_cours:
        d["displayed_section"] = "en_cours"
        out.append(d)
    if prochain:
        prochain["displayed_section"] = "prochain"
        out.append(prochain)
    for d in attente_autres:
        d["displayed_section"] = "autre"
        out.append(d)
    for d in termine_autres:
        d["displayed_section"] = "autre"
        out.append(d)
    return {"dossiers": out}


@router.get("/departs/jour")
def list_departs_jour(
    request: Request,
    date: Optional[str] = Query(None, description="YYYY-MM-DD (défaut : jour Paris)"),
):
    _require_expe(request)
    with get_db() as conn:
        rows = conn.execute(
            f"""{_DEPARTS_SELECT}
               WHERE d.statut = 'en_attente'
               ORDER BY d.date_enlevement ASC,
                        CASE WHEN COALESCE(NULLIF(TRIM(d.transporteur), ''), '') = '' THEN 1 ELSE 0 END,
                        LOWER(COALESCE(d.transporteur, '')) ASC,
                        d.id ASC""",
        ).fetchall()
    return [_depart_dict(r) for r in rows]


@router.post("/departs")
def create_depart(request: Request, body: dict = Body(...)):
    user = _require_expe_write(request)
    date_enl = _date_prefix(str(body.get("date_enlevement") or "").strip())
    if not date_enl or not re.match(r"^\d{4}-\d{2}-\d{2}$", date_enl):
        raise HTTPException(status_code=400, detail="Date d'enlèvement obligatoire (YYYY-MM-DD)")
    now = datetime.now(_PARIS).replace(tzinfo=None).isoformat(timespec="seconds")
    email = (user.get("email") or user.get("identifiant") or "").strip() or None

    def _f(key: str) -> Any:
        v = body.get(key)
        if v is None or v == "":
            return None
        return v

    def _float_opt(key: str) -> Any:
        v = body.get(key)
        if v is None or v == "":
            return None
        try:
            return float(str(v).replace(",", ".").replace("\u202f", "").replace(" ", ""))
        except ValueError:
            return None

    with get_db() as conn:
        type_palette_id = _validate_type_palette_matiere_id(
            conn, body.get("type_palette_matiere_id")
        )
        type_colis_val = (str(body.get("type_colis") or "").strip().lower() or None)
        if type_colis_val == "vrac":
            type_palette_id = None  # pas de matière première pour vrac
        planning_entry_id = _validate_planning_entry_id(
            conn, body.get("planning_entry_id")
        )
        palette_europe = _resolve_palette_europe_flag(conn, type_palette_id, body)
        palette_europe_statut = _validate_palette_europe_statut(
            body.get("palette_europe_statut") if palette_europe else "en_attente"
        )
        palette_europe_date_retour = (
            _date_prefix(str(body.get("palette_europe_date_retour") or "").strip())
            if palette_europe else None
        ) or None
        palette_europe_note = _f("palette_europe_note") if palette_europe else None
        if not body.get("no_bl_doublon_confirme"):
            _check_no_bl_unique(conn, body.get("no_bl"))
        cur = conn.execute(
            """INSERT INTO expe_departs (
                date_enlevement, affreteurs, transporteur, transporteur_id, client,
                code_postal_destination,
                ref_sifa, arc, no_cde_transport, no_bl, type_palette_matiere_id,
                type_colis, nb_palette, poids_total_kg, date_livraison,
                planning_entry_id, palette_europe, palette_europe_statut,
                palette_europe_date_retour, palette_europe_note,
                statut, created_at, created_by_email
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, 'en_attente', ?, ?)""",
            (
                date_enl,
                _f("affreteurs"),
                _f("transporteur"),
                _int_opt(body, "transporteur_id"),
                _f("client"),
                _f("code_postal_destination"),
                _f("ref_sifa"),
                _f("arc"),
                _f("no_cde_transport"),
                _f("no_bl"),
                type_palette_id,
                type_colis_val,
                _float_opt("nb_palette"),
                _float_opt("poids_total_kg"),
                _f("date_livraison"),
                planning_entry_id,
                palette_europe,
                palette_europe_statut,
                palette_europe_date_retour,
                palette_europe_note,
                now,
                email,
            ),
        )
        rid = cur.lastrowid
        # Avant le commit : le départ et sa référence de dossier entrent en base
        # dans la même transaction. Un départ enregistré sans son `no_dossier`,
        # même une fraction de seconde, est un trou dans la chaîne de contrôle.
        _sync_no_dossier(conn, rid)
        conn.commit()
        row = conn.execute(
            f"{_DEPARTS_SELECT} WHERE d.id=?", (rid,)
        ).fetchone()
    client_nom = (body.get("client") or "").strip() or "—"
    log_action(
        user=user,
        action="CREATE",
        module="expe",
        objet=f"Départ {client_nom} · {date_enl}",
        ip=request.client.host if request.client else None,
    )
    return _depart_dict(row)


@router.post("/departs/{depart_id}/valider")
def valider_depart(request: Request, depart_id: int):
    user = _require_expe_write(request)
    now = datetime.now(_PARIS).replace(tzinfo=None).isoformat(timespec="seconds")
    email = (user.get("email") or user.get("identifiant") or "").strip() or None
    with get_db() as conn:
        row = conn.execute(
            "SELECT id, statut FROM expe_departs WHERE id=?",
            (depart_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Départ introuvable")
        if row["statut"] != "en_attente":
            raise HTTPException(status_code=400, detail="Ce départ est déjà validé ou annulé")
        conn.execute(
            """UPDATE expe_departs SET statut='valide', validated_at=?, validated_by_email=?
               WHERE id=?""",
            (now, email, depart_id),
        )
        conn.commit()
        out = conn.execute(
            f"{_DEPARTS_SELECT} WHERE d.id=?", (depart_id,)
        ).fetchone()
    client_nom = (out["client"] or "").strip() if out else "—"
    log_action(
        user=user,
        action="VALIDATE",
        module="expe",
        objet=f"Départ #{depart_id} validé · {client_nom}",
        ip=request.client.host if request.client else None,
    )
    return _depart_dict(out)


@router.post("/departs/{depart_id}/invalider")
def invalider_depart(request: Request, depart_id: int):
    """Remet un départ validé dans le suivi du jour (statut en_attente)."""
    user = _require_expe_write(request)
    with get_db() as conn:
        row = conn.execute(
            "SELECT id, statut, client FROM expe_departs WHERE id=?",
            (depart_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Départ introuvable")
        if row["statut"] != "valide":
            raise HTTPException(
                status_code=400,
                detail="Seuls les départs validés peuvent être remis en suivi.",
            )
        conn.execute(
            """UPDATE expe_departs
               SET statut='en_attente', validated_at=NULL, validated_by_email=NULL
               WHERE id=?""",
            (depart_id,),
        )
        conn.commit()
        out = conn.execute(
            f"{_DEPARTS_SELECT} WHERE d.id=?", (depart_id,)
        ).fetchone()
    client_nom = (out["client"] or "").strip() if out else "—"
    log_action(
        user=user,
        action="UPDATE",
        module="expe",
        objet=f"Départ #{depart_id} remis en suivi · {client_nom}",
        ip=request.client.host if request.client else None,
    )
    return _depart_dict(out)


@router.put("/departs/{depart_id}")
async def update_depart(request: Request, depart_id: int, body: dict = Body(...)):
    """Modifie un départ (en attente ou validé)."""
    user = _require_expe_write(request)

    def _f(key: str) -> Any:
        v = body.get(key)
        if v is None or v == "":
            return None
        return v

    def _float_opt(key: str) -> Any:
        v = body.get(key)
        if v is None or v == "":
            return None
        try:
            return float(str(v).replace(",", ".").replace("\u202f", "").replace(" ", ""))
        except ValueError:
            return None

    sets = []
    args: list[Any] = []

    # Optionnel : permettre de modifier la date d'enlèvement si fournie
    if "date_enlevement" in body:
        date_enl = _date_prefix(str(body.get("date_enlevement") or "").strip())
        if not date_enl or not re.match(r"^\d{4}-\d{2}-\d{2}$", date_enl):
            raise HTTPException(status_code=400, detail="Date d'enlèvement invalide (YYYY-MM-DD)")
        sets.append("date_enlevement=?")
        args.append(date_enl)

    fields_text = [
        "affreteurs",
        "transporteur",
        "client",
        "code_postal_destination",
        "ref_sifa",
        "arc",
        "no_cde_transport",
        "no_bl",
        "date_livraison",
    ]
    for k in fields_text:
        if k in body:
            sets.append(f"{k}=?")
            args.append(_f(k))

    if "transporteur_id" in body:
        sets.append("transporteur_id=?")
        args.append(_int_opt(body, "transporteur_id"))

    if "type_palette_matiere_id" in body:
        sets.append("type_palette_matiere_id=?")
        args.append(None)  # remplacé après ouverture connexion

    if "type_colis" in body:
        sets.append("type_colis=?")
        tc = (str(body.get("type_colis") or "").strip().lower() or None)
        args.append(tc)

    fields_num = ["nb_palette", "poids_total_kg"]
    for k in fields_num:
        if k in body:
            sets.append(f"{k}=?")
            args.append(_float_opt(k))

    if "planning_entry_id" in body:
        sets.append("planning_entry_id=?")
        args.append(None)

    if "palette_europe" in body:
        raw = body.get("palette_europe")
        flag = 1 if raw in (1, True, "1", "true", "True") else 0
        sets.append("palette_europe=?")
        args.append(flag)
    if "palette_europe_statut" in body:
        sets.append("palette_europe_statut=?")
        args.append(_validate_palette_europe_statut(body.get("palette_europe_statut")))
    if "palette_europe_date_retour" in body:
        v = (body.get("palette_europe_date_retour") or "").strip()
        sets.append("palette_europe_date_retour=?")
        args.append(_date_prefix(v) or None)
    if "palette_europe_note" in body:
        sets.append("palette_europe_note=?")
        args.append(body.get("palette_europe_note") or None)

    if not sets:
        raise HTTPException(status_code=400, detail="Aucun champ à mettre à jour")

    with get_db() as conn:
        if "type_palette_matiere_id" in body:
            idx = next(i for i, s in enumerate(sets) if s.startswith("type_palette_matiere_id"))
            # Si type_colis=vrac envoyé en même temps, pas de matière première associée
            tc_in_body = (str(body.get("type_colis") or "").strip().lower() or None)
            if tc_in_body == "vrac":
                args[idx] = None
            else:
                args[idx] = _validate_type_palette_matiere_id(
                    conn, body.get("type_palette_matiere_id")
                )
        if "planning_entry_id" in body:
            idx = next(i for i, s in enumerate(sets) if s.startswith("planning_entry_id"))
            args[idx] = _validate_planning_entry_id(conn, body.get("planning_entry_id"))
        ex = conn.execute("SELECT id, statut FROM expe_departs WHERE id=?", (depart_id,)).fetchone()
        if not ex:
            raise HTTPException(status_code=404, detail="Départ introuvable")
        if ex["statut"] not in ("en_attente", "valide"):
            raise HTTPException(status_code=409, detail="Modification impossible : départ annulé")
        if "no_bl" in body and not body.get("no_bl_doublon_confirme"):
            _check_no_bl_unique(conn, body.get("no_bl"), exclure_id=depart_id)

        conn.execute(f"UPDATE expe_departs SET {', '.join(sets)} WHERE id=?", (*args, depart_id))
        if "planning_entry_id" in body:
            # Le dossier rattaché a changé (ou a été retiré) : la copie
            # textuelle doit suivre, sinon la chaîne FSC continue de pointer
            # vers l'ancien dossier.
            _sync_no_dossier(conn, depart_id)
        conn.commit()
        row = conn.execute(
            f"{_DEPARTS_SELECT} WHERE d.id=?", (depart_id,)
        ).fetchone()
    client_nom = (row["client"] or "").strip() if row else "—"
    log_action(
        user=user,
        action="UPDATE",
        module="expe",
        objet=f"Départ #{depart_id} · {client_nom}",
        ip=request.client.host if request.client else None,
    )
    return _depart_dict(row)


@router.patch("/departs/{depart_id}/palette-europe")
def update_depart_palette_europe(request: Request, depart_id: int, body: dict = Body(...)):
    """Met à jour le statut palette Europe d'un départ.

    Body : { palette_europe?: 0|1, statut?: 'en_attente'|'retournee'|'perdue',
             date_retour?: 'YYYY-MM-DD', note?: 'texte libre' }
    Seuls les champs présents sont modifiés. La date_retour est mise à NULL
    si le statut repasse à 'en_attente' et qu'aucune date n'est fournie.
    """
    user = _require_expe_write(request)

    sets: list = []
    args: list = []

    if "palette_europe" in body:
        raw = body.get("palette_europe")
        flag = 1 if raw in (1, True, "1", "true", "True") else 0
        sets.append("palette_europe=?")
        args.append(flag)

    new_statut = None
    if "statut" in body:
        new_statut = _validate_palette_europe_statut(body.get("statut"))
        sets.append("palette_europe_statut=?")
        args.append(new_statut)

    if "date_retour" in body:
        raw = (body.get("date_retour") or "").strip()
        sets.append("palette_europe_date_retour=?")
        args.append(_date_prefix(raw) or None)
    elif new_statut == "en_attente":
        sets.append("palette_europe_date_retour=?")
        args.append(None)

    if "note" in body:
        sets.append("palette_europe_note=?")
        args.append((body.get("note") or "").strip() or None)

    if not sets:
        raise HTTPException(status_code=400, detail="Aucun champ à mettre à jour")

    with get_db() as conn:
        ex = conn.execute(
            "SELECT id, client FROM expe_departs WHERE id=?", (depart_id,)
        ).fetchone()
        if not ex:
            raise HTTPException(status_code=404, detail="Départ introuvable")
        conn.execute(
            f"UPDATE expe_departs SET {', '.join(sets)} WHERE id=?",
            (*args, depart_id),
        )
        conn.commit()
        row = conn.execute(f"{_DEPARTS_SELECT} WHERE d.id=?", (depart_id,)).fetchone()
    client_nom = (row["client"] or "").strip() if row else "—"
    log_action(
        user=user,
        action="UPDATE",
        module="expe",
        objet=f"Départ #{depart_id} · palette Europe · {client_nom}",
        ip=request.client.host if request.client else None,
    )
    return _depart_dict(row)


@router.get("/palettes-europe")
def list_palettes_europe(
    request: Request,
    statut: Optional[str] = Query(None, description="en_attente | retournee | perdue"),
    client: Optional[str] = Query(None, description="Filtre client (LIKE insensible casse)"),
    q: Optional[str] = Query(None, description="Recherche libre (client, ARC, BL…)"),
):
    """Suivi des palettes Europe — liste détaillée + récap par client."""
    _require_expe(request)

    where = ["d.palette_europe = 1"]
    params: list = []

    if statut:
        st = _validate_palette_europe_statut(statut)
        where.append("d.palette_europe_statut = ?")
        params.append(st)
    if client:
        where.append("LOWER(COALESCE(d.client, '')) LIKE LOWER(?)")
        params.append(f"%{client.strip()}%")
    if q:
        search_sql, search_params = _historique_search_clause(q)
        if search_sql:
            where.append(f"({search_sql})")
            params.extend(search_params)

    where_sql = " AND ".join(where)

    with get_db() as conn:
        rows = conn.execute(
            f"""{_DEPARTS_SELECT}
                WHERE {where_sql}
                ORDER BY d.date_enlevement DESC, d.id DESC""",
            params,
        ).fetchall()

        recap_rows = conn.execute(
            """SELECT COALESCE(NULLIF(TRIM(client), ''), '— Sans client —') AS client,
                      COUNT(*) AS nb_departs,
                      COALESCE(SUM(CASE WHEN nb_palette IS NOT NULL THEN nb_palette ELSE 0 END), 0) AS nb_pal_envoyees,
                      COALESCE(SUM(CASE WHEN palette_europe_statut='retournee' AND nb_palette IS NOT NULL THEN nb_palette ELSE 0 END), 0) AS nb_pal_retournees,
                      COALESCE(SUM(CASE WHEN palette_europe_statut='perdue' AND nb_palette IS NOT NULL THEN nb_palette ELSE 0 END), 0) AS nb_pal_perdues,
                      COALESCE(SUM(CASE WHEN palette_europe_statut='en_attente' AND nb_palette IS NOT NULL THEN nb_palette ELSE 0 END), 0) AS nb_pal_en_attente
               FROM expe_departs
               WHERE palette_europe = 1
               GROUP BY client
               ORDER BY nb_pal_en_attente DESC, client COLLATE NOCASE ASC"""
        ).fetchall()

        recap_trp = _recap_palettes_transporteurs(conn)

    departs = [_depart_dict(r) for r in rows]
    recap = [dict(r) for r in recap_rows]
    # Les totaux restent calculés sur le récap CLIENT et non transporteur :
    # les deux populations donnent le même total de palettes envoyées, mais le
    # récap transporteur y ajoute reports et restitutions en vrac, qui n'ont
    # pas de contrepartie côté client. Mélanger les deux ferait un bandeau
    # dont aucune colonne ne s'additionne.
    totaux = {
        "nb_departs": sum(int(r["nb_departs"]) for r in recap),
        "nb_pal_envoyees": sum(float(r["nb_pal_envoyees"] or 0) for r in recap),
        "nb_pal_retournees": sum(float(r["nb_pal_retournees"] or 0) for r in recap),
        "nb_pal_perdues": sum(float(r["nb_pal_perdues"] or 0) for r in recap),
        "nb_pal_en_attente": sum(float(r["nb_pal_en_attente"] or 0) for r in recap),
    }
    totaux["solde_transporteurs"] = round(
        sum(float(t["solde"] or 0) for t in recap_trp), 2
    )
    totaux["nb_pal_contestees"] = round(
        sum(float(t["nb_pal_contestees"] or 0) for t in recap_trp), 2
    )
    return {
        "departs": departs,
        "recap_clients": recap,
        "recap_transporteurs": recap_trp,
        "totaux": totaux,
    }


# ─── Palettes Europe : compte courant par transporteur ─────────────
#
# Modèle repris du suivi métier historique (un onglet Excel par transporteur et
# par année, colonnes Données / Rendues / Solde). Le débiteur d'une palette
# Europe est le transporteur qui l'emporte, pas le client livré : c'est au
# transporteur qu'on la réclame, et c'est avec lui qu'on rapproche les comptes.
#
#   solde = report + données − rendues − perdues
#
# `perdues` sort du solde parce qu'une palette déclarée perdue est passée en
# perte : la laisser dedans reviendrait à la réclamer indéfiniment. Les
# palettes CONTESTÉES, elles, restent dans le solde — c'est tout l'intérêt
# d'ouvrir une contestation plutôt que de solder.

_PAL_SENS = ("report", "donnee", "rendue")


def _pal_norm_nom(nom: object) -> str:
    """Clé de rapprochement d'un nom de transporteur : casse, accents, espaces.

    « Coquelle TB », « COQUELLE TB » et « Coquelle  TB » désignent le même
    compte. Sans cette normalisation, le récap afficherait trois soldes
    partiels dont aucun n'est juste.
    """
    txt = str(nom or "").strip()
    if not txt:
        return ""
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    return " ".join(txt.upper().split())


def _pal_trp_key(trp_id: object, nom: object, id_par_nom: dict[str, int]) -> str:
    """Identité d'un compte transporteur, `transporteur_id` prioritaire.

    Les départs antérieurs au référentiel ne portent qu'un nom en texte libre.
    On les rattache au référentiel par le nom quand c'est possible, pour que
    l'historique et les mouvements récents tombent sur le même compte.
    """
    try:
        tid = int(trp_id) if trp_id not in (None, "") else None
    except (TypeError, ValueError):
        tid = None
    norm = _pal_norm_nom(nom)
    if tid is None and norm:
        tid = id_par_nom.get(norm)
    return f"id:{tid}" if tid is not None else (f"nom:{norm}" if norm else "nom:")


def _recap_palettes_transporteurs(conn) -> list[dict]:
    """Compte courant palettes Europe, un poste par transporteur."""
    id_par_nom: dict[str, int] = {}
    ref_nom: dict[int, str] = {}
    for t in conn.execute("SELECT id, nom FROM expe_transporteurs").fetchall():
        norm = _pal_norm_nom(t["nom"])
        if norm:
            id_par_nom.setdefault(norm, int(t["id"]))
        ref_nom[int(t["id"])] = (t["nom"] or "").strip()

    postes: dict[str, dict] = {}

    def poste(trp_id, nom) -> dict:
        key = _pal_trp_key(trp_id, nom, id_par_nom)
        p = postes.get(key)
        if p is None:
            tid = None
            if key.startswith("id:"):
                try:
                    tid = int(key[3:])
                except ValueError:
                    tid = None
            p = {
                "key": key,
                "transporteur_id": tid,
                "transporteur": (
                    ref_nom.get(tid) or (str(nom or "").strip() or "— Sans transporteur —")
                ),
                "report": 0.0,
                "donnees": 0.0,
                "rendues": 0.0,
                "perdues": 0.0,
                "nb_departs": 0,
                "nb_pal_contestees": 0.0,
                "nb_contestations": 0,
                "dernier_mouvement": None,
            }
            postes[key] = p
        return p

    for r in conn.execute(
        """SELECT transporteur_id, transporteur,
                  COUNT(*) AS nb_departs,
                  COALESCE(SUM(COALESCE(nb_palette,0)), 0) AS donnees,
                  COALESCE(SUM(CASE WHEN palette_europe_statut='retournee'
                                    THEN COALESCE(nb_palette,0) ELSE 0 END), 0) AS rendues,
                  COALESCE(SUM(CASE WHEN palette_europe_statut='perdue'
                                    THEN COALESCE(nb_palette,0) ELSE 0 END), 0) AS perdues,
                  MAX(date_enlevement) AS dernier
           FROM expe_departs
           WHERE palette_europe = 1
           GROUP BY transporteur_id, transporteur"""
    ).fetchall():
        p = poste(r["transporteur_id"], r["transporteur"])
        p["nb_departs"] += int(r["nb_departs"] or 0)
        p["donnees"] += float(r["donnees"] or 0)
        p["rendues"] += float(r["rendues"] or 0)
        p["perdues"] += float(r["perdues"] or 0)
        if r["dernier"] and (not p["dernier_mouvement"] or str(r["dernier"]) > p["dernier_mouvement"]):
            p["dernier_mouvement"] = str(r["dernier"])

    for r in conn.execute(
        """SELECT transporteur_id, transporteur_nom, sens,
                  COALESCE(SUM(COALESCE(nb_palette,0)), 0) AS n,
                  MAX(date_mvt) AS dernier
           FROM expe_palettes_mouvements
           GROUP BY transporteur_id, transporteur_nom, sens"""
    ).fetchall():
        p = poste(r["transporteur_id"], r["transporteur_nom"])
        sens = str(r["sens"] or "")
        if sens == "report":
            p["report"] += float(r["n"] or 0)
        elif sens == "donnee":
            p["donnees"] += float(r["n"] or 0)
        elif sens == "rendue":
            p["rendues"] += float(r["n"] or 0)
        if r["dernier"] and (not p["dernier_mouvement"] or str(r["dernier"]) > p["dernier_mouvement"]):
            p["dernier_mouvement"] = str(r["dernier"])

    for r in conn.execute(
        """SELECT transporteur_id, transporteur_nom,
                  COUNT(*) AS nb,
                  COALESCE(SUM(COALESCE(nb_palette,0)), 0) AS n
           FROM expe_palettes_contestations
           WHERE statut='ouverte'
           GROUP BY transporteur_id, transporteur_nom"""
    ).fetchall():
        p = poste(r["transporteur_id"], r["transporteur_nom"])
        p["nb_contestations"] += int(r["nb"] or 0)
        p["nb_pal_contestees"] += float(r["n"] or 0)

    out = []
    for p in postes.values():
        p["solde"] = round(
            p["report"] + p["donnees"] - p["rendues"] - p["perdues"], 2
        )
        for k in ("report", "donnees", "rendues", "perdues", "nb_pal_contestees"):
            p[k] = round(p[k], 2)
        out.append(p)
    # Le solde le plus lourd en premier : c'est le transporteur qu'on rappelle.
    out.sort(key=lambda p: (-p["solde"], p["transporteur"].upper()))
    return out


def _pal_resolve_transporteur(conn, body: dict) -> tuple[Optional[int], str]:
    """Extrait (transporteur_id, transporteur_nom) d'un body de saisie."""
    tid = body.get("transporteur_id")
    try:
        tid = int(tid) if tid not in (None, "") else None
    except (TypeError, ValueError):
        tid = None
    nom = (body.get("transporteur") or body.get("transporteur_nom") or "").strip()
    if tid is not None:
        row = conn.execute(
            "SELECT nom FROM expe_transporteurs WHERE id=?", (tid,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=400, detail="Transporteur introuvable.")
        nom = (row["nom"] or "").strip() or nom
    elif nom:
        # Saisie libre : on rattache au référentiel si le nom y correspond,
        # sinon on garde le texte. Un transporteur ponctuel n'a pas à être créé
        # dans le référentiel pour qu'on tienne son compte palettes.
        row = conn.execute(
            "SELECT id, nom FROM expe_transporteurs WHERE UPPER(TRIM(nom))=UPPER(TRIM(?)) LIMIT 1",
            (nom,),
        ).fetchone()
        if row:
            tid = int(row["id"])
            nom = (row["nom"] or "").strip()
    if tid is None and not nom:
        raise HTTPException(status_code=400, detail="Transporteur obligatoire.")
    return tid, nom


def _pal_nb(value: object, champ: str = "Nombre de palettes") -> float:
    try:
        n = float(value)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail=f"{champ} invalide.")
    if n <= 0:
        raise HTTPException(status_code=400, detail=f"{champ} doit être positif.")
    return n


@router.get("/palettes-europe/journal")
def journal_palettes_transporteur(
    request: Request,
    transporteur_id: Optional[int] = Query(None),
    transporteur: Optional[str] = Query(None),
):
    """Relevé de compte d'un transporteur : mouvements + départs, solde qui court.

    C'est la transposition d'un onglet de l'Excel : une ligne par événement,
    dans l'ordre chronologique, avec le solde recalculé à chaque ligne. Le
    solde n'est PAS stocké — le recalculer interdit qu'il diverge des
    écritures, ce qui est le défaut classique du tableur.
    """
    _require_expe(request)
    with get_db() as conn:
        id_par_nom = {
            _pal_norm_nom(t["nom"]): int(t["id"])
            for t in conn.execute("SELECT id, nom FROM expe_transporteurs").fetchall()
            if _pal_norm_nom(t["nom"])
        }
        cible = _pal_trp_key(transporteur_id, transporteur, id_par_nom)

        lignes: list[dict] = []
        for r in conn.execute(
            """SELECT id, date_enlevement, transporteur_id, transporteur, client,
                      arc, no_bl, nb_palette, palette_europe_statut,
                      palette_europe_date_retour, palette_europe_note
               FROM expe_departs WHERE palette_europe = 1"""
        ).fetchall():
            if _pal_trp_key(r["transporteur_id"], r["transporteur"], id_par_nom) != cible:
                continue
            nb = float(r["nb_palette"] or 0)
            statut = r["palette_europe_statut"] or "en_attente"
            lignes.append({
                "type": "depart",
                "id": int(r["id"]),
                "date": (r["date_enlevement"] or "")[:10],
                "libelle": (r["client"] or "—"),
                "reference": r["arc"] or r["no_bl"] or None,
                "donnees": nb,
                "rendues": nb if statut == "retournee" else 0.0,
                "perdues": nb if statut == "perdue" else 0.0,
                "statut": statut,
                "note": r["palette_europe_note"],
                "date_retour": (r["palette_europe_date_retour"] or "")[:10] or None,
            })
        for r in conn.execute(
            """SELECT id, date_mvt, transporteur_id, transporteur_nom, sens,
                      nb_palette, reference, client, note
               FROM expe_palettes_mouvements"""
        ).fetchall():
            if _pal_trp_key(r["transporteur_id"], r["transporteur_nom"], id_par_nom) != cible:
                continue
            nb = float(r["nb_palette"] or 0)
            sens = str(r["sens"] or "")
            lignes.append({
                "type": "mouvement",
                "id": int(r["id"]),
                "sens": sens,
                "date": (r["date_mvt"] or "")[:10],
                "libelle": (r["client"] or "").strip() or {
                    "report": "Solde d'ouverture",
                    "donnee": "Palettes remises",
                    "rendue": "Restitution",
                }.get(sens, sens),
                "reference": r["reference"],
                "donnees": nb if sens in ("donnee", "report") else 0.0,
                "rendues": nb if sens == "rendue" else 0.0,
                "perdues": 0.0,
                "report": nb if sens == "report" else 0.0,
                "note": r["note"],
            })
        contestations = [
            dict(r)
            for r in conn.execute(
                """SELECT * FROM expe_palettes_contestations
                   ORDER BY date_contestation DESC, id DESC"""
            ).fetchall()
            if _pal_trp_key(r["transporteur_id"], r["transporteur_nom"], id_par_nom) == cible
        ]

    # Le report d'abord quelle que soit sa date : il représente l'antériorité,
    # l'afficher au milieu du relevé rendrait la colonne solde illisible.
    lignes.sort(key=lambda x: (0 if x.get("sens") == "report" else 1, x["date"] or "", x["id"]))
    solde = 0.0
    for ligne in lignes:
        solde += ligne.get("donnees", 0.0) - ligne.get("rendues", 0.0) - ligne.get("perdues", 0.0)
        ligne["solde"] = round(solde, 2)
    return {
        "lignes": lignes,
        "contestations": contestations,
        "solde": round(solde, 2),
    }


@router.post("/palettes-europe/mouvements")
def create_palette_mouvement(request: Request, body: dict = Body(...)):
    """Saisie d'un mouvement : report d'ouverture, restitution, remise hors départ."""
    user = _require_expe_write(request)
    sens = (body.get("sens") or "").strip().lower()
    if sens not in _PAL_SENS:
        raise HTTPException(
            status_code=400,
            detail="Sens invalide (report, donnee ou rendue).",
        )
    nb = _pal_nb(body.get("nb_palette"))
    date_mvt = _date_prefix(str(body.get("date_mvt") or "").strip()) or datetime.now(
        _PARIS
    ).strftime("%Y-%m-%d")
    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    with get_db() as conn:
        tid, nom = _pal_resolve_transporteur(conn, body)
        cur = conn.execute(
            """INSERT INTO expe_palettes_mouvements
               (transporteur_id, transporteur_nom, date_mvt, sens, nb_palette,
                reference, client, note, created_at, created_by_email)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (
                tid,
                nom or None,
                date_mvt,
                sens,
                nb,
                (body.get("reference") or "").strip() or None,
                (body.get("client") or "").strip() or None,
                (body.get("note") or "").strip() or None,
                now,
                (user.get("email") or user.get("identifiant") or "").strip() or None,
            ),
        )
        conn.commit()
        mvt_id = int(cur.lastrowid)
    log_action(
        user=user,
        action="CREATE",
        module="expe",
        objet=f"Palettes Europe · {sens} {nb:g} · {nom or '—'}",
        ip=request.client.host if request.client else None,
    )
    return {"ok": True, "id": mvt_id}


@router.delete("/palettes-europe/mouvements/{mouvement_id}")
def delete_palette_mouvement(request: Request, mouvement_id: int):
    user = _require_expe_write(request)
    with get_db() as conn:
        row = conn.execute(
            "SELECT sens, nb_palette, transporteur_nom FROM expe_palettes_mouvements WHERE id=?",
            (mouvement_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Mouvement introuvable")
        conn.execute(
            "DELETE FROM expe_palettes_mouvements WHERE id=?", (mouvement_id,)
        )
        conn.commit()
    log_action(
        user=user,
        action="DELETE",
        module="expe",
        objet=(
            f"Palettes Europe · mouvement #{mouvement_id} "
            f"({row['sens']} {float(row['nb_palette'] or 0):g} · {row['transporteur_nom'] or '—'})"
        ),
        ip=request.client.host if request.client else None,
    )
    return {"ok": True}


@router.get("/palettes-europe/contestations")
def list_palette_contestations(
    request: Request,
    statut: Optional[str] = Query(None, description="ouverte | resolue | abandonnee"),
):
    """Registre des palettes non rendues / contestées — la base des réclamations."""
    _require_expe(request)
    where, params = [], []
    if statut:
        st = (statut or "").strip().lower()
        if st not in ("ouverte", "resolue", "abandonnee"):
            raise HTTPException(status_code=400, detail="Statut invalide.")
        where.append("statut=?")
        params.append(st)
    where_sql = (" WHERE " + " AND ".join(where)) if where else ""
    with get_db() as conn:
        rows = conn.execute(
            f"""SELECT * FROM expe_palettes_contestations{where_sql}
                ORDER BY date_contestation DESC, id DESC LIMIT 500""",
            params,
        ).fetchall()
        tot = conn.execute(
            """SELECT COUNT(*) AS nb, COALESCE(SUM(COALESCE(nb_palette,0)),0) AS n
               FROM expe_palettes_contestations WHERE statut='ouverte'"""
        ).fetchone()
    return {
        "contestations": [dict(r) for r in rows],
        "totaux": {
            "nb_ouvertes": int(tot["nb"] or 0),
            "nb_pal_ouvertes": round(float(tot["n"] or 0), 2),
        },
    }


@router.post("/palettes-europe/contestations")
def create_palette_contestation(request: Request, body: dict = Body(...)):
    user = _require_expe_write(request)
    nb = _pal_nb(body.get("nb_palette"))
    date_c = _date_prefix(str(body.get("date_contestation") or "").strip()) or datetime.now(
        _PARIS
    ).strftime("%Y-%m-%d")
    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    depart_id = body.get("depart_id")
    try:
        depart_id = int(depart_id) if depart_id not in (None, "") else None
    except (TypeError, ValueError):
        depart_id = None
    with get_db() as conn:
        tid, nom = _pal_resolve_transporteur(conn, body)
        cur = conn.execute(
            """INSERT INTO expe_palettes_contestations
               (transporteur_id, transporteur_nom, depart_id, date_contestation,
                recepisse, client, nb_palette, cause, statut, note,
                created_at, created_by_email)
               VALUES (?,?,?,?,?,?,?,?, 'ouverte', ?,?,?)""",
            (
                tid,
                nom or None,
                depart_id,
                date_c,
                (body.get("recepisse") or "").strip() or None,
                (body.get("client") or "").strip() or None,
                nb,
                (body.get("cause") or "Palette non rendue").strip() or None,
                (body.get("note") or "").strip() or None,
                now,
                (user.get("email") or user.get("identifiant") or "").strip() or None,
            ),
        )
        conn.commit()
        cid = int(cur.lastrowid)
    log_action(
        user=user,
        action="CREATE",
        module="expe",
        objet=f"Contestation palettes · {nb:g} pal. · {nom or '—'}",
        ip=request.client.host if request.client else None,
    )
    return {"ok": True, "id": cid}


@router.patch("/palettes-europe/contestations/{contestation_id}")
def update_palette_contestation(
    request: Request, contestation_id: int, body: dict = Body(...)
):
    """Met à jour une contestation (statut, cause, note, nombre)."""
    user = _require_expe_write(request)
    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    sets, args = [], []
    if "statut" in body:
        st = (body.get("statut") or "").strip().lower()
        if st not in ("ouverte", "resolue", "abandonnee"):
            raise HTTPException(status_code=400, detail="Statut invalide.")
        sets.append("statut=?")
        args.append(st)
        # `resolved_at` est remis à NULL quand on rouvre : une contestation
        # rouverte qui garde sa date de résolution se lit comme close.
        sets.append("resolved_at=?")
        args.append(now if st in ("resolue", "abandonnee") else None)
    for champ in ("recepisse", "client", "cause", "note"):
        if champ in body:
            sets.append(f"{champ}=?")
            args.append((body.get(champ) or "").strip() or None)
    if "nb_palette" in body:
        sets.append("nb_palette=?")
        args.append(_pal_nb(body.get("nb_palette")))
    if not sets:
        raise HTTPException(status_code=400, detail="Aucun champ à mettre à jour.")
    args.append(contestation_id)
    with get_db() as conn:
        ex = conn.execute(
            "SELECT id FROM expe_palettes_contestations WHERE id=?", (contestation_id,)
        ).fetchone()
        if not ex:
            raise HTTPException(status_code=404, detail="Contestation introuvable")
        conn.execute(
            f"UPDATE expe_palettes_contestations SET {', '.join(sets)} WHERE id=?", args
        )
        conn.commit()
        row = conn.execute(
            "SELECT * FROM expe_palettes_contestations WHERE id=?", (contestation_id,)
        ).fetchone()
    log_action(
        user=user,
        action="UPDATE",
        module="expe",
        objet=f"Contestation palettes #{contestation_id}",
        ip=request.client.host if request.client else None,
    )
    return dict(row)


@router.delete("/palettes-europe/contestations/{contestation_id}")
def delete_palette_contestation(request: Request, contestation_id: int):
    user = _require_expe_write(request)
    with get_db() as conn:
        ex = conn.execute(
            "SELECT id FROM expe_palettes_contestations WHERE id=?", (contestation_id,)
        ).fetchone()
        if not ex:
            raise HTTPException(status_code=404, detail="Contestation introuvable")
        conn.execute(
            "DELETE FROM expe_palettes_contestations WHERE id=?", (contestation_id,)
        )
        conn.commit()
    log_action(
        user=user,
        action="DELETE",
        module="expe",
        objet=f"Contestation palettes #{contestation_id}",
        ip=request.client.host if request.client else None,
    )
    return {"ok": True}


@router.patch("/matieres-palettes/{matiere_id}/europe")
def set_matiere_palette_europe(request: Request, matiere_id: int, body: dict = Body(...)):
    """Active / désactive le flag Europe sur une référence palette MyStock."""
    user = _require_expe_write(request)
    raw = body.get("is_europe")
    flag = 1 if raw in (1, True, "1", "true", "True") else 0
    with get_db() as conn:
        ex = conn.execute(
            "SELECT id, reference FROM matieres_premieres WHERE id=? AND categorie='palette'",
            (matiere_id,),
        ).fetchone()
        if not ex:
            raise HTTPException(status_code=404, detail="Référence palette introuvable")
        conn.execute(
            "UPDATE matieres_premieres SET is_europe=? WHERE id=?",
            (flag, matiere_id),
        )
        conn.commit()
    log_action(
        user=user,
        action="UPDATE",
        module="expe",
        objet=f"Réf palette {ex['reference']} · is_europe={flag}",
        ip=request.client.host if request.client else None,
    )
    return {"ok": True, "is_europe": flag}


@router.delete("/departs/{depart_id}")
def delete_depart(request: Request, depart_id: int):
    """Supprime un départ (en attente ou validé)."""
    user = _require_expe_write(request)
    client_nom = ""
    with get_db() as conn:
        ex = conn.execute(
            "SELECT id, statut, client FROM expe_departs WHERE id=?", (depart_id,)
        ).fetchone()
        if not ex:
            raise HTTPException(status_code=404, detail="Départ introuvable")
        if ex["statut"] not in ("en_attente", "valide"):
            raise HTTPException(status_code=409, detail="Suppression impossible : départ annulé")

        # Rétention FSC : un départ portant un claim est la PREUVE d'une vente
        # certifiée. FSC-STD-40-004 impose de conserver ces enregistrements
        # 5 ans ; les effacer rend la vente indémontrable en audit. En négoce
        # direct (A2) c'est même la seule trace existante — aucun lot, aucun
        # mouvement de stock ne subsiste ailleurs.
        try:
            fsc = conn.execute(
                """SELECT TRIM(COALESCE(fsc_claim_sortant,'')) AS claim,
                          TRIM(COALESCE(fsc_bl_fournisseur,'')) AS bl,
                          COALESCE(fsc_sans_transit,0) AS direct
                     FROM expe_departs WHERE id=?""",
                (depart_id,),
            ).fetchone()
        except Exception:
            fsc = None
        if fsc and (
            (fsc["claim"] and fsc["claim"] != "non_fsc") or fsc["bl"] or int(fsc["direct"] or 0)
        ):
            raise HTTPException(
                status_code=409,
                detail=(
                    "Suppression impossible : ce départ porte un claim FSC. "
                    "Ces enregistrements doivent être conservés 5 ans. "
                    "Annuler le départ plutôt que le supprimer."
                ),
            )

        client_nom = (ex["client"] or "").strip() or "—"
        conn.execute("DELETE FROM expe_departs WHERE id=?", (depart_id,))
        conn.commit()
    log_action(
        user=user,
        action="DELETE",
        module="expe",
        objet=f"Départ #{depart_id} supprimé · {client_nom}",
        ip=request.client.host if request.client else None,
    )
    return {"ok": True}


@router.get("/departs/historique")
def historique_departs(
    request: Request,
    q: str = "",
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
):
    _require_expe(request)
    search_sql, search_params = _historique_search_clause(q)
    where = "WHERE d.statut = 'valide'"
    if search_sql:
        where += f" AND ({search_sql})"
    offset = (page - 1) * limit
    with get_db() as conn:
        total = conn.execute(
            f"""SELECT COUNT(*) AS n
                FROM expe_departs d
                LEFT JOIN matieres_premieres mp ON mp.id = d.type_palette_matiere_id
                {where}""",
            search_params,
        ).fetchone()["n"]
        rows = conn.execute(
            f"""{_DEPARTS_SELECT}
                {where}
                ORDER BY datetime(COALESCE(d.validated_at, d.created_at)) DESC, d.id DESC
                LIMIT ? OFFSET ?""",
            (*search_params, limit, offset),
        ).fetchall()
    pages = max(1, (int(total) + limit - 1) // limit) if total else 1
    if page > pages:
        page = pages
    return {
        "rows": [_depart_dict(r) for r in rows],
        "total": int(total),
        "page": page,
        "limit": limit,
        "pages": pages,
    }


# ─── Transporteurs ───────────────────────────────────────────────────


@router.get("/transporteurs")
def list_transporteurs(request: Request):
    _require_expe(request)
    with get_db() as conn:
        seed_expe_transporteurs_if_empty(conn)
        conn.commit()
        rows = conn.execute(
            """SELECT * FROM expe_transporteurs
               ORDER BY actif DESC, nom ASC"""
        ).fetchall()
    return [_serialize_transporteur_row(r) for r in rows]


@router.post("/transporteurs")
def create_transporteur(request: Request, body: dict = Body(...)):
    user = _require_expe_write(request)
    nom = (body.get("nom") or "").strip()
    if not nom:
        raise HTTPException(status_code=400, detail="Nom du transporteur obligatoire")
    now = _now_paris_iso()
    taxe = _float_opt(body, "taxe_carburant_pct")
    if taxe is None:
        taxe = 0.0
    emails = _normalize_emails(body.get("contact_emails", body.get("contact_email")))
    portail = _normalize_portail(body.get("contact_portail_url"))
    # contact_email legacy = première adresse mail, sinon URL portail (back-compat lecture seule)
    legacy_email = emails[0] if emails else portail
    # Téléphones : liste [{numero, service}]. Fallback sur contact_tel string si présent.
    tels = _normalize_tels(body.get("contact_tels", body.get("contact_tel")))
    # contact_tel legacy = 1er numéro (back-compat lecture seule)
    legacy_tel = tels[0]["numero"] if tels else _f(body, "contact_tel")
    with get_db() as conn:
        cur = conn.execute(
            """INSERT INTO expe_transporteurs (
                nom, taxe_carburant_pct, contact_nom, contact_email, contact_tel,
                contact_portail_url, contact_emails, contact_tels,
                zone_france, zone_france_hors_paris, zone_affretement, zone_messagerie,
                palette_max, poids_max_kg, accepte_poids, accepte_palette,
                couleur, actif, created_at, langue
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                nom,
                taxe,
                _f(body, "contact_nom"),
                legacy_email,
                legacy_tel,
                portail,
                json.dumps(emails, ensure_ascii=False),
                json.dumps(tels, ensure_ascii=False),
                _int_flag(body, "zone_france", 1),
                _int_flag(body, "zone_france_hors_paris", 0),
                _int_flag(body, "zone_affretement", 0),
                _int_flag(body, "zone_messagerie", 0),
                _int_opt(body, "palette_max"),
                _float_opt(body, "poids_max_kg"),
                _int_flag(body, "accepte_poids", 1),
                _int_flag(body, "accepte_palette", 1),
                _f(body, "couleur"),
                _int_flag(body, "actif", 1),
                now,
                _normalize_langue(body.get("langue")),
            ),
        )
        conn.commit()
        rid = cur.lastrowid
        row = conn.execute(
            "SELECT * FROM expe_transporteurs WHERE id=?", (rid,)
        ).fetchone()
    log_action(
        user=user,
        action="CREATE",
        module="expe",
        objet=f"Transporteur {nom}",
        ip=request.client.host if request.client else None,
    )
    return _serialize_transporteur_row(row)


@router.put("/transporteurs/{transporteur_id}")
def update_transporteur(
    request: Request, transporteur_id: int, body: dict = Body(...)
):
    user = _require_expe_write(request)
    sets = []
    args: list[Any] = []

    if "nom" in body:
        nom = (body.get("nom") or "").strip()
        if not nom:
            raise HTTPException(status_code=400, detail="Nom du transporteur obligatoire")
        sets.append("nom=?")
        args.append(nom)

    if "taxe_carburant_pct" in body:
        taxe = _float_opt(body, "taxe_carburant_pct")
        sets.append("taxe_carburant_pct=?")
        args.append(0.0 if taxe is None else taxe)

    for k in ("contact_nom", "couleur"):
        if k in body:
            sets.append(f"{k}=?")
            args.append(_f(body, k))

    if "langue" in body:
        sets.append("langue=?")
        args.append(_normalize_langue(body.get("langue")))

    # Téléphones : liste [{numero, service}] — recalcule aussi contact_tel legacy
    new_tels: Optional[list[dict]] = None
    if "contact_tels" in body:
        new_tels = _normalize_tels(body.get("contact_tels"))
        sets.append("contact_tels=?")
        args.append(json.dumps(new_tels, ensure_ascii=False))
    elif "contact_tel" in body:
        # Compat ancien client : string simple -> liste normalisée
        new_tels = _normalize_tels(body.get("contact_tel"))
        sets.append("contact_tels=?")
        args.append(json.dumps(new_tels, ensure_ascii=False))

    if new_tels is not None:
        sets.append("contact_tel=?")
        args.append(new_tels[0]["numero"] if new_tels else None)

    # Portail (URL séparée du / des emails)
    if "contact_portail_url" in body:
        sets.append("contact_portail_url=?")
        args.append(_normalize_portail(body.get("contact_portail_url")))

    # Emails : nouveau champ (liste) — recalcule aussi contact_email legacy
    new_emails: Optional[list[str]] = None
    if "contact_emails" in body:
        new_emails = _normalize_emails(body.get("contact_emails"))
        sets.append("contact_emails=?")
        args.append(json.dumps(new_emails, ensure_ascii=False))
    elif "contact_email" in body:
        # Compat ancien client : on accepte une string et on convertit
        new_emails = _normalize_emails(body.get("contact_email"))
        sets.append("contact_emails=?")
        args.append(json.dumps(new_emails, ensure_ascii=False))

    if new_emails is not None:
        # Recalcule contact_email legacy = 1ʳᵉ adresse, sinon URL portail si fournie
        sets.append("contact_email=?")
        if new_emails:
            args.append(new_emails[0])
        elif "contact_portail_url" in body:
            args.append(_normalize_portail(body.get("contact_portail_url")))
        else:
            args.append(None)

    for k in (
        "zone_france",
        "zone_france_hors_paris",
        "zone_affretement",
        "zone_messagerie",
        "actif",
        "accepte_poids",
        "accepte_palette",
    ):
        if k in body:
            sets.append(f"{k}=?")
            args.append(_int_flag(body, k, 0))

    if "palette_max" in body:
        sets.append("palette_max=?")
        args.append(_int_opt(body, "palette_max"))

    if "poids_max_kg" in body:
        sets.append("poids_max_kg=?")
        args.append(_float_opt(body, "poids_max_kg"))

    if not sets:
        raise HTTPException(status_code=400, detail="Aucun champ à mettre à jour")

    sets.append("updated_at=?")
    args.append(_now_paris_iso())

    with get_db() as conn:
        ex = conn.execute(
            "SELECT id, nom FROM expe_transporteurs WHERE id=?",
            (transporteur_id,),
        ).fetchone()
        if not ex:
            raise HTTPException(status_code=404, detail="Transporteur introuvable")
        conn.execute(
            f"UPDATE expe_transporteurs SET {', '.join(sets)} WHERE id=?",
            (*args, transporteur_id),
        )
        conn.commit()
        row = conn.execute(
            "SELECT * FROM expe_transporteurs WHERE id=?", (transporteur_id,)
        ).fetchone()
    nom_log = (row["nom"] or ex["nom"] or "").strip() if row else f"#{transporteur_id}"
    log_action(
        user=user,
        action="UPDATE",
        module="expe",
        objet=f"Transporteur {nom_log}",
        ip=request.client.host if request.client else None,
    )
    return _serialize_transporteur_row(row)


@router.delete("/transporteurs/{transporteur_id}")
def delete_transporteur(request: Request, transporteur_id: int):
    user = _require_expe_write(request)
    now = _now_paris_iso()
    with get_db() as conn:
        ex = conn.execute(
            "SELECT id, nom FROM expe_transporteurs WHERE id=?",
            (transporteur_id,),
        ).fetchone()
        if not ex:
            raise HTTPException(status_code=404, detail="Transporteur introuvable")
        conn.execute(
            "UPDATE expe_transporteurs SET actif=0, updated_at=? WHERE id=?",
            (now, transporteur_id),
        )
        conn.commit()
    nom_log = (ex["nom"] or "").strip() or f"#{transporteur_id}"
    log_action(
        user=user,
        action="DELETE",
        module="expe",
        objet=f"Transporteur {nom_log} (désactivé)",
        ip=request.client.host if request.client else None,
    )
    return {"ok": True}


@router.post("/transporteurs/{transporteur_id}/tarif")
async def upload_transporteur_tarif(
    request: Request,
    transporteur_id: int,
    fichier: UploadFile = File(...),
):
    user = _require_expe_write(request)
    raw_name = fichier.filename or "tarif"
    safe_name = _safe_tarif_filename(raw_name)
    ext = Path(safe_name).suffix.lower()
    if ext not in _ALLOWED_TARIF_EXT:
        raise HTTPException(
            status_code=400,
            detail="Format non accepté (PDF, Excel, image).",
        )
    content = await fichier.read()
    if not content:
        raise HTTPException(status_code=400, detail="Fichier vide.")
    if len(content) > 15 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 15 Mo).")

    stored_name = f"{uuid.uuid4().hex}_{safe_name}"
    rel_url = f"{TARIF_UPLOAD_DIR}/{stored_name}"
    from config import BASE_DIR

    dest_path = os.path.join(BASE_DIR, rel_url)
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)

    with get_db() as conn:
        ex = conn.execute(
            "SELECT id, nom, tarif_url FROM expe_transporteurs WHERE id=?",
            (transporteur_id,),
        ).fetchone()
        if not ex:
            raise HTTPException(status_code=404, detail="Transporteur introuvable")
        old_url = ex["tarif_url"]
        try:
            with open(dest_path, "wb") as out:
                shutil.copyfileobj(BytesIO(content), out)
        except OSError:
            raise HTTPException(status_code=500, detail="Enregistrement du fichier impossible.")
        now = _now_paris_iso()
        conn.execute(
            """UPDATE expe_transporteurs
               SET tarif_filename=?, tarif_url=?, updated_at=?
               WHERE id=?""",
            (safe_name, rel_url, now, transporteur_id),
        )
        conn.commit()
    _unlink_tarif(old_url)
    nom_log = (ex["nom"] or "").strip() or f"#{transporteur_id}"
    log_action(
        user=user,
        action="UPDATE",
        module="expe",
        objet=f"Transporteur {nom_log} · tarif",
        ip=request.client.host if request.client else None,
    )
    return {"ok": True, "tarif_url": rel_url}


@router.delete("/transporteurs/{transporteur_id}/tarif")
def delete_transporteur_tarif(request: Request, transporteur_id: int):
    user = _require_expe_write(request)
    with get_db() as conn:
        ex = conn.execute(
            "SELECT id, nom, tarif_url FROM expe_transporteurs WHERE id=?",
            (transporteur_id,),
        ).fetchone()
        if not ex:
            raise HTTPException(status_code=404, detail="Transporteur introuvable")
        old_url = ex["tarif_url"]
        now = _now_paris_iso()
        conn.execute(
            """UPDATE expe_transporteurs
               SET tarif_filename=NULL, tarif_url=NULL, updated_at=?
               WHERE id=?""",
            (now, transporteur_id),
        )
        conn.commit()
    _unlink_tarif(old_url)
    nom_log = (ex["nom"] or "").strip() or f"#{transporteur_id}"
    log_action(
        user=user,
        action="UPDATE",
        module="expe",
        objet=f"Transporteur {nom_log} · tarif supprimé",
        ip=request.client.host if request.client else None,
    )
    return {"ok": True}


@router.get("/transporteurs/{transporteur_id}/tarif")
def get_transporteur_tarif(request: Request, transporteur_id: int):
    _require_expe(request)
    with get_db() as conn:
        ex = conn.execute(
            "SELECT tarif_url, tarif_filename FROM expe_transporteurs WHERE id=?",
            (transporteur_id,),
        ).fetchone()
    if not ex:
        raise HTTPException(status_code=404, detail="Transporteur introuvable")
    if not ex["tarif_url"]:
        raise HTTPException(status_code=404, detail="Aucun tarif enregistré")
    path = _resolve_tarif_path(ex["tarif_url"])
    if not path:
        raise HTTPException(status_code=404, detail="Fichier tarif introuvable")
    filename = (ex["tarif_filename"] or Path(path).name) or "tarif"
    return FileResponse(path=path, filename=filename)


# ─── Tarifs structurés ─────────────────────────────────────────────


@router.get("/transporteurs/{transporteur_id}/tarifs")
def list_tarifs(request: Request, transporteur_id: int):
    _require_expe(request)
    with get_db() as conn:
        if not conn.execute(
            "SELECT 1 FROM expe_transporteurs WHERE id=?", (transporteur_id,)
        ).fetchone():
            raise HTTPException(status_code=404, detail="Transporteur introuvable")
        lignes = conn.execute(
            """SELECT * FROM expe_tarifs WHERE transporteur_id=?
               ORDER BY type_envoi, zone_valeur, tranche_min""",
            (transporteur_id,),
        ).fetchall()
        frais = conn.execute(
            """SELECT * FROM expe_tarifs_frais WHERE transporteur_id=?
               ORDER BY libelle""",
            (transporteur_id,),
        ).fetchall()
    return {"lignes": [dict(r) for r in lignes], "frais": [dict(r) for r in frais]}


@router.post("/transporteurs/{transporteur_id}/tarifs/import-csv")
async def import_tarifs_csv(
    request: Request,
    transporteur_id: int,
    file: UploadFile = File(...),
):
    user = _require_expe_write(request)
    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    content = await file.read()
    reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
    cols_oblig = {
        "type_envoi",
        "base_calcul",
        "zone_type",
        "zone_valeur",
        "tranche_min",
        "prix",
        "unite",
    }
    rows = list(reader)
    if not rows:
        raise HTTPException(status_code=400, detail="CSV vide")
    if not cols_oblig.issubset(set(rows[0].keys())):
        raise HTTPException(
            status_code=400,
            detail=f"Colonnes manquantes. Attendu : {', '.join(sorted(cols_oblig))}",
        )

    def _csv_f(row: dict, k: str) -> Any:
        v = (row.get(k) or "").strip()
        return v or None

    def _csv_r(row: dict, k: str) -> Any:
        v = _csv_f(row, k)
        return float(v) if v is not None else None

    inserted = 0
    with get_db() as conn:
        if not conn.execute(
            "SELECT 1 FROM expe_transporteurs WHERE id=?", (transporteur_id,)
        ).fetchone():
            raise HTTPException(status_code=404, detail="Transporteur introuvable")
        for row in rows:
            conn.execute(
                """INSERT INTO expe_tarifs
                   (transporteur_id, type_envoi, base_calcul, zone_type, zone_valeur,
                    tranche_min, tranche_max, prix, unite, mini_perception,
                    valid_from, valid_to, actif, source_filename, created_at, created_by_email)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,0,?,?,?)""",
                (
                    transporteur_id,
                    _csv_f(row, "type_envoi"),
                    _csv_f(row, "base_calcul"),
                    _csv_f(row, "zone_type"),
                    _csv_f(row, "zone_valeur"),
                    float(row.get("tranche_min") or 0),
                    _csv_r(row, "tranche_max"),
                    float(row.get("prix") or 0),
                    _csv_f(row, "unite"),
                    _csv_r(row, "mini_perception"),
                    _csv_f(row, "valid_from"),
                    _csv_f(row, "valid_to"),
                    file.filename,
                    now,
                    user.get("email") or user.get("identifiant"),
                ),
            )
            inserted += 1
        conn.commit()
    return {
        "inserted": inserted,
        "actif": 0,
        "message": f"{inserted} lignes importées en brouillon — à valider.",
    }


@router.post("/transporteurs/{transporteur_id}/tarifs/valider")
def valider_tarifs(
    request: Request, transporteur_id: int, body: dict = Body(...)
):
    _require_expe_write(request)
    ids = body.get("ids") or []
    with get_db() as conn:
        if not conn.execute(
            "SELECT 1 FROM expe_transporteurs WHERE id=?", (transporteur_id,)
        ).fetchone():
            raise HTTPException(status_code=404, detail="Transporteur introuvable")
        if ids:
            placeholders = ",".join("?" * len(ids))
            conn.execute(
                f"""UPDATE expe_tarifs SET actif=1
                    WHERE transporteur_id=? AND id IN ({placeholders})""",
                (transporteur_id, *ids),
            )
        else:
            conn.execute(
                "UPDATE expe_tarifs SET actif=1 WHERE transporteur_id=? AND actif=0",
                (transporteur_id,),
            )
        conn.commit()
        updated = conn.execute(
            "SELECT COUNT(*) AS n FROM expe_tarifs WHERE transporteur_id=? AND actif=1",
            (transporteur_id,),
        ).fetchone()["n"]
    return {"actives": updated}


@router.delete("/transporteurs/{transporteur_id}/tarifs")
def vider_tarifs_transporteur(request: Request, transporteur_id: int):
    """Supprime toutes les lignes tarifaires importées (grille + frais annexes)."""
    user = _require_expe_write(request)
    with get_db() as conn:
        ex = conn.execute(
            "SELECT id, nom FROM expe_transporteurs WHERE id=?",
            (transporteur_id,),
        ).fetchone()
        if not ex:
            raise HTTPException(status_code=404, detail="Transporteur introuvable")
        n_lignes = conn.execute(
            "SELECT COUNT(*) AS n FROM expe_tarifs WHERE transporteur_id=?",
            (transporteur_id,),
        ).fetchone()["n"]
        n_frais = conn.execute(
            "SELECT COUNT(*) AS n FROM expe_tarifs_frais WHERE transporteur_id=?",
            (transporteur_id,),
        ).fetchone()["n"]
        conn.execute(
            "DELETE FROM expe_tarifs WHERE transporteur_id=?",
            (transporteur_id,),
        )
        conn.execute(
            "DELETE FROM expe_tarifs_frais WHERE transporteur_id=?",
            (transporteur_id,),
        )
        conn.commit()
    nom_log = (ex["nom"] or "").strip() or f"#{transporteur_id}"
    log_action(
        user=user,
        action="DELETE",
        module="expe",
        objet=f"Transporteur {nom_log} · tarifs vidés ({n_lignes} lignes, {n_frais} frais)",
        ip=request.client.host if request.client else None,
    )
    return {"deleted_lignes": n_lignes, "deleted_frais": n_frais}


_PROMPT_EXTRACTION_TARIF = """Tu es un expert en tarification transport en France.
Analyse cette grille tarifaire et extrait TOUTES les lignes tarifaires au format JSON strict.

Retourne UNIQUEMENT un objet JSON avec deux clés :
- "lignes" : liste de lignes tarifaires
- "frais" : liste de frais annexes (gasoil, sûreté, hayon, RDV, etc.)

Chaque ligne tarifaire a ces champs (tous requis sauf mention) :
{
  "type_envoi": "messagerie" | "ramasse" | "affretement" | "express_intl",
  "base_calcul": "poids" | "palette" | "metre_plancher",
  "zone_type": "departement" | "code_postal" | "zone_intl" | "pays",
  "zone_valeur": "59" (numéro département) | "59200" (CP) | "7" (zone intl) | "DE" (pays),
  "tranche_min": 0,
  "tranche_max": 10,
  "prix": 12.50,
  "unite": "forfait" | "au_100kg" | "au_kg",
  "mini_perception": 8.50
}

Chaque frais annexe a ces champs :
{
  "libelle": "Gasoil",
  "mode": "pct_transport" | "forfait_expedition" | "par_palette",
  "valeur": 12.8,
  "mini": null,
  "applique_defaut": 1
}

Règles importantes :
- Si la grille est par poids avec des tranches forfait puis au 100kg : utilise unite="forfait" pour les tranches ≤ 100 kg et unite="au_100kg" pour les tranches > 100 kg.
- Si la grille est par palette : base_calcul="palette", unite="forfait".
- zone_valeur pour les départements français : toujours en 2 caractères ("01".."95", "2A", "2B") ou 3 pour DOM ("971".."976").
- Si une cellule est vide ou marquée "NC" / "-" : ignorer cette ligne.
- Extraire les frais depuis les onglets "Conditions commerciales" ou équivalents.

Ne retourne rien d'autre que le JSON.
"""


def _parse_tarif_json_raw(raw: str) -> dict:
    text = (raw or "").strip()
    if text.startswith("```"):
        parts = text.split("```")
        if len(parts) >= 2:
            text = parts[1]
            if text.startswith("json"):
                text = text[4:]
    return json.loads(text.strip())


@router.post("/transporteurs/{transporteur_id}/tarif/parse")
async def parse_tarif_ia(request: Request, transporteur_id: int):
    user = _require_expe_write(request)
    from config import ANTHROPIC_API_KEY

    if not ANTHROPIC_API_KEY:
        raise HTTPException(
            status_code=503,
            detail="Clé Anthropic non configurée — ajouter ANTHROPIC_API_KEY dans .env",
        )

    with get_db() as conn:
        trp = conn.execute(
            "SELECT * FROM expe_transporteurs WHERE id=?", (transporteur_id,)
        ).fetchone()
    if not trp:
        raise HTTPException(status_code=404, detail="Transporteur introuvable")
    if not trp["tarif_url"]:
        raise HTTPException(status_code=400, detail="Aucun fichier tarif uploadé pour ce transporteur")

    filepath = _resolve_tarif_path(trp["tarif_url"])
    if not filepath:
        raise HTTPException(status_code=404, detail="Fichier tarif introuvable sur le disque")

    ext = os.path.splitext(filepath)[1].lower()

    if ext in (".xlsx", ".xls"):
        import openpyxl

        wb = openpyxl.load_workbook(filepath, data_only=True)
        parts = []
        for ws in wb.worksheets:
            parts.append(f"=== Feuille : {ws.title} ===")
            for row in ws.iter_rows(values_only=True):
                line = "\t".join("" if c is None else str(c) for c in row)
                if line.strip():
                    parts.append(line)
        file_text = "\n".join(parts)
        content_block = {
            "type": "text",
            "text": f"Voici la grille tarifaire au format texte (extrait Excel) :\n\n{file_text}",
        }
    elif ext == ".pdf":
        import base64

        with open(filepath, "rb") as f:
            b64 = base64.standard_b64encode(f.read()).decode("utf-8")
        content_block = {
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": b64,
            },
        }
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Format non supporté : {ext}. Uploader un .xlsx ou .pdf.",
        )

    import anthropic as _anthropic

    client = _anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    message = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=8192,
        messages=[
            {
                "role": "user",
                "content": [
                    content_block,
                    {"type": "text", "text": _PROMPT_EXTRACTION_TARIF},
                ],
            }
        ],
    )

    raw = message.content[0].text
    data = _parse_tarif_json_raw(raw)

    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    lignes_data = data.get("lignes", [])
    frais_data = data.get("frais", [])
    source_name = trp["tarif_filename"] or trp["tarif_url"]
    email = user.get("email") or user.get("identifiant")

    with get_db() as conn:
        for lg in lignes_data:
            conn.execute(
                """INSERT INTO expe_tarifs
                   (transporteur_id, type_envoi, base_calcul, zone_type, zone_valeur,
                    tranche_min, tranche_max, prix, unite, mini_perception,
                    actif, source_filename, created_at, created_by_email)
                   VALUES (?,?,?,?,?,?,?,?,?,?,0,?,?,?)""",
                (
                    transporteur_id,
                    lg.get("type_envoi"),
                    lg.get("base_calcul"),
                    lg.get("zone_type"),
                    lg.get("zone_valeur"),
                    lg.get("tranche_min", 0),
                    lg.get("tranche_max"),
                    lg.get("prix", 0),
                    lg.get("unite"),
                    lg.get("mini_perception"),
                    source_name,
                    now,
                    email,
                ),
            )
        for fr in frais_data:
            conn.execute(
                """INSERT OR IGNORE INTO expe_tarifs_frais
                   (transporteur_id, libelle, mode, valeur, mini, applique_defaut)
                   VALUES (?,?,?,?,?,?)""",
                (
                    transporteur_id,
                    fr.get("libelle"),
                    fr.get("mode"),
                    fr.get("valeur", 0),
                    fr.get("mini"),
                    fr.get("applique_defaut", 1),
                ),
            )
        conn.commit()

    return {
        "lignes_extraites": len(lignes_data),
        "frais_extraits": len(frais_data),
        "actif": 0,
        "apercu_lignes": lignes_data[:10],
        "message": (
            f"{len(lignes_data)} lignes et {len(frais_data)} frais extraits — "
            "à valider avant activation."
        ),
    }


def _tarif_float(v, default=None):
    """Convertit une valeur de cellule en float, None si impossible."""
    import math as _math

    try:
        f = float(str(v).strip().replace(",", "."))
        return None if _math.isnan(f) else f
    except Exception:
        return default


def _tarif_dept_from_label(label):
    """
    Extrait le code département depuis des formats variés :
      '(59) NORD'  →  '59'
      '59 - NORD'  →  '59'
      'FR59'       →  '59'
      '02'         →  '02'
    """
    s = str(label or "").strip()
    m = re.search(r"\((\w{1,3})\)", s)
    if m:
        code = m.group(1)
        return code.upper() if code.upper() in ("2A", "2B") else code.zfill(2)
    m = re.match(r"^FR(\w{2,3})$", s.upper())
    if m:
        code = m.group(1)
        return code.upper() if code.upper() in ("2A", "2B") else code.lstrip("0").zfill(2)
    m = re.match(r"^(\d{2,3})\s*[-–]?\s*", s)
    if m:
        code = m.group(1)
        return code.zfill(2) if len(code) <= 3 else None
    return None


def _tarif_unite_norm(v):
    s = str(v or "").strip().upper()
    if "100" in s:
        return "au_100kg"
    if "KG" in s and "100" not in s:
        return "au_kg"
    return "forfait"


def _tarif_find_header_row(ws, keywords, max_scan=40):
    """Retourne le numéro de la première ligne contenant un keyword (insensible à la casse)."""
    keywords_up = [k.upper() for k in keywords]
    for r in range(1, max_scan + 1):
        for c in range(1, min(ws.max_column + 1, 20)):
            val = str(ws.cell(row=r, column=c).value or "").upper()
            if any(k in val for k in keywords_up):
                return r
    return None


def _detect_tarif_format(wb):
    """
    Détecte le format de la grille tarifaire en examinant noms de feuilles + cellules clés.
    Retourne : 'compte100346' | 'ceva' | 'transbenelux' | 'generique'
    """
    sheet_names = " | ".join(ws.title.upper() for ws in wb.worksheets)

    if any(k in sheet_names for k in ("MESSAGERIE", "SMARTPAL", "SMART PAL", "CONDITIONS COMMERCIALES")):
        return "ceva"

    if any(k in sheet_names for k in ("BENELUX", "TRANSBENELUX", "SIFA VERS FRANCE")):
        return "transbenelux"

    for ws in wb.worksheets:
        a8 = str(ws["A8"].value or "").upper()
        if "POIDS" in a8 or "PALETTE" in a8:
            return "compte100346"

    return "generique"


def _parse_compte100346(wb, source_filename):
    """
    Format SIFA 010126 - P U (Compte 100346) :
    - Feuille avec A8 = "POIDS" ou "PALETTE"
    - Ligne 10 : bornes basses (DE)
    - Ligne 11 : bornes hautes (A)
    - Ligne 12 : unité (Forfait / Prx/100Kg)
    - Données à partir de la ligne 13
    - Col A : "(XX) NOM DÉPARTEMENT"
    """
    rows = []
    for ws in wb.worksheets:
        a8 = str(ws["A8"].value or "").upper()
        if "POIDS" in a8:
            base_calcul, type_envoi = "poids", "messagerie"
        elif "PALETTE" in a8:
            base_calcul, type_envoi = "palette", "messagerie"
        else:
            continue

        cols = []
        for c in range(3, ws.max_column + 1):
            tmax = _tarif_float(ws.cell(row=11, column=c).value)
            if tmax is None:
                continue
            tmin = _tarif_float(ws.cell(row=10, column=c).value, default=0)
            unite = _tarif_unite_norm(ws.cell(row=12, column=c).value)
            cols.append((c, tmin, tmax, unite))

        for r in range(13, ws.max_row + 1):
            dept = _tarif_dept_from_label(ws.cell(row=r, column=1).value)
            if not dept:
                continue
            for c, tmin, tmax, unite in cols:
                price = _tarif_float(ws.cell(row=r, column=c).value)
                if price is None:
                    continue
                rows.append(
                    {
                        "type_envoi": type_envoi,
                        "base_calcul": base_calcul,
                        "zone_type": "departement",
                        "zone_valeur": dept,
                        "tranche_min": tmin,
                        "tranche_max": int(tmax) if base_calcul == "palette" else tmax,
                        "prix": round(price, 4),
                        "unite": unite,
                        "mini_perception": None,
                        "source_filename": source_filename,
                    }
                )

    return rows, []


def _parse_ceva_messagerie(ws, source_filename):
    rows = []
    header_row = _tarif_find_header_row(
        ws, ["DÉPARTEMENT", "DEPARTEMENT", "ZONE", "CODE POSTAL", "CP"]
    )
    if header_row is None:
        return rows

    cols = []
    for c in range(2, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=c).value or "").strip()
        if not val:
            continue
        m = re.search(r"(\d+(?:[.,]\d+)?)\s*[-àaÀ]\s*(\d+(?:[.,]\d+)?)", val)
        if m:
            tmin = _tarif_float(m.group(1), 0)
            tmax = _tarif_float(m.group(2))
            unite = "forfait" if (tmax is not None and tmax <= 100) else "au_100kg"
            cols.append((c, tmin, tmax, unite))

    for r in range(header_row + 1, ws.max_row + 1):
        zone_lbl = str(ws.cell(row=r, column=1).value or "").strip()
        if not zone_lbl:
            continue
        dept = _tarif_dept_from_label(zone_lbl)
        cp_m = re.match(r"^(\d{5})\b", zone_lbl)
        if dept:
            zone_type, zone_valeur = "departement", dept
        elif cp_m:
            zone_type, zone_valeur = "code_postal", cp_m.group(1)
        else:
            continue
        for c, tmin, tmax, unite in cols:
            if tmin is None:
                continue
            price = _tarif_float(ws.cell(row=r, column=c).value)
            if price is None:
                continue
            rows.append(
                {
                    "type_envoi": "messagerie",
                    "base_calcul": "poids",
                    "zone_type": zone_type,
                    "zone_valeur": zone_valeur,
                    "tranche_min": tmin,
                    "tranche_max": tmax,
                    "prix": round(price, 4),
                    "unite": unite,
                    "mini_perception": None,
                    "source_filename": source_filename,
                }
            )
    return rows


def _parse_ceva_palettes(ws, source_filename):
    rows = []
    header_row = _tarif_find_header_row(ws, ["DÉPARTEMENT", "DEPARTEMENT", "ZONE", "PALETTE", "PAL"])
    if header_row is None:
        return rows
    cols = []
    for c in range(2, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=c).value or "").strip()
        m = re.match(r"^(\d+)\s*(?:palette|pal\.?)?$", val, re.IGNORECASE)
        if m:
            nb = int(m.group(1))
            if 1 <= nb <= 20:
                cols.append((c, nb))
    if not cols:
        cols = [(c, i) for i, c in enumerate(range(2, min(ws.max_column + 1, 7)), start=1)]
    for r in range(header_row + 1, ws.max_row + 1):
        dept = _tarif_dept_from_label(ws.cell(row=r, column=1).value)
        if not dept:
            continue
        for c, nb in cols:
            price = _tarif_float(ws.cell(row=r, column=c).value)
            if price is None:
                continue
            rows.append(
                {
                    "type_envoi": "messagerie",
                    "base_calcul": "palette",
                    "zone_type": "departement",
                    "zone_valeur": dept,
                    "tranche_min": nb,
                    "tranche_max": nb,
                    "prix": round(price, 4),
                    "unite": "forfait",
                    "mini_perception": None,
                    "source_filename": source_filename,
                }
            )
    return rows


def _parse_ceva_frais(ws):
    frais = []
    patterns = [
        (r"gasoil|carburant|fuel", "Gasoil", "pct_transport", 1),
        (r"sûreté|surete|sécurité|securite", "Taxe sûreté/sécurité", "forfait_expedition", 1),
        (r"prise.{0,10}rdv|rendez.{0,5}vous", "Prise de RDV", "forfait_expedition", 0),
        (r"hayon|tail.?lift", "Hayon", "par_palette", 0),
        (r"ville.{0,15}excentr", "Ville excentrée", "forfait_expedition", 0),
        (r"co2|contribution", "CO2", "forfait_expedition", 1),
        (r"centre.{0,10}urbain|urban", "Centres urbains", "forfait_expedition", 0),
    ]
    seen = set()
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(ws.max_column + 1, 10)):
            cell = str(ws.cell(row=r, column=c).value or "").strip()
            if not cell:
                continue
            for pattern, libelle, mode, defaut in patterns:
                if libelle in seen:
                    continue
                if re.search(pattern, cell, re.IGNORECASE):
                    for cc in range(c + 1, min(c + 6, ws.max_column + 1)):
                        val = _tarif_float(ws.cell(row=r, column=cc).value)
                        if val is not None and val > 0:
                            frais.append(
                                {
                                    "libelle": libelle,
                                    "mode": mode,
                                    "valeur": val,
                                    "mini": None,
                                    "applique_defaut": defaut,
                                }
                            )
                            seen.add(libelle)
                            break
                    break
    return frais


def _parse_ceva(wb, source_filename):
    rows = []
    frais = []
    for ws in wb.worksheets:
        t = ws.title.upper().replace(" ", "")
        if "MESSAGERIE" in t or ("TARIF" in t and "GN" in t):
            rows += _parse_ceva_messagerie(ws, source_filename)
        elif "PALETTE" in t or "SMARTPAL" in t or "SMART" in t:
            rows += _parse_ceva_palettes(ws, source_filename)
        elif "CONDITION" in t or "COMMERCIALE" in t or "ANNEXE" in t:
            frais += _parse_ceva_frais(ws)
    return rows, frais


def _parse_transbenelux(wb, source_filename):
    rows = []
    for ws in wb.worksheets:
        header_row = _tarif_find_header_row(ws, ["PALETTE", "PAL", "FRANCE", "DÉPARTEMENT"])
        if header_row is None:
            continue
        cols = []
        for c in range(2, ws.max_column + 1):
            val = str(ws.cell(row=header_row, column=c).value or "").strip()
            m = re.fullmatch(r"(\d{1,2})", val)
            if m:
                nb = int(m.group(1))
                if 1 <= nb <= 20:
                    cols.append((c, nb))
        if not cols:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            dept = _tarif_dept_from_label(ws.cell(row=r, column=1).value)
            if not dept:
                continue
            for c, nb in cols:
                raw = str(ws.cell(row=r, column=c).value or "").strip().upper()
                if raw in ("", "FO", "PU", "PP", "-", "NC", "N/A"):
                    continue
                price = _tarif_float(raw)
                if price is None or price <= 0:
                    continue
                rows.append(
                    {
                        "type_envoi": "affretement" if nb > 6 else "messagerie",
                        "base_calcul": "palette",
                        "zone_type": "departement",
                        "zone_valeur": dept,
                        "tranche_min": nb,
                        "tranche_max": nb,
                        "prix": round(price, 4),
                        "unite": "forfait",
                        "mini_perception": None,
                        "source_filename": source_filename,
                    }
                )
    return rows, []


@router.post("/transporteurs/{transporteur_id}/tarifs/parse-excel")
async def parse_tarif_excel(request: Request, transporteur_id: int):
    """
    Parser déterministe openpyxl pour grilles tarifaires Excel.
    Ne dépend pas de l'API Anthropic — traite les fichiers volumineux (2000+ lignes).
    Formats reconnus : Compte 100346, CEVA Logistics, TRANSBENELUX, générique.
    """
    user = _require_expe_write(request)

    with get_db() as conn:
        trp = conn.execute(
            "SELECT * FROM expe_transporteurs WHERE id=?", (transporteur_id,)
        ).fetchone()
    if not trp:
        raise HTTPException(status_code=404, detail="Transporteur introuvable")
    if not trp["tarif_url"]:
        raise HTTPException(status_code=400, detail="Aucun fichier tarif uploadé pour ce transporteur")

    filepath = _resolve_tarif_path(trp["tarif_url"])
    if not filepath:
        raise HTTPException(status_code=404, detail="Fichier tarif introuvable sur le disque")

    ext = os.path.splitext(filepath)[1].lower()
    if ext not in (".xlsx", ".xls"):
        raise HTTPException(
            status_code=400,
            detail=(
                f"Ce endpoint ne traite que les fichiers Excel (.xlsx). Fichier reçu : {ext}. "
                "Utilisez le bouton 'Parser avec IA' pour les PDFs."
            ),
        )

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=503, detail="openpyxl non installé — lancer : pip install openpyxl"
        )

    wb = openpyxl.load_workbook(filepath, data_only=True)
    source_name = trp["tarif_filename"] or os.path.basename(trp["tarif_url"])

    fmt = _detect_tarif_format(wb)

    if fmt == "compte100346":
        lignes_data, frais_data = _parse_compte100346(wb, source_name)
    elif fmt == "ceva":
        lignes_data, frais_data = _parse_ceva(wb, source_name)
    elif fmt == "transbenelux":
        lignes_data, frais_data = _parse_transbenelux(wb, source_name)
    else:
        structure = []
        for ws in wb.worksheets:
            preview = []
            for r in range(1, min(6, ws.max_row + 1)):
                row_vals = [
                    str(ws.cell(row=r, column=c).value or "")[:30]
                    for c in range(1, min(ws.max_column + 1, 8))
                ]
                preview.append(row_vals)
            structure.append({"sheet": ws.title, "preview": preview})
        raise HTTPException(
            status_code=422,
            detail={
                "message": "Format non reconnu automatiquement. Voici la structure du fichier.",
                "structure": structure,
                "hint": "Communiquer la structure à l'équipe pour ajouter le support de ce format.",
            },
        )

    if not lignes_data:
        raise HTTPException(
            status_code=422,
            detail=f"Format '{fmt}' détecté mais aucune ligne extraite. Vérifier le fichier.",
        )

    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    email = user.get("email") or user.get("identifiant")

    with get_db() as conn:
        for lg in lignes_data:
            conn.execute(
                """INSERT INTO expe_tarifs
                   (transporteur_id, type_envoi, base_calcul, zone_type, zone_valeur,
                    tranche_min, tranche_max, prix, unite, mini_perception,
                    actif, source_filename, created_at, created_by_email)
                   VALUES (?,?,?,?,?,?,?,?,?,?,0,?,?,?)""",
                (
                    transporteur_id,
                    lg.get("type_envoi"),
                    lg.get("base_calcul"),
                    lg.get("zone_type"),
                    lg.get("zone_valeur"),
                    lg.get("tranche_min", 0),
                    lg.get("tranche_max"),
                    lg.get("prix", 0),
                    lg.get("unite"),
                    lg.get("mini_perception"),
                    source_name,
                    now,
                    email,
                ),
            )
        for fr in frais_data:
            conn.execute(
                """INSERT OR IGNORE INTO expe_tarifs_frais
                   (transporteur_id, libelle, mode, valeur, mini, applique_defaut)
                   VALUES (?,?,?,?,?,?)""",
                (
                    transporteur_id,
                    fr.get("libelle"),
                    fr.get("mode"),
                    fr.get("valeur", 0),
                    fr.get("mini"),
                    fr.get("applique_defaut", 1),
                ),
            )
        conn.commit()

    return {
        "format_detecte": fmt,
        "lignes_extraites": len(lignes_data),
        "frais_extraits": len(frais_data),
        "actif": 0,
        "apercu_lignes": lignes_data[:10],
        "message": (
            f"{len(lignes_data)} lignes extraites (format {fmt}) — "
            "à valider avant activation."
        ),
    }


# ─── Comparateur de prix ───────────────────────────────────────────


def _deduire_departement(cp: str) -> str:
    """Déduit le département depuis le code postal (Corse, DOM, cas général)."""
    cp = (cp or "").strip().upper()
    if len(cp) < 2:
        return cp
    if cp.startswith("97") and len(cp) >= 3:
        return cp[:3]
    if cp.startswith("20") and len(cp) == 5 and cp.isdigit():
        num = int(cp)
        return "2A" if num <= 20190 else "2B"
    return cp[:2]


def _trouver_ligne_tarif(
    conn,
    transporteur_id: int,
    type_envoi: str,
    dept: str,
    cp: str,
    poids: float,
    nb_pal: float,
):
    """Cherche la ligne expe_tarifs la plus précise (CP → département, palette → poids)."""
    _MP_PAR_PALETTE = 0.4  # 1 palette 80x120 = 0.4 mètre plancher
    tentatives: list[tuple[str, float]] = []
    if nb_pal > 0:
        tentatives.append(("palette", nb_pal))
        tentatives.append(("metre_plancher", round(nb_pal * _MP_PAR_PALETTE, 4)))
    if poids > 0:
        tentatives.append(("poids", poids))

    zones_par_priorite = [
        ("code_postal", cp),
        ("departement", dept),
    ]

    for base_calcul, valeur_base in tentatives:
        for zone_type, zone_valeur in zones_par_priorite:
            ligne = conn.execute(
                """
                SELECT * FROM expe_tarifs
                WHERE transporteur_id=?
                  AND type_envoi=?
                  AND base_calcul=?
                  AND zone_type=?
                  AND zone_valeur=?
                  AND actif=1
                  AND tranche_min <= ?
                  AND (tranche_max IS NULL OR tranche_max >= ?)
                ORDER BY tranche_min DESC
                LIMIT 1
                """,
                (
                    transporteur_id,
                    type_envoi,
                    base_calcul,
                    zone_type,
                    zone_valeur,
                    valeur_base,
                    valeur_base,
                ),
            ).fetchone()
            if ligne:
                return ligne
    return None


_METHODE_TARIF_LIBELLE: dict[str, str] = {
    "poids": "Tarif au poids",
    "palette": "Tarif à la palette",
    "metre_plancher": "Tarif au mètre plancher",
}


def _trouver_toutes_lignes_tarif(
    conn,
    transporteur_id: int,
    type_envoi: str,
    dept: str,
    cp: str,
    poids: float,
    nb_pal: float,
) -> list[sqlite3.Row]:
    """Collecte toutes les lignes expe_tarifs applicables (palette, MP, poids)."""
    _MP_PAR_PALETTE = 0.4
    tentatives: list[tuple[str, float]] = []
    if nb_pal > 0:
        tentatives.append(("palette", nb_pal))
        tentatives.append(("metre_plancher", round(nb_pal * _MP_PAR_PALETTE, 4)))
    if poids > 0:
        tentatives.append(("poids", poids))

    zones_par_priorite = [
        ("code_postal", cp),
        ("departement", dept),
    ]

    result: list[sqlite3.Row] = []
    for base_calcul, valeur_base in tentatives:
        ligne: sqlite3.Row | None = None
        for zone_type, zone_valeur in zones_par_priorite:
            row = conn.execute(
                """
                SELECT * FROM expe_tarifs
                WHERE transporteur_id=?
                  AND type_envoi=?
                  AND base_calcul=?
                  AND zone_type=?
                  AND zone_valeur=?
                  AND actif=1
                  AND tranche_min <= ?
                  AND (tranche_max IS NULL OR tranche_max >= ?)
                ORDER BY tranche_min DESC
                LIMIT 1
                """,
                (
                    transporteur_id,
                    type_envoi,
                    base_calcul,
                    zone_type,
                    zone_valeur,
                    valeur_base,
                    valeur_base,
                ),
            ).fetchone()
            if row:
                ligne = row
                break
        if ligne:
            result.append(ligne)
    return result


def _calculer_prix_base(ligne, poids: float, nb_pal: float) -> tuple[float, str]:
    """Calcule le prix de base selon l'unité de la ligne tarifaire."""
    unite = ligne["unite"]
    prix = float(ligne["prix"] or 0)
    mini = float(ligne["mini_perception"] or 0)
    base_calcul = ligne["base_calcul"]

    if unite == "forfait":
        prix_calc = prix
        detail = f"forfait {prix:.2f} €"
    elif unite == "au_100kg":
        ref = poids if base_calcul == "poids" else nb_pal
        prix_calc = prix * ref / 100
        detail = f"{prix:.4f} €/100kg × {ref} = {prix_calc:.2f} €"
    elif unite == "au_kg":
        ref = poids if base_calcul == "poids" else nb_pal
        prix_calc = prix * ref
        detail = f"{prix:.4f} €/kg × {ref} = {prix_calc:.2f} €"
    else:
        prix_calc = prix
        detail = f"{prix:.2f} € (unité inconnue : {unite})"

    if mini and prix_calc < mini:
        detail += f" → mini perception {mini:.2f} €"
        prix_calc = mini

    return prix_calc, detail


def _appliquer_frais(
    conn, transporteur_id: int, prix_base: float, nb_pal: float = 0
) -> tuple[list[dict], float]:
    """Applique les frais par défaut du transporteur."""
    frais_rows = conn.execute(
        """
        SELECT * FROM expe_tarifs_frais
        WHERE transporteur_id=? AND applique_defaut=1
        ORDER BY libelle
        """,
        (transporteur_id,),
    ).fetchall()

    frais_list: list[dict] = []
    total_frais = 0.0

    for fr in frais_rows:
        mode = fr["mode"]
        valeur = float(fr["valeur"] or 0)
        mini_fr = float(fr["mini"] or 0)

        if mode == "pct_transport":
            montant = prix_base * valeur / 100
            if mini_fr and montant < mini_fr:
                montant = mini_fr
            detail = f"{valeur}% du transport = {montant:.2f} €"
        elif mode == "forfait_expedition":
            montant = valeur
            detail = f"forfait {valeur:.2f} €"
        elif mode == "par_palette":
            montant = valeur * nb_pal if nb_pal > 0 else valeur
            detail = (
                f"{valeur:.2f} €/pal × {nb_pal} = {montant:.2f} €"
                if nb_pal > 0
                else f"{valeur:.2f} €"
            )
        else:
            montant = valeur
            detail = f"{valeur:.2f} €"

        frais_list.append(
            {
                "libelle": fr["libelle"],
                "montant": round(montant, 2),
                "detail": detail,
            }
        )
        total_frais += montant

    return frais_list, total_frais


def _calculer_comparateur(
    conn,
    poids: float,
    nb_pal: float,
    dept: str,
    cp: str,
    type_envoi: str,
) -> tuple[list[dict], list[dict]]:
    """Éligibilité et prix pour chaque transporteur actif."""
    transporteurs = conn.execute(
        "SELECT * FROM expe_transporteurs WHERE actif=1"
    ).fetchall()

    eligibles: list[dict] = []
    non_eligibles: list[dict] = []

    zone_col = {
        "messagerie": "zone_messagerie",
        "ramasse": "zone_messagerie",
        "affretement": "zone_affretement",
        "express_intl": "zone_france",
    }.get(type_envoi, "zone_france")

    for trp in transporteurs:
        raisons_ineligibilite: list[str] = []

        if not trp[zone_col]:
            raisons_ineligibilite.append(f"hors zone ({type_envoi})")

        pal_max = trp["palette_max"]
        if pal_max is not None and nb_pal > 0 and nb_pal > float(pal_max):
            raisons_ineligibilite.append(
                f"capacité dépassée ({nb_pal:g} pal. > max {pal_max})"
            )

        if trp["accepte_poids"] == 0 and poids > 0 and nb_pal == 0:
            raisons_ineligibilite.append("n'accepte pas le tarif au poids")
        if trp["accepte_palette"] == 0 and nb_pal > 0:
            raisons_ineligibilite.append("n'accepte pas les palettes")

        lignes = _trouver_toutes_lignes_tarif(
            conn, trp["id"], type_envoi, dept, cp, poids, nb_pal
        )
        if not lignes and not raisons_ineligibilite:
            raisons_ineligibilite.append("aucune grille tarifaire pour ce poids/zone")

        if raisons_ineligibilite:
            non_eligibles.append(
                {
                    "transporteur_id": trp["id"],
                    "transporteur": trp["nom"],
                    "raison": " · ".join(raisons_ineligibilite),
                }
            )
            continue

        for ligne in lignes:
            prix_base, detail = _calculer_prix_base(ligne, poids, nb_pal)
            frais_list, prix_frais = _appliquer_frais(conn, trp["id"], prix_base, nb_pal)
            prix_total = prix_base + prix_frais
            base_calcul = ligne["base_calcul"] or ""

            eligibles.append(
                {
                    "transporteur_id": trp["id"],
                    "transporteur": trp["nom"],
                    "prix_ht": round(prix_total, 2),
                    "prix_base_ht": round(prix_base, 2),
                    "methode_tarification": _METHODE_TARIF_LIBELLE.get(
                        base_calcul, base_calcul
                    ),
                    "detail_calcul": {
                        "base": detail,
                        "frais": frais_list,
                    },
                    "delai_jours": None,
                }
            )

    eligibles.sort(key=lambda x: x["prix_ht"])
    if eligibles:
        eligibles[0]["moins_cher"] = True

    return eligibles, non_eligibles


@router.post("/comparateur")
def comparateur(request: Request, body: dict = Body(...)):
    """Calcule le prix de chaque transporteur éligible pour un envoi."""
    _require_expe(request)

    poids = float(body.get("poids_total_kg") or 0)
    nb_pal = float(body.get("nb_palette") or 0)
    cp = str(body.get("code_postal_destination") or "").strip()
    type_envoi = str(body.get("type_envoi") or "messagerie").strip()

    if not cp:
        raise HTTPException(
            status_code=400, detail="code_postal_destination est obligatoire"
        )
    if not poids and not nb_pal:
        raise HTTPException(
            status_code=400,
            detail="Saisir au moins un poids ou un nombre de palettes",
        )

    dept = _deduire_departement(cp)

    with get_db() as conn:
        eligibles, non_eligibles = _calculer_comparateur(
            conn, poids, nb_pal, dept, cp, type_envoi
        )

    return {
        "departement_deduit": dept,
        "eligibles": eligibles,
        "non_eligibles": non_eligibles,
    }


# ─── Demandes de devis (prospection parallèle) ─────────────────────

EXPE_DEVIS_CC = "expeditions@sifa.pro"


_ALLOWED_TYPE_PALETTES = {"europe", "perdue", "autre", "vrac"}


def _normalize_type_palette(value):
    """Renvoie une valeur de type_palette valide (parmi la whitelist) ou None."""
    if value is None:
        return None
    s = str(value).strip().lower()
    if not s:
        return None
    if s not in _ALLOWED_TYPE_PALETTES:
        # Tolerant : on ignore silencieusement une valeur inconnue plutot que
        # de bloquer la creation d'une demande.
        return None
    return s


def _next_demande_reference(conn, year: str) -> str:
    """Renvoie la prochaine référence YYYY-N pour l'année donnée."""
    rows = conn.execute(
        "SELECT reference FROM expe_demandes_devis "
        "WHERE reference LIKE ?",
        (f"{year}-%",),
    ).fetchall()
    n_max = 0
    for r in rows:
        ref = (r[0] or "").strip()
        if "-" in ref:
            _, num_str = ref.split("-", 1)
            try:
                n_max = max(n_max, int(num_str))
            except ValueError:
                continue
    return f"{year}-{n_max + 1}"


def _log_devis(request: Request, user: dict, action: str, objet: str) -> None:
    """Trace d'audit du cycle devis. Ne lève jamais.

    Il n'y avait qu'un seul `log_action` sur tout le module — à la retenue. Ni
    la création, ni l'envoi, ni la saisie d'une réponse, ni la clôture, ni la
    suppression ne laissaient de trace, ce qui se voit le jour où on cherche
    qui a envoyé quoi.
    """
    try:
        log_action(
            user=user,
            action=action,
            module="expe",
            objet=objet,
            ip=request.client.host if request.client else None,
        )
    except Exception:
        pass


def _get_demande_or_404(conn, demande_id: int, *, avec_corbeille: bool = False) -> dict:
    row = conn.execute(
        "SELECT * FROM expe_demandes_devis WHERE id=?", (int(demande_id),)
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Demande introuvable")
    d = dict(row)
    if d.get("deleted_at") and not avec_corbeille:
        raise HTTPException(status_code=404, detail="Demande dans la corbeille")
    return d


def _require_demande_ouverte(conn, demande_id: int) -> dict:
    """Garde serveur : refuse toute action métier sur une demande clôturée.

    L'interface masque déjà les boutons quand `statut !== 'ouverte'`, mais le
    masquage n'est pas un contrôle : un onglet resté ouvert, un double-clic ou
    un retour arrière suffisaient à renvoyer une demande de tarif sur une
    affaire déjà attribuée, ou à créer un second départ pour le même transport.
    """
    d = _get_demande_or_404(conn, demande_id)
    if (d.get("statut") or "") != "ouverte":
        raise HTTPException(
            status_code=409,
            detail="Cette demande est clôturée — rouvrir n'est pas prévu, créer une nouvelle demande.",
        )
    return d


def _valider_date_limite(value: object) -> Optional[str]:
    """Date ISO ou None. Refuse tout le reste plutôt que de le stocker tel quel.

    `_date_prefix` se contente de tronquer : « bientôt » ou « 9999-99-99 »
    passaient et repartaient dans l'email (« Réponse attendue avant le
    bientôt »), puis produisaient un `NaN` dans la pastille d'échéance côté
    front, où la comparaison de chaînes classait la demande en retard au
    hasard.
    """
    txt = _date_prefix(str(value or "").strip())
    if not txt:
        return None
    try:
        d = datetime.strptime(txt, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Date limite invalide (format AAAA-MM-JJ).")
    # Re-sérialisée depuis la date parsée : `strptime` accepte « 2026-2-8 »,
    # qui se compare mal en chaîne avec la date du jour ISO côté front.
    return d.strftime("%Y-%m-%d")


_DEVIS_LANGUES = ("fr", "en")

_DEVIS_STATUT_LABELS = {
    "envoyee": "Envoyée",
    "ouvert": "Ouverte",
    "recue": "Reçue",
    "retenue": "Retenue",
    "refusee": "Refusée",
    "echec": "Échec envoi",
}


def _normalize_langue(value: object, defaut: str = "fr") -> str:
    lg = str(value or "").strip().lower()[:2]
    return lg if lg in _DEVIS_LANGUES else defaut


def _langue_connue_pour_email(conn, email: str) -> str:
    """Langue déjà employée pour cette adresse, chaîne vide si inconnue.

    Sert au destinataire saisi à la main : si l'adresse appartient en réalité à
    un transporteur du référentiel, on lui réécrit dans SA langue plutôt que de
    repasser en bilingue.
    """
    mail = (email or "").strip().lower()
    if not mail:
        return ""
    row = conn.execute(
        """SELECT langue FROM expe_devis_reponses
           WHERE LOWER(TRIM(COALESCE(destinataire_email,''))) = ?
             AND COALESCE(TRIM(langue),'') <> ''
           ORDER BY id DESC LIMIT 1""",
        (mail,),
    ).fetchone()
    if row:
        return _normalize_langue(row["langue"], "")
    row = conn.execute(
        """SELECT langue FROM expe_transporteurs
           WHERE LOWER(TRIM(COALESCE(contact_email,''))) = ?
              OR LOWER(COALESCE(contact_emails,'')) LIKE ?
           LIMIT 1""",
        (mail, f'%"{mail}"%'),
    ).fetchone()
    return _normalize_langue(row["langue"], "") if row else ""


def _devis_langue_destinataire(conn, rep: dict) -> str:
    """Langue d'écriture d'un destinataire.

    Priorité à la langue figée sur la ligne de réponse : elle dit ce qui a
    réellement été envoyé. À défaut — première sollicitation, ou ligne
    antérieure à la colonne — on lit le référentiel transporteur.
    """
    fige = _normalize_langue(rep.get("langue"), "")
    if fige:
        return fige
    tid = rep.get("transporteur_id")
    if tid:
        row = conn.execute(
            "SELECT langue FROM expe_transporteurs WHERE id=?", (int(tid),)
        ).fetchone()
        if row:
            return _normalize_langue(row["langue"])
    return "fr"


@router.get("/devis/clients-suggestions")
def clients_suggestions_devis(request: Request, q: Optional[str] = Query(None)):
    """Noms de clients déjà employés — alimente la liste déroulante de saisie.

    On agrège trois sources plutôt que de brancher sur la seule table `clients` :
    le référentiel ne contient pas les destinataires ponctuels, et l'historique
    des devis et des départs porte l'orthographe réellement utilisée. Le but
    n'est pas de contraindre la saisie mais d'éviter qu'un même client existe
    en trois graphies.
    """
    _require_expe(request)
    like = f"%{(q or '').strip()}%"
    with get_db() as conn:
        rows = conn.execute(
            """SELECT nom, MAX(recent) AS recent, SUM(n) AS n FROM (
                   SELECT TRIM(client) AS nom, MAX(created_at) AS recent, COUNT(*) AS n
                   FROM expe_demandes_devis
                   WHERE COALESCE(TRIM(client),'') <> '' AND deleted_at IS NULL
                   GROUP BY TRIM(client)
                 UNION ALL
                   SELECT TRIM(client), MAX(created_at), COUNT(*)
                   FROM expe_departs
                   WHERE COALESCE(TRIM(client),'') <> ''
                   GROUP BY TRIM(client)
               )
               WHERE (? = '%%' OR nom LIKE ? COLLATE NOCASE)
               GROUP BY nom COLLATE NOCASE
               ORDER BY n DESC, recent DESC
               LIMIT 200""",
            (like, like),
        ).fetchall()
    return {"clients": [r["nom"] for r in rows if r["nom"]]}


@router.post("/devis/demandes")
def creer_demande_devis(request: Request, body: dict = Body(...)):
    user = _require_expe_write(request)
    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    email = (user.get("email") or user.get("identifiant") or "").strip() or None
    year = now[:4]
    client = (body.get("client") or "").strip() or None
    with get_db() as conn:
        reference = _next_demande_reference(conn, year)
        cur = conn.execute(
            """
            INSERT INTO expe_demandes_devis
            (depart_id, poids_total_kg, nb_palette, code_postal_destination,
             type_envoi, type_palette, contraintes, statut, created_at,
             created_by_email, reference, client, date_limite)
            VALUES (?,?,?,?,?,?,?,'ouverte',?,?,?,?,?)
            """,
            (
                body.get("depart_id"),
                body.get("poids_total_kg"),
                body.get("nb_palette"),
                (body.get("code_postal_destination") or "").strip(),
                (body.get("type_envoi") or "messagerie").strip(),
                _normalize_type_palette(body.get("type_palette")),
                (body.get("contraintes") or "").strip() or None,
                now,
                email,
                reference,
                client,
                _valider_date_limite(body.get("date_limite")),
            ),
        )
        conn.commit()
        demande = conn.execute(
            "SELECT * FROM expe_demandes_devis WHERE id=?", (cur.lastrowid,)
        ).fetchone()
    _log_devis(request, user, "CREATE", f"Demande devis {reference} · {client or '—'}")
    return dict(demande)


# Champs d'entête modifiables après création. `reference`, `statut`,
# `created_at` et `created_by_email` en sont volontairement absents : ce sont
# des faits, pas des paramètres.
_DEVIS_CHAMPS_EDITABLES = (
    "client",
    "code_postal_destination",
    "type_envoi",
    "poids_total_kg",
    "nb_palette",
    "contraintes",
)


@router.put("/devis/demandes/{demande_id}")
def modifier_demande_devis(request: Request, demande_id: int, body: dict = Body(...)):
    """Corrige l'entête d'une demande tant qu'aucun email n'est parti.

    La demande était figée à la création : une coquille dans le poids ou le code
    postal obligeait à supprimer et recréer, ce qui perdait la référence et les
    réponses déjà reçues. On autorise donc la correction, mais seulement avant
    le premier envoi : après, les transporteurs ont chiffré sur des données
    qu'on n'a plus le droit de changer dans leur dos.
    """
    user = _require_expe_write(request)
    sets, args = [], []
    with get_db() as conn:
        demande = _require_demande_ouverte(conn, demande_id)
        deja_envoye = conn.execute(
            "SELECT 1 FROM expe_devis_reponses WHERE demande_id=? AND sent_at IS NOT NULL LIMIT 1",
            (demande_id,),
        ).fetchone()
        if deja_envoye:
            raise HTTPException(
                status_code=409,
                detail=(
                    "Des demandes de tarif sont déjà parties : l'entête ne peut plus "
                    "changer. Seule la date limite reste modifiable."
                ),
            )
        for champ in _DEVIS_CHAMPS_EDITABLES:
            if champ not in body:
                continue
            val = body.get(champ)
            if champ in ("poids_total_kg", "nb_palette"):
                try:
                    val = float(val) if val not in (None, "") else None
                except (TypeError, ValueError):
                    raise HTTPException(status_code=400, detail=f"{champ} invalide.")
            else:
                val = (str(val or "").strip()) or None
            sets.append(f"{champ}=?")
            args.append(val)
        if "type_palette" in body:
            sets.append("type_palette=?")
            args.append(_normalize_type_palette(body.get("type_palette")))
        if "date_limite" in body:
            sets.append("date_limite=?")
            args.append(_valider_date_limite(body.get("date_limite")))
        if not sets:
            raise HTTPException(status_code=400, detail="Aucun champ à mettre à jour.")
        args.append(demande_id)
        conn.execute(
            f"UPDATE expe_demandes_devis SET {', '.join(sets)} WHERE id=?", args
        )
        conn.commit()
        row = conn.execute(
            "SELECT * FROM expe_demandes_devis WHERE id=?", (demande_id,)
        ).fetchone()
    _log_devis(
        request, user, "UPDATE", f"Demande devis {demande.get('reference') or demande_id}"
    )
    return dict(row)


@router.patch("/devis/demandes/{demande_id}/date-limite")
def modifier_date_limite_devis(request: Request, demande_id: int, body: dict = Body(...)):
    """Repousse la date limite, même après envoi.

    Séparé du PUT ci-dessus à dessein : décaler une échéance ne change pas ce
    sur quoi les transporteurs ont chiffré, c'est la seule modification qui
    reste légitime une fois les demandes parties.
    """
    user = _require_expe_write(request)
    dl = _valider_date_limite(body.get("date_limite"))
    with get_db() as conn:
        demande = _require_demande_ouverte(conn, demande_id)
        conn.execute(
            "UPDATE expe_demandes_devis SET date_limite=? WHERE id=?", (dl, demande_id)
        )
        conn.commit()
    _log_devis(
        request,
        user,
        "UPDATE",
        f"Demande devis {demande.get('reference') or demande_id} · date limite {dl or '—'}",
    )
    return {"ok": True, "date_limite": dl}


@router.post("/devis/demandes/{demande_id}/dupliquer")
def dupliquer_demande_devis(request: Request, demande_id: int, body: dict = Body(default={})):
    """Crée une demande neuve à partir d'une demande passée.

    Les demandes de tarif se répètent : même client, même destination, même
    gabarit, un mois plus tard. On copie donc l'entête et les pièces jointes,
    et on renvoie la liste des transporteurs sollicités la fois précédente pour
    que l'interface les re-coche — mais on ne recrée AUCUNE ligne de réponse :
    une demande dupliquée n'a encore été envoyée à personne, et pré-remplir des
    lignes vides ferait croire à des envois qui n'ont pas eu lieu.
    """
    user = _require_expe_write(request)
    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    email = (user.get("email") or user.get("identifiant") or "").strip() or None
    copier_pj = bool(body.get("copier_pieces_jointes", True))
    with get_db() as conn:
        src = _get_demande_or_404(conn, demande_id, avec_corbeille=True)
        reference = _next_demande_reference(conn, now[:4])
        cur = conn.execute(
            """INSERT INTO expe_demandes_devis
               (depart_id, poids_total_kg, nb_palette, code_postal_destination,
                type_envoi, type_palette, contraintes, statut, created_at,
                created_by_email, reference, client, date_limite)
               VALUES (NULL,?,?,?,?,?,?,'ouverte',?,?,?,?,NULL)""",
            (
                src.get("poids_total_kg"),
                src.get("nb_palette"),
                src.get("code_postal_destination"),
                src.get("type_envoi"),
                src.get("type_palette"),
                src.get("contraintes"),
                now,
                email,
                reference,
                src.get("client"),
            ),
        )
        new_id = int(cur.lastrowid)
        # `depart_id` et `date_limite` volontairement non copiés : le départ
        # d'origine n'a rien à voir avec le nouvel envoi, et une échéance
        # recopiée serait déjà passée.
        if copier_pj:
            for pj in conn.execute(
                """SELECT filename, path, taille_octets FROM expe_devis_pieces_jointes
                   WHERE demande_id=? AND origine='sifa'""",
                (demande_id,),
            ).fetchall():
                # On référence le MÊME fichier sur disque plutôt que de le
                # dupliquer : c'est le même document, et une copie doublerait
                # l'espace pour un contenu identique. La purge définitive en
                # tient compte (cf. _purger_demande).
                conn.execute(
                    """INSERT INTO expe_devis_pieces_jointes
                       (demande_id, reponse_id, origine, filename, path,
                        taille_octets, created_at, created_by_email)
                       VALUES (?,NULL,'sifa',?,?,?,?,?)""",
                    (new_id, pj["filename"], pj["path"], pj["taille_octets"], now, email),
                )
        trps = [
            {
                "transporteur_id": r["transporteur_id"],
                "nom": r["nom_transporteur"],
                "email": r["destinataire_email"],
            }
            for r in conn.execute(
                """SELECT DISTINCT transporteur_id, nom_transporteur, destinataire_email
                   FROM expe_devis_reponses WHERE demande_id=?""",
                (demande_id,),
            ).fetchall()
        ]
        conn.commit()
        row = conn.execute(
            "SELECT * FROM expe_demandes_devis WHERE id=?", (new_id,)
        ).fetchone()
    _log_devis(
        request,
        user,
        "CREATE",
        f"Demande devis {reference} dupliquée depuis {src.get('reference') or demande_id}",
    )
    out = dict(row)
    out["destinataires_precedents"] = trps
    return out


# Répertoire pour les pièces jointes des demandes de devis.
# Chemins relatifs stockés en DB ; le path absolu se construit avec BASE_DIR.
_DEVIS_UPLOAD_SUBDIR = "uploads/devis"


def _devis_upload_dir() -> Path:
    from config import BASE_DIR
    d = Path(BASE_DIR) / _DEVIS_UPLOAD_SUBDIR
    d.mkdir(parents=True, exist_ok=True)
    return d


def _devis_safe_filename(name: str) -> str:
    """Conserve la lettres/chiffres + . _ - et remplace le reste par _ ."""
    name = unicodedata.normalize("NFKD", name or "").encode("ascii", "ignore").decode("ascii")
    name = re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("._-")
    return name or "fichier"


@router.post("/devis/demandes/{demande_id}/piece-jointe")
async def upload_demande_devis_piece_jointe(
    request: Request,
    demande_id: int,
    file: UploadFile = File(...),
):
    _require_expe_write(request)
    # Limite raisonnable : 20 Mo
    MAX_BYTES = 20 * 1024 * 1024
    contents = await file.read()
    if len(contents) > MAX_BYTES:
        raise HTTPException(413, "Fichier trop volumineux (max 20 Mo).")
    orig = (file.filename or "fichier").strip()
    safe = _devis_safe_filename(orig)
    # Préfixe avec l'id de la demande + uuid court pour éviter les collisions
    unique = f"{demande_id}_{uuid.uuid4().hex[:8]}_{safe}"
    path_abs = _devis_upload_dir() / unique
    with open(path_abs, "wb") as out:
        out.write(contents)
    rel = f"{_DEVIS_UPLOAD_SUBDIR}/{unique}"
    with get_db() as conn:
        cur = conn.execute(
            "UPDATE expe_demandes_devis SET piece_jointe_path=?, piece_jointe_filename=? WHERE id=?",
            (rel, orig, demande_id),
        )
        conn.commit()
        if cur.rowcount == 0:
            try:
                path_abs.unlink()
            except Exception:
                pass
            raise HTTPException(404, "Demande introuvable.")
    return {"ok": True, "path": rel, "filename": orig}


@router.get("/devis/demandes/{demande_id}/piece-jointe")
def download_demande_devis_piece_jointe(request: Request, demande_id: int):
    get_current_user(request)
    with get_db() as conn:
        row = conn.execute(
            "SELECT piece_jointe_path, piece_jointe_filename FROM expe_demandes_devis WHERE id=?",
            (demande_id,),
        ).fetchone()
    if not row or not row["piece_jointe_path"]:
        raise HTTPException(404, "Pas de pièce jointe.")
    from config import BASE_DIR
    path_abs = Path(BASE_DIR) / row["piece_jointe_path"]
    if not path_abs.exists():
        raise HTTPException(404, "Fichier introuvable sur le disque.")
    return FileResponse(
        path=str(path_abs),
        filename=row["piece_jointe_filename"] or path_abs.name,
    )


_DEVIS_PJ_MAX_BYTES = 20 * 1024 * 1024


async def _devis_ecrire_pj(
    conn,
    *,
    demande_id: int,
    file: UploadFile,
    origine: str,
    reponse_id: Optional[int],
    email: Optional[str],
    now: str,
) -> dict:
    """Écrit un fichier sur disque et l'enregistre. Commun aux deux origines."""
    contents = await file.read()
    if len(contents) > _DEVIS_PJ_MAX_BYTES:
        raise HTTPException(413, "Fichier trop volumineux (max 20 Mo).")
    if not contents:
        raise HTTPException(400, "Fichier vide.")
    orig = (file.filename or "fichier").strip()
    unique = f"{demande_id}_{uuid.uuid4().hex[:8]}_{_devis_safe_filename(orig)}"
    path_abs = _devis_upload_dir() / unique
    with open(path_abs, "wb") as out:
        out.write(contents)
    rel = f"{_DEVIS_UPLOAD_SUBDIR}/{unique}"
    cur = conn.execute(
        """INSERT INTO expe_devis_pieces_jointes
           (demande_id, reponse_id, origine, filename, path, taille_octets,
            created_at, created_by_email)
           VALUES (?,?,?,?,?,?,?,?)""",
        (demande_id, reponse_id, origine, orig, rel, len(contents), now, email),
    )
    return {
        "id": int(cur.lastrowid),
        "filename": orig,
        "path": rel,
        "taille_octets": len(contents),
        "origine": origine,
    }


def _devis_pj_list(conn, demande_id: int) -> list[dict]:
    return [
        dict(r)
        for r in conn.execute(
            """SELECT id, demande_id, reponse_id, origine, filename,
                      taille_octets, created_at, created_by_email
               FROM expe_devis_pieces_jointes
               WHERE demande_id=? ORDER BY origine, id""",
            (int(demande_id),),
        ).fetchall()
    ]


@router.post("/devis/demandes/{demande_id}/pieces-jointes")
async def upload_devis_pieces_jointes(
    request: Request, demande_id: int, files: list[UploadFile] = File(...)
):
    """Ajoute une ou plusieurs pièces jointes à une demande.

    Remplace l'upload à colonne unique, qui écrasait silencieusement le fichier
    précédent : le second document ne prévenait pas qu'il chassait le premier.
    """
    user = _require_expe_write(request)
    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    email = (user.get("email") or user.get("identifiant") or "").strip() or None
    out = []
    with get_db() as conn:
        demande = _require_demande_ouverte(conn, demande_id)
        for f in files:
            out.append(
                await _devis_ecrire_pj(
                    conn,
                    demande_id=demande_id,
                    file=f,
                    origine="sifa",
                    reponse_id=None,
                    email=email,
                    now=now,
                )
            )
        conn.commit()
    _log_devis(
        request,
        user,
        "CREATE",
        f"Demande devis {demande.get('reference') or demande_id} · {len(out)} pièce(s) jointe(s)",
    )
    return {"ok": True, "pieces_jointes": out}


@router.get("/devis/demandes/{demande_id}/pieces-jointes")
def list_devis_pieces_jointes(request: Request, demande_id: int):
    _require_expe(request)
    with get_db() as conn:
        _get_demande_or_404(conn, demande_id, avec_corbeille=True)
        return {"pieces_jointes": _devis_pj_list(conn, demande_id)}


@router.get("/devis/pieces-jointes/{pj_id}")
def download_devis_piece_jointe(request: Request, pj_id: int):
    get_current_user(request)
    with get_db() as conn:
        row = conn.execute(
            "SELECT filename, path FROM expe_devis_pieces_jointes WHERE id=?", (pj_id,)
        ).fetchone()
    if not row:
        raise HTTPException(404, "Pièce jointe introuvable.")
    from config import BASE_DIR

    path_abs = Path(BASE_DIR) / row["path"]
    if not path_abs.exists():
        raise HTTPException(404, "Fichier introuvable sur le disque.")
    return FileResponse(path=str(path_abs), filename=row["filename"] or path_abs.name)


@router.delete("/devis/pieces-jointes/{pj_id}")
def delete_devis_piece_jointe(request: Request, pj_id: int):
    user = _require_expe_write(request)
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM expe_devis_pieces_jointes WHERE id=?", (pj_id,)
        ).fetchone()
        if not row:
            raise HTTPException(404, "Pièce jointe introuvable.")
        chemin = row["path"]
        conn.execute("DELETE FROM expe_devis_pieces_jointes WHERE id=?", (pj_id,))
        # Le fichier n'est effacé du disque que s'il n'est plus référencé :
        # la duplication d'une demande partage le même chemin, supprimer sans
        # vérifier viderait la pièce jointe de la copie.
        encore = conn.execute(
            "SELECT 1 FROM expe_devis_pieces_jointes WHERE path=? LIMIT 1", (chemin,)
        ).fetchone()
        conn.commit()
    if not encore:
        try:
            from config import BASE_DIR

            (Path(BASE_DIR) / chemin).unlink()
        except Exception:
            pass
    _log_devis(request, user, "DELETE", f"Pièce jointe devis #{pj_id} ({row['filename']})")
    return {"ok": True}


@router.get("/devis/reponses/{reponse_id}/retention-fichier")
def download_retention_fichier(request: Request, reponse_id: int):
    """Telecharge le fichier joint lors de la retenue d'une offre (si present)."""
    get_current_user(request)
    with get_db() as conn:
        row = conn.execute(
            """SELECT retention_file_path, retention_file_filename
                 FROM expe_devis_reponses WHERE id=?""",
            (reponse_id,),
        ).fetchone()
    if not row or not row["retention_file_path"]:
        raise HTTPException(404, "Aucun fichier de retenue pour cette reponse.")
    from config import BASE_DIR
    path_abs = Path(BASE_DIR) / row["retention_file_path"]
    if not path_abs.exists():
        raise HTTPException(404, "Fichier introuvable sur le disque.")
    return FileResponse(
        path=str(path_abs),
        filename=row["retention_file_filename"] or path_abs.name,
    )


@router.get("/devis/demandes")
def list_demandes_devis(request: Request, statut: str = "ouverte"):
    _require_expe(request)
    with get_db() as conn:
        # `deleted_at IS NULL` partout sauf sur la corbeille : une demande
        # supprimée ne doit réapparaître dans aucun autre filtre, pas même
        # « toutes » — sinon la corbeille ne sert à rien.
        if statut == "corbeille":
            rows = conn.execute(
                "SELECT * FROM expe_demandes_devis WHERE deleted_at IS NOT NULL "
                "ORDER BY deleted_at DESC LIMIT 100"
            ).fetchall()
        elif statut == "toutes":
            rows = conn.execute(
                "SELECT * FROM expe_demandes_devis WHERE deleted_at IS NULL "
                "ORDER BY created_at DESC LIMIT 100"
            ).fetchall()
        elif statut == "historique":
            # Toutes les demandes clôturées (manuellement ou via retenue).
            rows = conn.execute(
                "SELECT * FROM expe_demandes_devis WHERE statut='cloturee' AND deleted_at IS NULL "
                "ORDER BY created_at DESC LIMIT 100"
            ).fetchall()
        else:
            rows = conn.execute(
                """SELECT * FROM expe_demandes_devis WHERE statut=? AND deleted_at IS NULL
                   ORDER BY created_at DESC LIMIT 100""",
                (statut,),
            ).fetchall()
        result = []
        for d in rows:
            dd = dict(d)
            counts = conn.execute(
                """
                SELECT
                  SUM(CASE WHEN statut IN ('envoyee','ouvert','recue','retenue','refusee')
                      THEN 1 ELSE 0 END) AS envoyes,
                  SUM(CASE WHEN statut IN ('recue','retenue') THEN 1 ELSE 0 END) AS recues,
                  SUM(CASE WHEN statut='retenue' THEN 1 ELSE 0 END) AS retenues
                FROM expe_devis_reponses WHERE demande_id=?
                """,
                (dd["id"],),
            ).fetchone()
            dd["nb_envoyes"] = counts["envoyes"] or 0
            dd["nb_recus"] = counts["recues"] or 0
            dd["nb_retenus"] = counts["retenues"] or 0
            pj = conn.execute(
                "SELECT COUNT(*) AS n FROM expe_devis_pieces_jointes WHERE demande_id=?",
                (dd["id"],),
            ).fetchone()
            dd["nb_pieces_jointes"] = int(pj["n"] or 0)
            result.append(dd)
    return result


@router.get("/devis/demandes/{demande_id}")
def get_demande_devis(request: Request, demande_id: int):
    _require_expe(request)
    with get_db() as conn:
        demande = conn.execute(
            "SELECT * FROM expe_demandes_devis WHERE id=?", (demande_id,)
        ).fetchone()
        if not demande:
            raise HTTPException(status_code=404, detail="Demande introuvable")
        reponses = conn.execute(
            """SELECT * FROM expe_devis_reponses WHERE demande_id=?
               ORDER BY sent_at""",
            (demande_id,),
        ).fetchall()
        engagement = expe_ev.resume_par_reponse(conn, demande_id)
        pieces_jointes = _devis_pj_list(conn, demande_id)

    reponses_out = []
    for r in reponses:
        d = dict(r)
        # Le token de pixel ne sort jamais de l'API : c'est un identifiant de
        # suivi, il n'a rien à faire dans le JSON d'une page d'administration.
        d.pop("token_pixel", None)
        d["engagement"] = engagement.get(int(d["id"]), {})
        reponses_out.append(d)
    return {
        "demande": dict(demande),
        "reponses": reponses_out,
        "pieces_jointes": pieces_jointes,
    }


@router.get("/devis/reponses/{reponse_id}/evenements")
def evenements_reponse_devis(request: Request, reponse_id: int):
    """Timeline d'engagement d'un transporteur sur une demande de tarif."""
    _require_expe(request)
    with get_db() as conn:
        rep = conn.execute(
            """SELECT id, demande_id, nom_transporteur, destinataire_email, sent_at
               FROM expe_devis_reponses WHERE id=?""",
            (int(reponse_id),),
        ).fetchone()
        if not rep:
            raise HTTPException(status_code=404, detail="Réponse introuvable")
        evts = expe_ev.timeline(conn, int(reponse_id))
    return {
        "destinataire": {
            "id": int(rep["id"]),
            "nom": rep["nom_transporteur"],
            "email": rep["destinataire_email"],
            "sent_at": rep["sent_at"],
        },
        "evenements": evts,
    }


def _pr(v: object) -> str:
    """Échappement HTML pour la page d'impression."""
    return (
        str(v if v is not None else "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


@router.get("/devis/demandes/{demande_id}/imprimer", response_class=HTMLResponse)
def imprimer_comparatif_devis(request: Request, demande_id: int):
    """Comparatif des offres, en page A4 prête à imprimer.

    Rendu serveur plutôt qu'impression du DOM applicatif : la page de MyExpé
    est une SPA sombre avec sidebar et modales, dont l'impression donne un
    résultat illisible. Ici le document est autonome — on l'ouvre, on fait
    Ctrl+P, on le classe ou on le fait viser.

    Pas de dépendance PDF : le navigateur sait produire un PDF depuis une page
    HTML, ajouter reportlab pour un tableau de six colonnes serait payer une
    dépendance pour un problème que le poste de travail résout déjà.
    """
    _require_expe(request)
    with get_db() as conn:
        demande = _get_demande_or_404(conn, demande_id, avec_corbeille=True)
        reponses = [
            dict(r)
            for r in conn.execute(
                """SELECT * FROM expe_devis_reponses WHERE demande_id=?
                   ORDER BY CASE WHEN prix IS NULL THEN 1 ELSE 0 END, prix ASC, id ASC""",
                (demande_id,),
            ).fetchall()
        ]
    # `float()` défensif : une réponse saisie en interne a pu recevoir un prix
    # non numérique avant que la validation n'existe. Un comparatif qui plante
    # en 500 est pire qu'un comparatif qui ignore une ligne aberrante.
    def _prix_num(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    prix = [p for p in (_prix_num(r.get("prix")) for r in reponses) if p is not None]
    best = min(prix) if prix else None
    ref = demande.get("reference") or f"#{demande_id}"

    lignes = []
    for r in reponses:
        st = (r.get("statut") or "").strip()
        p = _prix_num(r.get("prix"))
        d_j = r.get("delai_jours")
        try:
            d_j = int(d_j) if d_j is not None else None
        except (TypeError, ValueError):
            d_j = None
        ecart = ""
        cls = ""
        # `best is not None` et non `best` : une offre gratuite (0 €, cas d'un
        # transport inclus) rendait le test faux et faisait disparaître toute
        # la colonne Écart du comparatif.
        if p is not None and best is not None:
            if p == best:
                cls = " class='best'"
                ecart = "référence"
            elif best > 0:
                # L'écart relatif, pas seulement le meilleur en gras : savoir
                # que le deuxième est à 3 % ou à 40 % ne se lit pas dans deux
                # nombres bruts, et c'est pourtant ce qui décide.
                ecart = f"+{(p - best) / best * 100:.1f} %"
            else:
                ecart = f"+{p - best:.2f} €"
        lignes.append(
            f"<tr{cls}>"
            f"<td class='nom'><span class='n'>{_pr(r.get('nom_transporteur') or '—')}</span>"
            + (
                f"<div class='sub'>{_pr(r.get('destinataire_email') or '')}</div>"
                if r.get("destinataire_email")
                else ""
            )
            + "</td>"
            f"<td>{_pr(_DEVIS_STATUT_LABELS.get(st, st or '—'))}</td>"
            f"<td class='num'>{('%.2f €' % p) if p is not None else '—'}</td>"
            f"<td class='num'>{_pr(ecart)}</td>"
            f"<td class='num'>{('J+%d' % d_j) if d_j is not None else '—'}</td>"
            f"<td class='cmt'>{_pr(r.get('commentaire') or '')}</td>"
            "</tr>"
        )
    if not lignes:
        lignes.append("<tr><td colspan='6' class='vide'>Aucun destinataire sollicité.</td></tr>")

    entete = [
        ("Client", demande.get("client")),
        ("Destination (CP)", demande.get("code_postal_destination")),
        ("Type d'envoi", demande.get("type_envoi")),
        ("Poids total", f"{demande['poids_total_kg']} kg" if demande.get("poids_total_kg") else None),
        ("Palettes", demande.get("nb_palette")),
        ("Date limite", demande.get("date_limite")),
        ("Créée le", (demande.get("created_at") or "")[:10]),
        ("Demandeur", demande.get("created_by_email")),
    ]
    entete_html = "".join(
        f"<div class='f'><dt>{_pr(k)}</dt><dd>{_pr(v)}</dd></div>"
        for k, v in entete
        if v not in (None, "")
    )
    contraintes = (demande.get("contraintes") or "").strip()
    retenu = next((r for r in reponses if (r.get("statut") or "") == "retenue"), None)
    bandeau = ""
    if retenu:
        bandeau = (
            "<div class='retenu'><strong>Offre retenue :</strong> "
            f"{_pr(retenu.get('nom_transporteur'))}"
            + (f" — {float(retenu['prix']):.2f} €" if retenu.get("prix") is not None else "")
            + "</div>"
        )

    html = f"""<!DOCTYPE html><html lang="fr"><head><meta charset="utf-8">
<title>Comparatif devis {_pr(ref)} — SIFA</title>
<style>
@page{{size:A4 portrait;margin:14mm 13mm}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:"Segoe UI",Arial,sans-serif;color:#0f172a;font-size:10pt;line-height:1.45;
  -webkit-print-color-adjust:exact;print-color-adjust:exact;padding:6mm}}
header{{border-bottom:2px solid #0f172a;padding-bottom:4mm;margin-bottom:5mm}}
.brand{{font-size:8pt;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:#0891b2}}
h1{{font-size:17pt;font-weight:800;margin-top:1mm;letter-spacing:-.3px}}
.meta{{font-size:9pt;color:#64748b;margin-top:1.5mm}}
dl{{display:flex;flex-wrap:wrap;gap:3mm 8mm;margin:4mm 0 5mm}}
.f{{min-width:34mm}}
dt{{font-size:7.5pt;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:#94a3b8}}
dd{{font-size:10pt;font-weight:600;margin-top:.5mm}}
.contraintes{{background:#f8fafc;border-left:3px solid #cbd5e1;padding:2.5mm 4mm;margin-bottom:5mm;font-size:9.5pt}}
.retenu{{background:#ecfdf5;border:1px solid #059669;color:#065f46;border-radius:2mm;
  padding:2.5mm 4mm;margin-bottom:4mm;font-size:10pt}}
table{{width:100%;border-collapse:collapse}}
th{{font-size:7.5pt;text-transform:uppercase;letter-spacing:.5px;color:#64748b;text-align:left;
  border-bottom:1.5px solid #0f172a;padding:2mm 2.5mm}}
td{{padding:2.4mm 2.5mm;border-bottom:1px solid #e2e8f0;font-size:9.5pt;vertical-align:top}}
.num{{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap}}
th.num{{text-align:right}}
.nom{{font-weight:700}}
.sub{{font-size:8pt;font-weight:400;color:#94a3b8}}
.cmt{{font-size:8.5pt;color:#475569}}
td:nth-child(2){{white-space:nowrap}}
tr.best td{{background:#ecfdf5}}
tr.best .n::after{{content:" ◆";color:#059669}}
.vide{{color:#94a3b8;text-align:center;padding:8mm}}
footer{{margin-top:6mm;padding-top:3mm;border-top:1px solid #e2e8f0;font-size:7.5pt;color:#94a3b8;
  display:flex;justify-content:space-between}}
@media screen{{body{{background:#e2e8f0}}
  .sheet{{background:#fff;max-width:190mm;margin:8mm auto;padding:12mm;box-shadow:0 2px 16px rgba(0,0,0,.18)}}}}
@media print{{.sheet{{padding:0;margin:0;box-shadow:none}} .noprint{{display:none}}}}
</style></head><body><div class="sheet">
<header>
  <div class="brand">SIFA · MyExpé</div>
  <h1>Comparatif des offres transport — {_pr(ref)}</h1>
  <div class="meta">{len(reponses)} destinataire(s) sollicité(s) · {len(prix)} offre(s) chiffrée(s)</div>
</header>
<dl>{entete_html}</dl>
{f"<div class='contraintes'><strong>Contraintes :</strong> {_pr(contraintes)}</div>" if contraintes else ""}
{bandeau}
<table>
  <thead><tr><th>Transporteur</th><th>Statut</th><th class="num">Prix HT</th>
  <th class="num">Écart</th><th class="num">Délai</th><th>Commentaire</th></tr></thead>
  <tbody>{''.join(lignes)}</tbody>
</table>
<footer><span>Document interne SIFA — comparatif à la date d'impression.</span>
<span>{_pr(ref)}</span></footer>
</div>
<script>window.addEventListener('load',function(){{setTimeout(function(){{window.print();}},250);}});</script>
</body></html>"""
    return HTMLResponse(html)


@router.post("/devis/demandes/{demande_id}/envoyer")
def envoyer_rfq(request: Request, demande_id: int, body: dict = Body(...)):
    user = _require_expe_write(request)
    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    reply_to = (user.get("email") or user.get("identifiant") or "").strip() or None

    with get_db() as conn:
        # Garde serveur : sans elle, un onglet resté ouvert renvoyait la
        # demande sur une affaire déjà attribuée.
        demande = _require_demande_ouverte(conn, demande_id)

        destinataires: list[dict] = []
        trp_ids = body.get("transporteur_ids") or []
        if trp_ids:
            placeholders = ",".join("?" * len(trp_ids))
            trps = conn.execute(
                f"""SELECT id, nom, contact_email, contact_emails, langue FROM expe_transporteurs
                    WHERE id IN ({placeholders}) AND actif=1""",
                trp_ids,
            ).fetchall()
            for t in trps:
                addrs = _normalize_emails(t["contact_emails"])
                if not addrs:
                    fallback = (t["contact_email"] or "").strip()
                    if fallback and "@" in fallback:
                        addrs = [fallback]
                for email_addr in addrs:
                    destinataires.append(
                        {
                            "transporteur_id": t["id"],
                            "nom": t["nom"],
                            "email": email_addr,
                            "langue": _normalize_langue(t["langue"]),
                        }
                    )

        for extra in body.get("transporteur_extras") or []:
            email_addr = (extra.get("email") or "").strip()
            if email_addr and "@" in email_addr:
                destinataires.append(
                    {
                        "transporteur_id": None,
                        "nom": extra.get("nom") or email_addr,
                        "email": email_addr,
                        # Destinataire ponctuel : langue inconnue, on double.
                        # Se tromper de langue coûte plus cher que doubler.
                        "langue": _normalize_langue(extra.get("langue"), ""),
                        # …sauf si cette adresse a déjà été servie dans une
                        # langue connue : re-saisir l'email d'un transporteur
                        # référencé ne doit pas lui renvoyer un mail bilingue.
                        "langue_heritee": True,
                    }
                )

        if not destinataires:
            raise HTTPException(
                status_code=400,
                detail="Aucun destinataire valide — vérifier les emails des transporteurs",
            )

        envois_ok: list[str] = []
        envois_ko: list[str] = []
        for dest in destinataires:
            import uuid as _uuid

            email_norm = (dest.get("email") or "").strip().lower()
            token_row = conn.execute(
                "SELECT token FROM expe_portal_transporteurs WHERE LOWER(email)=LOWER(?) AND actif=1 LIMIT 1",
                (email_norm,),
            ).fetchone()
            if token_row and token_row["token"]:
                token = str(token_row["token"])
            else:
                token = str(_uuid.uuid4())
                conn.execute(
                    """
                    INSERT OR IGNORE INTO expe_portal_transporteurs
                    (email, token, transporteur_id, prospect_id, created_at, actif)
                    VALUES (?,?,?,?,?,1)
                    """,
                    (
                        email_norm,
                        token,
                        dest.get("transporteur_id"),
                        None,
                        now,
                    ),
                )
                row2 = conn.execute(
                    "SELECT token FROM expe_portal_transporteurs WHERE LOWER(email)=LOWER(?) AND actif=1 LIMIT 1",
                    (email_norm,),
                ).fetchone()
                if row2 and row2["token"]:
                    token = str(row2["token"])
            portail_lien = f"{public_base_url()}/portail/expe/{token}"

            # La ligne de réponse est créée AVANT l'envoi, et non après comme
            # auparavant : le pixel de suivi a besoin de son id pour exister,
            # et un pixel posé après le départ du mail ne sert plus à rien.
            # L'ordre est donc : upsert de la ligne → token pixel → email →
            # envoi → mise à jour du statut. Une ligne créée dont l'envoi
            # échoue reste correcte : elle porte le statut `echec`, qui est
            # précisément l'information à afficher.
            existing = conn.execute(
                """
                SELECT id, statut, prix FROM expe_devis_reponses
                WHERE demande_id=?
                  AND LOWER(TRIM(COALESCE(destinataire_email,''))) = LOWER(TRIM(COALESCE(?,'')))
                ORDER BY id DESC
                LIMIT 1
                """,
                (demande_id, email_norm),
            ).fetchone()
            if existing:
                reponse_id = int(existing["id"])
                conn.execute(
                    """
                    UPDATE expe_devis_reponses
                    SET transporteur_id=?, nom_transporteur=?, destinataire_email=?
                    WHERE id=?
                    """,
                    (dest["transporteur_id"], dest["nom"], email_norm, reponse_id),
                )
                statut_precedent = existing["statut"]
            else:
                cur = conn.execute(
                    """
                    INSERT INTO expe_devis_reponses
                    (demande_id, transporteur_id, nom_transporteur, statut, destinataire_email)
                    VALUES (?,?,?,'envoyee',?)
                    """,
                    (
                        demande_id,
                        dest["transporteur_id"],
                        dest["nom"],
                        email_norm,
                    ),
                )
                reponse_id = int(cur.lastrowid)
                statut_precedent = None

            langue = dest.get("langue") or ""
            if not langue and dest.get("langue_heritee"):
                langue = _normalize_langue(
                    (existing["langue"] if existing and "langue" in existing.keys() else None), ""
                ) or _langue_connue_pour_email(conn, email_norm)
            px = expe_ev.url_pixel(expe_ev.token_pixel(conn, reponse_id), "rfq")
            sujet, corps_html = email_expe_rfq_transport(
                demande=demande,
                user=user,
                portail_lien=portail_lien,
                pixel_url=px,
                langue=langue or None,
                date_limite=(demande.get("date_limite") or None),
            )

            ok = send_email(
                to=dest["email"],
                subject=sujet,
                html_body=corps_html,
                reply_to=reply_to,
                cc=EXPE_DEVIS_CC,
            )
            statut_envoi = "envoyee" if ok else "echec"
            # Un transporteur qui a déjà chiffré ne repasse pas « envoyée »
            # parce qu'on lui renvoie la demande : sa réponse resterait
            # affichée mais le statut mentirait.
            keep_statut = statut_envoi
            if statut_precedent in ("recue", "retenue"):
                keep_statut = statut_precedent
            conn.execute(
                """
                UPDATE expe_devis_reponses
                SET statut=?, langue=?,
                    sent_at=CASE WHEN ? IS NOT NULL THEN ? ELSE sent_at END
                WHERE id=?
                """,
                (
                    keep_statut,
                    langue or None,
                    now if ok else None,
                    now if ok else None,
                    reponse_id,
                ),
            )

            expe_ev.log_evenement(
                conn,
                reponse_id=reponse_id,
                demande_id=demande_id,
                canal=expe_ev.CANAL_EMAIL,
                type_evenement=(
                    expe_ev.EV_EMAIL_ENVOYE if ok else expe_ev.EV_EMAIL_ECHEC
                ),
                date=now,
                fiable=bool(ok),
                motif=None if ok else "envoi refusé par le fournisseur d'email",
                meta={"destinataire": email_norm, "suivi": bool(px), "langue": langue or "fr+en"},
            )

            if ok:
                envois_ok.append(dest["nom"])
            else:
                envois_ko.append(dest["nom"])
        conn.commit()

    _log_devis(
        request,
        user,
        "UPDATE",
        f"Demande devis {demande.get('reference') or demande_id} envoyée · "
        f"{len(envois_ok)} OK / {len(envois_ko)} KO",
    )
    return {
        "envoyes": len(envois_ok),
        "echecs": len(envois_ko),
        "destinataires_ok": envois_ok,
        "destinataires_ko": envois_ko,
    }


@router.put("/devis/reponses/{reponse_id}")
def saisir_reponse_devis(request: Request, reponse_id: int, body: dict = Body(...)):
    user = _require_expe_write(request)
    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    with get_db() as conn:
        rep = conn.execute(
            "SELECT * FROM expe_devis_reponses WHERE id=?", (reponse_id,)
        ).fetchone()
        if not rep:
            raise HTTPException(status_code=404, detail="Réponse introuvable")
        _require_demande_ouverte(conn, int(rep["demande_id"]))
        # Validation alignée sur celle du portail. Sans elle, un prix « abc »
        # arrivait tel quel dans une colonne REAL de SQLite (typage souple) et
        # faisait ensuite planter le comparatif imprimable en 500.
        try:
            prix = float(body.get("prix"))
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Prix invalide.")
        if prix < 0:
            raise HTTPException(status_code=400, detail="Prix invalide.")
        try:
            delai = int(body.get("delai_jours"))
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Délai invalide.")
        if delai < 0 or delai > 365:
            raise HTTPException(status_code=400, detail="Délai invalide.")
        commentaire = (body.get("commentaire") or "").strip() or None
        if commentaire and len(commentaire) > 2000:
            raise HTTPException(status_code=400, detail="Commentaire trop long.")
        conn.execute(
            """
            UPDATE expe_devis_reponses
            SET prix=?, delai_jours=?, commentaire=?, statut='recue', recu_at=?
            WHERE id=?
            """,
            (prix, delai, commentaire, now, reponse_id),
        )
        # Distinguer la saisie interne du dépôt portail : sur la timeline, « le
        # transporteur a répondu » et « on a saisi son mail à sa place » ne
        # racontent pas la même chose sur son engagement réel.
        expe_ev.log_evenement(
            conn,
            reponse_id=int(reponse_id),
            demande_id=int(rep["demande_id"]) if rep["demande_id"] is not None else None,
            canal=expe_ev.CANAL_INTERNE,
            type_evenement=expe_ev.EV_REPONSE_SAISIE,
            date=now,
            meta={"prix": prix, "delai_jours": delai},
        )
        conn.commit()
        updated = conn.execute(
            "SELECT * FROM expe_devis_reponses WHERE id=?", (reponse_id,)
        ).fetchone()
    out = dict(updated)
    out.pop("token_pixel", None)
    _log_devis(
        request,
        user,
        "UPDATE",
        f"Réponse saisie · {rep['nom_transporteur'] or reponse_id} · demande #{rep['demande_id']}",
    )
    return out


@router.post("/devis/reponses/{reponse_id}/retenir")
async def retenir_reponse_devis(
    request: Request,
    reponse_id: int,
    commentaire: Optional[str] = Form(None),
    fichier: Optional[UploadFile] = File(None),
):
    """Retient une proposition de devis :
      1. Marque la réponse `retenue` et les autres `refusee`, clôture la demande.
      2. Crée automatiquement un départ pré-rempli (date = aujourd'hui,
         transporteur / CP / poids / palettes issus du devis).
      3. Envoie un email de confirmation au transporteur retenu, en CC service
         expéditions et reply-to l'utilisateur qui valide.
      4. Optionnel : joint un commentaire libre et une pièce jointe (bon de
         commande, instructions particulières…) à cet email et les archive en DB.

    Multipart/form-data attendu (les 2 champs sont facultatifs) :
      - commentaire : texte libre
      - fichier     : pièce jointe (max 20 Mo)
    """
    user = _require_expe_write(request)
    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    today = _today_paris_iso()
    email_user = (user.get("email") or user.get("identifiant") or "").strip() or None

    with get_db() as conn:
        rep = conn.execute(
            "SELECT * FROM expe_devis_reponses WHERE id=?", (reponse_id,)
        ).fetchone()
        if not rep:
            raise HTTPException(status_code=404, detail="Réponse introuvable")
        demande_id = rep["demande_id"]
        # Garde serveur : retenir deux fois sur la même demande créait un
        # second départ pour le même transport.
        demande = _require_demande_ouverte(conn, int(demande_id))
        rep_d = dict(rep)

        # Sauvegarde optionnelle : commentaire + piece jointe attaches a la
        # retenue. Persisted sur la ligne reponse pour tracabilite et joints
        # a l'email de confirmation transporteur ci-dessous.
        retention_comment = (commentaire or "").strip() or None
        retention_file_rel: str | None = None
        retention_file_name: str | None = None
        retention_file_bytes: bytes | None = None

        if fichier is not None:
            MAX_BYTES = 20 * 1024 * 1024
            retention_file_bytes = await fichier.read()
            if len(retention_file_bytes) > MAX_BYTES:
                raise HTTPException(413, "Fichier trop volumineux (max 20 Mo).")
            orig = (fichier.filename or "fichier").strip()
            safe = _devis_safe_filename(orig)
            unique = f"ret{reponse_id}_{uuid.uuid4().hex[:8]}_{safe}"
            sub_dir = _devis_upload_dir() / "retention"
            sub_dir.mkdir(parents=True, exist_ok=True)
            path_abs = sub_dir / unique
            with open(path_abs, "wb") as out:
                out.write(retention_file_bytes)
            retention_file_rel = f"{_DEVIS_UPLOAD_SUBDIR}/retention/{unique}"
            retention_file_name = orig

        conn.execute(
            """
            UPDATE expe_devis_reponses SET statut='refusee'
            WHERE demande_id=? AND id!=? AND statut NOT IN ('retenue','refusee')
            """,
            (demande_id, reponse_id),
        )
        conn.execute(
            """UPDATE expe_devis_reponses
                SET statut='retenue',
                    retention_comment=?,
                    retention_file_path=?,
                    retention_file_filename=?
                WHERE id=?""",
            (retention_comment, retention_file_rel, retention_file_name, reponse_id),
        )
        conn.execute(
            "UPDATE expe_demandes_devis SET statut='cloturee' WHERE id=?",
            (demande_id,),
        )

        cur = conn.execute(
            """INSERT INTO expe_departs (
                date_enlevement, transporteur, transporteur_id, client,
                code_postal_destination, nb_palette, poids_total_kg,
                statut, created_at, created_by_email,
                source_devis_reponse_id, source_devis_demande_id
            ) VALUES (?,?,?,?,?,?,?, 'en_attente', ?, ?, ?, ?)""",
            (
                today,
                rep_d.get("nom_transporteur"),
                rep_d.get("transporteur_id"),
                demande.get("client"),
                demande.get("code_postal_destination"),
                demande.get("nb_palette"),
                demande.get("poids_total_kg"),
                now,
                email_user,
                reponse_id,
                demande_id,
            ),
        )
        depart_id = cur.lastrowid
        depart_row = conn.execute(
            f"{_DEPARTS_SELECT} WHERE d.id=?", (depart_id,)
        ).fetchone()
        conn.commit()

    depart_dict = _depart_dict(depart_row) if depart_row else {}

    # Destinataire principal = adresse a laquelle la demande a ete envoyee.
    # Fallback : si le champ est vide (reponse creee manuellement via
    # "Saisir reponse" ou reponse portail sans email persiste), on retombe
    # sur les contact_emails du transporteur pour ne pas silencieusement
    # sauter la confirmation.
    dest_email = (rep_d.get("destinataire_email") or "").strip()
    if not (dest_email and "@" in dest_email):
        trp_id = rep_d.get("transporteur_id")
        if trp_id:
            with get_db() as conn:
                trow = conn.execute(
                    "SELECT contact_email, contact_emails FROM expe_transporteurs WHERE id=?",
                    (trp_id,),
                ).fetchone()
            if trow:
                addrs = _normalize_emails(trow["contact_emails"])
                if not addrs:
                    fb = (trow["contact_email"] or "").strip()
                    if fb and "@" in fb:
                        addrs = [fb]
                if addrs:
                    dest_email = addrs[0]
    email_sent = False
    email_error: str | None = None
    px_attr = None
    if dest_email and "@" in dest_email:
        try:
            with get_db() as conn:
                px_attr = expe_ev.url_pixel(
                    expe_ev.token_pixel(conn, reponse_id), "attr"
                )
                conn.commit()
        except Exception:
            px_attr = None
        try:
            subject, body_html = email_expe_devis_confirmation(
                demande=demande,
                reponse=rep_d,
                depart=depart_dict,
                user=user,
                retention_comment=retention_comment,
                retention_file_name=retention_file_name,
                pixel_url=px_attr,
            )
            atts = None
            if retention_file_bytes and retention_file_name:
                atts = [{
                    "filename": retention_file_name,
                    "content": retention_file_bytes,
                }]
            email_sent = bool(send_email(
                to=dest_email,
                subject=subject,
                html_body=body_html,
                reply_to=email_user,
                cc=EXPE_DEVIS_CC,
                attachments=atts,
            ))
            if not email_sent:
                email_error = "send_email_returned_false"
        except Exception as _e:  # noqa: BLE001
            email_sent = False
            email_error = f"{type(_e).__name__}: {_e}"[:200]
    else:
        email_error = "destinataire_email_absent"

    # Journal : l'attribution est un fait métier, la notification n'en est que
    # le véhicule. On trace donc les deux — « retenue » toujours, « email
    # d'attribution » seulement s'il est parti.
    try:
        with get_db() as conn:
            expe_ev.log_evenement(
                conn,
                reponse_id=reponse_id,
                demande_id=demande_id,
                canal=expe_ev.CANAL_INTERNE,
                type_evenement=expe_ev.EV_OFFRE_RETENUE,
                date=now,
                meta={"depart_id": depart_id, "par": email_user},
            )
            if email_sent:
                expe_ev.log_evenement(
                    conn,
                    reponse_id=reponse_id,
                    demande_id=demande_id,
                    canal=expe_ev.CANAL_EMAIL,
                    type_evenement=expe_ev.EV_EMAIL_ATTRIBUTION,
                    date=now,
                    meta={"destinataire": dest_email, "suivi": bool(px_attr)},
                )
            conn.commit()
    except Exception:
        pass

    try:
        log_action(
            user=user,
            action="CREATE",
            module="expe",
            objet=(
                f"Départ #{depart_id} créé depuis devis #{demande_id} "
                f"(réponse #{reponse_id}) · email transporteur "
                f"{'OK' if email_sent else 'KO'}"
            ),
            ip=request.client.host if request.client else None,
        )
    except Exception:
        pass

    return {
        "statut": "cloturee",
        "retenu": reponse_id,
        "depart_id": depart_id,
        "depart": depart_dict,
        "email_envoye": email_sent,
        "email_destinataire": dest_email or None,
        "email_error": email_error,
    }


@router.post("/devis/demandes/{demande_id}/cloturer")
def cloturer_demande_devis(request: Request, demande_id: int):
    """Clôture manuelle d'une demande : passe en statut 'cloturee' (archive)."""
    user = _require_expe_write(request)
    with get_db() as conn:
        demande = _require_demande_ouverte(conn, demande_id)
        conn.execute(
            "UPDATE expe_demandes_devis SET statut='cloturee' WHERE id=?",
            (demande_id,),
        )
        conn.commit()
    _log_devis(
        request, user, "UPDATE", f"Demande devis {demande.get('reference') or demande_id} clôturée"
    )
    return {"statut": "cloturee", "id": demande_id}


@router.delete("/devis/demandes/{demande_id}")
def supprimer_demande_devis(request: Request, demande_id: int):
    """Met la demande à la corbeille — réversible.

    La suppression était physique et immédiate : elle emportait la référence
    (2026-15, non réattribuée) et les offres déjà reçues, sans recours, et
    laissait les fichiers orphelins sur le disque. On marque désormais
    `deleted_at` ; la destruction réelle passe par `/purger`.
    """
    user = _require_expe_write(request)
    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    with get_db() as conn:
        demande = _get_demande_or_404(conn, demande_id)
        conn.execute(
            "UPDATE expe_demandes_devis SET deleted_at=? WHERE id=?", (now, demande_id)
        )
        conn.commit()
    _log_devis(
        request,
        user,
        "DELETE",
        f"Demande devis {demande.get('reference') or demande_id} → corbeille",
    )
    return {"deleted": demande_id, "corbeille": True}


@router.post("/devis/demandes/{demande_id}/restaurer")
def restaurer_demande_devis(request: Request, demande_id: int):
    user = _require_expe_write(request)
    with get_db() as conn:
        demande = _get_demande_or_404(conn, demande_id, avec_corbeille=True)
        if not demande.get("deleted_at"):
            raise HTTPException(status_code=400, detail="Cette demande n'est pas dans la corbeille.")
        conn.execute(
            "UPDATE expe_demandes_devis SET deleted_at=NULL WHERE id=?", (demande_id,)
        )
        conn.commit()
    _log_devis(
        request,
        user,
        "UPDATE",
        f"Demande devis {demande.get('reference') or demande_id} restaurée",
    )
    return {"ok": True, "id": demande_id}


@router.delete("/devis/demandes/{demande_id}/purger")
def purger_demande_devis(request: Request, demande_id: int):
    """Destruction définitive, réservée aux demandes déjà en corbeille.

    Deux gestes plutôt qu'un : le premier est rattrapable, le second demande
    d'aller le chercher dans la corbeille. C'est ici, et seulement ici, qu'on
    nettoie les fichiers sur disque — l'ancienne suppression les laissait
    s'accumuler indéfiniment.
    """
    user = _require_expe_write(request)
    with get_db() as conn:
        demande = _get_demande_or_404(conn, demande_id, avec_corbeille=True)
        if not demande.get("deleted_at"):
            raise HTTPException(
                status_code=409,
                detail="Mettre d'abord la demande à la corbeille.",
            )
        chemins = [
            r["path"]
            for r in conn.execute(
                "SELECT path FROM expe_devis_pieces_jointes WHERE demande_id=?",
                (demande_id,),
            ).fetchall()
        ]
        for r in conn.execute(
            """SELECT retention_file_path AS p FROM expe_devis_reponses
               WHERE demande_id=? AND COALESCE(TRIM(retention_file_path),'') <> ''""",
            (demande_id,),
        ).fetchall():
            chemins.append(r["p"])
        legacy = (demande.get("piece_jointe_path") or "").strip()
        if legacy:
            chemins.append(legacy)
        conn.execute("DELETE FROM expe_devis_pieces_jointes WHERE demande_id=?", (demande_id,))
        conn.execute("DELETE FROM expe_devis_evenements WHERE demande_id=?", (demande_id,))
        conn.execute("DELETE FROM expe_devis_reponses WHERE demande_id=?", (demande_id,))
        conn.execute("DELETE FROM expe_demandes_devis WHERE id=?", (demande_id,))
        # Un fichier partagé avec une demande dupliquée n'est pas effacé.
        restants = set()
        for c in set(chemins):
            if conn.execute(
                "SELECT 1 FROM expe_devis_pieces_jointes WHERE path=? LIMIT 1", (c,)
            ).fetchone():
                restants.add(c)
        conn.commit()
    from config import BASE_DIR

    for c in set(chemins) - restants:
        try:
            (Path(BASE_DIR) / c).unlink()
        except Exception:
            pass
    _log_devis(
        request,
        user,
        "DELETE",
        f"Demande devis {demande.get('reference') or demande_id} purgée définitivement",
    )
    return {"purged": demande_id}


@router.delete("/devis/reponses/{reponse_id}")
def retirer_destinataire_devis(request: Request, reponse_id: int):
    """Retire un destinataire d'une demande.

    Un email saisi de travers restait sur la demande jusqu'à la suppression de
    celle-ci. On refuse en revanche de retirer un transporteur qui a chiffré :
    son offre fait partie de l'histoire de la consultation, l'effacer
    reviendrait à réécrire la comparaison après coup.
    """
    user = _require_expe_write(request)
    with get_db() as conn:
        rep = conn.execute(
            "SELECT * FROM expe_devis_reponses WHERE id=?", (reponse_id,)
        ).fetchone()
        if not rep:
            raise HTTPException(status_code=404, detail="Destinataire introuvable")
        _require_demande_ouverte(conn, int(rep["demande_id"]))
        if (rep["statut"] or "") in ("recue", "retenue"):
            raise HTTPException(
                status_code=409,
                detail="Ce transporteur a répondu : sa réponse ne peut pas être retirée.",
            )
        # Cascade manuelle : `get_db()` n'active pas `PRAGMA foreign_keys`, le
        # ON DELETE CASCADE déclaré au schéma ne se déclenche donc jamais. Sans
        # ça, un transporteur ayant déposé un fichier sans chiffrer laissait sa
        # pièce jointe orpheline, listée sans nom dans le détail.
        fichiers = [
            r["path"]
            for r in conn.execute(
                "SELECT path FROM expe_devis_pieces_jointes WHERE reponse_id=?",
                (reponse_id,),
            ).fetchall()
        ]
        conn.execute("DELETE FROM expe_devis_pieces_jointes WHERE reponse_id=?", (reponse_id,))
        conn.execute("DELETE FROM expe_devis_evenements WHERE reponse_id=?", (reponse_id,))
        conn.execute("DELETE FROM expe_devis_reponses WHERE id=?", (reponse_id,))
        encore = {
            c
            for c in fichiers
            if conn.execute(
                "SELECT 1 FROM expe_devis_pieces_jointes WHERE path=? LIMIT 1", (c,)
            ).fetchone()
        }
        conn.commit()
    from config import BASE_DIR

    for c in set(fichiers) - encore:
        try:
            (Path(BASE_DIR) / c).unlink()
        except Exception:
            pass
    _log_devis(
        request,
        user,
        "DELETE",
        f"Destinataire {rep['nom_transporteur'] or rep['destinataire_email'] or reponse_id} "
        f"retiré de la demande #{rep['demande_id']}",
    )
    return {"ok": True}


@router.get("/devis/reponses/{reponse_id}/lien-portail")
def lien_portail_destinataire(request: Request, reponse_id: int):
    """Renvoie le lien portail du destinataire, pour le lui recopier.

    Utile quand un transporteur affirme n'avoir rien reçu : plutôt que de
    renvoyer l'email en aveugle, on lui donne le lien de vive voix.
    """
    _require_expe(request)
    with get_db() as conn:
        rep = conn.execute(
            "SELECT destinataire_email, transporteur_id FROM expe_devis_reponses WHERE id=?",
            (reponse_id,),
        ).fetchone()
        if not rep:
            raise HTTPException(status_code=404, detail="Destinataire introuvable")
        email = (rep["destinataire_email"] or "").strip().lower()
        row = conn.execute(
            "SELECT token FROM expe_portal_transporteurs WHERE LOWER(email)=? AND actif=1 LIMIT 1",
            (email,),
        ).fetchone()
    if not row:
        raise HTTPException(
            status_code=404,
            detail="Aucun accès portail : la demande n'a pas encore été envoyée à cette adresse.",
        )
    return {"lien": f"{public_base_url()}/portail/expe/{row['token']}", "email": email}


@router.post("/devis/reponses/{reponse_id}/relancer")
def relancer_destinataire_devis(request: Request, reponse_id: int, body: dict = Body(default={})):
    """Renvoie la demande de tarif à un transporteur resté silencieux.

    Le suivi d'ouverture rend le silence visible ; ce bouton est ce qui permet
    d'en faire quelque chose. Le pixel part avec le contexte `rel` — distinct
    de `rfq` — pour que la fenêtre anti-préchargement se recale sur la date de
    la relance et non sur celle de l'envoi initial, et pour que la timeline
    dise LEQUEL des deux emails a été ouvert.
    """
    user = _require_expe_write(request)
    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    reply_to = (user.get("email") or user.get("identifiant") or "").strip() or None
    message = (body.get("message") or "").strip() or None
    with get_db() as conn:
        rep = conn.execute(
            "SELECT * FROM expe_devis_reponses WHERE id=?", (reponse_id,)
        ).fetchone()
        if not rep:
            raise HTTPException(status_code=404, detail="Destinataire introuvable")
        rep = dict(rep)
        demande = _require_demande_ouverte(conn, int(rep["demande_id"]))
        if (rep.get("statut") or "") in ("recue", "retenue"):
            raise HTTPException(
                status_code=409, detail="Ce transporteur a déjà répondu."
            )
        email_dest = (rep.get("destinataire_email") or "").strip()
        if not email_dest or "@" not in email_dest:
            raise HTTPException(
                status_code=400, detail="Aucune adresse email pour ce destinataire."
            )
        token_row = conn.execute(
            "SELECT token FROM expe_portal_transporteurs WHERE LOWER(email)=LOWER(?) AND actif=1 LIMIT 1",
            (email_dest,),
        ).fetchone()
        if not token_row:
            raise HTTPException(
                status_code=409,
                detail="Aucun accès portail pour cette adresse — renvoyer la demande depuis « Envoyer ».",
            )
        portail_lien = f"{public_base_url()}/portail/expe/{token_row['token']}"
        langue = _devis_langue_destinataire(conn, rep)
        px = expe_ev.url_pixel(expe_ev.token_pixel(conn, reponse_id), "rel")
        sujet, corps = email_expe_rfq_transport(
            demande=demande,
            user=user,
            portail_lien=portail_lien,
            pixel_url=px,
            langue=langue,
            relance=True,
            message_perso=message,
            # La relance sert précisément à rappeler l'échéance : l'omettre
            # ferait un mail qui insiste sans dire jusqu'à quand.
            date_limite=(demande.get("date_limite") or None),
        )
        ok = send_email(
            to=email_dest,
            subject=sujet,
            html_body=corps,
            reply_to=reply_to,
            cc=EXPE_DEVIS_CC,
        )
        if ok:
            conn.execute(
                """UPDATE expe_devis_reponses
                   SET relances=COALESCE(relances,0)+1, last_relance_at=?, sent_at=?,
                       statut=CASE WHEN statut='echec' THEN 'envoyee' ELSE statut END
                   WHERE id=?""",
                (now, now, reponse_id),
            )
        expe_ev.log_evenement(
            conn,
            reponse_id=reponse_id,
            demande_id=int(rep["demande_id"]),
            canal=expe_ev.CANAL_EMAIL,
            type_evenement=(
                expe_ev.EV_EMAIL_RELANCE if ok else expe_ev.EV_EMAIL_ECHEC
            ),
            date=now,
            fiable=bool(ok),
            motif=None if ok else "envoi refusé par le fournisseur d'email",
            meta={"destinataire": email_dest, "suivi": bool(px), "langue": langue},
        )
        conn.commit()
    _log_devis(
        request,
        user,
        "UPDATE",
        f"Relance {rep.get('nom_transporteur') or email_dest} · demande {demande.get('reference') or demande['id']} · {'OK' if ok else 'KO'}",
    )
    if not ok:
        raise HTTPException(status_code=502, detail="L'email de relance n'est pas parti.")
    return {"ok": True, "relances": int(rep.get("relances") or 0) + 1}


# ─── Prospects transporteurs ───────────────────────────────────────


@router.get("/prospects")
def list_prospects(request: Request):
    _require_expe(request)
    with get_db() as conn:
        rows = conn.execute(
            """SELECT * FROM expe_transporteurs_prospects
               ORDER BY statut_demarchage, nom"""
        ).fetchall()
    return [dict(r) for r in rows]


@router.post("/prospects")
def creer_prospect(request: Request, body: dict = Body(...)):
    _require_expe_write(request)
    nom = (body.get("nom") or "").strip()
    if not nom:
        raise HTTPException(status_code=400, detail="Nom obligatoire")
    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    with get_db() as conn:
        cur = conn.execute(
            """
            INSERT INTO expe_transporteurs_prospects
            (nom, contact_nom, contact_email, contact_tel, zone_couverte,
             type_service, capacite_max_pal, statut_demarchage, notes, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?)
            """,
            (
                nom,
                (body.get("contact_nom") or "").strip() or None,
                (body.get("contact_email") or "").strip() or None,
                (body.get("contact_tel") or "").strip() or None,
                (body.get("zone_couverte") or "").strip() or None,
                (body.get("type_service") or "messagerie").strip(),
                body.get("capacite_max_pal"),
                (body.get("statut_demarchage") or "a_contacter").strip(),
                (body.get("notes") or "").strip() or None,
                now,
            ),
        )
        conn.commit()
        row = conn.execute(
            "SELECT * FROM expe_transporteurs_prospects WHERE id=?", (cur.lastrowid,)
        ).fetchone()
    return dict(row)


@router.put("/prospects/{prospect_id}")
def modifier_prospect(request: Request, prospect_id: int, body: dict = Body(...)):
    _require_expe_write(request)
    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    champs = [
        "nom",
        "contact_nom",
        "contact_email",
        "contact_tel",
        "zone_couverte",
        "type_service",
        "capacite_max_pal",
        "statut_demarchage",
        "notes",
    ]
    sets = ["updated_at=?"]
    args: list[Any] = [now]
    for c in champs:
        if c in body:
            sets.append(f"{c}=?")
            v = body[c]
            if c == "nom":
                v = (v or "").strip()
            elif isinstance(v, str):
                v = v.strip() or None
            args.append(v)
    args.append(prospect_id)
    with get_db() as conn:
        conn.execute(
            f"UPDATE expe_transporteurs_prospects SET {', '.join(sets)} WHERE id=?",
            args,
        )
        conn.commit()
        row = conn.execute(
            "SELECT * FROM expe_transporteurs_prospects WHERE id=?", (prospect_id,)
        ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Prospect introuvable")
    return dict(row)


@router.delete("/prospects/{prospect_id}")
def supprimer_prospect(request: Request, prospect_id: int):
    _require_expe_write(request)
    with get_db() as conn:
        conn.execute(
            "DELETE FROM expe_transporteurs_prospects WHERE id=?", (prospect_id,)
        )
        conn.commit()
    return {"deleted": prospect_id}


# ─── Délais carte France ───────────────────────────────────────────

_DELAIS_EDIT_ROLES = {"superadmin", "direction", "administration", "administration_ventes", "administration_technique", "expedition"}


def _delai_jours_from_texte(delai_texte: str) -> int:
    try:
        return int(str(delai_texte).replace("J+", "").strip())
    except (ValueError, AttributeError):
        return 2


@router.get("/delais")
def get_delais(request: Request, type_envoi: str = "default"):
    _require_expe(request)
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT departement, delai_texte, zone_label
            FROM expe_delais
            WHERE type_envoi=? AND transporteur_id IS NULL
            """,
            (type_envoi,),
        ).fetchall()

        if not rows and type_envoi != "default":
            rows = conn.execute(
                """
                SELECT departement, delai_texte, zone_label
                FROM expe_delais
                WHERE type_envoi='default' AND transporteur_id IS NULL
                """
            ).fetchall()

    from app.web.expe_france_delais_data import DELAIS_FRANCE_DEFAULT

    result: dict[str, dict] = {}
    for r in rows:
        dept = r["departement"]
        default_label = DELAIS_FRANCE_DEFAULT.get(dept, {}).get("label", dept)
        result[dept] = {
            "delai": r["delai_texte"],
            "zone": r["zone_label"],
            "label": default_label,
        }
    return result


@router.put("/delais")
def save_delais(request: Request, body: dict = Body(...)):
    user = _require_expe_write(request)
    role = (user.get("role") or "").strip()
    if role not in _DELAIS_EDIT_ROLES:
        raise HTTPException(
            status_code=403,
            detail="Accès refusé — rôle insuffisant pour modifier les délais",
        )

    overrides = body.get("overrides") or {}
    if not overrides:
        raise HTTPException(status_code=400, detail="overrides est vide")

    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    type_envoi = (body.get("type_envoi") or "default").strip()
    email = (user.get("email") or user.get("identifiant") or "").strip() or None

    with get_db() as conn:
        for dept, data in overrides.items():
            delai_texte = str(data.get("delai") or "J+2").strip()
            zone_label = str(data.get("zone") or "france").strip()
            delai_jours = _delai_jours_from_texte(delai_texte)
            conn.execute(
                """
                DELETE FROM expe_delais
                WHERE departement=? AND type_envoi=? AND transporteur_id IS NULL
                """,
                (dept, type_envoi),
            )
            conn.execute(
                """
                INSERT INTO expe_delais
                (departement, type_envoi, transporteur_id, delai_jours, zone_label,
                 delai_texte, updated_at, updated_by_email)
                VALUES (?, ?, NULL, ?, ?, ?, ?, ?)
                """,
                (dept, type_envoi, delai_jours, zone_label, delai_texte, now, email),
            )
        conn.commit()

    return {"updated": len(overrides)}


@router.post("/delais/reset")
def reset_delais(request: Request, body: dict = Body(default_factory=dict)):
    user = _require_expe_write(request)
    role = (user.get("role") or "").strip()
    if role not in _DELAIS_EDIT_ROLES:
        raise HTTPException(status_code=403, detail="Accès refusé")

    type_envoi = (body.get("type_envoi") or "default").strip()
    now = datetime.now(_PARIS).strftime("%Y-%m-%dT%H:%M:%S")
    email = (user.get("email") or user.get("identifiant") or "").strip() or None

    from app.web.expe_france_delais_data import DELAIS_FRANCE_DEFAULT

    with get_db() as conn:
        conn.execute(
            "DELETE FROM expe_delais WHERE type_envoi=? AND transporteur_id IS NULL",
            (type_envoi,),
        )
        for dept, data in DELAIS_FRANCE_DEFAULT.items():
            delai_texte = data.get("delai", "J+2")
            zone_label = data.get("zone", "france")
            delai_jours = _delai_jours_from_texte(str(delai_texte))
            conn.execute(
                """
                INSERT INTO expe_delais
                (departement, type_envoi, transporteur_id, delai_jours, zone_label,
                 delai_texte, updated_at, updated_by_email)
                VALUES (?, ?, NULL, ?, ?, ?, ?, ?)
                """,
                (dept, type_envoi, delai_jours, zone_label, delai_texte, now, email),
            )
        conn.commit()

    return {"reset": True, "type_envoi": type_envoi}
