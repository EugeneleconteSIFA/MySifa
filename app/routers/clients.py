"""Référentiel clients (ERP) — Paramètres super administrateur.

Gestion CRUD du référentiel clients utilisé par MyProd, MyExpé, MyCompta.
Import par fichier xlsx (export ERP, en-tête colonne 1 = entêtes).
"""

import io
import re
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, File, HTTPException, Request, UploadFile
from pydantic import BaseModel

from app.services.audit_service import log_action
from services.auth_service import require_superadmin

router = APIRouter(tags=["clients"])


# ─── Modèles ──────────────────────────────────────────────────────

class ClientIn(BaseModel):
    numero: Optional[int] = None
    code: Optional[str] = None
    raison_sociale: str
    adresse1: Optional[str] = None
    adresse2: Optional[str] = None
    bp: Optional[str] = None
    cp: Optional[str] = None
    ville: Optional[str] = None
    code_pays: Optional[str] = None
    pays: Optional[str] = None
    groupe: Optional[str] = None
    siret: Optional[str] = None
    rcs: Optional[str] = None
    tva: Optional[str] = None
    ean: Optional[str] = None
    nif: Optional[str] = None
    telephone: Optional[str] = None
    telecopie: Optional[str] = None
    email: Optional[str] = None
    representant: Optional[str] = None
    adv: Optional[str] = None
    categorie1: Optional[str] = None
    categorie2: Optional[str] = None
    categorie3: Optional[str] = None
    mode_livraison: Optional[str] = None
    mode_reglement: Optional[str] = None
    devise: Optional[str] = None
    encours_autorise: Optional[float] = None
    code_comptable: Optional[str] = None
    etat: Optional[str] = "Normal"
    contact_nom: Optional[str] = None
    contact_fonction: Optional[str] = None
    contact_email: Optional[str] = None
    contact_tel: Optional[str] = None
    notes: Optional[str] = None


# ─── Helpers ──────────────────────────────────────────────────────

_CLIENT_COLS = [
    "numero", "code", "raison_sociale", "adresse1", "adresse2", "bp", "cp",
    "ville", "code_pays", "pays", "groupe", "siret", "rcs", "tva", "ean",
    "nif", "telephone", "telecopie", "email", "representant", "adv",
    "categorie1", "categorie2", "categorie3", "mode_livraison",
    "mode_reglement", "devise", "encours_autorise", "code_comptable", "etat",
    "contact_nom", "contact_fonction", "contact_email", "contact_tel",
    "notes", "date_creation", "date_modification",
]


def _row_to_dict(row) -> dict:
    return dict(row) if row else {}


def _clean_str(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _clean_float(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return None


def _clean_int(v) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).replace(",", ".").replace(" ", "")))
    except (ValueError, TypeError):
        return None


def _norm_header(h: str) -> str:
    """Normalise un en-tête (sans accents, minuscules, espaces réduits)."""
    if not h:
        return ""
    s = str(h).strip().lower()
    # Suppression accents simples
    repl = {
        "à": "a", "â": "a", "ä": "a",
        "é": "e", "è": "e", "ê": "e", "ë": "e",
        "î": "i", "ï": "i",
        "ô": "o", "ö": "o",
        "ù": "u", "û": "u", "ü": "u",
        "ç": "c",
    }
    for a, b in repl.items():
        s = s.replace(a, b)
    s = re.sub(r"[\s\.]+", " ", s).strip()
    return s


# Mapping en-têtes export ERP → colonnes DB
_HEADER_MAP = {
    "no": "numero",
    "n°": "numero",
    "code": "code",
    "raison sociale": "raison_sociale",
    "adresse 1": "adresse1",
    "adresse1": "adresse1",
    "adresse 2": "adresse2",
    "adresse2": "adresse2",
    "b p": "bp",
    "bp": "bp",
    "c p": "cp",
    "cp": "cp",
    "ville": "ville",
    "c pays": "code_pays",
    "pays": "pays",
    "groupe": "groupe",
    "siret": "siret",
    "rcs": "rcs",
    "n tva": "tva",
    "tva": "tva",
    "n,tva": "tva",
    "ean": "ean",
    "nif": "nif",
    "telephone": "telephone",
    "telecopie": "telecopie",
    "email": "email",
    "representant": "representant",
    "adv": "adv",
    "categorie 1": "categorie1",
    "categorie 2": "categorie2",
    "categorie 3": "categorie3",
    "mode de livraison": "mode_livraison",
    "mode de reglement": "mode_reglement",
    "devise": "devise",
    "encours autorise": "encours_autorise",
    "code comptable": "code_comptable",
    "etat": "etat",
    "date creation": "date_creation",
    "date modification": "date_modification",
}


def _xlsx_to_rows(file_bytes: bytes) -> tuple[list[str], list[list]]:
    """Renvoie (headers, rows) en lisant un xlsx, tolérant aux fichiers mal formés."""
    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(500, f"openpyxl non disponible : {exc}")

    # Certains exports ERP contiennent des attributs corrompus dans styles.xml
    # (ex. "biltinId" au lieu de "builtinId"). On corrige à la volée.
    import zipfile
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as zin:
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.namelist():
                    data = zin.read(item)
                    if item.endswith("styles.xml"):
                        data = data.replace(b"biltinId", b"builtinId")
                    zout.writestr(item, data)
            buf.seek(0)
            wb = openpyxl.load_workbook(buf, data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Fichier xlsx illisible : {exc}")

    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(400, "Fichier vide.")
    headers = [str(h) if h is not None else "" for h in headers_row]
    rows: list[list] = []
    for r in rows_iter:
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        rows.append(list(r))
    return headers, rows


# ─── Endpoints ────────────────────────────────────────────────────

@router.get("/api/clients")
def list_clients(
    request: Request,
    search: str = "",
    etat: str = "",
    limit: int = 1000,
    offset: int = 0,
):
    """Liste paginée des clients avec recherche multi-champs."""
    require_superadmin(request)
    from database import get_db

    with get_db() as conn:
        conditions = ["1=1"]
        params: list = []
        if search:
            like = f"%{search.strip()}%"
            conditions.append(
                "(raison_sociale LIKE ? OR code LIKE ? OR ville LIKE ? "
                "OR siret LIKE ? OR tva LIKE ? OR email LIKE ? "
                "OR CAST(numero AS TEXT) LIKE ?)"
            )
            params.extend([like, like, like, like, like, like, like])
        if etat:
            conditions.append("etat = ?")
            params.append(etat)
        where = " AND ".join(conditions)
        total = conn.execute(
            f"SELECT COUNT(*) FROM clients WHERE {where}", params
        ).fetchone()[0]
        rows = conn.execute(
            f"""SELECT * FROM clients WHERE {where}
                ORDER BY raison_sociale COLLATE NOCASE ASC
                LIMIT ? OFFSET ?""",
            params + [limit, offset],
        ).fetchall()
        states = conn.execute(
            "SELECT DISTINCT etat FROM clients WHERE etat IS NOT NULL AND etat != '' ORDER BY etat"
        ).fetchall()

    return {
        "total": total,
        "limit": limit,
        "offset": offset,
        "items": [_row_to_dict(r) for r in rows],
        "etats": [r[0] for r in states],
    }


@router.get("/api/clients/{client_id}")
def get_client(client_id: int, request: Request):
    require_superadmin(request)
    from database import get_db
    with get_db() as conn:
        row = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Client introuvable")
    return _row_to_dict(row)


@router.post("/api/clients")
def create_client(payload: ClientIn, request: Request):
    user = require_superadmin(request)
    from database import get_db

    raison = (payload.raison_sociale or "").strip()
    if not raison:
        raise HTTPException(400, "Raison sociale requise")

    now = datetime.now().isoformat(timespec="seconds")
    data = payload.model_dump()
    data["raison_sociale"] = raison
    for k in ("code", "ville", "email", "telephone", "siret", "tva",
              "contact_nom", "contact_email", "contact_tel", "notes",
              "adresse1", "adresse2", "pays"):
        if isinstance(data.get(k), str):
            data[k] = data[k].strip() or None
    data["etat"] = data.get("etat") or "Normal"

    with get_db() as conn:
        if data.get("code"):
            ex = conn.execute(
                "SELECT id FROM clients WHERE code=? COLLATE NOCASE", (data["code"],)
            ).fetchone()
            if ex:
                raise HTTPException(409, f"Le code client « {data['code']} » existe déjà.")
        cur = conn.execute(
            """INSERT INTO clients (
                numero, code, raison_sociale, adresse1, adresse2, bp, cp, ville,
                code_pays, pays, groupe, siret, rcs, tva, ean, nif,
                telephone, telecopie, email, representant, adv,
                categorie1, categorie2, categorie3, mode_livraison, mode_reglement,
                devise, encours_autorise, code_comptable, etat,
                contact_nom, contact_fonction, contact_email, contact_tel, notes,
                created_at, updated_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                data.get("numero"), data.get("code"), data["raison_sociale"],
                data.get("adresse1"), data.get("adresse2"), data.get("bp"),
                data.get("cp"), data.get("ville"), data.get("code_pays"),
                data.get("pays"), data.get("groupe"), data.get("siret"),
                data.get("rcs"), data.get("tva"), data.get("ean"), data.get("nif"),
                data.get("telephone"), data.get("telecopie"), data.get("email"),
                data.get("representant"), data.get("adv"),
                data.get("categorie1"), data.get("categorie2"), data.get("categorie3"),
                data.get("mode_livraison"), data.get("mode_reglement"),
                data.get("devise"), data.get("encours_autorise"),
                data.get("code_comptable"), data.get("etat"),
                data.get("contact_nom"), data.get("contact_fonction"),
                data.get("contact_email"), data.get("contact_tel"), data.get("notes"),
                now, now,
            ),
        )
        conn.commit()
        new_id = cur.lastrowid

    log_action(
        user=user,
        action="CREATE",
        module="settings",
        objet=f"Client · {raison}",
        ip=request.client.host if request.client else None,
    )
    return {"success": True, "id": new_id}


@router.put("/api/clients/{client_id}")
def update_client(client_id: int, payload: ClientIn, request: Request):
    user = require_superadmin(request)
    from database import get_db

    raison = (payload.raison_sociale or "").strip()
    if not raison:
        raise HTTPException(400, "Raison sociale requise")

    data = payload.model_dump()
    data["raison_sociale"] = raison
    data["etat"] = data.get("etat") or "Normal"
    now = datetime.now().isoformat(timespec="seconds")

    with get_db() as conn:
        ex = conn.execute("SELECT id, raison_sociale FROM clients WHERE id=?", (client_id,)).fetchone()
        if not ex:
            raise HTTPException(404, "Client introuvable")
        if data.get("code"):
            dup = conn.execute(
                "SELECT id FROM clients WHERE code=? COLLATE NOCASE AND id<>?",
                (data["code"], client_id),
            ).fetchone()
            if dup:
                raise HTTPException(409, f"Le code client « {data['code']} » est déjà utilisé.")
        conn.execute(
            """UPDATE clients SET
                numero=?, code=?, raison_sociale=?, adresse1=?, adresse2=?, bp=?, cp=?, ville=?,
                code_pays=?, pays=?, groupe=?, siret=?, rcs=?, tva=?, ean=?, nif=?,
                telephone=?, telecopie=?, email=?, representant=?, adv=?,
                categorie1=?, categorie2=?, categorie3=?, mode_livraison=?, mode_reglement=?,
                devise=?, encours_autorise=?, code_comptable=?, etat=?,
                contact_nom=?, contact_fonction=?, contact_email=?, contact_tel=?, notes=?,
                updated_at=?
              WHERE id=?""",
            (
                data.get("numero"), data.get("code"), data["raison_sociale"],
                data.get("adresse1"), data.get("adresse2"), data.get("bp"),
                data.get("cp"), data.get("ville"), data.get("code_pays"),
                data.get("pays"), data.get("groupe"), data.get("siret"),
                data.get("rcs"), data.get("tva"), data.get("ean"), data.get("nif"),
                data.get("telephone"), data.get("telecopie"), data.get("email"),
                data.get("representant"), data.get("adv"),
                data.get("categorie1"), data.get("categorie2"), data.get("categorie3"),
                data.get("mode_livraison"), data.get("mode_reglement"),
                data.get("devise"), data.get("encours_autorise"),
                data.get("code_comptable"), data.get("etat"),
                data.get("contact_nom"), data.get("contact_fonction"),
                data.get("contact_email"), data.get("contact_tel"), data.get("notes"),
                now, client_id,
            ),
        )
        conn.commit()

    log_action(
        user=user,
        action="UPDATE",
        module="settings",
        objet=f"Client · {raison}",
        ip=request.client.host if request.client else None,
    )
    return {"success": True}


@router.delete("/api/clients/{client_id}")
def delete_client(client_id: int, request: Request):
    user = require_superadmin(request)
    from database import get_db
    with get_db() as conn:
        row = conn.execute("SELECT raison_sociale FROM clients WHERE id=?", (client_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Client introuvable")
        raison = row["raison_sociale"]
        conn.execute("DELETE FROM clients WHERE id=?", (client_id,))
        conn.commit()
    log_action(
        user=user,
        action="DELETE",
        module="settings",
        objet=f"Client · {raison}",
        ip=request.client.host if request.client else None,
    )
    return {"success": True}


@router.post("/api/clients/import-xlsx")
async def import_clients_xlsx(
    request: Request,
    file: UploadFile = File(...),
    mode: str = "merge",
):
    """Import xlsx — mode 'merge' (upsert par code) ou 'replace' (vide la table)."""
    user = require_superadmin(request)
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Le fichier doit être au format Excel (.xlsx).")
    content = await file.read()
    if not content:
        raise HTTPException(400, "Fichier vide.")
    if mode not in ("merge", "replace"):
        raise HTTPException(400, "Mode invalide (merge ou replace).")

    headers, rows = _xlsx_to_rows(content)

    # Construction du mapping colonne→idx
    col_idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        key = _norm_header(h)
        if key in _HEADER_MAP:
            col_idx[_HEADER_MAP[key]] = i

    if "raison_sociale" not in col_idx:
        raise HTTPException(
            400,
            "Colonne « Raison sociale » introuvable dans le fichier. "
            "Vérifiez que les en-têtes sont sur la première ligne.",
        )

    def cell(r, key):
        i = col_idx.get(key)
        if i is None or i >= len(r):
            return None
        return r[i]

    from database import get_db
    now = datetime.now().isoformat(timespec="seconds")
    inserted = updated = skipped = 0
    errors: list[str] = []

    with get_db() as conn:
        if mode == "replace":
            conn.execute("DELETE FROM clients")

        for ri, r in enumerate(rows, start=2):  # ligne 2 = 1ère ligne de données
            try:
                raison = _clean_str(cell(r, "raison_sociale"))
                if not raison:
                    skipped += 1
                    continue
                code = _clean_str(cell(r, "code"))
                payload = {
                    "numero": _clean_int(cell(r, "numero")),
                    "code": code,
                    "raison_sociale": raison,
                    "adresse1": _clean_str(cell(r, "adresse1")),
                    "adresse2": _clean_str(cell(r, "adresse2")),
                    "bp": _clean_str(cell(r, "bp")),
                    "cp": _clean_str(cell(r, "cp")),
                    "ville": _clean_str(cell(r, "ville")),
                    "code_pays": _clean_str(cell(r, "code_pays")),
                    "pays": _clean_str(cell(r, "pays")),
                    "groupe": _clean_str(cell(r, "groupe")),
                    "siret": _clean_str(cell(r, "siret")),
                    "rcs": _clean_str(cell(r, "rcs")),
                    "tva": _clean_str(cell(r, "tva")),
                    "ean": _clean_str(cell(r, "ean")),
                    "nif": _clean_str(cell(r, "nif")),
                    "telephone": _clean_str(cell(r, "telephone")),
                    "telecopie": _clean_str(cell(r, "telecopie")),
                    "email": _clean_str(cell(r, "email")),
                    "representant": _clean_str(cell(r, "representant")),
                    "adv": _clean_str(cell(r, "adv")),
                    "categorie1": _clean_str(cell(r, "categorie1")),
                    "categorie2": _clean_str(cell(r, "categorie2")),
                    "categorie3": _clean_str(cell(r, "categorie3")),
                    "mode_livraison": _clean_str(cell(r, "mode_livraison")),
                    "mode_reglement": _clean_str(cell(r, "mode_reglement")),
                    "devise": _clean_str(cell(r, "devise")),
                    "encours_autorise": _clean_float(cell(r, "encours_autorise")),
                    "code_comptable": _clean_str(cell(r, "code_comptable")),
                    "etat": _clean_str(cell(r, "etat")) or "Normal",
                    "date_creation": _clean_str(cell(r, "date_creation")),
                    "date_modification": _clean_str(cell(r, "date_modification")),
                }

                existing = None
                if mode == "merge" and code:
                    existing = conn.execute(
                        "SELECT id FROM clients WHERE code=? COLLATE NOCASE", (code,)
                    ).fetchone()

                if existing:
                    conn.execute(
                        """UPDATE clients SET
                            numero=?, raison_sociale=?, adresse1=?, adresse2=?, bp=?, cp=?, ville=?,
                            code_pays=?, pays=?, groupe=?, siret=?, rcs=?, tva=?, ean=?, nif=?,
                            telephone=?, telecopie=?, email=?, representant=?, adv=?,
                            categorie1=?, categorie2=?, categorie3=?, mode_livraison=?, mode_reglement=?,
                            devise=?, encours_autorise=?, code_comptable=?, etat=?,
                            date_creation=?, date_modification=?, updated_at=?
                           WHERE id=?""",
                        (
                            payload["numero"], payload["raison_sociale"],
                            payload["adresse1"], payload["adresse2"], payload["bp"],
                            payload["cp"], payload["ville"], payload["code_pays"],
                            payload["pays"], payload["groupe"], payload["siret"],
                            payload["rcs"], payload["tva"], payload["ean"], payload["nif"],
                            payload["telephone"], payload["telecopie"], payload["email"],
                            payload["representant"], payload["adv"],
                            payload["categorie1"], payload["categorie2"], payload["categorie3"],
                            payload["mode_livraison"], payload["mode_reglement"],
                            payload["devise"], payload["encours_autorise"],
                            payload["code_comptable"], payload["etat"],
                            payload["date_creation"], payload["date_modification"],
                            now, existing["id"],
                        ),
                    )
                    updated += 1
                else:
                    conn.execute(
                        """INSERT INTO clients (
                            numero, code, raison_sociale, adresse1, adresse2, bp, cp, ville,
                            code_pays, pays, groupe, siret, rcs, tva, ean, nif,
                            telephone, telecopie, email, representant, adv,
                            categorie1, categorie2, categorie3, mode_livraison, mode_reglement,
                            devise, encours_autorise, code_comptable, etat,
                            date_creation, date_modification, created_at, updated_at
                           ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (
                            payload["numero"], payload["code"], payload["raison_sociale"],
                            payload["adresse1"], payload["adresse2"], payload["bp"],
                            payload["cp"], payload["ville"], payload["code_pays"],
                            payload["pays"], payload["groupe"], payload["siret"],
                            payload["rcs"], payload["tva"], payload["ean"], payload["nif"],
                            payload["telephone"], payload["telecopie"], payload["email"],
                            payload["representant"], payload["adv"],
                            payload["categorie1"], payload["categorie2"], payload["categorie3"],
                            payload["mode_livraison"], payload["mode_reglement"],
                            payload["devise"], payload["encours_autorise"],
                            payload["code_comptable"], payload["etat"],
                            payload["date_creation"], payload["date_modification"],
                            now, now,
                        ),
                    )
                    inserted += 1
            except Exception as exc:
                skipped += 1
                if len(errors) < 10:
                    errors.append(f"Ligne {ri} : {exc}")
        conn.commit()

    log_action(
        user=user,
        action="IMPORT",
        module="settings",
        objet="Clients · import xlsx",
        detail={"mode": mode, "inserted": inserted, "updated": updated, "skipped": skipped},
        ip=request.client.host if request.client else None,
    )
    return {
        "success": True,
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }
