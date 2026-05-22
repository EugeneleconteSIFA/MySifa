import csv
import io
import re
import unicodedata
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import sqlite3
from fastapi import APIRouter, File, HTTPException, Request, UploadFile

from database import get_db
from services.auth_service import get_current_user, user_has_app_access

router = APIRouter()


def _require_compta(request: Request) -> dict:
    u = get_current_user(request)
    if not user_has_app_access(u, "compta"):
        raise HTTPException(status_code=403, detail="Accès réservé à MyCompta")
    return u


def _norm(s: str) -> str:
    s = str(s or "").strip().lower()
    s = "".join(
        c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c)
    )
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _libelle_key(libelle_condense: str) -> str:
    s = _norm(libelle_condense)
    # Exemple: "Achat de Factures No 2050267" -> "achat de factures"
    s = re.sub(r"\bno\b\s*\d+\b", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\bnum(ero)?\b\s*\d+\b", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _norm_code_vendeur(raw: Any) -> str:
    s = str(raw or "").strip()
    if not s:
        return ""
    try:
        f = float(s.replace(",", "."))
        if abs(f - round(f)) < 1e-9:
            return str(int(round(f)))
    except ValueError:
        pass
    return s


def _buyer_name(raw: str) -> str:
    s = str(raw or "").strip()
    if "/" in s:
        s = s.split("/", 1)[0]
    return _norm(s)


def _extract_siret(complement: str) -> Optional[str]:
    s = str(complement or "")
    # Cas "Siret: 399281468 00039"
    m = re.search(r"siret\s*:\s*([\d\s]{9,20})", s, flags=re.IGNORECASE)
    if m:
        digits = re.sub(r"\D", "", m.group(1))
        if len(digits) >= 14:
            return digits[:14]
    # Cas fallback: trouver un bloc de 14 chiffres n'importe où
    m2 = re.search(r"(\d[\d\s]{12,20}\d)", s)
    if m2:
        digits = re.sub(r"\D", "", m2.group(1))
        if len(digits) >= 14:
            return digits[:14]
    return None


def _parse_amount(x: Any) -> float:
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if not s:
        return 0.0
    s = s.replace("\u202f", "").replace("\xa0", "").replace(" ", "")
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _fmt_date(dt: Any) -> str:
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m-%d")
    s = str(dt or "").strip()
    if " " in s:
        return s.split(" ", 1)[0]
    return s


@router.get("/api/compta/acheteurs")
def list_acheteurs(request: Request, q: str = ""):
    _require_compta(request)
    qq = _norm(q)
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM compta_acheteurs ORDER BY raison_sociale ASC"
        ).fetchall()
    out = [dict(r) for r in rows]
    if qq:
        out = [
            r
            for r in out
            if qq in _norm(r.get("raison_sociale") or "")
            or qq in _norm(r.get("identifiant") or "")
        ]
    return out


@router.post("/api/compta/acheteurs")
async def upsert_acheteurs(request: Request):
    _require_compta(request)
    body = await request.json()
    items = body.get("items") if isinstance(body, dict) else None
    if not isinstance(items, list) or len(items) == 0:
        raise HTTPException(status_code=400, detail="items requis")
    now = datetime.now().isoformat()
    with get_db() as conn:
        for it in items:
            if not isinstance(it, dict):
                continue
            code_vendeur = (it.get("code_vendeur") or "").strip() or None
            identifiant = str(it.get("identifiant") or "").strip()
            rs = str(it.get("raison_sociale") or "").strip()
            if not identifiant or not rs:
                continue
            conn.execute(
                """INSERT INTO compta_acheteurs (code_vendeur,identifiant,raison_sociale,created_at,updated_at)
                   VALUES (?,?,?,?,?)
                   ON CONFLICT(code_vendeur,identifiant) DO UPDATE SET
                     raison_sociale=excluded.raison_sociale,
                     updated_at=excluded.updated_at
                """,
                (code_vendeur, identifiant, rs, now, now),
            )
        conn.commit()
    return {"success": True}


@router.delete("/api/compta/acheteurs/{acheteur_id}")
def delete_acheteur(acheteur_id: int, request: Request):
    _require_compta(request)
    with get_db() as conn:
        conn.execute("DELETE FROM compta_acheteurs WHERE id=?", (acheteur_id,))
        conn.commit()
    return {"success": True}


@router.put("/api/compta/acheteurs/{acheteur_id}")
async def update_acheteur(acheteur_id: int, request: Request):
    _require_compta(request)
    body = await request.json()
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="JSON invalide")
    code_vendeur = (body.get("code_vendeur") or "").strip() or None
    identifiant = str(body.get("identifiant") or "").strip()
    rs = str(body.get("raison_sociale") or "").strip()
    if not identifiant or not rs:
        raise HTTPException(status_code=400, detail="identifiant et raison_sociale requis")
    now = datetime.now().isoformat()
    with get_db() as conn:
        ex = conn.execute(
            "SELECT id FROM compta_acheteurs WHERE id=?",
            (acheteur_id,),
        ).fetchone()
        if not ex:
            raise HTTPException(status_code=404, detail="Acheteur introuvable")
        try:
            conn.execute(
                """UPDATE compta_acheteurs
                   SET code_vendeur=?, identifiant=?, raison_sociale=?, updated_at=?
                   WHERE id=?""",
                (code_vendeur, identifiant, rs, now, acheteur_id),
            )
            conn.commit()
        except sqlite3.IntegrityError:
            raise HTTPException(
                status_code=409,
                detail="Conflit: un acheteur existe déjà avec ce code vendeur + identifiant",
            )
    return {"success": True}


@router.get("/api/compta/banques")
def list_banques(request: Request, q: str = ""):
    _require_compta(request)
    qq = _norm(q)
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM compta_banques ORDER BY code_vendeur ASC"
        ).fetchall()
    out = [dict(r) for r in rows]
    if qq:
        out = [
            r
            for r in out
            if qq in _norm(r.get("code_vendeur") or "")
            or qq in _norm(r.get("numero_compte") or "")
            or qq in _norm(r.get("libelle") or "")
        ]
    return out


@router.post("/api/compta/banques")
async def upsert_banques(request: Request):
    _require_compta(request)
    body = await request.json()
    items = body.get("items") if isinstance(body, dict) else None
    if not isinstance(items, list) or len(items) == 0:
        raise HTTPException(status_code=400, detail="items requis")
    now = datetime.now().isoformat()
    with get_db() as conn:
        for it in items:
            if not isinstance(it, dict):
                continue
            code = _norm_code_vendeur(it.get("code_vendeur"))
            num = str(it.get("numero_compte") or "").strip()
            lib = str(it.get("libelle") or "").strip() or None
            if not code or not num:
                continue
            conn.execute(
                """INSERT INTO compta_banques (code_vendeur,numero_compte,libelle,created_at,updated_at)
                   VALUES (?,?,?,?,?)
                   ON CONFLICT(code_vendeur) DO UPDATE SET
                     numero_compte=excluded.numero_compte,
                     libelle=excluded.libelle,
                     updated_at=excluded.updated_at
                """,
                (code, num, lib, now, now),
            )
        conn.commit()
    return {"success": True}


@router.delete("/api/compta/banques/{banque_id}")
def delete_banque(banque_id: int, request: Request):
    _require_compta(request)
    with get_db() as conn:
        conn.execute("DELETE FROM compta_banques WHERE id=?", (banque_id,))
        conn.commit()
    return {"success": True}


@router.put("/api/compta/banques/{banque_id}")
async def update_banque(banque_id: int, request: Request):
    _require_compta(request)
    body = await request.json()
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="JSON invalide")
    code = _norm_code_vendeur(body.get("code_vendeur"))
    num = str(body.get("numero_compte") or "").strip()
    lib = str(body.get("libelle") or "").strip() or None
    if not code or not num:
        raise HTTPException(
            status_code=400, detail="code_vendeur et numero_compte requis"
        )
    now = datetime.now().isoformat()
    with get_db() as conn:
        ex = conn.execute(
            "SELECT id FROM compta_banques WHERE id=?", (banque_id,)
        ).fetchone()
        if not ex:
            raise HTTPException(status_code=404, detail="Code de banque introuvable")
        try:
            conn.execute(
                """UPDATE compta_banques
                   SET code_vendeur=?, numero_compte=?, libelle=?, updated_at=?
                   WHERE id=?""",
                (code, num, lib, now, banque_id),
            )
            conn.commit()
        except sqlite3.IntegrityError:
            raise HTTPException(
                status_code=409,
                detail="Conflit: ce code vendeur existe déjà",
            )
    return {"success": True}


@router.get("/api/compta/comptes")
def list_comptes(request: Request, q: str = ""):
    _require_compta(request)
    qq = _norm(q)
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM compta_comptes ORDER BY libelle_condense ASC"
        ).fetchall()
    out = [dict(r) for r in rows]
    if qq:
        out = [
            r
            for r in out
            if qq in _norm(r.get("libelle_condense") or "")
            or qq in _norm(r.get("numero_compte") or "")
        ]
    return out


@router.post("/api/compta/comptes")
async def upsert_comptes(request: Request):
    _require_compta(request)
    body = await request.json()
    items = body.get("items") if isinstance(body, dict) else None
    if not isinstance(items, list) or len(items) == 0:
        raise HTTPException(status_code=400, detail="items requis")
    now = datetime.now().isoformat()
    with get_db() as conn:
        for it in items:
            if not isinstance(it, dict):
                continue
            lib = str(it.get("libelle_condense") or "").strip()
            num = str(it.get("numero_compte") or "").strip()
            if not lib or not num:
                continue
            key = _libelle_key(lib)
            conn.execute(
                """INSERT INTO compta_comptes (libelle_condense,libelle_key,numero_compte,created_at,updated_at)
                   VALUES (?,?,?,?,?)
                   ON CONFLICT(libelle_key) DO UPDATE SET
                     libelle_condense=excluded.libelle_condense,
                     numero_compte=excluded.numero_compte,
                     updated_at=excluded.updated_at
                """,
                (lib, key, num, now, now),
            )
        conn.commit()
    return {"success": True}


@router.delete("/api/compta/comptes/{compte_id}")
def delete_compte(compte_id: int, request: Request):
    _require_compta(request)
    with get_db() as conn:
        conn.execute("DELETE FROM compta_comptes WHERE id=?", (compte_id,))
        conn.commit()
    return {"success": True}


@router.put("/api/compta/comptes/{compte_id}")
async def update_compte(compte_id: int, request: Request):
    _require_compta(request)
    body = await request.json()
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="JSON invalide")
    lib = str(body.get("libelle_condense") or "").strip()
    num = str(body.get("numero_compte") or "").strip()
    if not lib or not num:
        raise HTTPException(status_code=400, detail="libelle_condense et numero_compte requis")
    key = _libelle_key(lib)
    if not key:
        raise HTTPException(status_code=400, detail="Libellé invalide")
    now = datetime.now().isoformat()
    with get_db() as conn:
        ex = conn.execute("SELECT id FROM compta_comptes WHERE id=?", (compte_id,)).fetchone()
        if not ex:
            raise HTTPException(status_code=404, detail="Compte introuvable")
        try:
            conn.execute(
                """UPDATE compta_comptes
                   SET libelle_condense=?, libelle_key=?, numero_compte=?, updated_at=?
                   WHERE id=?""",
                (lib, key, num, now, compte_id),
            )
            conn.commit()
        except sqlite3.IntegrityError:
            raise HTTPException(
                status_code=409,
                detail="Conflit: un libellé existe déjà (clé de libellé déjà utilisée)",
            )
    return {"success": True}


def _load_banques(conn) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for r in conn.execute(
        "SELECT code_vendeur, numero_compte FROM compta_banques"
    ).fetchall():
        cv = _norm_code_vendeur(r["code_vendeur"])
        if cv:
            out[cv] = str(r["numero_compte"] or "").strip()
    return out


def _load_mappings(conn) -> Tuple[Dict[str, str], Dict[Tuple[Optional[str], str], str]]:
    comptes = {
        (r["libelle_key"] or ""): (r["numero_compte"] or "")
        for r in conn.execute("SELECT libelle_key, numero_compte FROM compta_comptes").fetchall()
    }
    acheteurs = {}
    for r in conn.execute(
        "SELECT code_vendeur, identifiant, raison_sociale FROM compta_acheteurs"
    ).fetchall():
        cv = _norm_code_vendeur(r["code_vendeur"]) or None
        rs_key = _norm(r["raison_sociale"] or "")
        if not rs_key:
            continue
        acheteurs[(cv, rs_key)] = str(r["identifiant"] or "").strip()
    return comptes, acheteurs


def _sheet_by_name(wb, candidates: List[str]) -> Optional[str]:
    for cand in candidates:
        if cand in wb.sheetnames:
            return cand
    # case-insensitive fallback
    lower = {str(n).strip().lower(): n for n in wb.sheetnames}
    for cand in candidates:
        n = lower.get(str(cand).strip().lower())
        if n:
            return n
    return None


def _find_header_row(ws, required_headers: List[str], max_scan_rows: int = 30) -> Tuple[int, Dict[str, int]]:
    req = [str(x) for x in required_headers]
    for r_i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        if not row:
            continue
        headers = [("" if v is None else str(v).strip()) for v in row]
        if not any(headers):
            continue
        idx = {h: i for i, h in enumerate(headers) if h}
        if all(h in idx for h in req):
            return r_i, idx
    raise HTTPException(status_code=400, detail="En-têtes introuvables dans la feuille Excel")


def _hdr_norm(s: str) -> str:
    # Normalize headers to tolerate casing/accents/quotes differences (’, ', etc.)
    s = _norm(str(s or ""))
    s = (
        s.replace("’", "'")
        .replace("‘", "'")
        .replace("", "'")  # latin-1 misread of cp1252 apostrophe
        .replace("`", "'")
    )
    return s


_FACTOR_REQUIRED = [
    "Code vendeur",
    "Date comptable de l'écriture",
    "Libellé condensé",
    "Montant du débit",
    "Montant du crédit",
    "Données de l'acheteur concerné par l'opération",
    "Complément sur l'acheteur concerné par l'opération",
]


def _find_header_row_from_list(
    rows: List[List[Any]], required_headers: List[str], max_scan_rows: int = 40
) -> Tuple[int, Dict[str, int]]:
    req = [_hdr_norm(x) for x in required_headers]
    for r_i, row in enumerate(rows[:max_scan_rows]):
        if not row:
            continue
        headers = [("" if v is None else str(v).strip()) for v in row]
        if not any(headers):
            continue
        idx = {_hdr_norm(h): i for i, h in enumerate(headers) if str(h or "").strip()}
        if all(h in idx for h in req):
            return r_i, idx
    raise HTTPException(status_code=400, detail="En-têtes Factor introuvables (vérifiez les colonnes du fichier ou du collage)")


def _decode_bytes(contents: bytes) -> str:
    # cp1252 before latin-1: latin-1 never fails but mangles Factor exports (apostrophe → U+0092)
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return contents.decode(enc)
        except UnicodeDecodeError:
            continue
    return contents.decode("utf-8", errors="replace")


def _parse_delimited_text(text: str) -> List[List[str]]:
    t = (text or "").strip()
    if not t:
        raise HTTPException(status_code=400, detail="Contenu vide")
    sample = t[:8192]
    delimiter = ";"
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        first = t.split("\n", 1)[0]
        if "\t" in first:
            delimiter = "\t"
        elif ";" in first:
            delimiter = ";"
        elif "," in first:
            delimiter = ","
    reader = csv.reader(io.StringIO(t), delimiter=delimiter)
    rows: List[List[str]] = []
    for row in reader:
        if not row or all(not str(c or "").strip() for c in row):
            continue
        rows.append([str(c or "").strip() for c in row])
    if not rows:
        raise HTTPException(status_code=400, detail="Aucune ligne lisible dans le contenu")
    return rows


def _load_table_from_bytes(contents: bytes, filename: str = "") -> List[List[Any]]:
    fn = (filename or "").lower()
    if fn.endswith(".csv") or fn.endswith(".txt"):
        return _parse_delimited_text(_decode_bytes(contents))
    if not fn.endswith((".xlsx", ".xlsm", ".xls")):
        head = contents[:2048]
        if b";" in head or b"\t" in head or (b"," in head and b"\n" in head):
            try:
                return _parse_delimited_text(_decode_bytes(contents))
            except HTTPException:
                pass
    return _load_table_from_excel(contents)


def _load_table_from_excel(contents: bytes) -> List[List[Any]]:
    wb = _load_wb_from_upload(contents)
    sheet_name = _sheet_by_name(wb, ["IMPORT DU FACTOR", "Import du factor", "IMPORT", "FACTOR"])
    if not sheet_name and len(getattr(wb, "sheetnames", []) or []) == 1:
        sheet_name = wb.sheetnames[0]
    if not sheet_name:
        raise HTTPException(
            status_code=400,
            detail="Onglet 'IMPORT DU FACTOR' introuvable (ou fichier multi-feuilles non reconnu)",
        )
    ws = wb[sheet_name]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _load_table_from_text(text: str) -> List[List[Any]]:
    return _parse_delimited_text(text)


def _transform_factor_table(table_rows: List[List[Any]]) -> Dict[str, Any]:
    header_row, idx = _find_header_row_from_list(table_rows, _FACTOR_REQUIRED, max_scan_rows=40)
    i_code = idx[_hdr_norm("Code vendeur")]
    i_date = idx[_hdr_norm("Date comptable de l'écriture")]
    i_lib = idx[_hdr_norm("Libellé condensé")]
    i_deb = idx[_hdr_norm("Montant du débit")]
    i_cred = idx[_hdr_norm("Montant du crédit")]
    i_buy = idx[_hdr_norm("Données de l'acheteur concerné par l'opération")]
    i_comp = idx[_hdr_norm("Complément sur l'acheteur concerné par l'opération")]

    with get_db() as conn:
        comptes, acheteurs = _load_mappings(conn)
        banques = _load_banques(conn)

    out_rows: List[Dict[str, Any]] = []
    missing_accounts: Dict[str, int] = {}
    missing_buyers: Dict[str, int] = {}
    missing_banques: Dict[str, int] = {}

    for row in table_rows[header_row + 1 :]:
        if not row or all(v in (None, "") for v in row):
            continue
        if len(row) <= max(i_code, i_date, i_lib, i_deb, i_cred, i_buy, i_comp):
            continue

        code_vendeur = _norm_code_vendeur(row[i_code])
        date_ecr = _fmt_date(row[i_date])
        lib_raw = str(row[i_lib] or "").strip()
        debit = _parse_amount(row[i_deb])
        credit = _parse_amount(row[i_cred])
        buyer_raw = str(row[i_buy] or "").strip()
        comp_raw = str(row[i_comp] or "").strip()

        if debit == 0.0 and credit == 0.0:
            continue

        main_compte = None
        main_libelle = lib_raw
        problem = None
        problem_detail = None
        if buyer_raw:
            main_libelle = buyer_raw
            bname = _buyer_name(buyer_raw)
            # Table Acheteurs = source de vérité (prioritaire sur le SIRET du complément Factor)
            main_compte = (
                acheteurs.get((code_vendeur or None, bname))
                or acheteurs.get((None, bname))
                or _extract_siret(comp_raw)
                or ""
            )
            if not main_compte:
                key_b = bname or buyer_raw
                missing_buyers[key_b] = missing_buyers.get(key_b, 0) + 1
                main_compte = ""
                problem = "acheteur_inconnu"
                problem_detail = key_b
        else:
            key = _libelle_key(lib_raw)
            main_compte = comptes.get(key)
            if not main_compte:
                missing_accounts[key or lib_raw] = missing_accounts.get(key or lib_raw, 0) + 1
                main_compte = ""
                problem = "compte_manquant"
                problem_detail = key or lib_raw

        caf_compte = banques.get(code_vendeur, "")
        caf_problem = problem
        caf_detail = problem_detail
        if not caf_compte:
            key_b = code_vendeur or "?"
            missing_banques[key_b] = missing_banques.get(key_b, 0) + 1
            caf_problem = "banque_manquante"
            caf_detail = key_b

        out_rows.append(
            {
                "date": date_ecr,
                "code_vendeur": code_vendeur,
                "compte": main_compte,
                "libelle": main_libelle,
                "debit": debit,
                "credit": credit,
                "problem": problem,
                "problem_detail": problem_detail,
            }
        )
        out_rows.append(
            {
                "date": date_ecr,
                "code_vendeur": code_vendeur,
                "compte": str(caf_compte),
                "libelle": main_libelle,
                "debit": credit,
                "credit": debit,
                "problem": caf_problem,
                "problem_detail": caf_detail,
            }
        )

    def f(x: float) -> str:
        if x is None:
            return "0"
        return ("%.2f" % float(x)).rstrip("0").rstrip(".") if float(x) % 1 else str(int(float(x)))

    cw_lines = [
        "\t".join(
            [
                str(r["date"] or ""),
                str(r["code_vendeur"] or ""),
                str(r["compte"] or ""),
                str(r["libelle"] or ""),
                f(r["debit"]),
                f(r["credit"]),
            ]
        )
        for r in out_rows
    ]

    return {
        "success": True,
        "rows": out_rows,
        "cw_text": "\n".join(cw_lines),
        "missing": {
            "comptes": [
                {"libelle_key": k, "count": v}
                for k, v in sorted(missing_accounts.items(), key=lambda kv: -kv[1])
            ],
            "acheteurs": [
                {"buyer": k, "count": v}
                for k, v in sorted(missing_buyers.items(), key=lambda kv: -kv[1])
            ],
            "banques": [
                {"code_vendeur": k, "count": v}
                for k, v in sorted(missing_banques.items(), key=lambda kv: -kv[1])
            ],
        },
    }


def _load_wb_from_upload(contents: bytes):
    from io import BytesIO
    from openpyxl import load_workbook

    try:
        return load_workbook(BytesIO(contents), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel invalide: {e}")


@router.post("/api/compta/transform")
async def transform_factor(
    request: Request,
    file: UploadFile = File(...),
):
    _require_compta(request)
    contents = await file.read()
    filename = file.filename or "import.xlsx"
    try:
        table = _load_table_from_bytes(contents, filename)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Fichier illisible: {e}")
    return _transform_factor_table(table)


@router.post("/api/compta/transform-paste")
async def transform_factor_paste(request: Request):
    _require_compta(request)
    body = await request.json()
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="JSON invalide")
    text = str(body.get("text") or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Collez au moins une ligne (en-têtes Factor inclus)")
    table = _load_table_from_text(text)
    return _transform_factor_table(table)


@router.post("/api/compta/import-acheteurs")
async def import_acheteurs_from_excel(request: Request, file: UploadFile = File(...)):
    _require_compta(request)
    contents = await file.read()
    wb = _load_wb_from_upload(contents)
    sheet_name = _sheet_by_name(wb, ["ACHETEURS", "acheteurs", "Acheteurs"])
    if not sheet_name:
        raise HTTPException(status_code=400, detail="Onglet 'ACHETEURS' introuvable")
    ws = wb[sheet_name]

    header_row, idx = _find_header_row(ws, ["Identifiant", "Raison sociale"])
    i_ident = idx["Identifiant"]
    i_rs = idx["Raison sociale"]
    # Code vendeur: soit une colonne explicitement, soit la 1ère colonne si différente des deux
    i_code = idx.get("Code vendeur", 0)

    now = datetime.now().isoformat()
    inserted = skipped = 0
    with get_db() as conn:
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or all(v in (None, "") for v in row):
                continue
            code_v = str(row[i_code] or "").strip() or None
            ident = str(row[i_ident] or "").strip()
            rs = str(row[i_rs] or "").strip()
            if not ident or not rs:
                skipped += 1
                continue
            # UPSERT
            cur = conn.execute(
                """INSERT INTO compta_acheteurs (code_vendeur,identifiant,raison_sociale,created_at,updated_at)
                   VALUES (?,?,?,?,?)
                   ON CONFLICT(code_vendeur,identifiant) DO UPDATE SET
                     raison_sociale=excluded.raison_sociale,
                     updated_at=excluded.updated_at
                """,
                (code_v, ident, rs, now, now),
            )
            inserted += 1 if cur.rowcount == 1 else 0
        conn.commit()
    return {"success": True, "inserted": inserted, "updated": None, "skipped": skipped}


@router.post("/api/compta/import-comptes")
async def import_comptes_from_excel(request: Request, file: UploadFile = File(...)):
    _require_compta(request)
    contents = await file.read()
    wb = _load_wb_from_upload(contents)
    sheet_name = _sheet_by_name(wb, ["TABLE DES COMPTES", "Table des comptes", "table des comptes"])
    if not sheet_name:
        raise HTTPException(status_code=400, detail="Onglet 'TABLE DES COMPTES' introuvable")
    ws = wb[sheet_name]

    header_row, idx = _find_header_row(ws, ["Libellé condensé", "Numéro de compte"])
    i_lib = idx["Libellé condensé"]
    i_num = idx["Numéro de compte"]

    now = datetime.now().isoformat()
    inserted = skipped = 0
    with get_db() as conn:
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or all(v in (None, "") for v in row):
                continue
            lib = str(row[i_lib] or "").strip()
            num = str(row[i_num] or "").strip()
            if not lib or not num:
                skipped += 1
                continue
            key = _libelle_key(lib)
            cur = conn.execute(
                """INSERT INTO compta_comptes (libelle_condense,libelle_key,numero_compte,created_at,updated_at)
                   VALUES (?,?,?,?,?)
                   ON CONFLICT(libelle_key) DO UPDATE SET
                     libelle_condense=excluded.libelle_condense,
                     numero_compte=excluded.numero_compte,
                     updated_at=excluded.updated_at
                """,
                (lib, key, num, now, now),
            )
            inserted += 1 if cur.rowcount == 1 else 0
        conn.commit()
    return {"success": True, "inserted": inserted, "updated": None, "skipped": skipped}

