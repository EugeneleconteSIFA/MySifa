"""SIFA — MyStock v2.0
FIFO lots, inventaire, recherche instantanée, mobile-first.
Accès : direction, administration, logistique.
"""
import csv
import io
import json
import re
import sqlite3
from datetime import datetime, timedelta
from typing import Any, Optional

import pandas as pd
from fastapi import APIRouter, File, HTTPException, Request, UploadFile
from pydantic import BaseModel
from fastapi.responses import StreamingResponse

from app.services.audit_service import log_action
from config import (
    STOCK_EMPLACEMENT_AU_SOL,
    STOCK_EMPLACEMENT_AU_SOL_LABEL,
    STOCK_EMPLACEMENT_SORTIE_PROD,
    STOCK_EMPLACEMENT_SORTIE_PROD_LABEL,
)
from database import get_db, parse_file
from services.auth_service import get_current_user, user_has_app_access

router = APIRouter()

INVENTAIRE_ALERTE_JOURS = 180  # 6 mois


def _normalize_emplacement(code: str) -> str:
    return (code or "").strip().upper()


def _is_valid_emplacement(code: str) -> bool:
    """Code grille (A121…) ou zones spéciales Z0 / Z1."""
    empl = _normalize_emplacement(code)
    if not empl:
        return False
    if empl in (STOCK_EMPLACEMENT_AU_SOL, STOCK_EMPLACEMENT_SORTIE_PROD):
        return True
    return empl[0].isalpha() and empl[1:].isdigit()

# Colonnes mouvements (évite m.* + collision avec join users)
_MVT_FIELDS = (
    "m.id, m.produit_id, m.emplacement, m.type_mouvement, m.quantite, "
    "m.quantite_avant, m.quantite_apres, m.note, m.created_at, m.created_by, m.created_by_name"
)

_STOCK_USER_JOIN = (
    "LEFT JOIN users u ON LOWER(TRIM(COALESCE(u.email,''))) = "
    "LOWER(TRIM(COALESCE(m.created_by,'')))"
)

# Références type XXXX-XXXX / XXXX/XXXX : même résultat si l’utilisateur omet ou change -, /, espaces.
_NORM_REF_SQL = (
    "REPLACE(REPLACE(REPLACE(UPPER(TRIM(p.reference)), '-', ''), '/', ''), ' ', '')"
)
_NORM_DES_SQL = (
    "REPLACE(REPLACE(REPLACE(UPPER(IFNULL(p.designation,'')), '-', ''), '/', ''), ' ', '')"
)


_CLIENT_REF_SQL = (
    "CASE WHEN INSTR(p.reference, '/') > 0 "
    "THEN UPPER(SUBSTR(p.reference, 1, INSTR(p.reference, '/') - 1)) "
    "ELSE UPPER(TRIM(p.reference)) END"
)


def _produit_client_search_where_args(client: str) -> tuple[str, list]:
    """Filtre par numéro client (segment avant le / dans la référence produit)."""
    client = (client or "").strip().upper()
    if not client:
        return "(1=0)", []
    return f"({_CLIENT_REF_SQL} LIKE ?)", [f"{client}%"]


def _produit_search_where_args(q: str) -> tuple[str, list]:
    """Clause WHERE + paramètres : concordance sur la chaîne tapée et sur la forme sans séparateurs."""
    q = (q or "").strip()
    pattern = f"%{q}%"
    if not q:
        return "(p.reference LIKE ? OR IFNULL(p.designation,'') LIKE ?)", [pattern, pattern]
    core = "".join(c for c in q.upper() if c not in "-/ ")
    if core:
        norm_pat = f"%{core}%"
        where = (
            f"({_NORM_REF_SQL} LIKE ? OR p.reference LIKE ? OR "
            f"{_NORM_DES_SQL} LIKE ? OR IFNULL(p.designation,'') LIKE ?)"
        )
        return where, [norm_pat, pattern, norm_pat, pattern]
    return "(p.reference LIKE ? OR IFNULL(p.designation,'') LIKE ?)", [pattern, pattern]


def require_stock(request: Request) -> dict:
    user = get_current_user(request)
    if not user_has_app_access(user, "stock"):
        raise HTTPException(403, "Accès réservé à la Direction, Administration et Logistique")
    return user


def require_stock_write(request: Request) -> dict:
    user = require_stock(request)
    if user.get("role") == "commercial":
        raise HTTPException(403, "Accès en lecture seule pour le rôle commercial")
    return user


_MP_CATEGORIES = frozenset({"mandrin", "palette", "adhesif", "carton", "frontal", "glassine", "complexe", "autre"})
_MP_CATEGORIES_LAIZEES = frozenset({"frontal", "glassine", "complexe"})


def _mp_is_laizee(categorie: str) -> bool:
    return (categorie or "").strip().lower() in _MP_CATEGORIES_LAIZEES


_MP_TYPES_MVT = frozenset({"entree", "sortie", "ajustement", "transfert"})
_STOCK_MATIERES_ADMIN_ROLES = frozenset({"superadmin", "direction", "administration", "administration_ventes", "administration_technique"})
_STOCK_VALORISATION_USD_ROLES = frozenset({"superadmin", "direction"})


def _mp_unite_gestion(categorie: str) -> str:
    """Unité de gestion du stock selon la catégorie matière première."""
    cat = (categorie or "").strip().lower()
    if cat in ("frontal", "glassine", "complexe"):
        return "bobine"
    if cat == "palette":
        return "palette"
    if cat == "carton":
        return "palette"
    if cat == "autre":
        return "unite"
    return "palette"


# Catégories qui ont une notion d'« unité d'achat » distincte de l'unité de gestion (palette).
# Le prix saisi est à l'unité d'achat ; le conditionnement (unites_par_palette) permet
# de calculer la valorisation = stock_palettes × unites_par_palette × prix_unitaire.
_MP_CATEGORIES_AVEC_CONDITIONNEMENT = frozenset({"carton", "adhesif", "mandrin"})


def _mp_unite_achat(categorie: str) -> str:
    """Unité d'achat (élément acheté à l'unité dans une palette).

    Renvoie la chaîne singulier — pluraliser au besoin côté UI.
    """
    cat = (categorie or "").strip().lower()
    if cat == "carton":
        return "carton"
    if cat == "adhesif":
        return "kg"
    if cat == "mandrin":
        return "tube"
    if cat == "palette":
        return "palette"
    if cat == "autre":
        return "unite"
    return ""


def _mp_a_conditionnement(categorie: str) -> bool:
    return (categorie or "").strip().lower() in _MP_CATEGORIES_AVEC_CONDITIONNEMENT


def _mp_row_dict(r, stock_par_laize: Optional[list[dict]] = None) -> dict:
    """Normalise une ligne matieres_premieres + stock pour l'API."""
    seuil = float(r["seuil_alerte"] or 0)
    qte = float(r["quantite"] or 0)
    cat = r["categorie"]
    ppp = r["palettes_par_pile"] if "palettes_par_pile" in r.keys() else None
    palettes_par_pile = float(ppp) if ppp is not None and float(ppp) > 0 else None
    keys = set(r.keys())
    metres = None
    prix_m2 = None
    if "metres_lineaires_par_bobine" in keys and r["metres_lineaires_par_bobine"] is not None:
        try:
            metres = float(r["metres_lineaires_par_bobine"])
        except (TypeError, ValueError):
            metres = None
    if "prix_eur_m2" in keys and r["prix_eur_m2"] is not None:
        try:
            prix_m2 = float(r["prix_eur_m2"])
        except (TypeError, ValueError):
            prix_m2 = None
    unites_par_palette = None
    if "unites_par_palette" in keys and r["unites_par_palette"] is not None:
        try:
            v = float(r["unites_par_palette"])
            unites_par_palette = v if v > 0 else None
        except (TypeError, ValueError):
            unites_par_palette = None
    laizee = _mp_is_laizee(cat)
    prix_par_laize = False
    if "prix_par_laize" in keys and r["prix_par_laize"] is not None:
        try:
            prix_par_laize = bool(int(r["prix_par_laize"]))
        except (TypeError, ValueError):
            prix_par_laize = False
    # Pour les matières laizées, la quantité totale = somme des stocks par laize
    if laizee and stock_par_laize is not None:
        qte = sum(float(x.get("quantite") or 0) for x in stock_par_laize)
    complete = True
    if laizee:
        has_laizes = bool(stock_par_laize) and len(stock_par_laize) > 0
        metres_ok = (metres or 0) > 0
        if prix_par_laize and stock_par_laize:
            # Mode prix par laize : chaque laize doit avoir un prix > 0
            prix_ok = all((float(spl.get("prix_eur_m2") or 0) > 0) for spl in stock_par_laize)
        else:
            prix_ok = (prix_m2 or 0) > 0
        complete = has_laizes and prix_ok and metres_ok
    sous_section = None
    if "sous_section" in r.keys() and r["sous_section"]:
        sous_section = str(r["sous_section"]).strip() or None
    return {
        "id": r["id"],
        "categorie": cat,
        "reference": r["reference"],
        "designation": r["designation"],
        "seuil_alerte": seuil,
        "actif": r["actif"],
        "quantite": qte,
        "en_alerte": seuil > 0 and qte <= seuil,
        "unite": _mp_unite_gestion(cat),
        "palettes_par_pile": palettes_par_pile,
        "couleur": (r["couleur"] or "").strip() if "couleur" in r.keys() and r["couleur"] else None,
        "laizee": laizee,
        "metres_lineaires_par_bobine": metres,
        "prix_eur_m2": prix_m2,
        "prix_par_laize": prix_par_laize,
        "stock_par_laize": stock_par_laize if laizee else None,
        "sous_section": sous_section,
        "avec_conditionnement": _mp_a_conditionnement(cat),
        "unites_par_palette": unites_par_palette,
        "unite_achat": _mp_unite_achat(cat) if _mp_a_conditionnement(cat) else None,
        "complete": complete,
    }


def require_stock_matieres_admin(request: Request) -> dict:
    user = require_stock(request)
    if user.get("role") not in _STOCK_MATIERES_ADMIN_ROLES:
        raise HTTPException(403, "Accès réservé à la Direction et Administration")
    return user


def _user_can_see_valorisation_usd(user: dict | None) -> bool:
    """Direction + superadmin uniquement voient la colonne « Prix unit. réel »
    et le KPI dédoublé en €/USD sur la page Valorisation."""
    return bool(user) and (user.get("role") in _STOCK_VALORISATION_USD_ROLES)


def require_valorisation_usd_admin(request: Request) -> dict:
    user = require_stock(request)
    if not _user_can_see_valorisation_usd(user):
        raise HTTPException(403, "Conversion EUR/USD réservée à la Direction.")
    return user


def _get_taux_eur_usd(conn) -> float:
    """Lit le paramètre Taux EUR / USD configuré dans MyCouts (mc_setting.eur_usd_rate).
    Retourne 0.0 si le paramètre n'est pas configuré — l'UI affichera alors « — »."""
    try:
        row = conn.execute(
            "SELECT value_decimal FROM mc_setting WHERE key='eur_usd_rate' LIMIT 1"
        ).fetchone()
        if not row:
            return 0.0
        return float(row["value_decimal"] or 0)
    except Exception:
        return 0.0


def _get_import_tax_pct(conn) -> float:
    """Lit le paramètre Taxe d'importation (mc_setting.import_tax_pct), en %.
    Retourne 0.0 si non configuré."""
    try:
        row = conn.execute(
            "SELECT value_decimal FROM mc_setting WHERE key='import_tax_pct' LIMIT 1"
        ).fetchone()
        if not row:
            return 0.0
        return float(row["value_decimal"] or 0)
    except Exception:
        return 0.0


def _get_transport_cost_fixed_eur(conn) -> float:
    """Lit le paramètre Coût transport forfaitaire (mc_setting.transport_cost_fixed_eur), en €.
    Retourne 0.0 si non configuré."""
    try:
        row = conn.execute(
            "SELECT value_decimal FROM mc_setting WHERE key='transport_cost_fixed_eur' LIMIT 1"
        ).fetchone()
        if not row:
            return 0.0
        return float(row["value_decimal"] or 0)
    except Exception:
        return 0.0


def _get_charge_production_pct(conn) -> float:
    """Lit le paramètre Charge de production (mc_setting.charge_production_pct), en %.
    Retourne 0.0 si non configuré. Bornée à [0, 99.9999[ (100 % impossible car diviseur).
    """
    try:
        row = conn.execute(
            "SELECT value_decimal FROM mc_setting WHERE key='charge_production_pct' LIMIT 1"
        ).fetchone()
        if not row:
            return 0.0
        val = float(row["value_decimal"] or 0)
        if val < 0:
            return 0.0
        if val >= 100:
            return 99.9999
        return val
    except Exception:
        return 0.0


def _get_container_params(conn) -> tuple[float, float, float, float]:
    """Retourne (container_cost_full_eur, container_cost_half_eur, qte_m2_full, qte_m2_half)
    pour le calcul du supplément transport container.

    Le libellé de default_container_cost_usd reste "usd" en base pour compat schéma,
    mais côté UI et calcul il est désormais interprété comme EUR (rename cosmétique
    demandé par Eugène le 2026-07-03).
    """
    def _read(key: str) -> float:
        try:
            row = conn.execute(
                "SELECT value_decimal FROM mc_setting WHERE key=? LIMIT 1", (key,)
            ).fetchone()
            if not row:
                return 0.0
            v = float(row["value_decimal"] or 0)
            return v if v >= 0 else 0.0
        except Exception:
            return 0.0
    return (
        _read("default_container_cost_usd"),           # coût container complet (EUR)
        _read("default_half_container_cost_eur"),      # coût demi-container (EUR)
        _read("logistique_qte_m2_container_complet"),  # qté m² container complet
        _read("logistique_qte_m2_demi_container"),     # qté m² demi-container
    )


# ─── Snapshot valorisation à une date passée ─────────────────────────────────
# Reconstitue les quantités en stock et les prix unitaires tels qu'ils étaient au
# soir de la date `date_iso` (YYYY-MM-DD). Les paramètres globaux (taux USD, taxe,
# containers, charges de prod) restent ceux d'aujourd'hui — c'est un choix produit :
# on veut « combien valait ce stock à l'époque, aux tarifs actuels ».
#
# Algo stock : pour chaque référence, on prend la `quantite_avant` du premier
# mouvement postérieur à la date. C'est par définition l'état exact du stock juste
# avant ce mouvement, donc l'état à la date. Aucune référence n'a de mouvement
# postérieur → stock actuel (pas de changement depuis).
#
# Algo prix : pour chaque référence, on prend le `prix_apres` du dernier changement
# historisé avant ou à la date. Aucun changement historisé avant la date → prix
# actuel (le prix courant s'appliquait déjà à l'époque, faute d'historique).


def _date_iso_end_of_day(date_iso: str) -> str:
    """YYYY-MM-DD → YYYY-MM-DDT23:59:59 (fin de journée locale, cohérent avec le
    format de created_at écrit par le code : strftime('%Y-%m-%dT%H:%M:%S','now','localtime'))."""
    return f"{date_iso}T23:59:59"


def _mp_stock_snapshot_at_date(conn, date_iso: str) -> tuple[dict[int, float], dict[tuple[int, int], float]]:
    """Retourne (stock_non_laize, stock_laize) au soir de date_iso.

    - stock_non_laize : dict {matiere_id → quantite}
    - stock_laize     : dict {(matiere_id, laize_id) → quantite}
    """
    end = _date_iso_end_of_day(date_iso)
    # Non-laizées : premier mouvement > end avec laize_id NULL
    rows = conn.execute(
        """
        WITH ranked AS (
          SELECT matiere_id, quantite_avant,
                 ROW_NUMBER() OVER (
                   PARTITION BY matiere_id
                   ORDER BY created_at ASC, id ASC
                 ) AS rn
          FROM mp_mouvements
          WHERE created_at > ? AND laize_id IS NULL
        )
        SELECT matiere_id, quantite_avant FROM ranked WHERE rn = 1
        """,
        (end,),
    ).fetchall()
    non_laize: dict[int, float] = {}
    for r in rows:
        try:
            non_laize[int(r["matiere_id"])] = float(r["quantite_avant"] or 0)
        except (TypeError, ValueError):
            pass
    # Fallback stock actuel pour matières sans mouvement postérieur
    all_stock = conn.execute("SELECT matiere_id, quantite FROM mp_stock").fetchall()
    for r in all_stock:
        try:
            mid = int(r["matiere_id"])
        except (TypeError, ValueError):
            continue
        if mid not in non_laize:
            non_laize[mid] = float(r["quantite"] or 0)

    # Laizées : premier mouvement > end par (matiere_id, laize_id)
    rows = conn.execute(
        """
        WITH ranked AS (
          SELECT matiere_id, laize_id, quantite_avant,
                 ROW_NUMBER() OVER (
                   PARTITION BY matiere_id, laize_id
                   ORDER BY created_at ASC, id ASC
                 ) AS rn
          FROM mp_mouvements
          WHERE created_at > ? AND laize_id IS NOT NULL
        )
        SELECT matiere_id, laize_id, quantite_avant FROM ranked WHERE rn = 1
        """,
        (end,),
    ).fetchall()
    laize: dict[tuple[int, int], float] = {}
    for r in rows:
        try:
            k = (int(r["matiere_id"]), int(r["laize_id"]))
            laize[k] = float(r["quantite_avant"] or 0)
        except (TypeError, ValueError):
            pass
    # Fallback stock actuel par (matiere, laize)
    try:
        all_laize = conn.execute(
            "SELECT matiere_id, laize_id, quantite FROM mp_stock_laize"
        ).fetchall()
    except Exception:
        all_laize = []
    for r in all_laize:
        try:
            k = (int(r["matiere_id"]), int(r["laize_id"]))
        except (TypeError, ValueError):
            continue
        if k not in laize:
            laize[k] = float(r["quantite"] or 0)
    return non_laize, laize


def _pf_stock_snapshot_at_date(conn, date_iso: str) -> dict[int, float]:
    """Retourne dict {produit_id → quantite totale (tous emplacements)} au soir de date_iso.
    Reconstitution : premier mouvement > end par (produit_id, emplacement) — on somme
    ensuite les quantite_avant pour retrouver le stock global.
    Fallback = stock actuel agrégé toutes emplacements pour les produits sans mouvement postérieur."""
    end = _date_iso_end_of_day(date_iso)
    rows = conn.execute(
        """
        WITH ranked AS (
          SELECT produit_id, emplacement, quantite_avant,
                 ROW_NUMBER() OVER (
                   PARTITION BY produit_id, emplacement
                   ORDER BY created_at ASC, id ASC
                 ) AS rn
          FROM mouvements_stock
          WHERE created_at > ?
        )
        SELECT produit_id, emplacement, quantite_avant FROM ranked WHERE rn = 1
        """,
        (end,),
    ).fetchall()
    # Reconstitue par (produit, emplacement) → somme par produit
    by_prod_empl: dict[tuple[int, str], float] = {}
    seen_prods: set[int] = set()
    for r in rows:
        try:
            pid = int(r["produit_id"])
        except (TypeError, ValueError):
            continue
        empl = r["emplacement"] or ""
        by_prod_empl[(pid, empl)] = float(r["quantite_avant"] or 0)
        seen_prods.add(pid)
    # Complète avec stock actuel pour tous les couples (produit, emplacement)
    # NON couverts par un mouvement postérieur.
    try:
        current = conn.execute(
            "SELECT produit_id, emplacement, quantite FROM stock_emplacements"
        ).fetchall()
    except Exception:
        current = []
    for r in current:
        try:
            pid = int(r["produit_id"])
        except (TypeError, ValueError):
            continue
        empl = r["emplacement"] or ""
        k = (pid, empl)
        if k not in by_prod_empl:
            by_prod_empl[k] = float(r["quantite"] or 0)
    # Agrège par produit_id
    out: dict[int, float] = {}
    for (pid, _empl), q in by_prod_empl.items():
        out[pid] = out.get(pid, 0.0) + q
    return out


def _mp_price_snapshot_at_date(conn, date_iso: str) -> dict[int, float]:
    """Retourne dict {matiere_id → prix} au soir de date_iso.

    - Matières non-laizées : prix = prix_unitaire courant à la date
    - Matières laizées     : prix = prix_eur_m2 courant à la date
    L'historique mp_valorisation_historique stocke les deux dans la même colonne
    prix_apres (le sens dépend du type de matière — géré par le caller)."""
    end = _date_iso_end_of_day(date_iso)
    rows = conn.execute(
        """
        WITH ranked AS (
          SELECT matiere_id, prix_apres,
                 ROW_NUMBER() OVER (
                   PARTITION BY matiere_id
                   ORDER BY created_at DESC, id DESC
                 ) AS rn
          FROM mp_valorisation_historique
          WHERE created_at <= ?
        )
        SELECT matiere_id, prix_apres FROM ranked WHERE rn = 1
        """,
        (end,),
    ).fetchall()
    out: dict[int, float] = {}
    for r in rows:
        try:
            out[int(r["matiere_id"])] = float(r["prix_apres"] or 0)
        except (TypeError, ValueError):
            pass
    return out


def _pf_price_snapshot_at_date(conn, date_iso: str) -> dict[int, float]:
    """Retourne dict {produit_id → prix_unitaire_ht} au soir de date_iso."""
    end = _date_iso_end_of_day(date_iso)
    rows = conn.execute(
        """
        WITH ranked AS (
          SELECT produit_id, prix_apres,
                 ROW_NUMBER() OVER (
                   PARTITION BY produit_id
                   ORDER BY created_at DESC, id DESC
                 ) AS rn
          FROM pf_valorisation_historique
          WHERE created_at <= ?
        )
        SELECT produit_id, prix_apres FROM ranked WHERE rn = 1
        """,
        (end,),
    ).fetchall()
    out: dict[int, float] = {}
    for r in rows:
        try:
            out[int(r["produit_id"])] = float(r["prix_apres"] or 0)
        except (TypeError, ValueError):
            pass
    return out


def _parse_snapshot_date(raw: str | None) -> str | None:
    """Valide un YYYY-MM-DD provenant d'un query param. Retourne la string validée
    ou None si absent/invalid. Bornes : 2020-01-01 <= date <= aujourd'hui."""
    if not raw:
        return None
    raw = raw.strip()
    if not raw:
        return None
    import re
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", raw):
        return None
    try:
        y, m, d = int(raw[:4]), int(raw[5:7]), int(raw[8:10])
        # Validation légère
        if y < 2020 or y > 2100 or m < 1 or m > 12 or d < 1 or d > 31:
            return None
        today = datetime.now().strftime("%Y-%m-%d")
        if raw > today:
            return None  # pas d'avenir
    except (TypeError, ValueError):
        return None
    return raw


def _get_storage_fees_pct(conn) -> float:
    """Lit le paramètre Frais de stockage (mc_setting.storage_fees_pct), en %.
    Retourne 0.0 si non configuré."""
    try:
        row = conn.execute(
            "SELECT value_decimal FROM mc_setting WHERE key='storage_fees_pct' LIMIT 1"
        ).fetchone()
        if not row:
            return 0.0
        val = float(row["value_decimal"] or 0)
        return val if val >= 0 else 0.0
    except Exception:
        return 0.0


_HISTORIQUE_TYPES_MVT = frozenset({"entree", "sortie", "ajustement", "inventaire", "transfert"})
_HISTORIQUE_TYPE_STOCK = frozenset({"tout", "mp", "produits"})

_HISTORIQUE_SQL_MP = """
    SELECT
        'mp-' || m.id AS id,
        'mp' AS type_stock,
        m.matiere_id,
        mp.categorie,
        mp.reference,
        mp.designation,
        CASE
            WHEN mp.categorie IN ('frontal', 'glassine') THEN 'bobine'
            WHEN mp.categorie = 'palette' THEN 'palette'
            WHEN mp.categorie = 'carton' THEN 'palette'
            ELSE 'palette'
        END AS unite,
        CASE
            WHEN m.type_mouvement = 'transfert'
                 AND TRIM(COALESCE(m.emplacement_source,'')) != ''
                 AND TRIM(COALESCE(m.emplacement_dest,'')) != ''
                THEN TRIM(m.emplacement_source) || ' → ' || TRIM(m.emplacement_dest)
            WHEN TRIM(COALESCE(m.emplacement_dest,'')) != '' THEN TRIM(m.emplacement_dest)
            WHEN TRIM(COALESCE(m.emplacement_source,'')) != '' THEN TRIM(m.emplacement_source)
            ELSE NULL
        END AS emplacement,
        m.type_mouvement,
        m.quantite,
        m.quantite_avant,
        m.quantite_apres,
        m.ref_bl,
        m.note,
        m.created_at,
        m.created_by_name
    FROM mp_mouvements m
    JOIN matieres_premieres mp ON mp.id = m.matiere_id
    WHERE {where}
"""

_HISTORIQUE_SQL_PF = """
    SELECT
        'pf-' || m.id AS id,
        'produit' AS type_stock,
        m.produit_id,
        NULL AS categorie,
        p.reference,
        p.designation,
        p.unite,
        m.emplacement,
        m.type_mouvement,
        m.quantite,
        m.quantite_avant,
        m.quantite_apres,
        NULL AS ref_bl,
        m.note,
        m.created_at,
        m.created_by_name
    FROM mouvements_stock m
    JOIN produits p ON p.id = m.produit_id
    WHERE {where}
"""


def _historique_where_clause(
    is_mp: bool,
    categorie: Optional[str],
    reference: Optional[str],
    type_mouvement: Optional[str],
    date_debut: Optional[str],
    date_fin: Optional[str],
) -> tuple[str, list]:
    parts = ["1=1"]
    params: list[Any] = []
    if is_mp and categorie:
        parts.append("mp.categorie=?")
        params.append(categorie)
    if reference:
        ref = reference.strip()
        if ref:
            pat = f"%{ref}%"
            if is_mp:
                parts.append(
                    "(LOWER(mp.reference) LIKE LOWER(?) "
                    "OR LOWER(IFNULL(mp.designation,'')) LIKE LOWER(?))"
                )
            else:
                parts.append(
                    "(LOWER(p.reference) LIKE LOWER(?) "
                    "OR LOWER(IFNULL(p.designation,'')) LIKE LOWER(?))"
                )
            params.extend([pat, pat])
    if type_mouvement:
        parts.append("m.type_mouvement=?")
        params.append(type_mouvement)
    if date_debut:
        d = date_debut.strip()
        parts.append("m.created_at >= ?")
        params.append(d if "T" in d else f"{d}T00:00:00")
    if date_fin:
        d = date_fin.strip()
        parts.append("m.created_at <= ?")
        params.append(d if "T" in d else f"{d}T23:59:59")
    return " AND ".join(parts), params


def _historique_row_dict(r) -> dict:
    def _f(v):
        return float(v) if v is not None else None

    return {
        "id": r["id"],
        "type_stock": r["type_stock"],
        "produit_id": r["produit_id"] if "produit_id" in r.keys() else None,
        "matiere_id": r["matiere_id"] if "matiere_id" in r.keys() else None,
        "categorie": r["categorie"],
        "reference": r["reference"],
        "designation": r["designation"],
        "unite": r["unite"],
        "emplacement": r["emplacement"],
        "type_mouvement": r["type_mouvement"],
        "quantite": _f(r["quantite"]),
        "quantite_avant": _f(r["quantite_avant"]),
        "quantite_apres": _f(r["quantite_apres"]),
        "ref_bl": r["ref_bl"],
        "note": r["note"],
        "created_at": r["created_at"],
        "created_by_name": r["created_by_name"],
    }


def _historique_type_stock_label(type_stock: str) -> str:
    return "Matières premières" if type_stock == "mp" else "Produits finis"


# ── Helpers FIFO ──────────────────────────────────────────────────
def get_stock_produit_total(conn, produit_id: int) -> dict:
    """Quantité totale + date FIFO (lot le plus ancien avec restant > 0)."""
    rows = conn.execute(
        """SELECT quantite_restante, date_entree, emplacement
           FROM lots_stock
           WHERE produit_id=? AND quantite_restante > 0
           ORDER BY date_entree ASC""",
        (produit_id,),
    ).fetchall()
    total = sum(r["quantite_restante"] for r in rows)
    date_fifo = rows[0]["date_entree"] if rows else None
    return {"total": total, "date_fifo": date_fifo, "nb_lots": len(rows)}


def apply_fifo_sortie(
    conn,
    produit_id: int,
    emplacement: str,
    quantite: float,
    user_email: str,
    user_name: Optional[str] = None,
    note: str = "",
    no_dossier: Optional[str] = None,
) -> dict:
    """
    Consomme les lots FIFO pour un emplacement donné.
    Retourne quantite_avant, quantite_apres.
    """
    lots = conn.execute(
        """SELECT id, quantite_restante, date_entree
           FROM lots_stock
           WHERE produit_id=? AND emplacement=? AND quantite_restante > 0
           ORDER BY date_entree ASC""",
        (produit_id, emplacement),
    ).fetchall()

    total_dispo = sum(l["quantite_restante"] for l in lots)
    if quantite > total_dispo:
        raise HTTPException(400, f"Stock insuffisant sur {emplacement} : {total_dispo} disponibles")

    quantite_avant = total_dispo
    restant = quantite
    now = datetime.now().isoformat()

    for lot in lots:
        if restant <= 0:
            break
        consomme = min(lot["quantite_restante"], restant)
        nouvelle_qte = lot["quantite_restante"] - consomme
        conn.execute(
            "UPDATE lots_stock SET quantite_restante=? WHERE id=?",
            (nouvelle_qte, lot["id"]),
        )
        restant -= consomme

    quantite_apres = total_dispo - quantite

    # Mettre à jour stock_emplacements (vue agrégée)
    conn.execute(
        """UPDATE stock_emplacements
           SET quantite=?, updated_at=?, updated_by=?, commentaire=?
           WHERE produit_id=? AND emplacement=?""",
        (quantite_apres, now, user_email, note or None, produit_id, emplacement),
    )

    # Historique
    cur = conn.execute(
        """INSERT INTO mouvements_stock
           (produit_id,emplacement,type_mouvement,quantite,quantite_avant,quantite_apres,note,created_at,created_by,created_by_name,no_dossier)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
        (produit_id, emplacement, "sortie", quantite, quantite_avant, quantite_apres, note, now, user_email, user_name, no_dossier),
    )
    return {"quantite_avant": quantite_avant, "quantite_apres": quantite_apres, "mouvement_id": cur.lastrowid}


def sortir_lot_fifo(
    conn,
    produit_id: int,
    emplacement: str,
    user_email: str,
    user_name: Optional[str] = None,
    note: str = "",
) -> dict:
    """Sortie du lot FIFO le plus ancien à un emplacement (quantité = lot entier)."""
    lot = conn.execute(
        """SELECT id, quantite_restante
           FROM lots_stock
           WHERE produit_id=? AND emplacement=? AND quantite_restante > 0
           ORDER BY date_entree ASC LIMIT 1""",
        (produit_id, emplacement),
    ).fetchone()
    if not lot:
        raise HTTPException(404, "Aucun lot actif à cet emplacement")
    qte_lot = float(lot["quantite_restante"])
    final_note = (note or "").strip() or "Sortie lot FIFO"
    result = apply_fifo_sortie(
        conn, produit_id, emplacement, qte_lot, user_email, user_name, final_note
    )
    return {**result, "quantite_sortie": qte_lot}


def deplacer_lot_fifo(
    conn,
    produit_id: int,
    emplacement_source: str,
    emplacement_destination: str,
    user_email: str,
    user_name: Optional[str] = None,
) -> dict:
    """Déplace le lot FIFO le plus ancien d'un emplacement vers un autre."""
    # Récupérer le lot FIFO source
    lot = conn.execute(
        """SELECT id, quantite_restante, date_entree
           FROM lots_stock
           WHERE produit_id=? AND emplacement=? AND quantite_restante > 0
           ORDER BY date_entree ASC LIMIT 1""",
        (produit_id, emplacement_source),
    ).fetchone()
    if not lot:
        raise HTTPException(404, "Aucun lot actif à l'emplacement source")
    
    qte_lot = float(lot["quantite_restante"])
    lot_id = lot["id"]
    date_entree = lot["date_entree"]
    now = datetime.now().isoformat()
    
    # Quantité avant le déplacement
    quantite_avant_source = qte_lot
    
    # Mettre à jour le lot source (quantite_restante = 0)
    conn.execute(
        "UPDATE lots_stock SET quantite_restante=? WHERE id=?",
        (0, lot_id),
    )
    
    # Créer un nouveau lot à la destination
    cursor = conn.execute(
        """INSERT INTO lots_stock (produit_id, emplacement, quantite_initiale, quantite_restante, date_entree, created_at)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (produit_id, emplacement_destination, qte_lot, qte_lot, date_entree, now),
    )
    
    # Mettre à jour stock_emplacements pour la source
    conn.execute(
        """UPDATE stock_emplacements
           SET quantite=?, updated_at=?, updated_by=?, commentaire='Déplacement vers ' || ?
           WHERE produit_id=? AND emplacement=?""",
        (0, now, user_email, emplacement_destination, produit_id, emplacement_source),
    )
    
    # Mettre à jour ou créer stock_emplacements pour la destination
    existing_dest = conn.execute(
        "SELECT quantite FROM stock_emplacements WHERE produit_id=? AND emplacement=?",
        (produit_id, emplacement_destination),
    ).fetchone()
    
    if existing_dest:
        quantite_apres_dest = float(existing_dest["quantite"]) + qte_lot
        conn.execute(
            """UPDATE stock_emplacements
               SET quantite=?, updated_at=?, updated_by=?, commentaire='Déplacement depuis ' || ?
               WHERE produit_id=? AND emplacement=?""",
            (quantite_apres_dest, now, user_email, emplacement_source, produit_id, emplacement_destination),
        )
    else:
        quantite_apres_dest = qte_lot
        conn.execute(
            """INSERT INTO stock_emplacements (produit_id, emplacement, quantite, updated_at, updated_by, commentaire)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (produit_id, emplacement_destination, quantite_apres_dest, now, user_email, f'Déplacement depuis {emplacement_source}'),
        )
    
    # Historique : sortie de la source
    conn.execute(
        """INSERT INTO mouvements_stock
           (produit_id,emplacement,type_mouvement,quantite,quantite_avant,quantite_apres,note,created_at,created_by,created_by_name)
           VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (produit_id, emplacement_source, "sortie", qte_lot, quantite_avant_source, 0, f"Déplacement vers {emplacement_destination}", now, user_email, user_name),
    )
    
    # Historique : entrée à la destination
    quantite_avant_dest = quantite_apres_dest - qte_lot
    conn.execute(
        """INSERT INTO mouvements_stock
           (produit_id,emplacement,type_mouvement,quantite,quantite_avant,quantite_apres,note,created_at,created_by,created_by_name)
           VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (produit_id, emplacement_destination, "entree", qte_lot, quantite_avant_dest, quantite_apres_dest, f"Déplacement depuis {emplacement_source}", now, user_email, user_name),
    )
    
    return {
        "quantite_avant": quantite_avant_source,
        "quantite_apres": 0,
        "quantite_deplacee": qte_lot,
    }


# ── Recherche instantanée ─────────────────────────────────────────
@router.get("/api/stock/search")
def search(request: Request, q: str = "", limit: int = 12):
    """Recherche unifiée : produits + emplacements. 'text contains'."""
    require_stock(request)
    q = q.strip()
    if not q:
        return {"produits": [], "emplacements": []}

    pattern = f"%{q}%"
    prod_where, prod_params = _produit_search_where_args(q)
    with get_db() as conn:
        produits = conn.execute(
            f"""SELECT p.id, p.reference, p.designation, p.unite,
                      COALESCE(p.type, 'fabrique') AS type,
                      COALESCE(SUM(l.quantite_restante),0) as stock_total,
                      MIN(CASE WHEN l.quantite_restante>0 THEN l.date_entree END) as date_fifo
               FROM produits p
               LEFT JOIN lots_stock l ON l.produit_id=p.id
               WHERE {prod_where}
               GROUP BY p.id
               ORDER BY p.reference
               LIMIT ?""",
            (*prod_params, limit),
        ).fetchall()

        # Emplacements : stock réel + référentiel plan (grille A/B/… + numéro)
        empls = conn.execute(
            """SELECT DISTINCT s.emplacement,
                      COUNT(DISTINCT s.produit_id) as nb_refs,
                      COALESCE(SUM(s.quantite),0) as total_unites,
                      s.derniere_inventaire
               FROM stock_emplacements s
               WHERE s.emplacement LIKE ? AND s.quantite > 0
               GROUP BY s.emplacement
               ORDER BY s.emplacement
               LIMIT ?""",
            (pattern, limit),
        ).fetchall()

        plan_empl = conn.execute(
            """SELECT code as emplacement,
                      0 as nb_refs,
                      0 as total_unites,
                      NULL as derniere_inventaire
               FROM emplacements_plan
               WHERE code LIKE ?
               ORDER BY code
               LIMIT ?""",
            (pattern, limit * 2),
        ).fetchall()

    by_code = {r["emplacement"]: dict(r) for r in empls}
    for r in plan_empl:
        d = dict(r)
        code = d["emplacement"]
        if code not in by_code:
            by_code[code] = d
    merged_empl = sorted(by_code.values(), key=lambda x: x["emplacement"])[:limit]

    return {
        "produits": [dict(r) for r in produits],
        "emplacements": merged_empl,
    }


# ── Vue tableau de bord (grille) ──────────────────────────────────
@router.get("/api/stock/vue-globale")
def vue_globale(request: Request):
    """Données du tableau de bord MyStock : une ligne par (emplacement × produit) avec lots actifs."""
    require_stock(request)
    with get_db() as conn:
        grille_rows = conn.execute(
            """SELECT l.emplacement, p.id AS id, p.reference, p.designation, p.unite,
                      SUM(l.quantite_restante) AS quantite
               FROM lots_stock l
               JOIN produits p ON p.id = l.produit_id
               WHERE l.quantite_restante > 0
               GROUP BY l.emplacement, p.id, p.reference, p.designation, p.unite
               ORDER BY l.emplacement, p.reference"""
        ).fetchall()
        stats_row = conn.execute(
            """SELECT
               COUNT(DISTINCT p.id) AS nb_refs,
               COUNT(DISTINCT CASE WHEN s.quantite > 0 THEN s.emplacement END) AS nb_empl,
               COALESCE(SUM(s.quantite), 0) AS total_unites
               FROM produits p
               LEFT JOIN stock_emplacements s ON s.produit_id = p.id"""
        ).fetchone()
        mvts = conn.execute(
            f"""SELECT {_MVT_FIELDS}, p.reference, p.designation, p.unite,
                      COALESCE(NULLIF(TRIM(m.created_by_name),''), u.nom) AS created_by_nom
               FROM mouvements_stock m
               JOIN produits p ON p.id = m.produit_id
               {_STOCK_USER_JOIN}
               ORDER BY m.created_at DESC LIMIT 15"""
        ).fetchall()
    return {
        "grille": [dict(r) for r in grille_rows],
        "stats": dict(stats_row) if stats_row else {},
        "derniers_mouvements": [dict(r) for r in mvts],
    }


# ── Produits CRUD ─────────────────────────────────────────────────
@router.get("/api/stock/produits")
def list_produits(
    request: Request,
    q: Optional[str] = None,
    client: Optional[str] = None,
    limit: int = 50,
):
    require_stock(request)
    client_qs = (client or "").strip()
    qs = (q or "").strip()
    if client_qs:
        prod_where, prod_params = _produit_client_search_where_args(client_qs)
        limit = min(max(limit, 1), 500)
    elif qs:
        prod_where, prod_params = _produit_search_where_args(qs)
    else:
        prod_where = "(p.reference LIKE ? OR IFNULL(p.designation,'') LIKE ?)"
        prod_params = ["%", "%"]
    with get_db() as conn:
        rows = conn.execute(
            f"""SELECT p.id, p.reference, p.designation, p.unite, p.description,
                      COALESCE(SUM(l.quantite_restante),0) as stock_total,
                      MIN(CASE WHEN l.quantite_restante>0 THEN l.date_entree END) as date_fifo,
                      COUNT(DISTINCT CASE WHEN l.quantite_restante>0 THEN l.emplacement END) as nb_emplacements
               FROM produits p
               LEFT JOIN lots_stock l ON l.produit_id=p.id
               WHERE {prod_where}
               GROUP BY p.id ORDER BY p.reference LIMIT ?""",
            (*prod_params, limit),
        ).fetchall()
    return [dict(r) for r in rows]


@router.get("/api/stock/produits/export")
def export_produits_referentiel(request: Request, format: str = "csv"):
    """Export CSV ou page d'impression du référentiel — avant /{produit_id} (sinon 'export' est parsé en id)."""
    require_stock(request)
    fmt = (format or "csv").strip().lower()
    with get_db() as conn:
        rows = conn.execute(
            "SELECT reference, unite, designation FROM produits ORDER BY reference"
        ).fetchall()
    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";")
        writer.writerow(["reference", "unite", "designation"])
        for r in rows:
            writer.writerow([r["reference"] or "", r["unite"] or "", r["designation"] or ""])
        data = buf.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="references_unites_vente.csv"'},
        )
    if fmt == "print":
        html_rows = "".join(
            f"<tr><td>{r['reference'] or ''}</td><td>{r['unite'] or ''}</td>"
            f"<td>{r['designation'] or ''}</td></tr>"
            for r in rows
        )
        html = f"""<!DOCTYPE html><html lang="fr"><head><meta charset="UTF-8">
<title>Références et unités de vente</title>
<style>
body{{font-family:system-ui,sans-serif;margin:24px;color:#111}}
h1{{font-size:18px;margin:0 0 4px}}
p{{color:#555;font-size:12px;margin:0 0 16px}}
table{{width:100%;border-collapse:collapse;font-size:12px}}
th,td{{border:1px solid #ccc;padding:8px 10px;text-align:left}}
th{{background:#f1f5f9;font-weight:700}}
tr:nth-child(even){{background:#f8fafc}}
@media print{{body{{margin:12px}}}}
</style></head><body>
<h1>Références et unités de vente</h1>
<p>Export MyStock — {datetime.now().strftime("%d/%m/%Y %H:%M")} — {len(rows)} référence(s)</p>
<table><thead><tr><th>Référence</th><th>Unité de vente</th><th>Désignation</th></tr></thead>
<tbody>{html_rows}</tbody></table>
<script>window.onload=function(){{window.print();}}</script>
</body></html>"""
        return StreamingResponse(
            io.BytesIO(html.encode("utf-8")),
            media_type="text/html; charset=utf-8",
        )
    raise HTTPException(400, "Format non supporté (csv ou print).")


@router.get("/api/stock/produits/{produit_id}")
def get_produit(produit_id: int, request: Request):
    require_stock(request)
    with get_db() as conn:
        p = conn.execute("SELECT * FROM produits WHERE id=?", (produit_id,)).fetchone()
        if not p:
            raise HTTPException(404, "Produit non trouvé")

        # Stock par emplacement avec date FIFO par emplacement
        empls = conn.execute(
            """SELECT l.emplacement,
                      SUM(l.quantite_restante) as quantite,
                      COUNT(*) as nb_lots,
                      MIN(CASE WHEN l.quantite_restante>0 THEN l.date_entree END) as date_fifo_empl,
                      (SELECT l2.quantite_restante FROM lots_stock l2
                       WHERE l2.produit_id=? AND l2.emplacement=l.emplacement
                         AND l2.quantite_restante > 0
                       ORDER BY l2.date_entree ASC LIMIT 1) as quantite_lot_fifo,
                      MAX(s.derniere_inventaire) as derniere_inventaire,
                      MAX(s.updated_at) as updated_at,
                      MAX(s.updated_by) as updated_by,
                      MAX(s.commentaire) as commentaire
               FROM lots_stock l
               LEFT JOIN stock_emplacements s ON s.produit_id=l.produit_id AND s.emplacement=l.emplacement
               WHERE l.produit_id=? AND l.quantite_restante>0
               GROUP BY l.emplacement
               ORDER BY l.emplacement""",
            (produit_id, produit_id),
        ).fetchall()

        # FIFO global
        fifo = get_stock_produit_total(conn, produit_id)

        # Historique mouvements
        mvts = conn.execute(
            f"""SELECT {_MVT_FIELDS}, p.reference, p.designation, p.unite,
                      COALESCE(NULLIF(TRIM(m.created_by_name),''), u.nom) AS created_by_nom
               FROM mouvements_stock m
               JOIN produits p ON p.id=m.produit_id
               {_STOCK_USER_JOIN}
               WHERE m.produit_id=?
               ORDER BY m.created_at DESC LIMIT 80""",
            (produit_id,),
        ).fetchall()

        # Alerte inventaire
        now = datetime.now()
        empl_data = []
        for e in empls:
            d = dict(e)
            if e["derniere_inventaire"]:
                try:
                    inv = str(e["derniere_inventaire"]).replace("Z", "")[:19]
                    delta = (now - datetime.fromisoformat(inv)).days
                    d["jours_depuis_inventaire"] = delta
                    d["alerte_inventaire"] = delta > INVENTAIRE_ALERTE_JOURS
                except Exception:
                    d["jours_depuis_inventaire"] = None
                    d["alerte_inventaire"] = True
            else:
                d["jours_depuis_inventaire"] = None
                d["alerte_inventaire"] = True
            empl_data.append(d)

    # Durée de stock en jours
    jours_stock = None
    if fifo["date_fifo"]:
        try:
            jours_stock = (datetime.now() - datetime.fromisoformat(fifo["date_fifo"][:19])).days
        except Exception:
            pass

    return {
        "produit": dict(p),
        "stock_total": fifo["total"],
        "date_fifo": fifo["date_fifo"],
        "jours_stock": jours_stock,
        "nb_lots": fifo["nb_lots"],
        "emplacements": empl_data,
        "mouvements": [dict(r) for r in mvts],
    }


@router.get("/api/stock/produits/{produit_id}/emplacements")
def get_produit_emplacements_compat(produit_id: int, request: Request):
    """Alias historique : l’app chargeait cette URL par erreur."""
    return get_produit(produit_id, request)


@router.post("/api/stock/produits")
async def create_produit(request: Request):
    user = require_stock_write(request)
    body = await request.json()
    ref = (body.get("reference") or "").strip().upper()
    if not ref:
        raise HTTPException(400, "Référence obligatoire")
    commentaire = (body.get("commentaire") or "").strip()
    des = (body.get("designation") or "").strip()
    if not des:
        des = commentaire if commentaire else ref
    q = body.get("quantite")
    description = ""
    if q is not None and str(q).strip() != "":
        description = f"Quantité: {str(q).strip()}"
    # Par défaut, l'unité de vente est "étiquette" (singulier ; le pluriel est géré côté frontend).
    unite = (body.get("unite") or "étiquette").strip() or "étiquette"
    now = datetime.now().isoformat()
    with get_db() as conn:
        try:
            cursor = conn.execute(
                "INSERT INTO produits (reference,designation,description,unite,created_at,updated_at) VALUES (?,?,?,?,?,?)",
                (ref, des, description, unite, now, now),
            )
            conn.commit()
            log_action(
                user=user,
                action="CREATE",
                module="stock",
                objet=f"Produit {ref} — {des}",
                detail={"reference": ref},
                ip=request.client.host if request.client else None,
            )
            return {"success": True, "id": cursor.lastrowid, "existing": False}
        except sqlite3.IntegrityError:
            conn.rollback()
            row = conn.execute("SELECT id FROM produits WHERE reference=?", (ref,)).fetchone()
            if row:
                return {"success": True, "id": row["id"], "existing": True}
            raise HTTPException(409, "Conflit en base pour cette référence")


@router.put("/api/stock/produits/{produit_id}")
async def update_produit(produit_id: int, request: Request):
    user = require_stock_write(request)
    body = await request.json()
    now = datetime.now().isoformat()
    ref_audit = ""
    with get_db() as conn:
        prow = conn.execute(
            "SELECT reference FROM produits WHERE id=?", (produit_id,)
        ).fetchone()
        if not prow:
            raise HTTPException(404, "Produit non trouvé")
        ref_audit = prow["reference"] or ""
        conn.execute(
            "UPDATE produits SET designation=?,description=?,unite=?,updated_at=? WHERE id=?",
            (body.get("designation"), body.get("description"), body.get("unite"), now, produit_id),
        )
        conn.commit()
    log_action(
        user=user,
        action="UPDATE",
        module="stock",
        objet=f"Produit {ref_audit}",
        ip=request.client.host if request.client else None,
    )
    return {"success": True}


@router.post("/api/stock/produits/{produit_id}/convertir-unite")
async def convertir_unite_produit(produit_id: int, request: Request):
    """Change l'unité de vente d'un produit et applique un facteur de conversion
    à toutes les quantités en stock (lots_stock + stock_emplacements).
    Les mouvements historiques (mouvements_stock) ne sont pas modifiés."""
    user = require_stock_write(request)
    body = await request.json()
    nouvelle_unite = (body.get("nouvelle_unite") or "").strip()
    if not nouvelle_unite:
        raise HTTPException(400, "Nouvelle unité obligatoire.")
    try:
        facteur = float(body.get("facteur", 0))
    except (TypeError, ValueError):
        raise HTTPException(400, "Facteur invalide.") from None
    if facteur <= 0:
        raise HTTPException(400, "Facteur doit être strictement positif.")
    now = datetime.now().isoformat()

    with get_db() as conn:
        prow = conn.execute(
            "SELECT id, reference, designation, unite FROM produits WHERE id=?",
            (produit_id,),
        ).fetchone()
        if not prow:
            raise HTTPException(404, "Produit non trouvé.")
        ancienne_unite = prow["unite"] or "—"

        old_total_row = conn.execute(
            "SELECT COALESCE(SUM(quantite_restante), 0) AS s FROM lots_stock WHERE produit_id=?",
            (produit_id,),
        ).fetchone()
        old_total = float(old_total_row["s"]) if old_total_row else 0.0

        # 1. Unité du produit
        conn.execute(
            "UPDATE produits SET unite=?, updated_at=? WHERE id=?",
            (nouvelle_unite, now, produit_id),
        )
        # 2. Lots actifs et historiques (quantite_initiale + quantite_restante)
        conn.execute(
            "UPDATE lots_stock SET quantite_restante=quantite_restante*?, "
            "quantite_initiale=quantite_initiale*? WHERE produit_id=?",
            (facteur, facteur, produit_id),
        )
        # 3. Résumé dénormalisé par emplacement
        conn.execute(
            "UPDATE stock_emplacements SET quantite=quantite*?, updated_at=? WHERE produit_id=?",
            (facteur, now, produit_id),
        )
        # 4. Trace dans mouvements_stock (type 'inventaire') par emplacement actif
        empls_actifs = conn.execute(
            "SELECT emplacement, SUM(quantite_restante) AS qte FROM lots_stock "
            "WHERE produit_id=? AND quantite_restante>0 GROUP BY emplacement",
            (produit_id,),
        ).fetchall()
        created_by_name = (user.get("nom") or "").strip() or None
        note_ajust = (
            f"Changement d'unité de vente : {ancienne_unite} → {nouvelle_unite} "
            f"(facteur ×{facteur:g})"
        )
        for r in empls_actifs:
            qte_apres = float(r["qte"] or 0)
            qte_avant = qte_apres / facteur if facteur else qte_apres
            conn.execute(
                """INSERT INTO mouvements_stock
                   (produit_id, emplacement, type_mouvement, quantite,
                    quantite_avant, quantite_apres, note, created_at,
                    created_by, created_by_name)
                   VALUES (?, ?, 'inventaire', ?, ?, ?, ?, ?, ?, ?)""",
                (
                    produit_id, r["emplacement"], qte_apres - qte_avant,
                    qte_avant, qte_apres, note_ajust, now,
                    user.get("email"), created_by_name,
                ),
            )

        new_total_row = conn.execute(
            "SELECT COALESCE(SUM(quantite_restante), 0) AS s FROM lots_stock WHERE produit_id=?",
            (produit_id,),
        ).fetchone()
        new_total = float(new_total_row["s"]) if new_total_row else 0.0
        conn.commit()

    log_action(
        user=user,
        action="UPDATE",
        module="stock",
        objet=f"Conversion unité {prow['reference']} : {ancienne_unite} → {nouvelle_unite} (x{facteur:g})",
        detail={
            "ancienne_unite": ancienne_unite,
            "nouvelle_unite": nouvelle_unite,
            "facteur": facteur,
            "stock_total_avant": old_total,
            "stock_total_apres": new_total,
        },
        ip=request.client.host if request.client else None,
    )
    return {
        "success": True,
        "reference": prow["reference"],
        "ancienne_unite": ancienne_unite,
        "nouvelle_unite": nouvelle_unite,
        "facteur": facteur,
        "stock_total_avant": old_total,
        "stock_total_apres": new_total,
    }


@router.delete("/api/stock/produits/{produit_id}")
def delete_produit(produit_id: int, request: Request):
    user = require_stock_write(request)
    ref_audit = ""
    with get_db() as conn:
        prow = conn.execute(
            "SELECT reference FROM produits WHERE id=?", (produit_id,)
        ).fetchone()
        if not prow:
            raise HTTPException(404, "Produit non trouvé")
        ref_audit = prow["reference"] or ""
        conn.execute("DELETE FROM lots_stock WHERE produit_id=?", (produit_id,))
        conn.execute("DELETE FROM stock_emplacements WHERE produit_id=?", (produit_id,))
        conn.execute("DELETE FROM mouvements_stock WHERE produit_id=?", (produit_id,))
        conn.execute("DELETE FROM produits WHERE id=?", (produit_id,))
        conn.commit()
    log_action(
        user=user,
        action="DELETE",
        module="stock",
        objet=f"Produit {ref_audit} supprimé",
        ip=request.client.host if request.client else None,
    )
    return {"success": True}


# ── Plan entrepôt (référentiel emplacements_plan) ─────────────────
@router.get("/api/stock/emplacements-plan")
def get_emplacements_plan(request: Request):
    """Liste des emplacements du plan (emplacements_plan). Lecture pour tous les rôles stock."""
    require_stock(request)
    with get_db() as conn:
        rows = conn.execute(
            "SELECT code FROM emplacements_plan ORDER BY code"
        ).fetchall()
    return [r["code"] for r in rows]


class _EmplacementPlanAdd(BaseModel):
    code: str

@router.post("/api/stock/emplacements-plan")
def add_emplacement_plan(payload: _EmplacementPlanAdd, request: Request):
    """Ajoute un emplacement au plan. Réservé direction / administration / superadmin."""
    user = require_stock(request)
    if user.get("role") not in {"superadmin", "direction", "administration", "administration_ventes", "administration_technique"}:
        raise HTTPException(403, "Ajout d'emplacement réservé aux administrateurs.")
    code = payload.code.strip().upper()
    if not code:
        raise HTTPException(400, "Code emplacement vide.")
    if len(code) > 20:
        raise HTTPException(400, "Code trop long (20 caractères max).")
    now = datetime.now().isoformat()
    with get_db() as conn:
        conn.execute(
            """CREATE TABLE IF NOT EXISTS emplacements_plan (
                code TEXT PRIMARY KEY NOT NULL,
                imported_at TEXT NOT NULL
            )"""
        )
        if conn.execute("SELECT 1 FROM emplacements_plan WHERE code=?", (code,)).fetchone():
            raise HTTPException(409, f"L'emplacement {code} existe déjà.")
        conn.execute(
            "INSERT INTO emplacements_plan (code, imported_at) VALUES (?, ?)", (code, now)
        )
        conn.commit()
    return {"code": code}


# ── Emplacements ──────────────────────────────────────────────────
@router.get("/api/stock/emplacements-list")
def list_emplacements(request: Request):
    """Retourne tous les emplacements connus : plan (référentiel) + stock réel."""
    require_stock(request)
    with get_db() as conn:
        plan = conn.execute(
            "SELECT code FROM emplacements_plan ORDER BY code"
        ).fetchall()
        reels = conn.execute(
            "SELECT DISTINCT emplacement FROM stock_emplacements ORDER BY emplacement"
        ).fetchall()
    codes_plan = {r["code"] for r in plan}
    codes_reels = {r["emplacement"] for r in reels}
    zones_speciales = {STOCK_EMPLACEMENT_AU_SOL, STOCK_EMPLACEMENT_SORTIE_PROD}
    tous = sorted(codes_plan | codes_reels | zones_speciales)
    ordered = (
        [STOCK_EMPLACEMENT_AU_SOL, STOCK_EMPLACEMENT_SORTIE_PROD]
        + [c for c in tous if c not in zones_speciales]
    )
    return {
        "emplacements": ordered,
        "emplacement_au_sol": STOCK_EMPLACEMENT_AU_SOL,
        "emplacement_au_sol_label": STOCK_EMPLACEMENT_AU_SOL_LABEL,
        "emplacement_sortie_prod": STOCK_EMPLACEMENT_SORTIE_PROD,
        "emplacement_sortie_prod_label": STOCK_EMPLACEMENT_SORTIE_PROD_LABEL,
    }


@router.get("/api/stock/emplacements/{emplacement}")
def get_emplacement(emplacement: str, request: Request):
    require_stock(request)
    emplacement = emplacement.upper()
    now = datetime.now()
    with get_db() as conn:
        refs = conn.execute(
            """SELECT p.id, p.reference, p.designation, p.unite,
                      SUM(l.quantite_restante) as quantite,
                      COUNT(*) as nb_lots,
                      MIN(CASE WHEN l.quantite_restante>0 THEN l.date_entree END) as date_fifo,
                      (SELECT l2.quantite_restante FROM lots_stock l2
                       WHERE l2.produit_id=l.produit_id AND l2.emplacement=?
                         AND l2.quantite_restante > 0
                       ORDER BY l2.date_entree ASC LIMIT 1) as quantite_lot_fifo,
                      MAX(s.derniere_inventaire) as derniere_inventaire,
                      MAX(s.updated_at) as updated_at,
                      MAX(s.updated_by) as updated_by
               FROM lots_stock l
               JOIN produits p ON p.id=l.produit_id
               LEFT JOIN stock_emplacements s ON s.produit_id=l.produit_id AND s.emplacement=l.emplacement
               WHERE l.emplacement=? AND l.quantite_restante>0
               GROUP BY l.produit_id
               ORDER BY p.reference""",
            (emplacement, emplacement),
        ).fetchall()

        mvts = conn.execute(
            f"""SELECT {_MVT_FIELDS}, p.reference, p.designation, p.unite,
                      COALESCE(NULLIF(TRIM(m.created_by_name),''), u.nom) AS created_by_nom
               FROM mouvements_stock m
               JOIN produits p ON p.id=m.produit_id
               {_STOCK_USER_JOIN}
               WHERE m.emplacement=? ORDER BY m.created_at DESC LIMIT 80""",
            (emplacement,),
        ).fetchall()

        # Dernière session d'inventaire complète (table inventaires_sessions)
        last_inv_row = conn.execute(
            """SELECT date_validation, operateur_nom, operateur_email,
                      nb_produits, nb_modifications
               FROM inventaires_sessions
               WHERE emplacement = ?
               ORDER BY date_validation DESC LIMIT 1""",
            (emplacement,),
        ).fetchone()

    last_inventaire = None
    inv_jours_depuis = None
    inv_couleur = "rouge"
    if last_inv_row and last_inv_row["date_validation"]:
        last_inventaire = dict(last_inv_row)
        try:
            d_iso = str(last_inv_row["date_validation"])[:19]
            inv_jours_depuis = (now - datetime.fromisoformat(d_iso)).days
            inv_couleur = _inv_v2_couleur(inv_jours_depuis)
        except Exception:
            inv_jours_depuis = None
            inv_couleur = "rouge"

    refs_data = []
    for r in refs:
        d = dict(r)
        try:
            delta = (now - datetime.fromisoformat(d["date_fifo"][:19])).days if d["date_fifo"] else None
            d["jours_stock"] = delta
        except Exception:
            d["jours_stock"] = None
        if d["derniere_inventaire"]:
            try:
                inv = str(d["derniere_inventaire"]).replace("Z", "")[:19]
                d["alerte_inventaire"] = (
                    now - datetime.fromisoformat(inv)
                ).days > INVENTAIRE_ALERTE_JOURS
            except Exception:
                d["alerte_inventaire"] = True
        else:
            d["alerte_inventaire"] = True
        refs_data.append(d)

    if emplacement == STOCK_EMPLACEMENT_AU_SOL:
        label = STOCK_EMPLACEMENT_AU_SOL_LABEL
    elif emplacement == STOCK_EMPLACEMENT_SORTIE_PROD:
        label = STOCK_EMPLACEMENT_SORTIE_PROD_LABEL
    else:
        label = emplacement
    return {
        "emplacement": emplacement,
        "label": label,
        "est_au_sol": emplacement == STOCK_EMPLACEMENT_AU_SOL,
        "est_sortie_prod": emplacement == STOCK_EMPLACEMENT_SORTIE_PROD,
        "refs": refs_data,
        "total_unites": sum(r["quantite"] for r in refs),
        "nb_refs": len(refs),
        "mouvements": [dict(r) for r in mvts],
        "last_inventaire": last_inventaire,
        "inv_jours_depuis": inv_jours_depuis,
        "inv_couleur": inv_couleur,
    }


@router.get("/api/stock/a-expedier")
def stock_a_expedier(request: Request):
    """Produits finis en zone Au sol - à expédier — stock à expédier prochainement."""
    require_stock(request)
    empl = STOCK_EMPLACEMENT_AU_SOL
    with get_db() as conn:
        refs = conn.execute(
            """SELECT p.id, p.reference, p.designation, p.unite,
                      SUM(l.quantite_restante) as quantite,
                      COUNT(*) as nb_lots,
                      MIN(CASE WHEN l.quantite_restante>0 THEN l.date_entree END) as date_fifo
               FROM lots_stock l
               JOIN produits p ON p.id=l.produit_id
               WHERE l.emplacement=? AND l.quantite_restante>0
               GROUP BY p.id
               ORDER BY p.reference""",
            (empl,),
        ).fetchall()
    refs_data = [dict(r) for r in refs]
    total = sum(float(r["quantite"] or 0) for r in refs_data)
    return {
        "emplacement": empl,
        "label": STOCK_EMPLACEMENT_AU_SOL_LABEL,
        "refs": refs_data,
        "total_unites": total,
        "nb_refs": len(refs_data),
    }


@router.get("/api/stock/sortie-prod")
def stock_sortie_prod(request: Request):
    """Produits en zone En attente - sortie de prod (Z1) — stock fraîchement sorti de production.
    Inclut la date de dernière entrée et le nom de l'opérateur qui l'a posée (pour la vue Production)."""
    require_stock(request)
    empl = STOCK_EMPLACEMENT_SORTIE_PROD
    with get_db() as conn:
        refs = conn.execute(
            """SELECT p.id, p.reference, p.designation, p.unite,
                      SUM(l.quantite_restante) as quantite,
                      COUNT(*) as nb_lots,
                      MIN(CASE WHEN l.quantite_restante>0 THEN l.date_entree END) as date_fifo,
                      MAX(CASE WHEN l.quantite_restante>0 THEN l.date_entree END) as derniere_entree,
                      (SELECT COALESCE(NULLIF(TRIM(u.nom),''), l2.created_by)
                         FROM lots_stock l2
                         LEFT JOIN users u ON LOWER(TRIM(COALESCE(u.email,''))) =
                                              LOWER(TRIM(COALESCE(l2.created_by,'')))
                         WHERE l2.produit_id = p.id AND l2.emplacement = ?
                           AND l2.quantite_restante > 0
                         ORDER BY l2.date_entree DESC LIMIT 1) as dernier_operateur
               FROM lots_stock l
               JOIN produits p ON p.id=l.produit_id
               WHERE l.emplacement=? AND l.quantite_restante>0
               GROUP BY p.id
               ORDER BY MAX(l.date_entree) DESC, p.reference""",
            (empl, empl),
        ).fetchall()
    refs_data = [dict(r) for r in refs]
    total = sum(float(r["quantite"] or 0) for r in refs_data)
    return {
        "emplacement": empl,
        "label": STOCK_EMPLACEMENT_SORTIE_PROD_LABEL,
        "refs": refs_data,
        "total_unites": total,
        "nb_refs": len(refs_data),
    }


def _resolve_created_by_name(conn: sqlite3.Connection, user: dict) -> Optional[str]:
    created_by_name = (user.get("nom") or "").strip() or None
    if created_by_name:
        return created_by_name
    try:
        if user.get("id") is not None:
            r = conn.execute(
                "SELECT nom FROM users WHERE id=? LIMIT 1", (int(user["id"]),)
            ).fetchone()
        else:
            r = conn.execute(
                "SELECT nom FROM users WHERE LOWER(TRIM(COALESCE(email,'')))=? LIMIT 1",
                (str(user.get("email") or "").strip().lower(),),
            ).fetchone()
        return (str(r["nom"] or "").strip() if r else "") or None
    except Exception:
        return None


def _apply_stock_mouvement(
    conn: sqlite3.Connection,
    user: dict,
    produit_id: int,
    emplacement: str,
    type_mvt: str,
    quantite: float,
    note: str,
    date_entree: Optional[str] = None,
    no_dossier: Optional[str] = None,
) -> tuple[dict, str, str]:
    """Applique un mouvement PF (entree / sortie / inventaire) sur la base existante."""
    now = datetime.now().isoformat()
    date_entree = date_entree or datetime.now().strftime("%Y-%m-%d")
    created_by_name = _resolve_created_by_name(conn, user)

    p = conn.execute(
        "SELECT id, reference FROM produits WHERE id=?", (produit_id,)
    ).fetchone()
    if not p:
        raise HTTPException(404, "Produit non trouvé")
    ref_audit = p["reference"] or ""
    if type_mvt == "entree":
        audit_action = "CREATE"
    elif type_mvt == "sortie":
        audit_action = "DELETE"
    else:
        audit_action = "UPDATE"

    ex = conn.execute(
        "SELECT quantite FROM stock_emplacements WHERE produit_id=? AND emplacement=?",
        (produit_id, emplacement),
    ).fetchone()
    qte_avant = float(ex["quantite"]) if ex else 0.0

    if type_mvt == "entree":
        conn.execute(
            """INSERT INTO lots_stock
               (produit_id,emplacement,quantite_initiale,quantite_restante,date_entree,note,created_by,created_at)
               VALUES (?,?,?,?,?,?,?,?)""",
            (produit_id, emplacement, quantite, quantite, date_entree, note, user["email"], now),
        )
        qte_apres = qte_avant + quantite
        if ex:
            conn.execute(
                """UPDATE stock_emplacements
                   SET quantite=?,updated_at=?,updated_by=?,derniere_inventaire=?,commentaire=?
                   WHERE produit_id=? AND emplacement=?""",
                (qte_apres, now, user["email"], now, note or None, produit_id, emplacement),
            )
        else:
            conn.execute(
                """INSERT INTO stock_emplacements
                   (produit_id,emplacement,quantite,updated_at,updated_by,derniere_inventaire,commentaire)
                   VALUES (?,?,?,?,?,?,?)""",
                (produit_id, emplacement, qte_apres, now, user["email"], now, note or None),
            )
        cur = conn.execute(
            """INSERT INTO mouvements_stock
               (produit_id,emplacement,type_mouvement,quantite,quantite_avant,quantite_apres,note,created_at,created_by,created_by_name,no_dossier)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (
                produit_id,
                emplacement,
                "entree",
                quantite,
                qte_avant,
                qte_apres,
                note,
                now,
                user["email"],
                created_by_name,
                no_dossier,
            ),
        )
        result = {"quantite_avant": qte_avant, "quantite_apres": qte_apres, "mouvement_id": cur.lastrowid}

    elif type_mvt == "sortie":
        result = apply_fifo_sortie(
            conn, produit_id, emplacement, quantite, user["email"], created_by_name, note,
            no_dossier=no_dossier,
        )

    elif type_mvt == "inventaire":
        conn.execute(
            "UPDATE lots_stock SET quantite_restante=0 WHERE produit_id=? AND emplacement=?",
            (produit_id, emplacement),
        )
        conn.execute(
            """INSERT INTO lots_stock
               (produit_id,emplacement,quantite_initiale,quantite_restante,date_entree,note,created_by,created_at)
               VALUES (?,?,?,?,?,?,?,?)""",
            (
                produit_id,
                emplacement,
                quantite,
                quantite,
                date_entree,
                f"Inventaire — {note}",
                user["email"],
                now,
            ),
        )
        if ex:
            conn.execute(
                """UPDATE stock_emplacements
                   SET quantite=?,updated_at=?,updated_by=?,derniere_inventaire=?,commentaire=?
                   WHERE produit_id=? AND emplacement=?""",
                (quantite, now, user["email"], now, note or None, produit_id, emplacement),
            )
        else:
            conn.execute(
                """INSERT INTO stock_emplacements
                   (produit_id,emplacement,quantite,updated_at,updated_by,derniere_inventaire,commentaire)
                   VALUES (?,?,?,?,?,?,?)""",
                (produit_id, emplacement, quantite, now, user["email"], now, note or None),
            )
        cur = conn.execute(
            """INSERT INTO mouvements_stock
               (produit_id,emplacement,type_mouvement,quantite,quantite_avant,quantite_apres,note,created_at,created_by,created_by_name,no_dossier)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (
                produit_id,
                emplacement,
                "inventaire",
                quantite,
                qte_avant,
                quantite,
                note,
                now,
                user["email"],
                created_by_name,
                no_dossier,
            ),
        )
        result = {"quantite_avant": qte_avant, "quantite_apres": quantite, "mouvement_id": cur.lastrowid}
    else:
        raise HTTPException(400, "Type de mouvement invalide.")

    return result, ref_audit, audit_action


# ── Mouvement de stock ────────────────────────────────────────────
@router.post("/api/stock/mouvement")
async def mouvement_stock(request: Request):
    user = require_stock_write(request)
    body = await request.json()

    produit_id = body.get("produit_id")
    emplacement = (body.get("emplacement") or "").strip().upper()
    type_mvt = body.get("type_mouvement", "entree")
    quantite = float(body.get("quantite", 0))
    note = (body.get("note") or "").strip()
    date_entree = (body.get("date_entree") or datetime.now().strftime("%Y-%m-%d"))
    no_dossier = (body.get("no_dossier") or "").strip() or None
    palettes_in = body.get("palettes") or []

    if not produit_id or not emplacement:
        raise HTTPException(400, "produit_id et emplacement obligatoires")
    if not _is_valid_emplacement(emplacement):
        raise HTTPException(400, f"Format emplacement invalide : {emplacement}")
    if quantite <= 0:
        raise HTTPException(400, "Quantité doit être positive")

    # Validation palettes : liste d'objets {matiere_id:int, nombre:int>0}.
    # Refusee si type != entree ET no_dossier vide (cas peu utile, on bloque pour clarte).
    palettes_clean: list[tuple[int, int]] = []
    if palettes_in:
        if not isinstance(palettes_in, list):
            raise HTTPException(400, "palettes doit etre une liste")
        for it in palettes_in:
            if not isinstance(it, dict):
                continue
            try:
                mid = int(it.get("matiere_id"))
                nb = int(it.get("nombre"))
            except (TypeError, ValueError):
                raise HTTPException(400, "palette : matiere_id et nombre numeriques requis")
            if nb <= 0:
                continue
            palettes_clean.append((mid, nb))

    with get_db() as conn:
        # Verifier que les matieres palettes existent et sont bien de categorie palette
        if palettes_clean:
            ids = list({mid for mid, _ in palettes_clean})
            placeholders = ",".join("?" * len(ids))
            rows = conn.execute(
                f"SELECT id, categorie FROM matieres_premieres WHERE id IN ({placeholders})",
                ids,
            ).fetchall()
            found = {r["id"]: (r["categorie"] or "") for r in rows}
            for mid, _ in palettes_clean:
                if mid not in found:
                    raise HTTPException(400, f"Reference palette inconnue (matiere_id={mid})")
                if (found[mid] or "").strip().lower() != "palette":
                    raise HTTPException(
                        400,
                        f"Reference {mid} n'est pas une palette (categorie={found[mid]!r})",
                    )

        result, ref_audit, audit_action = _apply_stock_mouvement(
            conn,
            user,
            int(produit_id),
            emplacement,
            type_mvt,
            quantite,
            note,
            date_entree,
            no_dossier=no_dossier,
        )

        # Insertion des palettes (meme transaction)
        mvt_id = result.get("mouvement_id")
        if palettes_clean and mvt_id:
            now = datetime.now().isoformat()
            conn.executemany(
                """INSERT INTO mouvement_palettes (mouvement_id, matiere_id, nombre, created_at)
                   VALUES (?,?,?,?)""",
                [(mvt_id, mid, nb, now) for mid, nb in palettes_clean],
            )
        conn.commit()

    log_action(
        user=user,
        action=audit_action,
        module="stock",
        objet=f"{type_mvt} · {ref_audit} · {emplacement} · {quantite}"
              + (f" · dossier {no_dossier}" if no_dossier else "")
              + (f" · {sum(n for _, n in palettes_clean)} palette(s)" if palettes_clean else ""),
        detail={
            "type_mouvement": type_mvt,
            "quantite": quantite,
            "no_dossier": no_dossier,
            "palettes": [{"matiere_id": m, "nombre": n} for m, n in palettes_clean],
        },
        ip=request.client.host if request.client else None,
    )
    return {"success": True, **result}


@router.post("/api/stock/sortir-lot")
async def api_sortir_lot(request: Request):
    """Sortie du lot FIFO le plus ancien (produit + emplacement)."""
    user = require_stock_write(request)
    body = await request.json()
    produit_id = body.get("produit_id")
    emplacement = (body.get("emplacement") or "").strip().upper()
    note = (body.get("note") or "").strip()

    if not produit_id or not emplacement:
        raise HTTPException(400, "produit_id et emplacement obligatoires")
    if not _is_valid_emplacement(emplacement):
        raise HTTPException(400, f"Format emplacement invalide : {emplacement}")

    with get_db() as conn:
        created_by_name = _resolve_created_by_name(conn, user)
        p = conn.execute(
            "SELECT id, reference FROM produits WHERE id=?", (produit_id,)
        ).fetchone()
        if not p:
            raise HTTPException(404, "Produit non trouvé")
        ref_audit = p["reference"] or ""

        result = sortir_lot_fifo(
            conn, int(produit_id), emplacement, user["email"], created_by_name, note
        )
        conn.commit()

    log_action(
        user=user,
        action="UPDATE",
        module="stock",
        objet=f"Sortie lot FIFO · {ref_audit} · {emplacement} · {result['quantite_sortie']}",
        detail={"emplacement": emplacement, "quantite": result["quantite_sortie"]},
        ip=request.client.host if request.client else None,
    )
    return {"success": True, **result}


@router.post("/api/stock/deplacer-lot")
async def api_deplacer_lot(request: Request):
    """Déplace le lot FIFO le plus ancien d'un emplacement vers un autre."""
    user = require_stock_write(request)
    body = await request.json()
    produit_id = body.get("produit_id")
    emplacement_source = (body.get("emplacement_source") or "").strip().upper()
    emplacement_destination = (body.get("emplacement_destination") or "").strip().upper()

    if not produit_id or not emplacement_source or not emplacement_destination:
        raise HTTPException(400, "produit_id, emplacement_source et emplacement_destination obligatoires")
    if not _is_valid_emplacement(emplacement_source):
        raise HTTPException(400, f"Format emplacement source invalide : {emplacement_source}")
    if not _is_valid_emplacement(emplacement_destination):
        raise HTTPException(400, f"Format emplacement destination invalide : {emplacement_destination}")
    if emplacement_source == emplacement_destination:
        raise HTTPException(400, "Les emplacements source et destination doivent être différents")

    with get_db() as conn:
        created_by_name = _resolve_created_by_name(conn, user)
        p = conn.execute(
            "SELECT id, reference FROM produits WHERE id=?", (produit_id,)
        ).fetchone()
        if not p:
            raise HTTPException(404, "Produit non trouvé")
        ref_audit = p["reference"] or ""

        result = deplacer_lot_fifo(
            conn, int(produit_id), emplacement_source, emplacement_destination, user["email"], created_by_name
        )
        conn.commit()

    log_action(
        user=user,
        action="UPDATE",
        module="stock",
        objet=f"Déplacement lot · {ref_audit} · {emplacement_source} → {emplacement_destination} · {result['quantite_deplacee']}",
        detail={"emplacement_source": emplacement_source, "emplacement_destination": emplacement_destination, "quantite": result["quantite_deplacee"]},
        ip=request.client.host if request.client else None,
    )
    return {"success": True, **result}


# ── Inventaire en chaîne ──────────────────────────────────────────
@router.get("/api/stock/inventaire/priorites")
def inventaire_priorites(request: Request):
    """Toutes les lignes en stock (quantite > 0), tri : jamais / le plus ancien inventaire d'abord."""
    require_stock(request)
    with get_db() as conn:
        rows = conn.execute(
            """SELECT p.id AS produit_id, p.reference, p.designation, p.unite,
                      s.emplacement, s.quantite, s.derniere_inventaire,
                      CASE
                        WHEN s.derniere_inventaire IS NULL
                          OR TRIM(COALESCE(s.derniere_inventaire, '')) = ''
                          THEN 999999
                        ELSE CAST(
                          julianday('now') - julianday(SUBSTR(s.derniere_inventaire, 1, 10))
                          AS INTEGER
                        )
                      END AS jours_depuis_inv
               FROM stock_emplacements s
               JOIN produits p ON p.id = s.produit_id
               WHERE s.quantite > 0
               ORDER BY jours_depuis_inv DESC, s.quantite DESC, p.reference, s.emplacement"""
        ).fetchall()
    return [dict(r) for r in rows]


@router.get("/api/stock/inventaire/produits-a-inventorier")
def produits_a_inventorier(request: Request, jours: int = 180):
    """Produits avec emplacements non inventoriés depuis > jours."""
    require_stock(request)
    limite = (datetime.now() - timedelta(days=jours)).isoformat()
    with get_db() as conn:
        rows = conn.execute(
            """SELECT p.id, p.reference, p.designation, p.unite,
                      s.emplacement, s.quantite, s.derniere_inventaire,
                      COALESCE(CAST(julianday('now') - julianday(s.derniere_inventaire) AS INTEGER), 9999) as jours_depuis
               FROM stock_emplacements s
               JOIN produits p ON p.id=s.produit_id
               WHERE s.quantite > 0
                 AND (s.derniere_inventaire IS NULL OR s.derniere_inventaire < ?)
               ORDER BY jours_depuis DESC, p.reference""",
            (limite,),
        ).fetchall()
    return [dict(r) for r in rows]


# ── Inventaire v2 (par emplacement) ──────────────────────────────
# Seuils en jours depuis le dernier inventaire complet d'un emplacement.
INV_V2_SEUIL_VERT = 15   # < 15 jours = vert
INV_V2_SEUIL_JAUNE = 30  # 15-30 jours = jaune
INV_V2_SEUIL_ORANGE = 60 # 30-60 jours = orange, > 60 ou jamais = rouge


def _inv_v2_couleur(jours: Optional[int]) -> str:
    if jours is None:
        return "rouge"
    if jours < INV_V2_SEUIL_VERT:
        return "vert"
    if jours < INV_V2_SEUIL_JAUNE:
        return "jaune"
    if jours < INV_V2_SEUIL_ORANGE:
        return "orange"
    return "rouge"


def _inv_v2_empl_label(empl: str) -> str:
    if empl == STOCK_EMPLACEMENT_AU_SOL:
        return STOCK_EMPLACEMENT_AU_SOL_LABEL
    if empl == STOCK_EMPLACEMENT_SORTIE_PROD:
        return STOCK_EMPLACEMENT_SORTIE_PROD_LABEL
    return empl


@router.get("/api/stock/inventaire-v2/emplacements")
def inventaire_v2_emplacements(request: Request):
    """Liste des emplacements avec stock + jours depuis dernier inventaire complet.

    Triés du plus ancien (rouge / jamais) au plus récent (vert).
    """
    require_stock(request)
    now = datetime.now()
    with get_db() as conn:
        rows = conn.execute(
            """SELECT ep.code AS emplacement,
                      COUNT(DISTINCT l.produit_id) AS nb_refs,
                      COALESCE(SUM(l.quantite_restante), 0) AS total_qte
               FROM emplacements_plan ep
               LEFT JOIN lots_stock l
                 ON l.emplacement = ep.code AND l.quantite_restante > 0
               GROUP BY ep.code
               ORDER BY ep.code"""
        ).fetchall()
        # Dernier inventaire par emplacement
        last_invs = {
            r["emplacement"]: r
            for r in conn.execute(
                """SELECT s.emplacement, s.date_validation, s.operateur_nom, s.operateur_email
                   FROM inventaires_sessions s
                   JOIN (
                       SELECT emplacement, MAX(date_validation) AS d
                       FROM inventaires_sessions
                       GROUP BY emplacement
                   ) m ON m.emplacement=s.emplacement AND m.d=s.date_validation"""
            ).fetchall()
        }
    out = []
    for r in rows:
        empl = r["emplacement"]
        inv = last_invs.get(empl)
        jours = None
        d_iso = None
        op_nom = None
        if inv and inv["date_validation"]:
            d_iso = inv["date_validation"]
            try:
                jours = (now - datetime.fromisoformat(str(d_iso)[:19])).days
            except Exception:
                jours = None
            op_nom = inv["operateur_nom"]
        out.append({
            "emplacement": empl,
            "label": _inv_v2_empl_label(empl),
            "nb_refs": int(r["nb_refs"] or 0),
            "total_qte": float(r["total_qte"] or 0),
            "derniere_date": d_iso,
            "dernier_operateur": op_nom,
            "jours_depuis": jours,
            "couleur": _inv_v2_couleur(jours),
        })
    # Tri : jamais inventorié d'abord, puis plus anciens d'abord
    out.sort(key=lambda x: (
        0 if x["jours_depuis"] is None else 1,
        -(x["jours_depuis"] or 0),
        x["emplacement"],
    ))
    return out


@router.get("/api/stock/inventaire-v2/emplacement/{emplacement}")
def inventaire_v2_emplacement(emplacement: str, request: Request):
    """Détail d'un emplacement pour l'inventaire : produits + historique."""
    require_stock(request)
    emplacement = _normalize_emplacement(emplacement)
    if not emplacement:
        raise HTTPException(400, "Code emplacement vide.")
    with get_db() as conn:
        # Valider : soit dans le plan, soit dans le stock réel, soit zone spéciale
        in_plan = conn.execute(
            "SELECT 1 FROM emplacements_plan WHERE code=? LIMIT 1", (emplacement,)
        ).fetchone()
        in_stock = conn.execute(
            "SELECT 1 FROM lots_stock WHERE emplacement=? AND quantite_restante>0 LIMIT 1", (emplacement,)
        ).fetchone()
        if not in_plan and not in_stock and not _is_valid_emplacement(emplacement):
            raise HTTPException(404, f"Emplacement inconnu : {emplacement}")
    with get_db() as conn:
        refs = conn.execute(
            """SELECT p.id AS produit_id, p.reference, p.designation, p.unite,
                      SUM(l.quantite_restante) AS quantite,
                      COUNT(*) AS nb_lots,
                      GROUP_CONCAT(l.id) AS lot_ids,
                      MIN(CASE WHEN l.quantite_restante>0 THEN l.date_entree END) AS date_fifo
               FROM lots_stock l
               JOIN produits p ON p.id = l.produit_id
               WHERE l.emplacement = ? AND l.quantite_restante > 0
               GROUP BY p.id
               ORDER BY p.reference""",
            (emplacement,),
        ).fetchall()

        # Lots détaillés par produit (pour affichage référence / lots)
        refs_data = []
        for r in refs:
            d = dict(r)
            lots = conn.execute(
                """SELECT id, quantite_restante, date_entree
                   FROM lots_stock
                   WHERE produit_id=? AND emplacement=? AND quantite_restante > 0
                   ORDER BY date_entree ASC""",
                (d["produit_id"], emplacement),
            ).fetchall()
            d["lots"] = [dict(l) for l in lots]
            refs_data.append(d)

        history = conn.execute(
            """SELECT id, operateur_nom, operateur_email, date_validation,
                      nb_produits, nb_modifications, commentaires_json
               FROM inventaires_sessions
               WHERE emplacement = ?
               ORDER BY date_validation DESC
               LIMIT 100""",
            (emplacement,),
        ).fetchall()

    history_list = []
    for h in history:
        d = dict(h)
        try:
            d["commentaires"] = json.loads(d.pop("commentaires_json") or "[]")
            if not isinstance(d["commentaires"], list):
                d["commentaires"] = []
        except Exception:
            d["commentaires"] = []
        history_list.append(d)
    return {
        "emplacement": emplacement,
        "label": _inv_v2_empl_label(emplacement),
        "refs": refs_data,
        "nb_refs": len(refs_data),
        "total_qte": sum(float(r["quantite"] or 0) for r in refs_data),
        "history": history_list,
        "last_inventaire": history_list[0] if history_list else None,
    }


@router.post("/api/stock/inventaire-v2/valider")
async def inventaire_v2_valider(request: Request):
    """Valide un inventaire complet d'un emplacement.

    Body :
      {
        emplacement: "A121",
        nb_produits: 4,
        modifications: [{produit_id: 12, qte_apres: 250}, ...]
      }

    - Applique chaque modification via un mouvement de stock 'inventaire'
    - Crée une ligne dans inventaires_sessions (avec snapshot des modifs)
    """
    user = require_stock_write(request)
    body = await request.json()
    emplacement = _normalize_emplacement(body.get("emplacement") or "")
    if not _is_valid_emplacement(emplacement):
        raise HTTPException(400, f"Format emplacement invalide : {emplacement}")

    try:
        nb_produits = int(body.get("nb_produits") or 0)
    except Exception:
        nb_produits = 0
    modifications = body.get("modifications") or []
    if not isinstance(modifications, list):
        raise HTTPException(400, "modifications doit être une liste")
    commentaires_in = body.get("commentaires") or []
    if not isinstance(commentaires_in, list):
        commentaires_in = []

    now = datetime.now().isoformat()
    modifs_applied: list[dict] = []
    with get_db() as conn:
        for mod in modifications:
            try:
                pid = int(mod.get("produit_id"))
                qte_apres = float(mod.get("qte_apres"))
            except Exception:
                continue
            if qte_apres < 0:
                qte_apres = 0.0

            pref_row = conn.execute(
                "SELECT reference FROM produits WHERE id=?", (pid,)
            ).fetchone()
            if not pref_row:
                continue
            ref_str = pref_row["reference"] or f"#{pid}"

            ex = conn.execute(
                "SELECT quantite FROM stock_emplacements WHERE produit_id=? AND emplacement=?",
                (pid, emplacement),
            ).fetchone()
            qte_avant = float(ex["quantite"]) if ex else 0.0
            if abs(qte_apres - qte_avant) < 1e-9:
                # Aucun écart : on ne crée pas de mouvement, mais on marque la ligne
                # comme inventoriée plus bas (update derniere_inventaire).
                continue
            try:
                _apply_stock_mouvement(
                    conn,
                    user,
                    pid,
                    emplacement,
                    "inventaire",
                    qte_apres,
                    "Inventaire emplacement",
                )
                modifs_applied.append({
                    "produit_id": pid,
                    "reference": ref_str,
                    "qte_avant": qte_avant,
                    "qte_apres": qte_apres,
                })
            except Exception:
                pass

        # Tag derniere_inventaire pour TOUS les produits actuellement
        # présents à cet emplacement (y compris ceux non modifiés mais validés).
        conn.execute(
            """UPDATE stock_emplacements
               SET derniere_inventaire=?, updated_at=?, updated_by=?
               WHERE emplacement=?""",
            (now, now, user.get("email"), emplacement),
        )

        # Normalise les commentaires (un seul par produit_id, vides ignorés)
        commentaires_clean: list[dict] = []
        seen_pids: set[int] = set()
        for c in commentaires_in:
            try:
                pid = int(c.get("produit_id"))
            except Exception:
                continue
            if pid in seen_pids:
                continue
            txt = (c.get("commentaire") or "").strip()
            if not txt:
                continue
            # Charge réf / désignation si non fournie
            ref_str = (c.get("reference") or "").strip()
            des_str = (c.get("designation") or "").strip()
            if not ref_str or not des_str:
                row = conn.execute(
                    "SELECT reference, designation FROM produits WHERE id=?", (pid,)
                ).fetchone()
                if row:
                    ref_str = ref_str or (row["reference"] or "")
                    des_str = des_str or (row["designation"] or "")
            commentaires_clean.append({
                "produit_id": pid,
                "reference": ref_str,
                "designation": des_str,
                "commentaire": txt[:1000],
            })
            seen_pids.add(pid)

        operateur_nom = _resolve_created_by_name(conn, user) or (user.get("email") or "")
        conn.execute(
            """INSERT INTO inventaires_sessions
               (emplacement, operateur_email, operateur_nom, date_validation,
                nb_produits, nb_modifications, modifications_json, commentaires_json)
               VALUES (?,?,?,?,?,?,?,?)""",
            (
                emplacement,
                user.get("email"),
                operateur_nom,
                now,
                nb_produits,
                len(modifs_applied),
                json.dumps(modifs_applied, ensure_ascii=False),
                json.dumps(commentaires_clean, ensure_ascii=False),
            ),
        )
        conn.commit()

    log_action(
        user=user,
        action="UPDATE",
        module="stock",
        objet=f"Inventaire emplacement {emplacement}",
        detail={"nb_produits": nb_produits, "nb_modifs": len(modifs_applied)},
        ip=request.client.host if request.client else None,
    )
    return {
        "success": True,
        "nb_modifications": len(modifs_applied),
        "date_validation": now,
    }


# ── Historique unifié ─────────────────────────────────────────────
@router.get("/api/stock/historique-mouvements")
def historique_mouvements(
    request: Request,
    type_stock: str = "tout",
    categorie: Optional[str] = None,
    reference: Optional[str] = None,
    type_mouvement: Optional[str] = None,
    date_debut: Optional[str] = None,
    date_fin: Optional[str] = None,
    limit: int = 200,
    offset: int = 0,
    format: str = "json",
):
    require_stock(request)
    ts = (type_stock or "tout").strip().lower()
    if ts not in _HISTORIQUE_TYPE_STOCK:
        raise HTTPException(400, "type_stock invalide — valeurs : tout, mp, produits.")
    cat = (categorie or "").strip().lower() or None
    if cat and cat not in _MP_CATEGORIES:
        raise HTTPException(400, "Catégorie invalide.")
    tm = (type_mouvement or "").strip().lower() or None
    if tm and tm not in _HISTORIQUE_TYPES_MVT:
        raise HTTPException(400, "Type de mouvement invalide.")
    lim = max(1, min(int(limit), 500))
    off = max(0, int(offset))
    fmt = (format or "json").strip().lower()

    rows_out: list[dict] = []
    with get_db() as conn:
        if ts in ("tout", "mp"):
            where, params = _historique_where_clause(
                True, cat, reference, tm, date_debut, date_fin
            )
            mp_rows = conn.execute(
                _HISTORIQUE_SQL_MP.format(where=where), params
            ).fetchall()
            rows_out.extend(_historique_row_dict(r) for r in mp_rows)

        if ts in ("tout", "produits"):
            where, params = _historique_where_clause(
                False, None, reference, tm, date_debut, date_fin
            )
            pf_rows = conn.execute(
                _HISTORIQUE_SQL_PF.format(where=where), params
            ).fetchall()
            rows_out.extend(_historique_row_dict(r) for r in pf_rows)

    rows_out.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    rows_out = rows_out[off : off + lim]

    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";")
        writer.writerow([
            "Date",
            "Type stock",
            "Catégorie",
            "Référence",
            "Désignation",
            "Unité",
            "Mouvement",
            "Quantité",
            "Avant",
            "Après",
            "Ref BL",
            "Note",
            "Opérateur",
        ])
        for r in rows_out:
            writer.writerow([
                r.get("created_at") or "",
                _historique_type_stock_label(r.get("type_stock") or ""),
                r.get("categorie") or "",
                r.get("reference") or "",
                r.get("designation") or "",
                r.get("unite") or "",
                r.get("type_mouvement") or "",
                r.get("quantite") if r.get("quantite") is not None else "",
                r.get("quantite_avant") if r.get("quantite_avant") is not None else "",
                r.get("quantite_apres") if r.get("quantite_apres") is not None else "",
                r.get("ref_bl") or "",
                r.get("note") or "",
                r.get("created_by_name") or "",
            ])
        data = buf.getvalue().encode("utf-8-sig")
        fname = f"historique_stock_{datetime.now().strftime('%Y%m%d')}.csv"
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    if fmt != "json":
        raise HTTPException(400, "Format non supporté (json ou csv).")
    return rows_out


# ── Stats dashboard ───────────────────────────────────────────────
@router.get("/api/stock/dashboard")
def dashboard(request: Request):
    require_stock(request)
    now = datetime.now()
    limite_alerte = (now - timedelta(days=INVENTAIRE_ALERTE_JOURS)).isoformat()
    with get_db() as conn:
        stats = conn.execute(
            """SELECT
               COUNT(DISTINCT p.id) as nb_refs,
               COUNT(DISTINCT CASE WHEN s.quantite>0 THEN s.emplacement END) as nb_empl_occupes,
               COALESCE(SUM(s.quantite),0) as total_unites,
               COUNT(DISTINCT CASE WHEN (s.derniere_inventaire IS NULL OR s.derniere_inventaire < ?) AND s.quantite>0 THEN s.produit_id END) as nb_a_inventorier
               FROM produits p
               LEFT JOIN stock_emplacements s ON s.produit_id=p.id""",
            (limite_alerte,),
        ).fetchone()

        derniers_mvts = conn.execute(
            f"""SELECT {_MVT_FIELDS}, p.reference, p.designation, p.unite,
                      COALESCE(NULLIF(TRIM(m.created_by_name),''), u.nom) AS created_by_nom
               FROM mouvements_stock m
               JOIN produits p ON p.id=m.produit_id
               {_STOCK_USER_JOIN}
               ORDER BY m.created_at DESC LIMIT 15""",
        ).fetchall()

        top_refs = conn.execute(
            """SELECT p.id, p.reference, p.designation, p.unite,
                      COALESCE(SUM(l.quantite_restante),0) as stock_total
               FROM produits p
               LEFT JOIN lots_stock l ON l.produit_id=p.id AND l.quantite_restante>0
               GROUP BY p.id ORDER BY stock_total DESC LIMIT 8""",
        ).fetchall()

        alertes_mp = conn.execute(
            """
            SELECT mp.id, mp.categorie, mp.reference, mp.designation, mp.seuil_alerte,
                   mp.palettes_par_pile,
                   COALESCE(s.quantite, 0) AS quantite
            FROM matieres_premieres mp
            LEFT JOIN mp_stock s ON s.matiere_id = mp.id
            WHERE mp.actif = 1 AND mp.seuil_alerte > 0
              AND COALESCE(s.quantite, 0) <= mp.seuil_alerte
            ORDER BY mp.categorie, mp.reference
            """
        ).fetchall()

        derniers_mouvements_mp = conn.execute(
            """
            SELECT m.type_mouvement, m.quantite, m.quantite_apres, m.created_at,
                   m.created_by_name, mp.reference, mp.designation, mp.categorie
            FROM mp_mouvements m
            JOIN matieres_premieres mp ON mp.id = m.matiere_id
            ORDER BY m.created_at DESC
            LIMIT 5
            """
        ).fetchall()

        refs_a_expedier = conn.execute(
            """
            SELECT COUNT(DISTINCT l.produit_id) AS nb
            FROM lots_stock l
            WHERE l.emplacement = ? AND l.quantite_restante > 0
            """,
            (STOCK_EMPLACEMENT_AU_SOL,),
        ).fetchone()

        today_prefix = datetime.now().strftime("%Y-%m-%d")
        departs_jour = conn.execute(
            """
            SELECT COUNT(*) AS nb
            FROM expe_departs
            WHERE date_enlevement LIKE ?
            """,
            (today_prefix + "%",),
        ).fetchone()

    return {
        "stats": dict(stats),
        "derniers_mouvements": [dict(r) for r in derniers_mvts],
        "top_refs": [dict(r) for r in top_refs],
        "alertes_mp": [dict(r) for r in alertes_mp],
        "derniers_mouvements_mp": [dict(r) for r in derniers_mouvements_mp],
        "nb_mp_a_approvisionner": len(alertes_mp),
        "nb_refs_a_expedier": refs_a_expedier["nb"] if refs_a_expedier else 0,
        "nb_departs_aujourd_hui": departs_jour["nb"] if departs_jour else 0,
    }


# ─── Réception matière ─────────────────────────────────────────────────────────

_FSC_TYPE_CLAIM_VALUES = frozenset({
    "fsc_100",
    "fsc_mix_credit",
    "fsc_mix",
    "fsc_recycled",
    "non_fsc",
})


def _parse_fsc_type_claim(raw: Any, default: str = "non_fsc") -> str:
    t = (raw or default).strip() if raw is not None else default
    if t not in _FSC_TYPE_CLAIM_VALUES:
        raise HTTPException(
            status_code=400,
            detail="Type FSC invalide — valeurs : fsc_100, fsc_mix_credit, fsc_mix, fsc_recycled, non_fsc.",
        )
    return t


_FSC_CLAIM_SHORT = {
    "fsc_100":        "100",
    "fsc_mix_credit": "MIXC",
    "fsc_mix":        "MIX",
    "fsc_recycled":   "REC",
    "non_fsc":        "NFSC",
}


def _slug_fournisseur(nom: Optional[str]) -> str:
    """Slug court MAJUSCULE alphanum pour numéro de lot (5 chars max)."""
    if not nom:
        return "SANS"
    cleaned = re.sub(r"[^A-Za-z0-9]", "", nom).upper()
    return cleaned[:5] if cleaned else "SANS"


def _build_lot_numero(fournisseur: Optional[str], dt: datetime, fsc_type_claim: str) -> str:
    """Format : LOT-YYYYMMDD-HH-FOURN-FSC."""
    return "LOT-{d}-{h}-{f}-{c}".format(
        d=dt.strftime("%Y%m%d"),
        h=dt.strftime("%H"),
        f=_slug_fournisseur(fournisseur),
        c=_FSC_CLAIM_SHORT.get(fsc_type_claim, "NFSC"),
    )


@router.get("/api/stock/fournisseurs")
def list_fournisseurs_stock(request: Request):
    """Liste des fournisseurs FSC pour la réception matière et le guide traça (lecture publique interne)."""
    get_current_user(request)
    with get_db() as conn:
        rows = conn.execute(
            """SELECT id, nom, licence, certificat, traca_photo_url, traca_explication, traca_exemple_code
               FROM fournisseurs_fsc ORDER BY nom COLLATE NOCASE ASC"""
        ).fetchall()
    return [dict(r) for r in rows]


@router.get("/api/stock/receptions")
def list_receptions(request: Request, limit: int = 50):
    """Historique des réceptions de bobines."""
    user = require_stock(request)
    with get_db() as conn:
        lots = conn.execute(
            """SELECT r.*, GROUP_CONCAT(i.code_barre, '||') as codes
               FROM stock_receptions r
               LEFT JOIN stock_reception_items i ON i.reception_id = r.id
               GROUP BY r.id
               ORDER BY r.created_at DESC LIMIT ?""",
            (limit,),
        ).fetchall()
    result = []
    for lot in lots:
        d = dict(lot)
        raw = d.pop("codes", None)
        d["items"] = raw.split("||") if raw else []
        result.append(d)
    return {"receptions": result}


@router.post("/api/stock/receptions")
async def create_reception(request: Request):
    """Enregistre une réception de bobines (lot de codes-barres).

    Règle de fusion : si un lot existe déjà avec exactement les mêmes
    (fournisseur, jour, heure, type FSC), les bobines sont ajoutées à ce
    lot au lieu d'en créer un nouveau. Le numéro de lot est déterministe :
    LOT-YYYYMMDD-HH-FOURN-FSC.
    """
    user = require_stock(request)
    body = await request.json()
    codes = [str(c).strip() for c in (body.get("codes") or []) if str(c).strip()]
    if not codes:
        raise HTTPException(status_code=400, detail="Aucun code-barres fourni")
    note = (body.get("note") or "").strip() or None
    fournisseur = (body.get("fournisseur") or "").strip() or None
    certificat_fsc = (body.get("certificat_fsc") or "").strip() or None
    fsc_type_claim = _parse_fsc_type_claim(body.get("fsc_type_claim"), "non_fsc")
    if fsc_type_claim != "non_fsc" and not certificat_fsc:
        raise HTTPException(
            status_code=400,
            detail="Certificat FSC requis pour une réception certifiée FSC.",
        )
    now_dt = datetime.now()
    now = now_dt.isoformat()
    lot_numero = _build_lot_numero(fournisseur, now_dt, fsc_type_claim)

    with get_db() as conn:
        # Fusion : chercher un lot identique du jour/heure/fournisseur/FSC.
        existing = conn.execute(
            """SELECT id, nb_bobines FROM stock_receptions
               WHERE lot_numero = ?
               ORDER BY id DESC LIMIT 1""",
            (lot_numero,),
        ).fetchone()

        if existing:
            reception_id = existing["id"]
            merged_note = _merge_note(note, conn, reception_id)
            conn.executemany(
                "INSERT INTO stock_reception_items (reception_id, code_barre, scanned_at) VALUES (?,?,?)",
                [(reception_id, code, now) for code in codes],
            )
            conn.execute(
                """UPDATE stock_receptions
                   SET nb_bobines = COALESCE(nb_bobines, 0) + ?,
                       note = COALESCE(?, note)
                   WHERE id = ?""",
                (len(codes), merged_note, reception_id),
            )
            conn.commit()
            merged = True
            new_total = (existing["nb_bobines"] or 0) + len(codes)
        else:
            cur = conn.execute(
                """INSERT INTO stock_receptions
                   (created_at, created_by, created_by_name, note, nb_bobines,
                    fournisseur, certificat_fsc, fsc_type_claim, lot_numero)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (
                    now,
                    user.get("email"),
                    user.get("nom"),
                    note,
                    len(codes),
                    fournisseur,
                    certificat_fsc,
                    fsc_type_claim,
                    lot_numero,
                ),
            )
            reception_id = cur.lastrowid
            conn.executemany(
                "INSERT INTO stock_reception_items (reception_id, code_barre, scanned_at) VALUES (?,?,?)",
                [(reception_id, code, now) for code in codes],
            )
            conn.commit()
            merged = False
            new_total = len(codes)

    return {
        "success": True,
        "id": reception_id,
        "nb_bobines": new_total,
        "nb_bobines_ajoutees": len(codes),
        "lot_numero": lot_numero,
        "merged": merged,
    }


def _merge_note(new_note: Optional[str], conn, reception_id: int) -> Optional[str]:
    """Concatène la nouvelle note à la note existante si distinctes."""
    if not new_note:
        return None
    row = conn.execute(
        "SELECT note FROM stock_receptions WHERE id=?", (reception_id,)
    ).fetchone()
    prev = (row["note"] or "").strip() if row else ""
    if not prev:
        return new_note
    if new_note in prev:
        return prev
    return f"{prev} | {new_note}"


@router.patch("/api/stock/receptions/{reception_id}")
async def patch_reception(reception_id: int, request: Request):
    """Corrige a posteriori fournisseur, certificat FSC, type de claim ou note."""
    user = require_stock_write(request)
    body = await request.json()

    with get_db() as conn:
        ex = conn.execute(
            "SELECT * FROM stock_receptions WHERE id=?", (reception_id,)
        ).fetchone()
        if not ex:
            raise HTTPException(status_code=404, detail="Réception introuvable")

        exd = dict(ex)
        fournisseur = (
            (body.get("fournisseur") or "").strip() or None
            if "fournisseur" in body
            else exd.get("fournisseur")
        )
        note = (
            (body.get("note") or "").strip() or None
            if "note" in body
            else exd.get("note")
        )
        certificat_fsc = (
            (body.get("certificat_fsc") or "").strip() or None
            if "certificat_fsc" in body
            else exd.get("certificat_fsc")
        )
        fsc_type_claim = (
            _parse_fsc_type_claim(body.get("fsc_type_claim"))
            if "fsc_type_claim" in body
            else _parse_fsc_type_claim(exd.get("fsc_type_claim"), "non_fsc")
        )
        if fsc_type_claim != "non_fsc" and not certificat_fsc:
            raise HTTPException(
                status_code=400,
                detail="Certificat FSC requis pour une réception certifiée FSC.",
            )

        audit_detail: dict = {}
        if fournisseur != exd.get("fournisseur"):
            audit_detail["fournisseur"] = fournisseur
        if note != exd.get("note"):
            audit_detail["note"] = note
        if certificat_fsc != exd.get("certificat_fsc"):
            audit_detail["certificat_fsc"] = certificat_fsc
        if fsc_type_claim != (exd.get("fsc_type_claim") or "non_fsc"):
            audit_detail["fsc_type_claim"] = fsc_type_claim

        conn.execute(
            """UPDATE stock_receptions
               SET fournisseur=?, certificat_fsc=?, fsc_type_claim=?, note=?
               WHERE id=?""",
            (fournisseur, certificat_fsc, fsc_type_claim, note, reception_id),
        )
        conn.commit()

    if audit_detail:
        log_action(
            user=user,
            action="UPDATE",
            module="stock",
            objet=f"Réception bobines #{reception_id}",
            detail=audit_detail,
            ip=request.client.host if request.client else None,
        )

    return {"success": True, "id": reception_id}


# ── Référentiel produits (référence + unité de vente) ─────────────

_REF_HEADER_KEYS = frozenset({
    "reference", "référence", "ref", "ref produit", "référence produit",
    "reference produit", "code", "code produit",
})
_UNITE_HEADER_KEYS = frozenset({
    "unite", "unité", "unite de vente", "unité de vente", "uv", "unité vente",
    "unite vente", "unit",
})
_DES_HEADER_KEYS = frozenset({
    "designation", "désignation", "description", "libelle", "libellé",
})


def _norm_header(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _map_produits_import_columns(df: pd.DataFrame) -> tuple[Optional[str], Optional[str], Optional[str]]:
    ref_col = unite_col = des_col = None
    for col in df.columns:
        key = _norm_header(col)
        if key in _REF_HEADER_KEYS:
            ref_col = col
        elif key in _UNITE_HEADER_KEYS:
            unite_col = col
        elif key in _DES_HEADER_KEYS:
            des_col = col
    if ref_col is None and len(df.columns) >= 1:
        ref_col = df.columns[0]
    if unite_col is None and len(df.columns) >= 2:
        for col in df.columns:
            if col != ref_col:
                unite_col = col
                break
    return ref_col, unite_col, des_col


def _parse_produits_import_df(df: pd.DataFrame) -> list[dict]:
    ref_col, unite_col, des_col = _map_produits_import_columns(df)
    if not ref_col or not unite_col:
        raise HTTPException(
            400,
            "Colonnes introuvables : le fichier doit contenir au minimum « référence » et « unité de vente ».",
        )
    rows: list[dict] = []
    seen_refs: set[str] = set()
    for idx, row in df.iterrows():
        line = int(idx) + 2
        ref_raw = row.get(ref_col)
        unite_raw = row.get(unite_col)
        des_raw = row.get(des_col) if des_col else None
        ref = "" if pd.isna(ref_raw) else str(ref_raw).strip().upper()
        unite = "" if pd.isna(unite_raw) else str(unite_raw).strip()
        designation = ""
        if des_raw is not None and not pd.isna(des_raw):
            designation = str(des_raw).strip()
        if not ref and not unite:
            continue
        if not ref:
            rows.append({
                "line": line, "reference": "", "unite": unite, "designation": designation,
                "action": "error", "message": "Référence manquante",
            })
            continue
        if not unite:
            rows.append({
                "line": line, "reference": ref, "unite": "", "designation": designation,
                "action": "error", "message": "Unité de vente manquante",
            })
            continue
        if ref in seen_refs:
            rows.append({
                "line": line, "reference": ref, "unite": unite, "designation": designation,
                "action": "error", "message": "Référence en double dans le fichier",
            })
            continue
        seen_refs.add(ref)
        rows.append({
            "line": line, "reference": ref, "unite": unite, "designation": designation,
            "action": "pending", "message": "",
        })
    return rows


def _enrich_produits_import_preview(conn, rows: list[dict]) -> dict:
    stats = {"create": 0, "update": 0, "unchanged": 0, "error": 0}
    for r in rows:
        if r["action"] == "error":
            stats["error"] += 1
            continue
        existing = conn.execute(
            "SELECT id, unite, designation FROM produits WHERE reference=?",
            (r["reference"],),
        ).fetchone()
        if not existing:
            r["action"] = "create"
            r["message"] = "Nouvelle référence"
            stats["create"] += 1
        else:
            old_u = (existing["unite"] or "").strip()
            old_d = (existing["designation"] or "").strip()
            new_d = (r["designation"] or "").strip() or old_d or r["reference"]
            if old_u == r["unite"] and (not r["designation"] or old_d == new_d):
                r["action"] = "unchanged"
                r["message"] = "Déjà à jour"
                stats["unchanged"] += 1
            else:
                r["action"] = "update"
                r["message"] = "Mise à jour unité / désignation"
                stats["update"] += 1
    return stats


def _apply_produits_import_row(conn, r: dict, now: str) -> str:
    ref = r["reference"]
    unite = r["unite"]
    des = (r.get("designation") or "").strip() or ref
    existing = conn.execute("SELECT id FROM produits WHERE reference=?", (ref,)).fetchone()
    if existing:
        conn.execute(
            "UPDATE produits SET unite=?, designation=?, updated_at=? WHERE id=?",
            (unite, des, now, existing["id"]),
        )
        return "update"
    conn.execute(
        "INSERT INTO produits (reference, designation, description, unite, created_at, updated_at) "
        "VALUES (?,?,?,?,?,?)",
        (ref, des, "", unite, now, now),
    )
    return "create"


# ── Matières premières ────────────────────────────────────────────
@router.get("/api/stock/matieres")
def list_matieres_premieres(request: Request, all: int = 0):
    user = require_stock(request)
    include_inactive = bool(all) and user.get("role") in _STOCK_MATIERES_ADMIN_ROLES
    with get_db() as conn:
        where_actif = "" if include_inactive else "WHERE mp.actif = 1"
        rows = conn.execute(
            f"""
            SELECT mp.id, mp.categorie, mp.reference, mp.designation,
                   mp.seuil_alerte, mp.actif, mp.palettes_par_pile, mp.couleur,
                   mp.metres_lineaires_par_bobine, mp.prix_eur_m2,
                   COALESCE(mp.prix_par_laize, 0) AS prix_par_laize,
                   mp.unites_par_palette,
                   mp.sous_section,
                   COALESCE(mp.intervalle_inventaire_jours, 180) AS intervalle_inventaire_jours,
                   COALESCE(s.quantite, 0) AS quantite
            FROM matieres_premieres mp
            LEFT JOIN mp_stock s ON s.matiere_id = mp.id
            {where_actif}
            ORDER BY mp.categorie ASC, mp.reference ASC
            """
        ).fetchall()
        # Récupère stocks par laize pour toutes les matières laizées en un coup
        laize_rows = conn.execute(
            """
            SELECT ml.matiere_id, l.id AS laize_id, l.valeur_mm, l.label, l.ordre, l.actif,
                   COALESCE(sl.quantite, 0) AS quantite,
                   ml.prix_eur_m2 AS laize_prix_eur_m2
            FROM mp_matiere_laizes ml
            JOIN mp_laizes l ON l.id = ml.laize_id
            LEFT JOIN mp_stock_laize sl ON sl.matiere_id = ml.matiere_id AND sl.laize_id = ml.laize_id
            ORDER BY l.ordre ASC, l.valeur_mm ASC
            """
        ).fetchall()
        fourn_by_mat: dict[int, dict[int, list[dict]]] = {}
        fourn_rows = conn.execute("""
            SELECT mlf.matiere_id, mlf.laize_id, f.id, f.nom, COALESCE(f.has_fsc, 1) AS has_fsc,
                   f.licence, f.certificat
              FROM matiere_laize_fournisseurs mlf
              JOIN fournisseurs_fsc f ON f.id = mlf.fournisseur_id
             ORDER BY f.nom COLLATE NOCASE ASC
        """).fetchall()
        for fr in fourn_rows:
            fourn_by_mat.setdefault(int(fr["matiere_id"]), {}).setdefault(int(fr["laize_id"]), []).append({
                "id": int(fr["id"]),
                "nom": fr["nom"],
                "has_fsc": int(fr["has_fsc"] or 0),
                "licence": fr["licence"],
                "certificat": fr["certificat"],
            })
    by_mat: dict[int, list[dict]] = {}
    for r in laize_rows:
        try:
            l_prix = float(r["laize_prix_eur_m2"]) if r["laize_prix_eur_m2"] is not None else None
        except (TypeError, ValueError):
            l_prix = None
        mat_id = r["matiere_id"]
        laize_id = r["laize_id"]
        by_mat.setdefault(mat_id, []).append({
            "laize_id": laize_id,
            "valeur_mm": float(r["valeur_mm"] or 0),
            "label": r["label"],
            "actif": int(r["actif"] or 0),
            "quantite": float(r["quantite"] or 0),
            "prix_eur_m2": l_prix,
            "fournisseurs": fourn_by_mat.get(int(mat_id), {}).get(int(laize_id), []),
        })
    out = []
    for r in rows:
        spl = by_mat.get(r["id"]) if _mp_is_laizee(r["categorie"]) else None
        out.append(_mp_row_dict(r, spl))
    return out


_MP_CATEGORIES_WITH_SOUS_SECTION = frozenset({"autre", "frontal"})


@router.get("/api/stock/matieres/sous-sections")
def list_sous_sections(request: Request, categorie: Optional[str] = None):
    """Liste des sous-sections distinctes utilisees pour 'autre' (par defaut) ou 'frontal'.

    Param query `categorie` accepte: 'autre', 'frontal', ou 'all' pour les deux.
    """
    require_stock(request)
    cat = (categorie or "autre").strip().lower()
    if cat == "all":
        cats = tuple(_MP_CATEGORIES_WITH_SOUS_SECTION)
        placeholders = ",".join("?" * len(cats))
        sql = (
            f"SELECT DISTINCT TRIM(sous_section) AS sous_section "
            f"FROM matieres_premieres "
            f"WHERE categorie IN ({placeholders}) "
            f"  AND sous_section IS NOT NULL "
            f"  AND TRIM(sous_section) != '' "
            f"ORDER BY sous_section COLLATE NOCASE ASC"
        )
        params: tuple = cats
    else:
        if cat not in _MP_CATEGORIES_WITH_SOUS_SECTION:
            cat = "autre"
        sql = (
            "SELECT DISTINCT TRIM(sous_section) AS sous_section "
            "FROM matieres_premieres "
            "WHERE categorie = ? "
            "  AND sous_section IS NOT NULL "
            "  AND TRIM(sous_section) != '' "
            "ORDER BY sous_section COLLATE NOCASE ASC"
        )
        params = (cat,)
    with get_db() as conn:
        rows = conn.execute(sql, params).fetchall()
    return [r["sous_section"] for r in rows]


@router.post("/api/stock/matieres")
async def create_matiere_premiere(request: Request):
    require_stock_matieres_admin(request)
    body = await request.json()
    categorie = (body.get("categorie") or "").strip().lower()
    reference = (body.get("reference") or "").strip()
    designation = (body.get("designation") or "").strip()
    seuil_alerte = float(body.get("seuil_alerte") or 0)
    palettes_par_pile = None
    if categorie == "palette":
        try:
            palettes_par_pile = float(body.get("palettes_par_pile") or 0)
        except (TypeError, ValueError):
            palettes_par_pile = 0.0
        if palettes_par_pile <= 0:
            raise HTTPException(
                400,
                "Palettes par pile obligatoire (valeur positive).",
            )

    if categorie not in _MP_CATEGORIES:
        raise HTTPException(400, "Catégorie invalide.")
    if not reference:
        raise HTTPException(400, "Référence obligatoire.")
    if not designation:
        raise HTTPException(400, "Désignation obligatoire.")
    couleur = (body.get("couleur") or "").strip() or None
    if categorie == "glassine" and not couleur:
        couleur = designation

    sous_section = (body.get("sous_section") or "").strip() or None
    if categorie not in _MP_CATEGORIES_WITH_SOUS_SECTION:
        sous_section = None

    unites_par_palette = None
    if _mp_a_conditionnement(categorie) and body.get("unites_par_palette") not in (None, ""):
        try:
            unites_par_palette = float(body.get("unites_par_palette"))
        except (TypeError, ValueError):
            raise HTTPException(400, "Unités par palette invalide.") from None
        if unites_par_palette <= 0:
            raise HTTPException(400, "Unités par palette doit être > 0.")

    with get_db() as conn:
        try:
            cur = conn.execute(
                """
                INSERT INTO matieres_premieres (
                    categorie, reference, designation, seuil_alerte, palettes_par_pile, couleur, sous_section, unites_par_palette
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (categorie, reference, designation, seuil_alerte, palettes_par_pile, couleur, sous_section, unites_par_palette),
            )
            matiere_id = cur.lastrowid
            conn.execute(
                "INSERT INTO mp_stock (matiere_id, quantite) VALUES (?, 0)",
                (matiere_id,),
            )
            conn.commit()
            return {"ok": True, "id": matiere_id}
        except sqlite3.IntegrityError:
            conn.rollback()
            raise HTTPException(
                400,
                "Cette référence existe déjà dans cette catégorie.",
            ) from None


@router.put("/api/stock/matieres/{matiere_id}")
async def update_matiere_premiere(matiere_id: int, request: Request):
    require_stock_matieres_admin(request)
    body = await request.json()
    sets = []
    params: list[Any] = []

    if "reference" in body:
        reference = (body.get("reference") or "").strip()
        if not reference:
            raise HTTPException(400, "Référence obligatoire.")
        sets.append("reference=?")
        params.append(reference)
    if "designation" in body:
        des = (body.get("designation") or "").strip()
        if not des:
            raise HTTPException(400, "Désignation obligatoire.")
        sets.append("designation=?")
        params.append(des)
    if "seuil_alerte" in body:
        sets.append("seuil_alerte=?")
        params.append(float(body.get("seuil_alerte") or 0))
    if "palettes_par_pile" in body:
        sets.append("palettes_par_pile=?")
        params.append(float(body.get("palettes_par_pile") or 0))
    if "actif" in body:
        sets.append("actif=?")
        params.append(1 if body.get("actif") else 0)
    if "couleur" in body:
        sets.append("couleur=?")
        params.append((body.get("couleur") or "").strip() or None)
    if "metres_lineaires_par_bobine" in body:
        v = body.get("metres_lineaires_par_bobine")
        try:
            v = float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            raise HTTPException(400, "Métrage invalide.") from None
        if v is not None and v < 0:
            raise HTTPException(400, "Métrage négatif.")
        sets.append("metres_lineaires_par_bobine=?")
        params.append(v)
    if "prix_eur_m2" in body:
        v = body.get("prix_eur_m2")
        try:
            v = float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            raise HTTPException(400, "Prix m² invalide.") from None
        if v is not None and v < 0:
            raise HTTPException(400, "Prix m² négatif.")
        sets.append("prix_eur_m2=?")
        params.append(v)
    if "unites_par_palette" in body:
        v = body.get("unites_par_palette")
        try:
            v = float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            raise HTTPException(400, "Unités par palette invalide.") from None
        if v is not None and v < 0:
            raise HTTPException(400, "Unités par palette négatif.")
        sets.append("unites_par_palette=?")
        params.append(v)
    if "sous_section" in body:
        sv = (body.get("sous_section") or "").strip() or None
        sets.append("sous_section=?")
        params.append(sv)
    if "intervalle_inventaire_jours" in body:
        v = body.get("intervalle_inventaire_jours")
        try:
            v_int = int(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            raise HTTPException(400, "Intervalle inventaire invalide.") from None
        if v_int is not None and (v_int < 1 or v_int > 3650):
            raise HTTPException(400, "Intervalle inventaire hors bornes (1–3650 jours).")
        sets.append("intervalle_inventaire_jours=?")
        params.append(v_int)

    if not sets:
        raise HTTPException(400, "Aucun champ à mettre à jour.")

    sets.append("updated_at=strftime('%Y-%m-%dT%H:%M:%S','now','localtime')")
    params.append(matiere_id)

    with get_db() as conn:
        row = conn.execute(
            "SELECT id, categorie FROM matieres_premieres WHERE id=?", (matiere_id,)
        ).fetchone()
        if not row:
            raise HTTPException(404, "Matière non trouvée.")
        if "palettes_par_pile" in body:
            if row["categorie"] != "palette":
                raise HTTPException(
                    400,
                    "Palettes par pile uniquement pour les références palette.",
                )
            try:
                ppp = float(body.get("palettes_par_pile") or 0)
            except (TypeError, ValueError):
                raise HTTPException(400, "Palettes par pile invalide.") from None
            if ppp <= 0:
                raise HTTPException(
                    400,
                    "Palettes par pile doit être une valeur positive.",
                )
        if "unites_par_palette" in body and not _mp_a_conditionnement(row["categorie"]):
            raise HTTPException(
                400,
                "Unités par palette uniquement pour cartons, adhésifs et mandrins.",
            )
        try:
            conn.execute(
                f"UPDATE matieres_premieres SET {', '.join(sets)} WHERE id=?",
                params,
            )
            conn.commit()
        except sqlite3.IntegrityError:
            conn.rollback()
            raise HTTPException(
                400,
                "Cette référence existe déjà dans cette catégorie.",
            ) from None
    return {"ok": True}


@router.delete("/api/stock/matieres/{matiere_id}")
def delete_matiere_premiere(matiere_id: int, request: Request):
    require_stock_matieres_admin(request)
    with get_db() as conn:
        row = conn.execute(
            "SELECT id FROM matieres_premieres WHERE id=?", (matiere_id,)
        ).fetchone()
        if not row:
            raise HTTPException(404, "Matière non trouvée.")

        has_mvt = conn.execute(
            "SELECT 1 FROM mp_mouvements WHERE matiere_id=? LIMIT 1",
            (matiere_id,),
        ).fetchone()
        if has_mvt:
            stock = conn.execute(
                "SELECT quantite FROM mp_stock WHERE matiere_id=?",
                (matiere_id,),
            ).fetchone()
            qte = float(stock["quantite"]) if stock else 0.0
            if qte > 0:
                raise HTTPException(
                    400,
                    "Impossible de désactiver une matière avec du stock en cours.",
                )

        conn.execute(
            """
            UPDATE matieres_premieres
            SET actif=0,
                updated_at=strftime('%Y-%m-%dT%H:%M:%S','now','localtime')
            WHERE id=?
            """,
            (matiere_id,),
        )
        conn.commit()
    return {"ok": True}


@router.post("/api/stock/matieres/mouvement")
async def mouvement_matiere_premiere(request: Request):
    user = require_stock_write(request)
    body = await request.json()

    matiere_id = body.get("matiere_id")
    type_mvt = (body.get("type_mouvement") or "").strip().lower()
    if matiere_id is None:
        raise HTTPException(400, "matiere_id obligatoire")
    matiere_id = int(matiere_id)
    if type_mvt not in _MP_TYPES_MVT:
        raise HTTPException(400, "Type de mouvement invalide.")

    try:
        quantite = float(body.get("quantite", 0))
    except (TypeError, ValueError):
        raise HTTPException(400, "Quantité invalide.") from None

    ref_bl = (body.get("ref_bl") or "").strip() or None
    note = (body.get("note") or "").strip() or None
    emplacement_source = body.get("emplacement_source")
    emplacement_dest = body.get("emplacement_dest")
    if emplacement_source is not None:
        emplacement_source = str(emplacement_source).strip().upper() or None
    if emplacement_dest is not None:
        emplacement_dest = str(emplacement_dest).strip().upper() or None

    is_fabrication_user = user.get("role") == "fabrication"

    def _check_emplacement_code(code: Optional[str], allow_null: bool = False) -> Optional[str]:
        if not code:
            if allow_null or is_fabrication_user:
                # Service fabrication ou matière laizée (frontal/glassine/complexe) : emplacement non requis.
                return None
            raise HTTPException(400, "Emplacement obligatoire.")
        if not _is_valid_emplacement(code):
            raise HTTPException(400, f"Format emplacement invalide : {code}")
        return _normalize_emplacement(code)

    # Les contrôles d'emplacement dépendent de la catégorie (laizée ou non) et
    # sont donc effectués une fois que la matière a été récupérée dans le bloc DB
    # ci-dessous.

    if type_mvt in ("entree", "sortie") and quantite <= 0:
        raise HTTPException(400, "Quantité doit être positive.")
    if type_mvt == "ajustement" and quantite < 0:
        raise HTTPException(400, "Quantité invalide pour un ajustement.")

    created_by = user.get("id")
    created_by_name = (user.get("nom") or "").strip() or None

    # Laize (obligatoire pour les matières laizées)
    laize_id_raw = body.get("laize_id")
    laize_id: Optional[int] = None
    if laize_id_raw not in (None, ""):
        try:
            laize_id = int(laize_id_raw)
        except (TypeError, ValueError):
            raise HTTPException(400, "laize_id invalide.") from None

    # Prix €/m² de la réception (optionnel, uniquement pour bobines laizées)
    prix_eur_m2_mvt_raw = body.get("prix_eur_m2")
    prix_eur_m2_mvt: Optional[float] = None
    if prix_eur_m2_mvt_raw not in (None, ""):
        try:
            prix_eur_m2_mvt = float(prix_eur_m2_mvt_raw)
        except (TypeError, ValueError):
            raise HTTPException(400, "prix_eur_m2 invalide.") from None
        if prix_eur_m2_mvt < 0:
            raise HTTPException(400, "prix_eur_m2 négatif.")
        if prix_eur_m2_mvt > 1_000_000:
            raise HTTPException(400, "prix_eur_m2 hors limites.")

    with get_db() as conn:
        mp = conn.execute(
            """SELECT id, categorie, COALESCE(prix_par_laize, 0) AS prix_par_laize,
                      COALESCE(prix_eur_m2, 0) AS prix_matiere_eur_m2
               FROM matieres_premieres WHERE id=? AND actif=1""",
            (matiere_id,),
        ).fetchone()
        if not mp:
            raise HTTPException(404, "Matière non trouvée.")
        unite_mp = _mp_unite_gestion(mp["categorie"])
        laizee = _mp_is_laizee(mp["categorie"])
        prix_par_laize_mode = bool(int(mp["prix_par_laize"] or 0))

        if type_mvt == "entree":
            emplacement_dest = _check_emplacement_code(emplacement_dest, allow_null=laizee)
        elif type_mvt == "sortie":
            emplacement_source = _check_emplacement_code(emplacement_source, allow_null=laizee)

        if laizee:
            if laize_id is None:
                raise HTTPException(400, "Laize obligatoire pour cette catégorie.")
            # Vérifier que la laize est bien associée à la matière
            assoc = conn.execute(
                "SELECT 1 FROM mp_matiere_laizes WHERE matiere_id=? AND laize_id=?",
                (matiere_id, laize_id),
            ).fetchone()
            if not assoc:
                raise HTTPException(400, "Cette laize n'est pas associée à cette matière.")
            stock = conn.execute(
                "SELECT quantite FROM mp_stock_laize WHERE matiere_id=? AND laize_id=?",
                (matiere_id, laize_id),
            ).fetchone()
        else:
            stock = conn.execute(
                "SELECT quantite FROM mp_stock WHERE matiere_id=?",
                (matiere_id,),
            ).fetchone()
        quantite_avant = float(stock["quantite"]) if stock else 0.0
        mvt_quantite = quantite

        if type_mvt == "entree":
            quantite_apres = quantite_avant + quantite
        elif type_mvt == "sortie":
            quantite_apres = quantite_avant - quantite
            if quantite_apres < 0:
                raise HTTPException(
                    400,
                    f"Stock insuffisant — stock actuel : {quantite_avant:g} {unite_mp}.",
                )
        elif type_mvt == "ajustement":
            quantite_apres = quantite
            mvt_quantite = abs(quantite_apres - quantite_avant)
        else:  # transfert
            quantite_apres = quantite_avant

        if laizee:
            # Met à jour mp_stock_laize
            conn.execute(
                """
                INSERT INTO mp_stock_laize (matiere_id, laize_id, quantite, updated_at, updated_by_name)
                VALUES (?,?,?,strftime('%Y-%m-%dT%H:%M:%S','now','localtime'),?)
                ON CONFLICT(matiere_id, laize_id) DO UPDATE SET
                    quantite=excluded.quantite,
                    updated_at=excluded.updated_at,
                    updated_by_name=excluded.updated_by_name
                """,
                (matiere_id, laize_id, quantite_apres, created_by_name),
            )
            # Met à jour mp_stock global comme somme des stocks par laize
            total = conn.execute(
                "SELECT COALESCE(SUM(quantite), 0) AS s FROM mp_stock_laize WHERE matiere_id=?",
                (matiere_id,),
            ).fetchone()
            total_q = float(total["s"] or 0)
            conn.execute(
                """
                INSERT OR REPLACE INTO mp_stock(matiere_id, quantite, updated_at, updated_by_name)
                VALUES (?, ?, strftime('%Y-%m-%dT%H:%M:%S','now','localtime'), ?)
                """,
                (matiere_id, total_q, created_by_name),
            )
        else:
            conn.execute(
                """
                INSERT OR REPLACE INTO mp_stock(matiere_id, quantite, updated_at, updated_by_name)
                VALUES (
                    ?,
                    ?,
                    strftime('%Y-%m-%dT%H:%M:%S','now','localtime'),
                    ?
                )
                """,
                (matiere_id, quantite_apres, created_by_name),
            )

        # Prix moyen pondéré (PMP) — uniquement sur entrée d'une matière laizée avec un prix fourni
        if laizee and type_mvt == "entree" and prix_eur_m2_mvt is not None and prix_eur_m2_mvt > 0:
            if prix_par_laize_mode:
                # PMP au niveau de la laize
                ancien_prix_row = conn.execute(
                    "SELECT prix_eur_m2 FROM mp_matiere_laizes WHERE matiere_id=? AND laize_id=?",
                    (matiere_id, laize_id),
                ).fetchone()
                ancien_prix = 0.0
                if ancien_prix_row and ancien_prix_row["prix_eur_m2"] is not None:
                    try:
                        ancien_prix = float(ancien_prix_row["prix_eur_m2"])
                    except (TypeError, ValueError):
                        ancien_prix = 0.0
                if quantite_avant > 0 and ancien_prix > 0:
                    nouveau_prix = (
                        (quantite_avant * ancien_prix) + (quantite * prix_eur_m2_mvt)
                    ) / (quantite_avant + quantite)
                else:
                    # Stock vide ou prix inconnu → on adopte le prix de cette entrée
                    nouveau_prix = prix_eur_m2_mvt
                conn.execute(
                    "UPDATE mp_matiere_laizes SET prix_eur_m2=? WHERE matiere_id=? AND laize_id=?",
                    (round(nouveau_prix, 6), matiere_id, laize_id),
                )
            else:
                # PMP au niveau matière (agrège toutes les laizes)
                ancien_prix = float(mp["prix_matiere_eur_m2"] or 0)
                stock_mat_row = conn.execute(
                    "SELECT COALESCE(quantite, 0) AS q FROM mp_stock WHERE matiere_id=?",
                    (matiere_id,),
                ).fetchone()
                # On utilise le stock matière AVANT l'entrée : le mp_stock vient d'être mis à jour,
                # donc on soustrait la quantité entrée pour retrouver l'état ancien
                stock_mat_apres = float(stock_mat_row["q"]) if stock_mat_row else 0.0
                stock_mat_avant = max(0.0, stock_mat_apres - quantite)
                if stock_mat_avant > 0 and ancien_prix > 0:
                    nouveau_prix = (
                        (stock_mat_avant * ancien_prix) + (quantite * prix_eur_m2_mvt)
                    ) / (stock_mat_avant + quantite)
                else:
                    nouveau_prix = prix_eur_m2_mvt
                conn.execute(
                    "UPDATE matieres_premieres SET prix_eur_m2=? WHERE id=?",
                    (round(nouveau_prix, 6), matiere_id),
                )
        conn.execute(
            """
            INSERT INTO mp_mouvements (
                matiere_id, type_mouvement, quantite,
                quantite_avant, quantite_apres,
                ref_bl, note, emplacement_source, emplacement_dest,
                created_by, created_by_name, laize_id, prix_eur_m2
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                matiere_id,
                type_mvt,
                mvt_quantite,
                quantite_avant,
                quantite_apres,
                ref_bl,
                note,
                emplacement_source,
                emplacement_dest,
                created_by,
                created_by_name,
                laize_id,
                prix_eur_m2_mvt,
            ),
        )
        conn.commit()

    return {"ok": True, "quantite_apres": quantite_apres, "laize_id": laize_id}


@router.get("/api/stock/matieres/{matiere_id}/mouvements")
def list_matiere_mouvements(matiere_id: int, request: Request):
    require_stock(request)
    with get_db() as conn:
        mp = conn.execute(
            "SELECT id FROM matieres_premieres WHERE id=?",
            (matiere_id,),
        ).fetchone()
        if not mp:
            raise HTTPException(404, "Matière non trouvée.")
        rows = conn.execute(
            """
            SELECT m.*, mp.reference, mp.designation
            FROM mp_mouvements m
            JOIN matieres_premieres mp ON mp.id = m.matiere_id
            WHERE m.matiere_id=?
            ORDER BY m.created_at DESC
            LIMIT 50
            """,
            (matiere_id,),
        ).fetchall()
    return [dict(r) for r in rows]


# ── Inventaire matière (par référence, aligné sur invV2 emplacement) ─────
_INVMAT_DEFAULT_INTERVAL_DAYS = 180


def _invmat_status(jours_depuis: Optional[float], intervalle: int) -> str:
    """Retourne 'vert' | 'orange' | 'rouge' selon le seuil configuré."""
    if jours_depuis is None:
        return "rouge"  # jamais inventoriée
    seuil = max(1, int(intervalle or _INVMAT_DEFAULT_INTERVAL_DAYS))
    ratio = jours_depuis / seuil
    if ratio < 0.5:
        return "vert"
    if ratio < 1.0:
        return "orange"
    return "rouge"


@router.get("/api/stock/matieres/inventaire")
def matieres_inventaire_liste(request: Request):
    """Liste des matières avec dernier inventaire, jours écoulés et statut couleur."""
    require_stock(request)
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT mp.id, mp.reference, mp.designation, mp.categorie, mp.couleur,
                   COALESCE(mp.intervalle_inventaire_jours, ?) AS intervalle_jours,
                   COALESCE(s.quantite, 0) AS stock_actuel,
                   (SELECT MAX(date_validation) FROM inventaires_matieres
                    WHERE matiere_id = mp.id) AS derniere_date,
                   (SELECT COUNT(*) FROM inventaires_matieres
                    WHERE matiere_id = mp.id) AS nb_inventaires
            FROM matieres_premieres mp
            LEFT JOIN mp_stock s ON s.matiere_id = mp.id
            WHERE mp.actif = 1
            ORDER BY mp.categorie, mp.reference COLLATE NOCASE
            """,
            (_INVMAT_DEFAULT_INTERVAL_DAYS,),
        ).fetchall()

    now = datetime.now()
    result = []
    for r in rows:
        d = dict(r)
        derniere = d.get("derniere_date")
        jours = None
        if derniere:
            try:
                dt = datetime.fromisoformat(derniere.split(".")[0])
                jours = (now - dt).total_seconds() / 86400.0
            except (ValueError, TypeError):
                jours = None
        d["jours_depuis"] = round(jours, 1) if jours is not None else None
        d["statut"] = _invmat_status(jours, d["intervalle_jours"])
        d["unite"] = _mp_unite_gestion(d["categorie"])
        d["laizee"] = _mp_is_laizee(d["categorie"])
        result.append(d)
    return {"items": result}


@router.get("/api/stock/matieres/{matiere_id}/inventaire")
def matiere_inventaire_detail(matiere_id: int, request: Request):
    """Détail pour saisir l'inventaire d'une matière (une ligne par laize si laizée).

    Retourne : matiere, intervalle_jours, dernier_inventaire, lignes (stock actuel par laize).
    """
    require_stock(request)
    with get_db() as conn:
        mp = conn.execute(
            """SELECT id, reference, designation, categorie, couleur,
                      COALESCE(intervalle_inventaire_jours, ?) AS intervalle_jours
               FROM matieres_premieres WHERE id=? AND actif=1""",
            (_INVMAT_DEFAULT_INTERVAL_DAYS, matiere_id),
        ).fetchone()
        if not mp:
            raise HTTPException(404, "Matière non trouvée.")
        mp_d = dict(mp)
        laizee = _mp_is_laizee(mp_d["categorie"])
        mp_d["laizee"] = laizee
        mp_d["unite"] = _mp_unite_gestion(mp_d["categorie"])

        lignes: list[dict] = []
        if laizee:
            rows = conn.execute(
                """SELECT ml.laize_id, l.valeur_mm, l.label,
                          COALESCE(sl.quantite, 0) AS quantite
                   FROM mp_matiere_laizes ml
                   JOIN laizes l ON l.id = ml.laize_id
                   LEFT JOIN mp_stock_laize sl
                     ON sl.matiere_id = ml.matiere_id AND sl.laize_id = ml.laize_id
                   WHERE ml.matiere_id = ? AND COALESCE(l.actif, 1) = 1
                   ORDER BY COALESCE(l.ordre, 999), l.valeur_mm""",
                (matiere_id,),
            ).fetchall()
            for r in rows:
                lignes.append({
                    "laize_id": r["laize_id"],
                    "valeur_mm": r["valeur_mm"],
                    "label": r["label"] or (str(r["valeur_mm"]) + " mm" if r["valeur_mm"] else "—"),
                    "stock_actuel": float(r["quantite"] or 0),
                })
        else:
            stock = conn.execute(
                "SELECT COALESCE(quantite, 0) AS quantite FROM mp_stock WHERE matiere_id=?",
                (matiere_id,),
            ).fetchone()
            lignes.append({
                "laize_id": None,
                "valeur_mm": None,
                "label": mp_d["reference"],
                "stock_actuel": float(stock["quantite"]) if stock else 0.0,
            })

        derniere = conn.execute(
            """SELECT MAX(date_validation) AS d FROM inventaires_matieres WHERE matiere_id=?""",
            (matiere_id,),
        ).fetchone()
        mp_d["derniere_date"] = derniere["d"] if derniere else None

    return {"matiere": mp_d, "lignes": lignes}


@router.post("/api/stock/matieres/{matiere_id}/inventaire")
async def matiere_inventaire_valider(matiere_id: int, request: Request):
    """Valide l'inventaire d'une matière.

    Body : { lignes: [{ laize_id, quantite_comptee, commentaire }] }.
    Pour chaque ligne : calcule l'écart, crée un mp_mouvements type='ajustement'
    si écart != 0 (avec note auto), met à jour mp_stock/mp_stock_laize,
    et insère un enregistrement dans inventaires_matieres.
    """
    user = require_stock_write(request)
    body = await request.json()
    lignes_in = body.get("lignes") or []
    if not isinstance(lignes_in, list) or not lignes_in:
        raise HTTPException(400, "Aucune ligne d'inventaire fournie.")

    created_by = user.get("id")
    operateur_email = user.get("email")
    operateur_nom = (user.get("nom") or "").strip() or None
    now_dt = datetime.now()
    now = now_dt.isoformat()
    note_date = now_dt.strftime("%d/%m/%Y")

    with get_db() as conn:
        mp = conn.execute(
            """SELECT id, reference, categorie FROM matieres_premieres WHERE id=? AND actif=1""",
            (matiere_id,),
        ).fetchone()
        if not mp:
            raise HTTPException(404, "Matière non trouvée.")
        laizee = _mp_is_laizee(mp["categorie"])
        unite = _mp_unite_gestion(mp["categorie"])

        results = []
        for idx, li in enumerate(lignes_in):
            laize_id_raw = li.get("laize_id")
            laize_id: Optional[int] = None
            if laize_id_raw not in (None, ""):
                try:
                    laize_id = int(laize_id_raw)
                except (TypeError, ValueError):
                    raise HTTPException(400, f"Ligne {idx + 1} : laize_id invalide.") from None

            if laizee and laize_id is None:
                raise HTTPException(400, f"Ligne {idx + 1} : laize obligatoire (matière laizée).")
            if not laizee and laize_id is not None:
                # Ignorer silencieusement le laize_id sur matière non laizée
                laize_id = None

            try:
                q_comptee = float(li.get("quantite_comptee"))
            except (TypeError, ValueError):
                raise HTTPException(400, f"Ligne {idx + 1} : quantite_comptee invalide.") from None
            if q_comptee < 0:
                raise HTTPException(400, f"Ligne {idx + 1} : quantité négative.")

            commentaire = (li.get("commentaire") or "").strip() or None

            # Stock actuel avant
            if laizee:
                assoc = conn.execute(
                    "SELECT 1 FROM mp_matiere_laizes WHERE matiere_id=? AND laize_id=?",
                    (matiere_id, laize_id),
                ).fetchone()
                if not assoc:
                    raise HTTPException(400, f"Ligne {idx + 1} : laize non associée à la matière.")
                stk_row = conn.execute(
                    "SELECT quantite FROM mp_stock_laize WHERE matiere_id=? AND laize_id=?",
                    (matiere_id, laize_id),
                ).fetchone()
            else:
                stk_row = conn.execute(
                    "SELECT quantite FROM mp_stock WHERE matiere_id=?",
                    (matiere_id,),
                ).fetchone()
            q_avant = float(stk_row["quantite"]) if stk_row and stk_row["quantite"] is not None else 0.0
            ecart = q_comptee - q_avant

            mouvement_id = None
            if abs(ecart) > 1e-9:
                # Note auto : "Inventaire du 09/07/2026 — écart : +2 bob."
                sign = "+" if ecart > 0 else ""
                # Format écart : entier si round, sinon 3 décimales max
                if abs(ecart - round(ecart)) < 1e-6:
                    ecart_txt = f"{sign}{int(round(ecart))}"
                else:
                    ecart_txt = f"{sign}{ecart:.3f}"
                note_auto = f"Inventaire du {note_date} — écart : {ecart_txt} {unite}"
                if commentaire:
                    note_auto += f" | {commentaire}"

                # Update stock (idempotent avec INSERT OR REPLACE, aligné sur code existant)
                if laizee:
                    conn.execute(
                        """INSERT INTO mp_stock_laize (matiere_id, laize_id, quantite, updated_at, updated_by_name)
                           VALUES (?,?,?,strftime('%Y-%m-%dT%H:%M:%S','now','localtime'),?)
                           ON CONFLICT(matiere_id, laize_id) DO UPDATE SET
                               quantite=excluded.quantite,
                               updated_at=excluded.updated_at,
                               updated_by_name=excluded.updated_by_name""",
                        (matiere_id, laize_id, q_comptee, operateur_nom),
                    )
                    total = conn.execute(
                        "SELECT COALESCE(SUM(quantite), 0) AS s FROM mp_stock_laize WHERE matiere_id=?",
                        (matiere_id,),
                    ).fetchone()
                    conn.execute(
                        """INSERT OR REPLACE INTO mp_stock(matiere_id, quantite, updated_at, updated_by_name)
                           VALUES (?, ?, strftime('%Y-%m-%dT%H:%M:%S','now','localtime'), ?)""",
                        (matiere_id, float(total["s"] or 0), operateur_nom),
                    )
                else:
                    conn.execute(
                        """INSERT OR REPLACE INTO mp_stock(matiere_id, quantite, updated_at, updated_by_name)
                           VALUES (?, ?, strftime('%Y-%m-%dT%H:%M:%S','now','localtime'), ?)""",
                        (matiere_id, q_comptee, operateur_nom),
                    )

                cur = conn.execute(
                    """INSERT INTO mp_mouvements (
                           matiere_id, type_mouvement, quantite,
                           quantite_avant, quantite_apres,
                           ref_bl, note, emplacement_source, emplacement_dest,
                           created_by, created_by_name, laize_id, prix_eur_m2
                       ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        matiere_id,
                        "ajustement",
                        abs(ecart),
                        q_avant,
                        q_comptee,
                        None,
                        note_auto,
                        None,
                        None,
                        created_by,
                        operateur_nom,
                        laize_id,
                        None,
                    ),
                )
                mouvement_id = cur.lastrowid

            # Trace inventaire (toujours, écart nul inclus)
            conn.execute(
                """INSERT INTO inventaires_matieres (
                       matiere_id, laize_id, quantite_avant, quantite_comptee,
                       ecart, commentaire, operateur_email, operateur_nom,
                       date_validation, mouvement_id
                   ) VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (
                    matiere_id,
                    laize_id,
                    q_avant,
                    q_comptee,
                    ecart,
                    commentaire,
                    operateur_email,
                    operateur_nom,
                    now,
                    mouvement_id,
                ),
            )
            results.append({
                "laize_id": laize_id,
                "quantite_avant": q_avant,
                "quantite_comptee": q_comptee,
                "ecart": ecart,
                "mouvement_id": mouvement_id,
            })
        conn.commit()

    log_action(
        user=user,
        action="INVENTAIRE",
        module="stock",
        objet=f"Matière #{matiere_id} ({mp['reference']})",
        detail={"lignes": len(results), "total_ecart": sum(r["ecart"] for r in results)},
        ip=request.client.host if request.client else None,
    )
    return {"success": True, "matiere_id": matiere_id, "lignes": results}


@router.get("/api/stock/matieres/{matiere_id}/inventaires")
def matiere_inventaires_historique(matiere_id: int, request: Request, limit: int = 100):
    """Historique des inventaires d'une matière."""
    require_stock(request)
    with get_db() as conn:
        rows = conn.execute(
            """SELECT im.*, l.valeur_mm, l.label AS laize_label
               FROM inventaires_matieres im
               LEFT JOIN laizes l ON l.id = im.laize_id
               WHERE im.matiere_id = ?
               ORDER BY im.date_validation DESC
               LIMIT ?""",
            (matiere_id, limit),
        ).fetchall()
    return [dict(r) for r in rows]


@router.patch("/api/stock/matieres/{matiere_id}/intervalle-inventaire")
async def matiere_set_intervalle(matiere_id: int, request: Request):
    """Ajuste l'intervalle d'inventaire (jours) d'une matière — admin matière."""
    user = require_stock_write(request)
    if user.get("role") not in _STOCK_MATIERES_ADMIN_ROLES:
        raise HTTPException(403, "Réservé aux administrateurs matières.")
    body = await request.json()
    try:
        jours = int(body.get("intervalle_inventaire_jours"))
    except (TypeError, ValueError):
        raise HTTPException(400, "intervalle_inventaire_jours invalide.") from None
    if jours < 1 or jours > 3650:
        raise HTTPException(400, "Intervalle hors bornes (1–3650 jours).")
    with get_db() as conn:
        ex = conn.execute("SELECT id FROM matieres_premieres WHERE id=?", (matiere_id,)).fetchone()
        if not ex:
            raise HTTPException(404, "Matière non trouvée.")
        conn.execute(
            "UPDATE matieres_premieres SET intervalle_inventaire_jours=? WHERE id=?",
            (jours, matiere_id),
        )
        conn.commit()
    return {"success": True, "matiere_id": matiere_id, "intervalle_inventaire_jours": jours}


# ── Produits finis (onglet dédié — source : produits / mouvements_stock) ──
_PF_MVT_SQL = f"""
    SELECT m.id, p.id AS produit_id, p.reference, p.designation, m.type_mouvement AS type,
           m.quantite, m.quantite_avant, m.quantite_apres,
           p.unite, m.emplacement, m.note AS commentaire,
           COALESCE(NULLIF(TRIM(m.created_by_name),''), u.nom, m.created_by) AS user_login,
           m.created_at AS date_mouvement
    FROM mouvements_stock m
    JOIN produits p ON p.id = m.produit_id
    {_STOCK_USER_JOIN}
"""


def _pf_compose_note(body: dict) -> str:
    parts: list[str] = []
    no_of = (body.get("no_of") or "").strip()
    if no_of:
        parts.append("OF " + no_of)
    motif = (body.get("motif_destinataire") or "").strip()
    if motif:
        parts.append("Motif : " + motif)
    com = (body.get("commentaire") or "").strip()
    if com:
        parts.append(com)
    return " | ".join(parts)


def _pf_ensure_produit_id(
    conn: sqlite3.Connection,
    reference: str,
    designation: str,
    unite: str,
    now: str,
    produit_type: str = "fabrique",
) -> int:
    ref = reference.strip().upper()
    row = conn.execute(
        "SELECT id, designation, unite, type FROM produits WHERE reference=?", (ref,)
    ).fetchone()
    des = designation.strip() or ref
    u = (unite or "").strip() or "étiquette"
    if row:
        existing_type = row["type"] or "fabrique"
        if existing_type != produit_type:
            label_existing = "produit fabriqué" if existing_type == "fabrique" else "produit de négoce"
            label_wanted = "produit de négoce" if produit_type == "negoce" else "produit fabriqué"
            raise HTTPException(
                409,
                f"La référence {ref} est déjà enregistrée comme {label_existing}. "
                f"Impossible de l'utiliser comme {label_wanted}.",
            )
        conn.execute(
            "UPDATE produits SET designation=?, unite=?, updated_at=? WHERE id=?",
            (des, u, now, row["id"]),
        )
        return int(row["id"])
    cur = conn.execute(
        "INSERT INTO produits (reference, designation, description, unite, type, created_at, updated_at) "
        "VALUES (?,?,?,?,?,?,?)",
        (ref, des, "", u, produit_type, now, now),
    )
    return int(cur.lastrowid)


def _pf_list_stock_rows(conn: sqlite3.Connection, produit_type: str = "fabrique") -> list:
    """Stock PF/négoce : lots FIFO, repli sur stock_emplacements. Filtré par type."""
    rows = conn.execute(
        """
        SELECT
            p.id AS produit_id,
            p.reference,
            p.designation,
            p.unite,
            l.emplacement,
            SUM(l.quantite_restante) AS quantite,
            MAX(l.date_entree) AS derniere_entree
        FROM lots_stock l
        JOIN produits p ON p.id = l.produit_id
        WHERE l.quantite_restante > 0.0001
          AND COALESCE(p.type, 'fabrique') = ?
        GROUP BY l.emplacement, p.id, p.reference, p.designation, p.unite
        ORDER BY p.reference ASC, l.emplacement ASC
        """,
        (produit_type,),
    ).fetchall()
    if rows:
        return rows
    return conn.execute(
        """
        SELECT
            p.id AS produit_id,
            p.reference,
            p.designation,
            p.unite,
            s.emplacement,
            s.quantite,
            (
                SELECT MAX(m.created_at)
                FROM mouvements_stock m
                WHERE m.produit_id = s.produit_id
                  AND m.emplacement = s.emplacement
                  AND m.type_mouvement = 'entree'
            ) AS derniere_entree
        FROM stock_emplacements s
        JOIN produits p ON p.id = s.produit_id
        WHERE s.quantite > 0.0001
          AND COALESCE(p.type, 'fabrique') = ?
        ORDER BY p.reference ASC, s.emplacement ASC
        """,
        (produit_type,),
    ).fetchall()


def _pf_kpis(conn: sqlite3.Connection, produit_type: str = "fabrique") -> dict:
    today = datetime.now().strftime("%Y-%m-%d")
    refs = conn.execute(
        """
        SELECT COUNT(DISTINCT l.produit_id) AS n
        FROM (
            SELECT produit_id FROM lots_stock WHERE quantite_restante > 0.0001
            UNION
            SELECT produit_id FROM stock_emplacements WHERE quantite > 0.0001
        ) l
        JOIN produits p ON p.id = l.produit_id
        WHERE COALESCE(p.type, 'fabrique') = ?
        """,
        (produit_type,),
    ).fetchone()
    empl = conn.execute(
        """
        SELECT COUNT(DISTINCT l.emplacement) AS n
        FROM (
            SELECT ls.emplacement, ls.produit_id FROM lots_stock ls WHERE ls.quantite_restante > 0.0001
            UNION
            SELECT s.emplacement, s.produit_id FROM stock_emplacements s WHERE s.quantite > 0.0001
        ) l
        JOIN produits p ON p.id = l.produit_id
        WHERE COALESCE(p.type, 'fabrique') = ?
        """,
        (produit_type,),
    ).fetchone()
    mvt_today = conn.execute(
        """
        SELECT COUNT(*) AS n
        FROM mouvements_stock m
        JOIN produits p ON p.id = m.produit_id
        WHERE m.created_at >= ?
          AND m.type_mouvement IN ('entree', 'sortie')
          AND COALESCE(p.type, 'fabrique') = ?
        """,
        (today + "T00:00:00", produit_type),
    ).fetchone()
    total_mvt = conn.execute(
        """
        SELECT COUNT(*) AS n
        FROM mouvements_stock m
        JOIN produits p ON p.id = m.produit_id
        WHERE COALESCE(p.type, 'fabrique') = ?
        """,
        (produit_type,),
    ).fetchone()
    return {
        "references": int(refs["n"] or 0) if refs else 0,
        "mouvements_aujourdhui": int(mvt_today["n"] or 0) if mvt_today else 0,
        "emplacements_occupes": int(empl["n"] or 0) if empl else 0,
        "total_mouvements": int(total_mvt["n"] or 0) if total_mvt else 0,
    }


def _pf_normalize_mvt_row(r) -> dict:
    d = dict(r)
    t = (d.get("type") or d.get("type_mouvement") or "").strip().lower()
    if t in ("entree", "sortie"):
        d["type"] = t
    d["date_mouvement"] = d.get("date_mouvement") or d.get("created_at")
    d["user_login"] = d.get("user_login") or d.get("created_by_name") or d.get("created_by")
    return d


@router.get("/api/stock/produits-finis")
def list_produits_finis(request: Request):
    require_stock(request)
    with get_db() as conn:
        stock_rows = _pf_list_stock_rows(conn, produit_type="fabrique")
        catalogue = conn.execute(
            "SELECT reference, designation, unite FROM produits "
            "WHERE COALESCE(type,'fabrique')='fabrique' ORDER BY reference ASC"
        ).fetchall()
        kpis = _pf_kpis(conn, produit_type="fabrique")
    return {
        "stock": [
            {
                "produit_id": r["produit_id"],
                "reference": r["reference"],
                "designation": r["designation"],
                "unite": r["unite"],
                "emplacement": r["emplacement"],
                "quantite": float(r["quantite"] or 0),
                "derniere_entree": r["derniere_entree"],
            }
            for r in stock_rows
        ],
        "catalogue": [dict(r) for r in catalogue],
        "kpis": kpis,
    }


@router.get("/api/stock/produits-finis/mouvements")
def list_produits_finis_mouvements(request: Request, limit: int = 20):
    require_stock(request)
    limit = max(1, min(int(limit or 20), 100))
    with get_db() as conn:
        rows = conn.execute(
            _PF_MVT_SQL
            + """
            WHERE m.type_mouvement IN ('entree', 'sortie')
            ORDER BY m.created_at DESC, m.id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [_pf_normalize_mvt_row(r) for r in rows]


@router.get("/api/stock/produits-finis/{reference}")
def get_produit_fini_detail(reference: str, request: Request):
    require_stock(request)
    ref = (reference or "").strip().upper()
    if not ref:
        raise HTTPException(400, "Référence obligatoire.")
    with get_db() as conn:
        cat = conn.execute(
            "SELECT id, reference, designation, unite FROM produits WHERE reference=?",
            (ref,),
        ).fetchone()
        if not cat:
            raise HTTPException(404, "Référence non trouvée.")
        produit_id = cat["id"]
        stock_rows = conn.execute(
            """
            SELECT emplacement, SUM(quantite_restante) AS quantite
            FROM lots_stock
            WHERE produit_id=? AND quantite_restante > 0.0001
            GROUP BY emplacement
            ORDER BY emplacement ASC
            """,
            (produit_id,),
        ).fetchall()
        if not stock_rows:
            stock_rows = conn.execute(
                """
                SELECT emplacement, quantite
                FROM stock_emplacements
                WHERE produit_id=? AND quantite > 0.0001
                ORDER BY emplacement ASC
                """,
                (produit_id,),
            ).fetchall()
        historique = conn.execute(
            _PF_MVT_SQL
            + """
            WHERE p.id=?
            ORDER BY m.created_at DESC, m.id DESC
            """,
            (produit_id,),
        ).fetchall()
    stock_total = sum(float(r["quantite"] or 0) for r in stock_rows)
    return {
        "reference": ref,
        "designation": cat["designation"],
        "unite": cat["unite"],
        "stock_total": stock_total,
        "emplacements": [
            {
                "emplacement": r["emplacement"],
                "quantite": float(r["quantite"] or 0),
                "unite": cat["unite"],
            }
            for r in stock_rows
        ],
        "historique": [_pf_normalize_mvt_row(r) for r in historique],
    }


def _pf_validate_mouvement_body(body: dict) -> tuple[str, str, str, float, str]:
    reference = (body.get("reference") or "").strip().upper()
    designation = (body.get("designation") or "").strip()
    emplacement = (body.get("emplacement") or "").strip().upper()
    unite = (body.get("unite") or "étiquette").strip() or "étiquette"
    if not reference:
        raise HTTPException(400, "Référence obligatoire.")
    if not emplacement:
        raise HTTPException(400, "Emplacement obligatoire.")
    if not _is_valid_emplacement(emplacement):
        raise HTTPException(400, f"Format emplacement invalide : {emplacement}")
    try:
        quantite = float(body.get("quantite", 0))
    except (TypeError, ValueError):
        raise HTTPException(400, "Quantité invalide.") from None
    if quantite < 0.01:
        raise HTTPException(400, "Quantité invalide — minimum 0,01.")
    if not designation:
        designation = reference
    return reference, designation, _normalize_emplacement(emplacement), quantite, unite


@router.post("/api/stock/produits-finis/entree")
async def produit_fini_entree(request: Request):
    user = require_stock_write(request)
    body = await request.json()
    reference, designation, emplacement, quantite, unite = _pf_validate_mouvement_body(body)
    note = _pf_compose_note(body)
    now = datetime.now().isoformat()

    with get_db() as conn:
        produit_id = _pf_ensure_produit_id(conn, reference, designation, unite, now)
        result, ref_audit, audit_action = _apply_stock_mouvement(
            conn, user, produit_id, emplacement, "entree", quantite, note
        )
        conn.commit()

    log_action(
        user=user,
        action=audit_action,
        module="stock",
        objet=f"PF entrée {ref_audit} +{quantite:g} @ {emplacement}",
        detail={"type_mouvement": "entree", "quantite": quantite},
        ip=request.client.host if request.client else None,
    )
    return {"ok": True, **result}


@router.post("/api/stock/produits-finis/sortie")
async def produit_fini_sortie(request: Request):
    user = require_stock_write(request)
    body = await request.json()
    reference, designation, emplacement, quantite, unite = _pf_validate_mouvement_body(body)
    note = _pf_compose_note(body)
    now = datetime.now().isoformat()

    with get_db() as conn:
        produit_id = _pf_ensure_produit_id(conn, reference, designation, unite, now)
        ex = conn.execute(
            "SELECT quantite FROM stock_emplacements WHERE produit_id=? AND emplacement=?",
            (produit_id, emplacement),
        ).fetchone()
        stock = float(ex["quantite"]) if ex else 0.0
        if quantite > stock + 0.0001:
            raise HTTPException(
                400,
                f"Stock insuffisant — disponible : {stock:g} {unite}.",
            )
        result, ref_audit, audit_action = _apply_stock_mouvement(
            conn, user, produit_id, emplacement, "sortie", quantite, note
        )
        conn.commit()

    log_action(
        user=user,
        action=audit_action,
        module="stock",
        objet=f"PF sortie {ref_audit} -{quantite:g} @ {emplacement}",
        detail={"type_mouvement": "sortie", "quantite": quantite},
        ip=request.client.host if request.client else None,
    )
    return {"ok": True, **result}


# ── Produits de négoce ─────────────────────────────────────────────
# Même mécanique que produits-finis, filtrée sur type='negoce'.

@router.get("/api/stock/negoce")
def list_negoce(request: Request):
    require_stock(request)
    with get_db() as conn:
        stock_rows = _pf_list_stock_rows(conn, produit_type="negoce")
        catalogue = conn.execute(
            "SELECT reference, designation, unite FROM produits "
            "WHERE COALESCE(type,'fabrique')='negoce' ORDER BY reference ASC"
        ).fetchall()
        kpis = _pf_kpis(conn, produit_type="negoce")
    return {
        "stock": [
            {
                "produit_id": r["produit_id"],
                "reference": r["reference"],
                "designation": r["designation"],
                "unite": r["unite"],
                "emplacement": r["emplacement"],
                "quantite": float(r["quantite"] or 0),
                "derniere_entree": r["derniere_entree"],
            }
            for r in stock_rows
        ],
        "catalogue": [dict(r) for r in catalogue],
        "kpis": kpis,
    }


@router.get("/api/stock/negoce/mouvements")
def list_negoce_mouvements(request: Request, limit: int = 20):
    require_stock(request)
    limit = max(1, min(int(limit or 20), 100))
    with get_db() as conn:
        rows = conn.execute(
            _PF_MVT_SQL
            + """
            WHERE m.type_mouvement IN ('entree', 'sortie')
              AND COALESCE(p.type, 'fabrique') = 'negoce'
            ORDER BY m.created_at DESC, m.id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [_pf_normalize_mvt_row(r) for r in rows]


@router.get("/api/stock/negoce/{reference}")
def get_negoce_detail(reference: str, request: Request):
    require_stock(request)
    ref = (reference or "").strip().upper()
    if not ref:
        raise HTTPException(400, "Référence obligatoire.")
    with get_db() as conn:
        cat = conn.execute(
            "SELECT id, reference, designation, unite FROM produits "
            "WHERE reference=? AND COALESCE(type,'fabrique')='negoce'",
            (ref,),
        ).fetchone()
        if not cat:
            raise HTTPException(404, "Référence non trouvée.")
        produit_id = cat["id"]
        stock_rows = conn.execute(
            """
            SELECT emplacement, SUM(quantite_restante) AS quantite
            FROM lots_stock
            WHERE produit_id=? AND quantite_restante > 0.0001
            GROUP BY emplacement
            ORDER BY emplacement ASC
            """,
            (produit_id,),
        ).fetchall()
        if not stock_rows:
            stock_rows = conn.execute(
                """
                SELECT emplacement, quantite
                FROM stock_emplacements
                WHERE produit_id=? AND quantite > 0.0001
                ORDER BY emplacement ASC
                """,
                (produit_id,),
            ).fetchall()
        historique = conn.execute(
            _PF_MVT_SQL
            + """
            WHERE p.id=?
            ORDER BY m.created_at DESC, m.id DESC
            """,
            (produit_id,),
        ).fetchall()
    stock_total = sum(float(r["quantite"] or 0) for r in stock_rows)
    return {
        "reference": ref,
        "designation": cat["designation"],
        "unite": cat["unite"],
        "stock_total": stock_total,
        "emplacements": [
            {
                "emplacement": r["emplacement"],
                "quantite": float(r["quantite"] or 0),
                "unite": cat["unite"],
            }
            for r in stock_rows
        ],
        "historique": [_pf_normalize_mvt_row(r) for r in historique],
    }


@router.post("/api/stock/negoce/entree")
async def negoce_entree(request: Request):
    user = require_stock_write(request)
    body = await request.json()
    reference, designation, emplacement, quantite, unite = _pf_validate_mouvement_body(body)
    # Unité par défaut rouleau pour le négoce
    if not (body.get("unite") or "").strip():
        unite = "rouleau"
    note = _pf_compose_note(body)
    now = datetime.now().isoformat()

    with get_db() as conn:
        produit_id = _pf_ensure_produit_id(conn, reference, designation, unite, now, produit_type="negoce")
        result, ref_audit, audit_action = _apply_stock_mouvement(
            conn, user, produit_id, emplacement, "entree", quantite, note
        )
        conn.commit()

    log_action(
        user=user,
        action=audit_action,
        module="stock",
        objet=f"Négoce entrée {ref_audit} +{quantite:g} @ {emplacement}",
        detail={"type_mouvement": "entree", "quantite": quantite, "produit_type": "negoce"},
        ip=request.client.host if request.client else None,
    )
    return {"ok": True, **result}


@router.post("/api/stock/negoce/sortie")
async def negoce_sortie(request: Request):
    user = require_stock_write(request)
    body = await request.json()
    reference, designation, emplacement, quantite, unite = _pf_validate_mouvement_body(body)
    note = _pf_compose_note(body)
    now = datetime.now().isoformat()

    with get_db() as conn:
        produit_id = _pf_ensure_produit_id(conn, reference, designation, unite, now, produit_type="negoce")
        ex = conn.execute(
            "SELECT quantite FROM stock_emplacements WHERE produit_id=? AND emplacement=?",
            (produit_id, emplacement),
        ).fetchone()
        stock = float(ex["quantite"]) if ex else 0.0
        if quantite > stock + 0.0001:
            raise HTTPException(
                400,
                f"Stock insuffisant — disponible : {stock:g} {unite}.",
            )
        result, ref_audit, audit_action = _apply_stock_mouvement(
            conn, user, produit_id, emplacement, "sortie", quantite, note
        )
        conn.commit()

    log_action(
        user=user,
        action=audit_action,
        module="stock",
        objet=f"Négoce sortie {ref_audit} -{quantite:g} @ {emplacement}",
        detail={"type_mouvement": "sortie", "quantite": quantite, "produit_type": "negoce"},
        ip=request.client.host if request.client else None,
    )
    return {"ok": True, **result}


@router.post("/api/stock/negoce/produit")
async def upsert_negoce_produit(request: Request):
    """Créer ou mettre à jour un produit de négoce dans le catalogue."""
    require_stock_write(request)
    body = await request.json()
    reference = (body.get("reference") or "").strip().upper()
    designation = (body.get("designation") or "").strip()
    unite = (body.get("unite") or "rouleau").strip() or "rouleau"
    if not reference:
        raise HTTPException(400, "Référence obligatoire.")
    now = datetime.now().isoformat()
    with get_db() as conn:
        existing = conn.execute(
            "SELECT id, type FROM produits WHERE reference=?", (reference,)
        ).fetchone()
        if existing:
            if (existing["type"] or "fabrique") != "negoce":
                raise HTTPException(
                    409,
                    f"La référence {reference} existe déjà comme produit fabriqué.",
                )
            conn.execute(
                "UPDATE produits SET designation=?, unite=?, updated_at=? WHERE id=?",
                (designation or reference, unite, now, existing["id"]),
            )
            produit_id = existing["id"]
        else:
            cur = conn.execute(
                "INSERT INTO produits (reference, designation, description, unite, type, created_at, updated_at) "
                "VALUES (?,?,?,?,?,?,?)",
                (reference, designation or reference, "", unite, "negoce", now, now),
            )
            produit_id = cur.lastrowid
        conn.commit()
    return {"ok": True, "produit_id": produit_id}


@router.delete("/api/stock/negoce/produit/{reference}")
def delete_negoce_produit(reference: str, request: Request):
    """Supprimer un produit de négoce (seulement si stock = 0)."""
    require_stock_write(request)
    ref = (reference or "").strip().upper()
    with get_db() as conn:
        cat = conn.execute(
            "SELECT id, type FROM produits WHERE reference=?", (ref,)
        ).fetchone()
        if not cat:
            raise HTTPException(404, "Référence non trouvée.")
        if (cat["type"] or "fabrique") != "negoce":
            raise HTTPException(403, "Ce produit n'est pas un produit de négoce.")
        has_stock = conn.execute(
            "SELECT 1 FROM lots_stock WHERE produit_id=? AND quantite_restante > 0.0001 LIMIT 1",
            (cat["id"],),
        ).fetchone()
        if has_stock:
            raise HTTPException(
                409, "Impossible de supprimer : stock non nul. Soldez le stock d'abord."
            )
        conn.execute("DELETE FROM stock_emplacements WHERE produit_id=?", (cat["id"],))
        conn.execute("DELETE FROM lots_stock WHERE produit_id=?", (cat["id"],))
        conn.execute("DELETE FROM produits WHERE id=?", (cat["id"],))
        conn.commit()
    return {"ok": True}


@router.post("/api/stock/produits/import/preview")
async def preview_produits_import(request: Request, file: UploadFile = File(...)):
    require_stock_write(request)
    contents = await file.read()
    filename = file.filename or "import.csv"
    try:
        df = parse_file(contents, filename)
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    df.columns = [str(c).strip() for c in df.columns]
    parsed = _parse_produits_import_df(df)
    if not parsed:
        raise HTTPException(400, "Aucune ligne valide dans le fichier.")
    with get_db() as conn:
        stats = _enrich_produits_import_preview(conn, parsed)
    return {"filename": filename, "rows": parsed, "stats": stats}


@router.post("/api/stock/produits/import/confirm")
async def confirm_produits_import(request: Request):
    require_stock_write(request)
    body = await request.json()
    rows = body.get("rows") or []
    if not isinstance(rows, list) or not rows:
        raise HTTPException(400, "Aucune ligne à importer.")
    now = datetime.now().isoformat()
    applied = {"create": 0, "update": 0, "skipped": 0, "error": 0}
    with get_db() as conn:
        for item in rows:
            ref = (item.get("reference") or "").strip().upper()
            unite = (item.get("unite") or "").strip()
            action = (item.get("action") or "").strip()
            if action in ("error", "unchanged", "skip"):
                applied["skipped"] += 1
                continue
            if not ref or not unite:
                applied["error"] += 1
                continue
            row = {
                "reference": ref,
                "unite": unite,
                "designation": (item.get("designation") or "").strip(),
            }
            try:
                result = _apply_produits_import_row(conn, row, now)
                applied[result] += 1
            except sqlite3.IntegrityError:
                applied["error"] += 1
        conn.commit()
    return {"success": True, "applied": applied}


# ─────────────────────────────────────────────────────────────────────────
# Valorisation matières premières (Contrôle → Valorisation)
# ─────────────────────────────────────────────────────────────────────────

_MP_CATEGORIE_LABELS = {
    "frontal": "Frontal",
    "glassine": "Glassine",
    "adhesif": "Adhésif",
    "mandrin": "Mandrin",
    "carton": "Carton",
    "palette": "Palette",
    "autre": "Autre",
}


def _mp_categorie_order(cat: str) -> int:
    order = ["frontal", "glassine", "adhesif", "mandrin", "carton", "palette", "autre"]
    try:
        return order.index((cat or "").lower())
    except ValueError:
        return len(order)


def _valorisation_query(conn) -> list[dict]:
    """Retourne la liste des lignes de valorisation.

    - Catégories non laizées : 1 ligne par matière, valorisation = qté × prix_unitaire.
    - Catégories laizées (frontal/glassine/complexe) : 1 ligne par (matière × laize),
      valorisation_bobine = (laize_mm/1000) × metres_lineaires × prix_eur_m2,
      valorisation = qté_bobines × valorisation_bobine.
    """
    # Lignes "classiques" (non laizées)
    rows_simple = conn.execute(
        """
        SELECT mp.id, mp.categorie, mp.reference, mp.designation, mp.actif,
               mp.unites_par_palette, mp.sous_section,
               COALESCE(s.quantite, 0) AS quantite,
               COALESCE(v.prix_unitaire, 0) AS prix_unitaire,
               COALESCE(v.prix_en_usd, 0) AS prix_en_usd,
               COALESCE(v.taxe_importation, 0) AS taxe_importation,
               COALESCE(v.cout_transport_inclus, 0) AS cout_transport_inclus,
               v.updated_at AS prix_updated_at,
               v.updated_by_name AS prix_updated_by_name
        FROM matieres_premieres mp
        LEFT JOIN mp_stock s ON s.matiere_id = mp.id
        LEFT JOIN mp_valorisation v ON v.matiere_id = mp.id
        WHERE mp.actif = 1 AND mp.categorie NOT IN ('frontal','glassine','complexe')
        ORDER BY mp.categorie ASC, mp.reference ASC
        """
    ).fetchall()
    # Lignes laizées (une par couple matière × laize associée)
    rows_laizees = conn.execute(
        """
        SELECT mp.id AS matiere_id, mp.categorie, mp.reference, mp.designation,
               mp.sous_section,
               COALESCE(mp.metres_lineaires_par_bobine, 0) AS metres,
               COALESCE(mp.prix_eur_m2, 0) AS prix_eur_m2,
               COALESCE(mp.prix_par_laize, 0) AS prix_par_laize,
               ml.prix_eur_m2 AS laize_prix_eur_m2,
               COALESCE(v.prix_en_usd, 0) AS prix_en_usd,
               COALESCE(v.taxe_importation, 0) AS taxe_importation,
               COALESCE(v.cout_transport_inclus, 0) AS cout_transport_inclus,
               l.id AS laize_id, l.valeur_mm, l.label AS laize_label, l.ordre AS laize_ordre,
               COALESCE(sl.quantite, 0) AS quantite
        FROM matieres_premieres mp
        JOIN mp_matiere_laizes ml ON ml.matiere_id = mp.id
        JOIN mp_laizes l ON l.id = ml.laize_id
        LEFT JOIN mp_stock_laize sl ON sl.matiere_id = mp.id AND sl.laize_id = l.id
        LEFT JOIN mp_valorisation v ON v.matiere_id = mp.id
        WHERE mp.actif = 1 AND mp.categorie IN ('frontal','glassine','complexe')
        ORDER BY mp.categorie ASC, mp.reference ASC, l.ordre ASC, l.valeur_mm ASC
        """
    ).fetchall()
    # Matières laizées SANS aucune laize associée (à configurer)
    rows_laizees_vides = conn.execute(
        """
        SELECT mp.id AS matiere_id, mp.categorie, mp.reference, mp.designation,
               mp.sous_section,
               COALESCE(mp.metres_lineaires_par_bobine, 0) AS metres,
               COALESCE(mp.prix_eur_m2, 0) AS prix_eur_m2,
               COALESCE(v.prix_en_usd, 0) AS prix_en_usd,
               COALESCE(v.taxe_importation, 0) AS taxe_importation,
               COALESCE(v.cout_transport_inclus, 0) AS cout_transport_inclus
        FROM matieres_premieres mp
        LEFT JOIN mp_valorisation v ON v.matiere_id = mp.id
        WHERE mp.actif = 1 AND mp.categorie IN ('frontal','glassine','complexe')
          AND NOT EXISTS (
            SELECT 1 FROM mp_matiere_laizes ml WHERE ml.matiere_id = mp.id
          )
        ORDER BY mp.categorie ASC, mp.reference ASC
        """
    ).fetchall()
    out: list[dict] = []
    for r in rows_simple:
        cat = r["categorie"]
        qte = float(r["quantite"] or 0)
        prix = float(r["prix_unitaire"] or 0)
        avec_cond = _mp_a_conditionnement(cat)
        upp_raw = r["unites_par_palette"] if "unites_par_palette" in r.keys() else None
        try:
            unites_par_palette = float(upp_raw) if upp_raw is not None else None
        except (TypeError, ValueError):
            unites_par_palette = None
        if avec_cond and unites_par_palette and unites_par_palette > 0:
            # Prix saisi = prix à l'unité d'achat (€/carton, €/tube, €/kg)
            # Prix palette calculé pour l'affichage
            prix_palette = prix * unites_par_palette
            valorisation = round(qte * prix_palette, 2)
            incomplete = False
        elif avec_cond:
            # Cat avec conditionnement mais unites_par_palette pas encore saisi : valorisation neutralisée
            prix_palette = 0.0
            valorisation = 0.0
            incomplete = True
        else:
            # Cat sans conditionnement (palette, autre) : prix_unitaire = prix par unité de gestion
            prix_palette = prix
            valorisation = round(qte * prix, 2)
            incomplete = False
        try:
            transport_state = int(r["cout_transport_inclus"] or 0)
        except (TypeError, ValueError):
            transport_state = 0
        out.append({
            "id": r["id"],
            "row_key": f"m{r['id']}",
            "matiere_id": r["id"],
            "laize_id": None,
            "laize_label": None,
            "categorie": cat,
            "categorie_label": _MP_CATEGORIE_LABELS.get(cat, cat or ""),
            "sous_section": (r["sous_section"] if "sous_section" in r.keys() else None),
            "reference": r["reference"],
            "designation": r["designation"],
            "quantite": qte,
            "unite": _mp_unite_gestion(cat),
            "prix_unitaire": prix,
            "valorisation": valorisation,
            "laizee": False,
            "incomplete": incomplete,
            "metres_lineaires_par_bobine": None,
            "laize_mm": None,
            "surface_bobine_m2": None,
            "prix_eur_m2": None,
            "valorisation_bobine": None,
            "avec_conditionnement": avec_cond,
            "unites_par_palette": unites_par_palette,
            "unite_achat": _mp_unite_achat(cat) if avec_cond else None,
            "prix_palette": round(prix_palette, 4),
            "prix_updated_at": r["prix_updated_at"],
            "prix_updated_by_name": r["prix_updated_by_name"],
            "prix_en_usd": bool(r["prix_en_usd"] or 0),
            "taxe_importation": bool(r["taxe_importation"] or 0),
            "cout_transport_inclus": transport_state,  # 0/1/2 (int)
        })
    for r in rows_laizees:
        cat = r["categorie"]
        qte = float(r["quantite"] or 0)
        metres = float(r["metres"] or 0)
        prix_matiere = float(r["prix_eur_m2"] or 0)
        prix_par_laize = int(r["prix_par_laize"] or 0) == 1
        laize_prix_raw = r["laize_prix_eur_m2"] if "laize_prix_eur_m2" in r.keys() else None
        try:
            prix_laize_val = float(laize_prix_raw) if laize_prix_raw is not None else None
        except (TypeError, ValueError):
            prix_laize_val = None
        prix_m2 = prix_laize_val if (prix_par_laize and prix_laize_val is not None) else prix_matiere
        laize_mm = float(r["valeur_mm"] or 0)
        surface_bobine = (laize_mm / 1000.0) * metres
        valo_bobine = surface_bobine * prix_m2
        try:
            transport_state = int(r["cout_transport_inclus"] or 0)
        except (TypeError, ValueError):
            transport_state = 0
        out.append({
            "id": r["matiere_id"],
            "row_key": f"m{r['matiere_id']}_l{r['laize_id']}",
            "matiere_id": r["matiere_id"],
            "laize_id": r["laize_id"],
            "laize_label": r["laize_label"],
            "categorie": cat,
            "categorie_label": _MP_CATEGORIE_LABELS.get(cat, cat or ""),
            "sous_section": (r["sous_section"] if "sous_section" in r.keys() else None),
            "reference": r["reference"],
            "designation": r["designation"] + (f" — {r['laize_label']}" if r["laize_label"] else ""),
            "quantite": qte,
            "unite": _mp_unite_gestion(cat),
            "prix_unitaire": round(valo_bobine, 4),  # prix unitaire calculé par bobine (lecture seule)
            "valorisation": round(qte * valo_bobine, 2),
            "laizee": True,
            "incomplete": False,
            "metres_lineaires_par_bobine": metres,
            "laize_mm": laize_mm,
            "surface_bobine_m2": round(surface_bobine, 4),
            "prix_eur_m2": prix_m2,
            "valorisation_bobine": round(valo_bobine, 4),
            "avec_conditionnement": False,
            "unites_par_palette": None,
            "unite_achat": None,
            "prix_palette": None,
            "prix_updated_at": None,
            "prix_updated_by_name": None,
            "prix_en_usd": bool(r["prix_en_usd"] or 0),
            "taxe_importation": bool(r["taxe_importation"] or 0),
            "cout_transport_inclus": transport_state,  # 0/1/2 (int, 3 états)
        })
    # Matières laizées sans laizes : affichées avec un placeholder "À configurer"
    for r in rows_laizees_vides:
        cat = r["categorie"]
        try:
            transport_state = int(r["cout_transport_inclus"] or 0)
        except (TypeError, ValueError):
            transport_state = 0
        out.append({
            "id": r["matiere_id"],
            "row_key": f"m{r['matiere_id']}_empty",
            "matiere_id": r["matiere_id"],
            "laize_id": None,
            "laize_label": None,
            "categorie": cat,
            "categorie_label": _MP_CATEGORIE_LABELS.get(cat, cat or ""),
            "sous_section": (r["sous_section"] if "sous_section" in r.keys() else None),
            "reference": r["reference"],
            "designation": r["designation"],
            "quantite": 0.0,
            "unite": _mp_unite_gestion(cat),
            "prix_unitaire": 0.0,
            "valorisation": 0.0,
            "laizee": True,
            "incomplete": True,
            "metres_lineaires_par_bobine": float(r["metres"] or 0),
            "laize_mm": None,
            "surface_bobine_m2": None,
            "prix_eur_m2": float(r["prix_eur_m2"] or 0),
            "valorisation_bobine": 0.0,
            "avec_conditionnement": False,
            "unites_par_palette": None,
            "unite_achat": None,
            "prix_palette": None,
            "prix_updated_at": None,
            "prix_updated_by_name": None,
            "prix_en_usd": bool(r["prix_en_usd"] or 0),
            "taxe_importation": bool(r["taxe_importation"] or 0),
            "cout_transport_inclus": transport_state,  # 0/1/2 (int)
        })
    return out


# Catégories bobines éligibles au supplément transport container (Q3 : frontal + glassine + complexe).
_TRANSPORT_BOBINE_CATEGORIES = frozenset({"frontal", "glassine", "complexe"})


def _valorisation_query_at_date(conn, snapshot_date: str | None) -> list[dict]:
    """Idem _valorisation_query mais avec quantités et prix reconstitués à la date.

    Si snapshot_date est None → délègue tel quel à _valorisation_query.
    Sinon override quantite + prix_unitaire (non-laizées) OU quantite + prix_eur_m2
    (laizées), puis recalcule les champs dérivés (valorisation, prix_palette,
    valorisation_bobine)."""
    items = _valorisation_query(conn)
    if not snapshot_date:
        return items
    stock_nl, stock_l = _mp_stock_snapshot_at_date(conn, snapshot_date)
    prix_map = _mp_price_snapshot_at_date(conn, snapshot_date)
    for it in items:
        mid = it.get("matiere_id")
        try:
            mid_int = int(mid) if mid is not None else None
        except (TypeError, ValueError):
            mid_int = None
        # Override quantité (snapshot) — laizée vs non-laizée
        if it.get("laizee") and it.get("laize_id") is not None:
            try:
                lid_int = int(it["laize_id"])
            except (TypeError, ValueError):
                lid_int = None
            k = (mid_int, lid_int) if (mid_int is not None and lid_int is not None) else None
            new_qte = float(stock_l.get(k, it.get("quantite") or 0)) if k else float(it.get("quantite") or 0)
        elif mid_int is not None:
            new_qte = float(stock_nl.get(mid_int, it.get("quantite") or 0))
        else:
            new_qte = float(it.get("quantite") or 0)
        it["quantite"] = round(new_qte, 4)
        # Override prix (snapshot)
        if mid_int is not None and mid_int in prix_map:
            historique_prix = float(prix_map[mid_int])
        else:
            historique_prix = None
        if it.get("laizee"):
            # Laizées : le prix historisé est prix_eur_m2. Recalcule valorisation_bobine et valorisation.
            if historique_prix is not None:
                it["prix_eur_m2"] = round(historique_prix, 4)
            surface = float(it.get("surface_bobine_m2") or 0)
            prix_m2 = float(it.get("prix_eur_m2") or 0)
            valo_bobine = surface * prix_m2
            it["valorisation_bobine"] = round(valo_bobine, 4)
            it["prix_unitaire"] = round(valo_bobine, 4)
            it["valorisation"] = round(new_qte * valo_bobine, 2)
        else:
            # Non-laizées : le prix historisé est prix_unitaire.
            if historique_prix is not None:
                it["prix_unitaire"] = round(historique_prix, 4)
            prix = float(it.get("prix_unitaire") or 0)
            avec_cond = bool(it.get("avec_conditionnement"))
            upp = it.get("unites_par_palette")
            if avec_cond and upp and float(upp) > 0:
                prix_palette = prix * float(upp)
                it["prix_palette"] = round(prix_palette, 4)
                it["valorisation"] = round(new_qte * prix_palette, 2)
                it["incomplete"] = False
            elif avec_cond:
                it["prix_palette"] = 0.0
                it["valorisation"] = 0.0
                it["incomplete"] = True
            else:
                it["prix_palette"] = round(prix, 4)
                it["valorisation"] = round(new_qte * prix, 2)
                it["incomplete"] = False
    return items


def _row_tax_multiplier(item: dict, import_tax_pct: float) -> float:
    """Multiplicateur Taxe d'importation seul (1 + taxe%/100).

    Restreint aux adhésifs (Q2) : ignoré pour toute autre catégorie même si le flag
    taxe_importation est actif en base (rétrocompat sans breakage des données)."""
    cat = str(item.get("categorie") or "").lower()
    if cat != "adhesif":
        return 1.0
    if item.get("taxe_importation") and import_tax_pct > 0:
        return 1.0 + (import_tax_pct / 100.0)
    return 1.0


def _row_usd_multiplier(item: dict, taux_eur_usd: float) -> float:
    """Multiplicateur USD seul (taux EUR/USD). Neutre si le flag est off."""
    if item.get("prix_en_usd") and taux_eur_usd > 0:
        return float(taux_eur_usd)
    return 1.0


def _row_multiplier(item: dict, taux_eur_usd: float, import_tax_pct: float) -> float:
    """Multiplicateur composé hors transport : taxe puis USD.
    Formule : ((prix × (1+taxe%)) × taux_USD). Le supplément transport est additif
    en €/m², appliqué séparément — voir _row_transport_supplement_eur_per_m2."""
    return _row_tax_multiplier(item, import_tax_pct) * _row_usd_multiplier(item, taux_eur_usd)


def _row_transport_supplement_eur_per_m2(
    item: dict,
    container_cost_full_eur: float,
    container_cost_half_eur: float,
    qte_m2_full: float,
    qte_m2_half: float,
) -> float:
    """Supplément transport en €/m² selon l'état du flag cout_transport_inclus.

    - État 0 : off → 0
    - État 1 : container complet → container_cost_full_eur / qte_m2_full
    - État 2 : demi-container    → container_cost_half_eur / qte_m2_half

    Restreint aux catégories bobines (frontal, glassine, complexe). Retourne 0
    pour toute autre catégorie même si le flag est actif en base."""
    cat = str(item.get("categorie") or "").lower()
    if cat not in _TRANSPORT_BOBINE_CATEGORIES:
        return 0.0
    try:
        state = int(item.get("cout_transport_inclus") or 0)
    except (TypeError, ValueError):
        state = 0
    if state == 1 and qte_m2_full > 0 and container_cost_full_eur > 0:
        return float(container_cost_full_eur) / float(qte_m2_full)
    if state == 2 and qte_m2_half > 0 and container_cost_half_eur > 0:
        return float(container_cost_half_eur) / float(qte_m2_half)
    return 0.0


def _row_surface_bobine_m2(item: dict) -> float:
    """Surface m² d'une bobine — dérivée de laize_mm × metres_lineaires_par_bobine.
    Retourne 0 si la matière n'est pas laizée ou si laize/métrage non renseigné."""
    if not item.get("laizee"):
        return 0.0
    metres = float(item.get("metres_lineaires_par_bobine") or 0)
    laize_mm = float(item.get("laize_mm") or 0)
    if metres <= 0 or laize_mm <= 0:
        return 0.0
    return (laize_mm / 1000.0) * metres


def _valorisation_summary(
    items: list[dict],
    taux_eur_usd: float = 0.0,
    import_tax_pct: float = 0.0,
    container_cost_full_eur: float = 0.0,
    container_cost_half_eur: float = 0.0,
    qte_m2_full: float = 0.0,
    qte_m2_half: float = 0.0,
) -> dict:
    """Calcule les totaux. Formule par ligne bobine :
        prix_m2_reel = prix_m2 × (1 + taxe%/100) × taux_USD + supplement_transport_m²
        val_reel     = quantite × prix_m2_reel × surface_bobine_m²

    - Taxe : uniquement adhésifs (Q2).
    - Transport container/demi-container : uniquement bobines (Q3).
    - Le forfait transport historique n'existe plus (Q1).
    """
    totals_by_cat: dict[str, dict] = {}
    total = 0.0
    total_reel = 0.0
    seen_usd: set[int] = set()
    seen_tax: set[int] = set()
    seen_both: set[int] = set()
    seen_transport: set[int] = set()
    for it in items:
        cat = it["categorie"]
        if cat not in totals_by_cat:
            totals_by_cat[cat] = {
                "categorie": cat,
                "categorie_label": it["categorie_label"],
                "total": 0.0,
                "nb_refs": 0,
                "nb_refs_valorisees": 0,
            }
        bucket = totals_by_cat[cat]
        bucket["total"] += it["valorisation"]
        bucket["nb_refs"] += 1
        is_valued = (it.get("prix_unitaire") or 0) > 0
        if it.get("avec_conditionnement"):
            is_valued = is_valued and (it.get("unites_par_palette") or 0) > 0
        if is_valued:
            bucket["nb_refs_valorisees"] += 1
        total += it["valorisation"]

        tax_mult = _row_tax_multiplier(it, import_tax_pct)
        usd_mult = _row_usd_multiplier(it, taux_eur_usd)
        supp_per_m2 = _row_transport_supplement_eur_per_m2(
            it, container_cost_full_eur, container_cost_half_eur, qte_m2_full, qte_m2_half
        )
        mid = it.get("matiere_id")
        is_usd = bool(it.get("prix_en_usd")) and taux_eur_usd > 0
        is_tax = bool(it.get("taxe_importation")) and import_tax_pct > 0 and str(it.get("categorie") or "").lower() == "adhesif"
        is_transport = supp_per_m2 > 0

        # val_reel : (val × tax × usd) + supp_per_m² × surface_bobine × qte
        val_reel_base = it["valorisation"] * tax_mult * usd_mult
        qte = float(it.get("quantite") or 0)
        surface = _row_surface_bobine_m2(it)
        val_reel = val_reel_base + supp_per_m2 * surface * qte
        total_reel += val_reel

        if isinstance(mid, int):
            if is_transport and mid not in seen_transport:
                seen_transport.add(mid)
            if is_usd and is_tax and mid not in seen_both:
                seen_both.add(mid)
            elif is_usd and not is_tax and mid not in seen_usd:
                seen_usd.add(mid)
            elif is_tax and not is_usd and mid not in seen_tax:
                seen_tax.add(mid)
    cats = sorted(totals_by_cat.values(), key=lambda c: _mp_categorie_order(c["categorie"]))
    for c in cats:
        c["total"] = round(c["total"], 2)
    nb_usd_only = len(seen_usd)
    nb_tax_only = len(seen_tax)
    nb_both = len(seen_both)
    nb_transport = len(seen_transport)
    # €/m² supp pour indication UI (fonction du paramétrage)
    supp_per_m2_full = (container_cost_full_eur / qte_m2_full) if (qte_m2_full > 0 and container_cost_full_eur > 0) else 0.0
    supp_per_m2_half = (container_cost_half_eur / qte_m2_half) if (qte_m2_half > 0 and container_cost_half_eur > 0) else 0.0
    return {
        "total_mp": round(total, 2),
        "total_mp_reel": round(total_reel, 2),
        "total_pf": 0.0,
        "total_global": round(total, 2),
        "categories": cats,
        "nb_refs": sum(c["nb_refs"] for c in cats),
        "nb_refs_valorisees": sum(c["nb_refs_valorisees"] for c in cats),
        "nb_refs_usd": nb_usd_only + nb_both,
        "nb_refs_tax": nb_tax_only + nb_both,
        "nb_refs_usd_only": nb_usd_only,
        "nb_refs_tax_only": nb_tax_only,
        "nb_refs_usd_and_tax": nb_both,
        "nb_refs_transport": nb_transport,
        "taux_eur_usd": round(taux_eur_usd, 6) if taux_eur_usd > 0 else 0,
        "import_tax_pct": round(import_tax_pct, 4) if import_tax_pct > 0 else 0,
        # Nouveau : supplément container en €/m² (plein et demi) pour affichage
        "container_supplement_eur_per_m2_full": round(supp_per_m2_full, 4) if supp_per_m2_full > 0 else 0,
        "container_supplement_eur_per_m2_half": round(supp_per_m2_half, 4) if supp_per_m2_half > 0 else 0,
        "container_cost_full_eur": round(container_cost_full_eur, 2) if container_cost_full_eur > 0 else 0,
        "container_cost_half_eur": round(container_cost_half_eur, 2) if container_cost_half_eur > 0 else 0,
        "qte_m2_container_complet": round(qte_m2_full, 2) if qte_m2_full > 0 else 0,
        "qte_m2_demi_container": round(qte_m2_half, 2) if qte_m2_half > 0 else 0,
    }


def _enrich_items_with_usd(
    items: list[dict],
    taux_eur_usd: float,
    import_tax_pct: float = 0.0,
    container_cost_full_eur: float = 0.0,
    container_cost_half_eur: float = 0.0,
    qte_m2_full: float = 0.0,
    qte_m2_half: float = 0.0,
) -> None:
    """Ajoute prix_unitaire_reel / valorisation_reelle / prix_eur_m2_reel /
    transport_addon_eur in-place pour chaque item.

    Formule par ligne bobine (frontal/glassine/complexe) :
        prix_m2_reel = prix_m2 × (1 + taxe%/100) × taux_USD + supp_transport_m²
        prix_bobine_reel = prix_m2_reel × surface_bobine_m²
        val_reel        = quantite × prix_bobine_reel

    Formule pour les autres catégories :
        prix_reel = prix × (1 + taxe%/100) × taux_USD  (taxe uniquement si adhesif)
        val_reel  = val × (1 + taxe%/100) × taux_USD
    """
    for it in items:
        tax_mult = _row_tax_multiplier(it, import_tax_pct)
        usd_mult = _row_usd_multiplier(it, taux_eur_usd)
        mult = tax_mult * usd_mult
        supp_per_m2 = _row_transport_supplement_eur_per_m2(
            it, container_cost_full_eur, container_cost_half_eur, qte_m2_full, qte_m2_half
        )
        surface = _row_surface_bobine_m2(it)
        supp_per_bobine = supp_per_m2 * surface

        # Prix €/m² réel (bobines uniquement — sinon prix_eur_m2 est None)
        base_prix_m2 = float(it.get("prix_eur_m2") or 0)
        if it.get("prix_eur_m2") is not None:
            prix_m2_reel = base_prix_m2 * mult + supp_per_m2
            it["prix_eur_m2_reel"] = round(prix_m2_reel, 4)
        else:
            it["prix_eur_m2_reel"] = None

        # Prix unitaire (par bobine ou par unité de gestion) réel
        base_prix_unit = float(it.get("prix_unitaire") or 0)
        prix_unitaire_reel = base_prix_unit * mult + supp_per_bobine
        # Si aucun multiplicateur ni supplément → recopie tel quel (pour cohérence historique)
        if mult == 1.0 and supp_per_bobine == 0.0:
            it["prix_unitaire_reel"] = base_prix_unit
        else:
            it["prix_unitaire_reel"] = round(prix_unitaire_reel, 4)

        # Valorisation réelle : val_base × mult + supplement_bobine × qté
        qte = float(it.get("quantite") or 0)
        val_reel = float(it.get("valorisation") or 0) * mult + supp_per_bobine * qte
        it["valorisation_reelle"] = round(val_reel, 2)
        it["reel_multiplier"] = round(mult, 6)
        # Total supplément transport pour cette ligne (= supp × surface × quantité, en €)
        it["transport_addon_eur"] = round(supp_per_bobine * qte, 2)
        it["transport_addon_raw"] = round(supp_per_bobine, 4)
        it["transport_supplement_eur_per_m2"] = round(supp_per_m2, 4)


@router.get("/api/stock/valorisation")
def get_valorisation(request: Request, date: str | None = None):
    """Valorisation MP. Query optionnel `date=YYYY-MM-DD` → figée à cette date
    (quantités reconstituées via l'historique des mouvements + prix historiques).
    Paramètres globaux (taux USD, taxe, containers) restent actuels."""
    user = require_stock_matieres_admin(request)
    can_see_usd = _user_can_see_valorisation_usd(user)
    snapshot_date = _parse_snapshot_date(date)
    with get_db() as conn:
        items = _valorisation_query_at_date(conn, snapshot_date)
        taux = _get_taux_eur_usd(conn) if can_see_usd else 0.0
        tax_pct = _get_import_tax_pct(conn) if can_see_usd else 0.0
        c_full, c_half, q_full, q_half = _get_container_params(conn) if can_see_usd else (0.0, 0.0, 0.0, 0.0)
    _enrich_items_with_usd(items, taux, tax_pct, c_full, c_half, q_full, q_half)
    summary = _valorisation_summary(items, taux, tax_pct, c_full, c_half, q_full, q_half)
    summary["can_see_usd"] = can_see_usd
    summary["snapshot_date"] = snapshot_date  # None si aujourd'hui
    return {
        "items": items,
        "summary": summary,
    }


@router.put("/api/stock/valorisation/{matiere_id}")
async def update_valorisation(matiere_id: int, request: Request):
    user = require_stock_matieres_admin(request)
    body = await request.json()
    now = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    user_name = (user.get("nom") or user.get("email") or "").strip() or None
    user_id = user.get("id")
    with get_db() as conn:
        mat = conn.execute(
            "SELECT id, categorie, prix_eur_m2, metres_lineaires_par_bobine FROM matieres_premieres WHERE id=?",
            (matiere_id,),
        ).fetchone()
        if not mat:
            raise HTTPException(404, "Matière introuvable.")
        cat = (mat["categorie"] or "").lower()
        if _mp_is_laizee(cat):
            # Matière laizée → on met à jour prix_eur_m2 et/ou metres_lineaires_par_bobine
            sets: list[str] = []
            args: list[Any] = []
            note_parts: list[str] = []
            prev_prix_m2 = float(mat["prix_eur_m2"] or 0) if mat["prix_eur_m2"] is not None else None
            prev_metres = float(mat["metres_lineaires_par_bobine"] or 0) if mat["metres_lineaires_par_bobine"] is not None else None
            new_prix_m2 = prev_prix_m2
            new_metres = prev_metres
            if "prix_eur_m2" in body:
                v = body.get("prix_eur_m2")
                try:
                    v = float(v) if v not in (None, "") else 0.0
                except (TypeError, ValueError):
                    raise HTTPException(400, "Prix m² invalide.") from None
                if v < 0:
                    raise HTTPException(400, "Prix m² négatif.")
                if v > 1_000_000:
                    raise HTTPException(400, "Prix m² hors limites.")
                sets.append("prix_eur_m2=?"); args.append(v)
                new_prix_m2 = v
                if (prev_prix_m2 or 0) != v:
                    note_parts.append(
                        f"Prix m²: {prev_prix_m2 if prev_prix_m2 is not None else '—'} → {v}"
                    )
            if "metres_lineaires_par_bobine" in body:
                v = body.get("metres_lineaires_par_bobine")
                try:
                    v = float(v) if v not in (None, "") else 0.0
                except (TypeError, ValueError):
                    raise HTTPException(400, "Métrage invalide.") from None
                if v < 0:
                    raise HTTPException(400, "Métrage négatif.")
                sets.append("metres_lineaires_par_bobine=?"); args.append(v)
                new_metres = v
                if (prev_metres or 0) != v:
                    note_parts.append(
                        f"Métrage: {prev_metres if prev_metres is not None else '—'} → {v}"
                    )
            if not sets:
                raise HTTPException(400, "Aucun champ à mettre à jour.")
            args.append(matiere_id)
            conn.execute(
                f"UPDATE matieres_premieres SET {', '.join(sets)} WHERE id=?", args
            )
            # Historique : log si le prix_eur_m2 a changé
            if (prev_prix_m2 or 0) != (new_prix_m2 or 0) or (prev_metres or 0) != (new_metres or 0):
                user_note = (body.get("note") or "").strip()
                note_full = " · ".join(note_parts) + (f" — {user_note}" if user_note else "")
                conn.execute(
                    """INSERT INTO mp_valorisation_historique
                       (matiere_id, prix_avant, prix_apres, note, created_at, created_by, created_by_name)
                       VALUES (?,?,?,?,?,?,?)""",
                    (matiere_id, prev_prix_m2, new_prix_m2, note_full, now, user_id, user_name),
                )
            conn.commit()
        else:
            # Matière non laizée → comportement classique avec prix_unitaire
            try:
                prix = float(body.get("prix_unitaire") or 0)
            except (TypeError, ValueError):
                raise HTTPException(400, "Prix unitaire invalide.") from None
            if prix < 0:
                raise HTTPException(400, "Prix unitaire négatif interdit.")
            if prix > 1_000_000:
                raise HTTPException(400, "Prix unitaire hors limites (max 1 000 000).")
            note = (body.get("note") or "").strip() or None
            prev = conn.execute(
                "SELECT prix_unitaire FROM mp_valorisation WHERE matiere_id=?", (matiere_id,)
            ).fetchone()
            prix_avant = float(prev["prix_unitaire"]) if prev else None
            if prev:
                conn.execute(
                    """UPDATE mp_valorisation
                       SET prix_unitaire=?, updated_at=?, updated_by_name=?
                       WHERE matiere_id=?""",
                    (prix, now, user_name, matiere_id),
                )
            else:
                conn.execute(
                    """INSERT INTO mp_valorisation (matiere_id, prix_unitaire, updated_at, updated_by_name)
                       VALUES (?,?,?,?)""",
                    (matiere_id, prix, now, user_name),
                )
            if prix_avant != prix:
                conn.execute(
                    """INSERT INTO mp_valorisation_historique
                       (matiere_id, prix_avant, prix_apres, note, created_at, created_by, created_by_name)
                       VALUES (?,?,?,?,?,?,?)""",
                    (matiere_id, prix_avant, prix, note, now, user_id, user_name),
                )
            conn.commit()
        items = _valorisation_query(conn)
        can_see_usd = _user_can_see_valorisation_usd(user)
        taux = _get_taux_eur_usd(conn) if can_see_usd else 0.0
        tax_pct = _get_import_tax_pct(conn) if can_see_usd else 0.0
        c_full, c_half, q_full, q_half = _get_container_params(conn) if can_see_usd else (0.0, 0.0, 0.0, 0.0)
    _enrich_items_with_usd(items, taux, tax_pct, c_full, c_half, q_full, q_half)
    summary = _valorisation_summary(items, taux, tax_pct, c_full, c_half, q_full, q_half)
    summary["can_see_usd"] = can_see_usd
    # Pour les matières laizées, plusieurs lignes ont le même matiere_id → on renvoie toutes
    matching = [x for x in items if x["matiere_id"] == matiere_id]
    return {
        "ok": True,
        "item": matching[0] if matching else None,
        "items_matiere": matching,
        "summary": summary,
    }


@router.put("/api/stock/valorisation/{matiere_id}/cout-transport")
async def toggle_valorisation_cout_transport(matiere_id: int, request: Request):
    """Cycle le flag cout_transport_inclus pour une matière (3 états).

    - 0 : off
    - 1 : container complet (supplément = coût container / m² × surface_bobine)
    - 2 : demi-container    (supplément = coût demi-container / m² × surface_bobine)

    Réservé Direction / superadmin. Sans body → cycle 0→1→2→0. Avec body
    `{"cout_transport_inclus": N}` (N ∈ {0,1,2}) → fixe la valeur explicitement.
    Une valeur booléenne True est mappée à 1 (rétrocompat clients v1)."""
    user = require_valorisation_usd_admin(request)
    body = await request.json() if request.headers.get("content-length") else {}
    requested = body.get("cout_transport_inclus") if isinstance(body, dict) else None
    user_name = (user.get("nom") or user.get("email") or "").strip() or None
    now = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    with get_db() as conn:
        mat = conn.execute(
            "SELECT id FROM matieres_premieres WHERE id=?", (matiere_id,)
        ).fetchone()
        if not mat:
            raise HTTPException(404, "Matière introuvable.")
        prev = conn.execute(
            "SELECT cout_transport_inclus FROM mp_valorisation WHERE matiere_id=?",
            (matiere_id,),
        ).fetchone()
        # Détermine la nouvelle valeur (0/1/2)
        if requested is None:
            try:
                cur = int(prev["cout_transport_inclus"] or 0) if prev else 0
            except (TypeError, ValueError):
                cur = 0
            new_val = (cur + 1) % 3
        elif isinstance(requested, bool):
            new_val = 1 if requested else 0
        else:
            try:
                new_val = int(requested)
            except (TypeError, ValueError):
                new_val = 0
            if new_val not in (0, 1, 2):
                raise HTTPException(400, "cout_transport_inclus doit être 0, 1 ou 2.")
        if prev:
            conn.execute(
                "UPDATE mp_valorisation SET cout_transport_inclus=?, updated_at=?, updated_by_name=? WHERE matiere_id=?",
                (new_val, now, user_name, matiere_id),
            )
        else:
            conn.execute(
                """INSERT INTO mp_valorisation (matiere_id, prix_unitaire, cout_transport_inclus, updated_at, updated_by_name)
                   VALUES (?,0,?,?,?)""",
                (matiere_id, new_val, now, user_name),
            )
        conn.commit()
        items = _valorisation_query(conn)
        taux = _get_taux_eur_usd(conn)
        tax_pct = _get_import_tax_pct(conn)
        c_full, c_half, q_full, q_half = _get_container_params(conn)
    _enrich_items_with_usd(items, taux, tax_pct, c_full, c_half, q_full, q_half)
    summary = _valorisation_summary(items, taux, tax_pct, c_full, c_half, q_full, q_half)
    summary["can_see_usd"] = True
    matching = [x for x in items if x["matiere_id"] == matiere_id]
    return {
        "ok": True,
        "cout_transport_inclus": new_val,
        "item": matching[0] if matching else None,
        "items_matiere": matching,
        "summary": summary,
    }


@router.put("/api/stock/valorisation/{matiere_id}/taxe-importation")
async def toggle_valorisation_taxe_importation(matiere_id: int, request: Request):
    """Bascule le flag taxe_importation pour une matière. Réservé Direction / superadmin.
    Crée la ligne mp_valorisation si nécessaire."""
    user = require_valorisation_usd_admin(request)
    body = await request.json() if request.headers.get("content-length") else {}
    requested = body.get("taxe_importation") if isinstance(body, dict) else None
    user_name = (user.get("nom") or user.get("email") or "").strip() or None
    now = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    with get_db() as conn:
        mat = conn.execute(
            "SELECT id FROM matieres_premieres WHERE id=?", (matiere_id,)
        ).fetchone()
        if not mat:
            raise HTTPException(404, "Matière introuvable.")
        prev = conn.execute(
            "SELECT taxe_importation FROM mp_valorisation WHERE matiere_id=?",
            (matiere_id,),
        ).fetchone()
        if requested is None:
            current = bool(prev["taxe_importation"]) if prev else False
            new_val = not current
        else:
            new_val = bool(requested)
        if prev:
            conn.execute(
                "UPDATE mp_valorisation SET taxe_importation=?, updated_at=?, updated_by_name=? WHERE matiere_id=?",
                (1 if new_val else 0, now, user_name, matiere_id),
            )
        else:
            conn.execute(
                """INSERT INTO mp_valorisation (matiere_id, prix_unitaire, taxe_importation, updated_at, updated_by_name)
                   VALUES (?,0,?,?,?)""",
                (matiere_id, 1 if new_val else 0, now, user_name),
            )
        conn.commit()
        items = _valorisation_query(conn)
        taux = _get_taux_eur_usd(conn)
        tax_pct = _get_import_tax_pct(conn)
        c_full, c_half, q_full, q_half = _get_container_params(conn)
    _enrich_items_with_usd(items, taux, tax_pct, c_full, c_half, q_full, q_half)
    summary = _valorisation_summary(items, taux, tax_pct, c_full, c_half, q_full, q_half)
    summary["can_see_usd"] = True
    matching = [x for x in items if x["matiere_id"] == matiere_id]
    return {
        "ok": True,
        "taxe_importation": new_val,
        "item": matching[0] if matching else None,
        "items_matiere": matching,
        "summary": summary,
    }


@router.put("/api/stock/valorisation/{matiere_id}/prix-en-usd")
async def toggle_valorisation_prix_en_usd(matiere_id: int, request: Request):
    """Bascule le flag prix_en_usd pour une matière. Réservé Direction / superadmin.
    Crée la ligne mp_valorisation si elle n'existe pas (utile pour les matières laizées
    qui n'ont pas forcément de prix_unitaire stocké)."""
    user = require_valorisation_usd_admin(request)
    body = await request.json() if request.headers.get("content-length") else {}
    requested = body.get("prix_en_usd") if isinstance(body, dict) else None
    user_name = (user.get("nom") or user.get("email") or "").strip() or None
    now = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    with get_db() as conn:
        mat = conn.execute(
            "SELECT id FROM matieres_premieres WHERE id=?", (matiere_id,)
        ).fetchone()
        if not mat:
            raise HTTPException(404, "Matière introuvable.")
        prev = conn.execute(
            "SELECT prix_unitaire, prix_en_usd FROM mp_valorisation WHERE matiere_id=?",
            (matiere_id,),
        ).fetchone()
        if requested is None:
            # Toggle si pas de valeur explicite
            current = bool(prev["prix_en_usd"]) if prev else False
            new_val = not current
        else:
            new_val = bool(requested)
        if prev:
            conn.execute(
                "UPDATE mp_valorisation SET prix_en_usd=?, updated_at=?, updated_by_name=? WHERE matiere_id=?",
                (1 if new_val else 0, now, user_name, matiere_id),
            )
        else:
            conn.execute(
                """INSERT INTO mp_valorisation (matiere_id, prix_unitaire, prix_en_usd, updated_at, updated_by_name)
                   VALUES (?,0,?,?,?)""",
                (matiere_id, 1 if new_val else 0, now, user_name),
            )
        conn.commit()
        items = _valorisation_query(conn)
        taux = _get_taux_eur_usd(conn)
        tax_pct = _get_import_tax_pct(conn)
        c_full, c_half, q_full, q_half = _get_container_params(conn)
    _enrich_items_with_usd(items, taux, tax_pct, c_full, c_half, q_full, q_half)
    summary = _valorisation_summary(items, taux, tax_pct, c_full, c_half, q_full, q_half)
    summary["can_see_usd"] = True
    matching = [x for x in items if x["matiere_id"] == matiere_id]
    return {
        "ok": True,
        "prix_en_usd": new_val,
        "item": matching[0] if matching else None,
        "items_matiere": matching,
        "summary": summary,
    }


@router.get("/api/stock/valorisation/{matiere_id}/historique")
def get_valorisation_historique(matiere_id: int, request: Request):
    require_stock_matieres_admin(request)
    with get_db() as conn:
        mat = conn.execute(
            """SELECT mp.id, mp.categorie, mp.reference, mp.designation
               FROM matieres_premieres mp WHERE mp.id=?""",
            (matiere_id,),
        ).fetchone()
        if not mat:
            raise HTTPException(404, "Matière introuvable.")
        rows = conn.execute(
            """SELECT id, prix_avant, prix_apres, note, created_at, created_by_name
               FROM mp_valorisation_historique
               WHERE matiere_id=?
               ORDER BY datetime(created_at) DESC, id DESC""",
            (matiere_id,),
        ).fetchall()
    return {
        "matiere": {
            "id": mat["id"],
            "categorie": mat["categorie"],
            "categorie_label": _MP_CATEGORIE_LABELS.get(mat["categorie"], mat["categorie"] or ""),
            "reference": mat["reference"],
            "designation": mat["designation"],
        },
        "historique": [
            {
                "id": r["id"],
                "prix_avant": float(r["prix_avant"]) if r["prix_avant"] is not None else None,
                "prix_apres": float(r["prix_apres"]) if r["prix_apres"] is not None else 0.0,
                "note": r["note"],
                "created_at": r["created_at"],
                "created_by_name": r["created_by_name"],
            }
            for r in rows
        ],
    }


@router.get("/api/stock/valorisation/export")
def export_valorisation(request: Request):
    require_stock_matieres_admin(request)
    with get_db() as conn:
        items = _valorisation_query(conn)
    summary = _valorisation_summary(items)
    # Construction d'un classeur Excel avec une feuille de détail + une feuille récap
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise HTTPException(500, f"openpyxl indisponible : {e}") from None
    wb = Workbook()
    ws = wb.active
    ws.title = "Matières premières"
    head_fill = PatternFill("solid", fgColor="1E293B")
    head_font = Font(bold=True, color="FFFFFF")
    total_font = Font(bold=True)
    headers = [
        "Catégorie", "Référence", "Désignation", "Laize",
        "Quantité", "Unité",
        "Prix m² (€)", "Métrage (m)",
        "Prix unitaire (€)", "Valorisation (€)",
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    grand_total = 0.0
    cur_cat: Optional[str] = None
    cat_total = 0.0
    cat_count = 0
    n_cols = len(headers)

    def _write_cat_total(label: str, tot: float) -> None:
        row = [f"Sous-total {label}"] + [""] * (n_cols - 2) + [round(tot, 2)]
        ws.append(row)
        r = ws.max_row
        for col_idx in range(1, n_cols + 1):
            ws.cell(row=r, column=col_idx).font = total_font
        ws.cell(row=r, column=n_cols).number_format = "#,##0.00 €"

    for it in items:
        if cur_cat is not None and it["categorie"] != cur_cat and cat_count > 0:
            _write_cat_total(_MP_CATEGORIE_LABELS.get(cur_cat, cur_cat), cat_total)
            cat_total = 0.0
            cat_count = 0
        cur_cat = it["categorie"]
        ws.append([
            it["categorie_label"],
            it["reference"],
            it["designation"],
            it.get("laize_label") or "",
            it["quantite"],
            it["unite"],
            it.get("prix_eur_m2") if it.get("laizee") else "",
            it.get("metres_lineaires_par_bobine") if it.get("laizee") else "",
            it["prix_unitaire"],
            it["valorisation"],
        ])
        r = ws.max_row
        if it.get("laizee"):
            ws.cell(row=r, column=7).number_format = "#,##0.0000 €"
            ws.cell(row=r, column=8).number_format = "#,##0.0 m"
        ws.cell(row=r, column=9).number_format = "#,##0.0000 €"
        ws.cell(row=r, column=10).number_format = "#,##0.00 €"
        cat_total += it["valorisation"]
        cat_count += 1
        grand_total += it["valorisation"]
    if cur_cat is not None and cat_count > 0:
        _write_cat_total(_MP_CATEGORIE_LABELS.get(cur_cat, cur_cat), cat_total)
    # Ligne grand total
    ws.append([])
    ws.append(["TOTAL MATIÈRES PREMIÈRES"] + [""] * (n_cols - 2) + [round(grand_total, 2)])
    r = ws.max_row
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=r, column=col_idx)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill("solid", fgColor="DBEAFE")
    ws.cell(row=r, column=n_cols).number_format = "#,##0.00 €"
    # Largeurs colonnes
    widths = [14, 22, 40, 10, 12, 10, 14, 14, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    # Feuille récap par catégorie
    ws2 = wb.create_sheet("Récapitulatif")
    ws2.append(["Catégorie", "Nb références", "Nb références valorisées", "Total (€)"])
    for col_idx in range(1, 5):
        c = ws2.cell(row=1, column=col_idx)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    for c in summary["categories"]:
        ws2.append([c["categorie_label"], c["nb_refs"], c["nb_refs_valorisees"], c["total"]])
        ws2.cell(row=ws2.max_row, column=4).number_format = "#,##0.00 €"
    ws2.append([])
    ws2.append(["TOTAL", summary["nb_refs"], summary["nb_refs_valorisees"], summary["total_mp"]])
    r = ws2.max_row
    for col_idx in range(1, 5):
        cell = ws2.cell(row=r, column=col_idx)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill("solid", fgColor="DBEAFE")
    ws2.cell(row=r, column=4).number_format = "#,##0.00 €"
    for i, w in enumerate([20, 16, 24, 18], start=1):
        ws2.column_dimensions[chr(64 + i)].width = w
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"valorisation-stock-{date_str}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ═════════════════════════════════════════════════════════════════════════
# Valorisation Produits Finis (PF + Négoce) — Contrôle → Valorisation
# Source d'import : feuille "Valorisation" de Lignes_de_facturation.xlsx
# ═════════════════════════════════════════════════════════════════════════


def _pf_valo_normalize_ref(ref: str) -> str:
    return (ref or "").strip().upper()


def _pf_valo_pick_stock(conn: sqlite3.Connection, produit_type: str) -> dict[int, float]:
    """Quantité totale par produit_id (somme lots_stock + fallback stock_emplacements)."""
    out: dict[int, float] = {}
    rows = conn.execute(
        """
        SELECT p.id AS produit_id, COALESCE(SUM(l.quantite_restante),0) AS q
        FROM produits p
        LEFT JOIN lots_stock l ON l.produit_id = p.id AND l.quantite_restante > 0.0001
        WHERE COALESCE(p.type, 'fabrique') = ?
        GROUP BY p.id
        """,
        (produit_type,),
    ).fetchall()
    for r in rows:
        out[int(r["produit_id"])] = float(r["q"] or 0)
    # Compléter avec stock_emplacements pour les refs sans lots (legacy)
    rows2 = conn.execute(
        """
        SELECT p.id AS produit_id, COALESCE(SUM(s.quantite),0) AS q
        FROM produits p
        JOIN stock_emplacements s ON s.produit_id = p.id AND s.quantite > 0.0001
        WHERE COALESCE(p.type, 'fabrique') = ?
        GROUP BY p.id
        """,
        (produit_type,),
    ).fetchall()
    for r in rows2:
        pid = int(r["produit_id"])
        if out.get(pid, 0) <= 0:
            out[pid] = float(r["q"] or 0)
    return out


_PF_UNITE_PER_MILLE = frozenset({"etiquette", "etiquettes"})


def _pf_is_unite_par_mille(unite: str) -> bool:
    """Vrai si l'unité est facturée au mille (le prix HT stocké est /1000 unités)."""
    u = (unite or "").strip().lower()
    # Normalise les accents pour matcher "étiquette" / "étiquettes"
    u = u.replace("é", "e").replace("è", "e").replace("ê", "e")
    return u in _PF_UNITE_PER_MILLE


def _pf_valo_query(conn: sqlite3.Connection) -> list[dict]:
    """Liste de toutes les références (fabrique + négoce) avec leur prix HT courant + stock.

    Pour les produits facturés au mille (unité = étiquette[s]), le prix HT stocké
    correspond au prix pour 1000 unités. La valorisation = qte * prix / 1000.
    """
    rows = conn.execute(
        """
        SELECT p.id, p.reference, p.designation, p.unite,
               COALESCE(p.type, 'fabrique') AS type,
               v.prix_unitaire_ht, v.source_prix, v.statut, v.date_derniere_cmd,
               v.updated_at, v.updated_by_name
        FROM produits p
        LEFT JOIN pf_valorisation v ON v.produit_id = p.id
        ORDER BY p.reference ASC
        """
    ).fetchall()
    stock_fab = _pf_valo_pick_stock(conn, "fabrique")
    stock_neg = _pf_valo_pick_stock(conn, "negoce")
    out: list[dict] = []
    for r in rows:
        ptype = r["type"]
        pid = int(r["id"])
        stock = stock_fab.get(pid, 0.0) if ptype == "fabrique" else stock_neg.get(pid, 0.0)
        prix = float(r["prix_unitaire_ht"] or 0) if r["prix_unitaire_ht"] is not None else 0.0
        unite = r["unite"] or ""
        par_mille = _pf_is_unite_par_mille(unite)
        divisor = 1000.0 if par_mille else 1.0
        valorisation = round(stock * prix / divisor, 2)
        out.append({
            "id": pid,
            "reference": r["reference"],
            "designation": r["designation"],
            "unite": unite,
            "par_mille": par_mille,
            "type": ptype,  # 'fabrique' ou 'negoce'
            "type_label": "Négoce" if ptype == "negoce" else "Fabriqué",
            "quantite": round(stock, 4),
            "prix_unitaire_ht": prix,
            "valorisation": valorisation,
            "source_prix": r["source_prix"],
            "statut": r["statut"],
            "date_derniere_cmd": r["date_derniere_cmd"],
            "updated_at": r["updated_at"],
            "updated_by_name": r["updated_by_name"],
            "has_price": prix > 0,
        })
    return out


def _pf_valo_query_at_date(conn: sqlite3.Connection, snapshot_date: str | None) -> list[dict]:
    """Idem _pf_valo_query mais reconstitue quantité et prix HT à la date."""
    items = _pf_valo_query(conn)
    if not snapshot_date:
        return items
    stock_pf = _pf_stock_snapshot_at_date(conn, snapshot_date)
    prix_map = _pf_price_snapshot_at_date(conn, snapshot_date)
    for it in items:
        try:
            pid = int(it["id"])
        except (TypeError, ValueError, KeyError):
            continue
        new_qte = float(stock_pf.get(pid, it.get("quantite") or 0))
        new_prix = float(prix_map[pid]) if pid in prix_map else float(it.get("prix_unitaire_ht") or 0)
        divisor = 1000.0 if it.get("par_mille") else 1.0
        it["quantite"] = round(new_qte, 4)
        it["prix_unitaire_ht"] = new_prix
        it["valorisation"] = round(new_qte * new_prix / divisor, 2)
        it["has_price"] = new_prix > 0
    return items


def _pf_valo_summary(
    items: list[dict],
    *,
    charge_prod_pct: float = 0.0,
    storage_fees_pct: float = 0.0,
) -> dict:
    total_fab = 0.0
    total_neg = 0.0
    nb_fab = 0
    nb_fab_valued = 0
    nb_neg = 0
    nb_neg_valued = 0
    nb_sans_prix = 0
    for it in items:
        if it["type"] == "negoce":
            nb_neg += 1
            total_neg += it["valorisation"]
            if it["has_price"]:
                nb_neg_valued += 1
            else:
                nb_sans_prix += 1
        else:
            nb_fab += 1
            total_fab += it["valorisation"]
            if it["has_price"]:
                nb_fab_valued += 1
            else:
                nb_sans_prix += 1
    total_pf = total_fab + total_neg
    # Valo PF avec charges = valo_pf * (1 + storage/100) * (1 - charge/100)
    # (charge_production_pct est un multiplicateur (1 - x/100), pas un diviseur)
    charge_mult = 1.0 - (charge_prod_pct / 100.0)
    if charge_mult < 0:
        charge_mult = 0.0  # garde-fou : borné à 100 % max côté validator
    total_pf_avec_charges = total_pf * (1.0 + storage_fees_pct / 100.0) * charge_mult
    return {
        "total_pf": round(total_pf, 2),
        "total_fabrique": round(total_fab, 2),
        "total_negoce": round(total_neg, 2),
        "nb_refs": nb_fab + nb_neg,
        "nb_refs_fabrique": nb_fab,
        "nb_refs_negoce": nb_neg,
        "nb_refs_valorisees": nb_fab_valued + nb_neg_valued,
        "nb_refs_sans_prix": nb_sans_prix,
        "charge_production_pct": round(charge_prod_pct, 4),
        "storage_fees_pct": round(storage_fees_pct, 4),
        "total_pf_avec_charges": round(total_pf_avec_charges, 2),
        "total_charges_prod": round(total_pf_avec_charges - total_pf, 2),
    }


def _pf_enrich_charges(
    items: list[dict],
    *,
    charge_prod_pct: float,
    storage_fees_pct: float,
) -> None:
    """Ajoute sur chaque item :
    - valorisation_avec_charges : valo * (1 + storage/100) * (1 - charge/100)
    - charges_prod : valorisation_avec_charges - valorisation
    """
    charge_mult = 1.0 - (charge_prod_pct / 100.0)
    if charge_mult < 0:
        charge_mult = 0.0
    mult = (1.0 + storage_fees_pct / 100.0) * charge_mult
    for it in items:
        v = float(it.get("valorisation") or 0)
        v_charged = v * mult
        it["valorisation_avec_charges"] = round(v_charged, 2)
        it["charges_prod"] = round(v_charged - v, 2)


@router.get("/api/stock/valorisation/pf")
def list_valorisation_pf(request: Request, date: str | None = None):
    """Liste tous les produits finis (fabrique + négoce) avec prix HT courant et stock.
    Query optionnel `date=YYYY-MM-DD` → snapshot à la date (stock + prix historisés)."""
    require_stock_matieres_admin(request)
    snapshot_date = _parse_snapshot_date(date)
    with get_db() as conn:
        items = _pf_valo_query_at_date(conn, snapshot_date)
        charge = _get_charge_production_pct(conn)
        storage = _get_storage_fees_pct(conn)
    _pf_enrich_charges(items, charge_prod_pct=charge, storage_fees_pct=storage)
    summary = _pf_valo_summary(items, charge_prod_pct=charge, storage_fees_pct=storage)
    summary["snapshot_date"] = snapshot_date
    return {"items": items, "summary": summary}


@router.put("/api/stock/valorisation/pf/{produit_id}")
async def update_valorisation_pf(produit_id: int, request: Request):
    """Met à jour le prix unitaire HT d'un produit fini (ou négoce). Trace l'historique."""
    user = require_stock_matieres_admin(request)
    body = await request.json()
    raw = body.get("prix_unitaire_ht", body.get("prix_unitaire"))
    try:
        new_price = float(raw) if raw not in (None, "") else 0.0
    except (TypeError, ValueError):
        raise HTTPException(400, "Prix invalide.") from None
    if new_price < 0:
        raise HTTPException(400, "Prix négatif.")
    if new_price > 1_000_000:
        raise HTTPException(400, "Prix hors limites.")
    now = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    user_name = (user.get("nom") or user.get("email") or "").strip() or None
    user_id = user.get("id")
    with get_db() as conn:
        prod = conn.execute(
            "SELECT id, reference, designation, unite, type FROM produits WHERE id=?",
            (produit_id,),
        ).fetchone()
        if not prod:
            raise HTTPException(404, "Produit introuvable.")
        prev = conn.execute(
            "SELECT prix_unitaire_ht FROM pf_valorisation WHERE produit_id=?",
            (produit_id,),
        ).fetchone()
        prev_price = float(prev["prix_unitaire_ht"]) if prev and prev["prix_unitaire_ht"] is not None else None
        if (prev_price or 0) == new_price:
            # Pas de changement réel : on rafraîchit juste l'item sans logger
            pass
        else:
            conn.execute(
                """
                INSERT INTO pf_valorisation (produit_id, prix_unitaire_ht, updated_at, updated_by, updated_by_name)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(produit_id) DO UPDATE SET
                    prix_unitaire_ht = excluded.prix_unitaire_ht,
                    updated_at = excluded.updated_at,
                    updated_by = excluded.updated_by,
                    updated_by_name = excluded.updated_by_name
                """,
                (produit_id, new_price, now, user_id, user_name),
            )
            conn.execute(
                """
                INSERT INTO pf_valorisation_historique
                  (produit_id, prix_avant, prix_apres, source, note, created_at, created_by, created_by_name)
                VALUES (?, ?, ?, 'edition', ?, ?, ?, ?)
                """,
                (produit_id, prev_price, new_price, None, now, user_id, user_name),
            )
            conn.commit()
        items = _pf_valo_query(conn)
        charge = _get_charge_production_pct(conn)
        storage = _get_storage_fees_pct(conn)
    _pf_enrich_charges(items, charge_prod_pct=charge, storage_fees_pct=storage)
    item = next((x for x in items if x["id"] == produit_id), None)
    return {
        "item": item,
        "summary": _pf_valo_summary(
            items, charge_prod_pct=charge, storage_fees_pct=storage
        ),
    }


@router.get("/api/stock/valorisation/pf/{produit_id}/historique")
def get_valorisation_pf_historique(produit_id: int, request: Request):
    require_stock_matieres_admin(request)
    with get_db() as conn:
        prod = conn.execute(
            "SELECT id, reference, designation FROM produits WHERE id=?",
            (produit_id,),
        ).fetchone()
        if not prod:
            raise HTTPException(404, "Produit introuvable.")
        rows = conn.execute(
            """SELECT id, prix_avant, prix_apres, source, note, created_at, created_by_name
               FROM pf_valorisation_historique WHERE produit_id=?
               ORDER BY datetime(created_at) DESC, id DESC""",
            (produit_id,),
        ).fetchall()
    return {
        "produit": {
            "id": prod["id"],
            "reference": prod["reference"],
            "designation": prod["designation"],
        },
        "historique": [
            {
                "id": r["id"],
                "prix_avant": float(r["prix_avant"]) if r["prix_avant"] is not None else None,
                "prix_apres": float(r["prix_apres"]) if r["prix_apres"] is not None else 0.0,
                "source": r["source"],
                "note": r["note"],
                "created_at": r["created_at"],
                "created_by_name": r["created_by_name"],
            }
            for r in rows
        ],
    }


# ─────────── Import Excel (feuille "Valorisation") ───────────

def _pf_valo_parse_excel(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """Parse la feuille Valorisation. Retourne (rows, warnings).
    Format attendu : Référence, Désignation, Unité (MySIFA), …, Prix unitaire HT, …, Statut, Source prix.
    Recherche la feuille 'Valorisation' (insensible à la casse) ou prend la 1re feuille en fallback.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise HTTPException(500, f"openpyxl indisponible : {e}") from None
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Fichier Excel illisible : {e}") from None
    sheet_name = None
    for sn in wb.sheetnames:
        if sn.strip().lower() == "valorisation":
            sheet_name = sn
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    # En-têtes (ligne 1)
    headers_raw = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        headers_raw.append(str(cell).strip() if cell is not None else "")
    def find_col(*candidates: str) -> int | None:
        norm = lambda s: re.sub(r"\s+", " ", (s or "").strip().lower())
        targets = {norm(c) for c in candidates}
        for i, h in enumerate(headers_raw):
            if norm(h) in targets:
                return i
        return None
    idx_ref = find_col("Référence", "Reference", "Ref", "Référence produit")
    idx_des = find_col("Désignation", "Designation")
    idx_unite = find_col("Unité (MySIFA)", "Unité MySIFA", "Unite (MySIFA)", "Unité")
    idx_qte = find_col("Quantité stock", "Quantite stock", "Quantité")
    idx_date = find_col("Date dernière cmd", "Date derniere cmd", "Date dernière commande")
    idx_ht = find_col("Prix unitaire HT", "Prix HT", "PU HT")
    idx_statut = find_col("Statut")
    idx_source = find_col("Source prix", "Source")
    warnings: list[str] = []
    if idx_ref is None:
        raise HTTPException(400, "Colonne 'Référence' introuvable dans la feuille Valorisation.")
    if idx_ht is None:
        raise HTTPException(400, "Colonne 'Prix unitaire HT' introuvable dans la feuille Valorisation.")
    rows_out: list[dict] = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue
        ref = row[idx_ref] if idx_ref < len(row) else None
        if ref is None:
            continue
        ref_s = _pf_valo_normalize_ref(str(ref))
        if not ref_s:
            continue
        prix_raw = row[idx_ht] if (idx_ht is not None and idx_ht < len(row)) else None
        try:
            prix = float(prix_raw) if prix_raw not in (None, "") else 0.0
        except (TypeError, ValueError):
            warnings.append(f"Ligne {r_idx}: prix HT illisible « {prix_raw} » — ignoré (mis à 0)")
            prix = 0.0
        des = row[idx_des] if (idx_des is not None and idx_des < len(row)) else None
        unite = row[idx_unite] if (idx_unite is not None and idx_unite < len(row)) else None
        date_cmd = row[idx_date] if (idx_date is not None and idx_date < len(row)) else None
        statut = row[idx_statut] if (idx_statut is not None and idx_statut < len(row)) else None
        source = row[idx_source] if (idx_source is not None and idx_source < len(row)) else None
        # Normaliser la date
        date_s = None
        if date_cmd is not None and date_cmd != "":
            if isinstance(date_cmd, datetime):
                date_s = date_cmd.strftime("%Y-%m-%d")
            else:
                date_s = str(date_cmd).strip() or None
        rows_out.append({
            "reference": ref_s,
            "designation": (str(des).strip() if des is not None else ""),
            "unite": (str(unite).strip() if unite is not None else ""),
            "prix_unitaire_ht": prix,
            "date_derniere_cmd": date_s,
            "statut": (str(statut).strip() if statut is not None else None),
            "source_prix": (str(source).strip() if source is not None else None),
        })
    return rows_out, warnings


@router.post("/api/stock/valorisation/pf/import")
async def import_valorisation_pf(request: Request, file: UploadFile = File(...)):
    """Import du fichier Excel — onglet 'Valorisation'.
    - Crée les références fabrique manquantes (produits.type='fabrique').
    - Met à jour le prix HT pour toutes les refs trouvées.
    - Log un historique avec source='import'.
    """
    user = require_stock_matieres_admin(request)
    if not file or not file.filename:
        raise HTTPException(400, "Aucun fichier reçu.")
    fname = (file.filename or "").lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".xlsm")):
        raise HTTPException(400, "Format attendu : .xlsx ou .xlsm")
    content = await file.read()
    rows, warnings = _pf_valo_parse_excel(content)
    if not rows:
        raise HTTPException(400, "Aucune ligne exploitable dans la feuille Valorisation.")
    now = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    user_name = (user.get("nom") or user.get("email") or "").strip() or None
    user_id = user.get("id")
    created = 0
    updated = 0
    unchanged = 0
    skipped: list[dict] = []
    with get_db() as conn:
        for row in rows:
            ref = row["reference"]
            des = row["designation"] or ref
            unite = row["unite"] or "étiquette"
            prix = float(row["prix_unitaire_ht"] or 0)
            statut = row["statut"]
            source_prix = row["source_prix"] or "Import Excel"
            date_cmd = row["date_derniere_cmd"]
            existing = conn.execute(
                "SELECT id, designation, unite, type FROM produits WHERE reference=?",
                (ref,),
            ).fetchone()
            if not existing:
                # Création comme produit fabriqué
                cur = conn.execute(
                    "INSERT INTO produits (reference, designation, description, unite, type, created_at, updated_at) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (ref, des, "", unite, "fabrique", now, now),
                )
                pid = int(cur.lastrowid)
                created += 1
            else:
                pid = int(existing["id"])
                # Si le type est negoce, on ne touche pas — l'utilisateur l'a marqué exprès
                # Sinon, on met à jour designation / unité si vides
                if (not existing["designation"]) and des:
                    conn.execute("UPDATE produits SET designation=?, updated_at=? WHERE id=?", (des, now, pid))
                if (not existing["unite"]) and unite:
                    conn.execute("UPDATE produits SET unite=?, updated_at=? WHERE id=?", (unite, now, pid))
            prev = conn.execute(
                "SELECT prix_unitaire_ht FROM pf_valorisation WHERE produit_id=?",
                (pid,),
            ).fetchone()
            prev_price = float(prev["prix_unitaire_ht"]) if prev and prev["prix_unitaire_ht"] is not None else None
            conn.execute(
                """
                INSERT INTO pf_valorisation
                  (produit_id, prix_unitaire_ht, source_prix, statut, date_derniere_cmd, updated_at, updated_by, updated_by_name)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(produit_id) DO UPDATE SET
                  prix_unitaire_ht = excluded.prix_unitaire_ht,
                  source_prix = excluded.source_prix,
                  statut = excluded.statut,
                  date_derniere_cmd = excluded.date_derniere_cmd,
                  updated_at = excluded.updated_at,
                  updated_by = excluded.updated_by,
                  updated_by_name = excluded.updated_by_name
                """,
                (pid, prix, source_prix, statut, date_cmd, now, user_id, user_name),
            )
            if (prev_price or 0) == prix:
                unchanged += 1
            else:
                updated += 1
                conn.execute(
                    """
                    INSERT INTO pf_valorisation_historique
                      (produit_id, prix_avant, prix_apres, source, note, created_at, created_by, created_by_name)
                    VALUES (?, ?, ?, 'import', ?, ?, ?, ?)
                    """,
                    (pid, prev_price, prix, source_prix, now, user_id, user_name),
                )
        conn.commit()
        items = _pf_valo_query(conn)
        charge = _get_charge_production_pct(conn)
        storage = _get_storage_fees_pct(conn)
    _pf_enrich_charges(items, charge_prod_pct=charge, storage_fees_pct=storage)
    return {
        "success": True,
        "stats": {
            "lignes_lues": len(rows),
            "refs_creees": created,
            "refs_mises_a_jour": updated,
            "refs_inchangees": unchanged,
            "avertissements": warnings[:50],
        },
        "summary": _pf_valo_summary(
            items, charge_prod_pct=charge, storage_fees_pct=storage
        ),
    }


@router.get("/api/stock/valorisation/export-pdf")
def export_valorisation_pdf(
    request: Request,
    date: str | None = None,
    type: str = "sommaire",
):
    """Export PDF de la valorisation (MP + PF).

    Query params :
      - date : YYYY-MM-DD optionnel → snapshot à cette date
      - type : lignes | sommaire | insights (par défaut sommaire)

    Réservé Direction + superadmin (contient des chiffres réels/avec charges)."""
    user = require_valorisation_usd_admin(request)
    snapshot_date = _parse_snapshot_date(date)
    type_vue = str(type or "sommaire").strip().lower()
    if type_vue not in ("lignes", "sommaire", "insights"):
        raise HTTPException(400, "Type de vue PDF invalide (attendu : lignes | sommaire | insights).")
    with get_db() as conn:
        items_mp = _valorisation_query_at_date(conn, snapshot_date)
        items_pf = _pf_valo_query_at_date(conn, snapshot_date)
        taux = _get_taux_eur_usd(conn)
        tax_pct = _get_import_tax_pct(conn)
        c_full, c_half, q_full, q_half = _get_container_params(conn)
        charge = _get_charge_production_pct(conn)
        storage = _get_storage_fees_pct(conn)
    _enrich_items_with_usd(items_mp, taux, tax_pct, c_full, c_half, q_full, q_half)
    summary_mp = _valorisation_summary(items_mp, taux, tax_pct, c_full, c_half, q_full, q_half)
    _pf_enrich_charges(items_pf, charge_prod_pct=charge, storage_fees_pct=storage)
    summary_pf = _pf_valo_summary(items_pf, charge_prod_pct=charge, storage_fees_pct=storage)
    try:
        from app.services.valorisation_pdf import build_valorisation_pdf
    except ImportError as e:
        raise HTTPException(500, f"Générateur PDF indisponible : {e}") from None
    try:
        pdf_bytes = build_valorisation_pdf(
            type_vue=type_vue,
            items_mp=items_mp, items_pf=items_pf,
            summary_mp=summary_mp, summary_pf=summary_pf,
            snapshot_date=snapshot_date,
            generated_by=(user.get("nom") or user.get("email") or "").strip() or None,
        )
    except Exception as e:
        raise HTTPException(500, f"Erreur génération PDF : {e}") from None
    date_slug = snapshot_date or datetime.now().strftime("%Y-%m-%d")
    filename = f"valorisation_{type_vue}_{date_slug}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/api/stock/valorisation/pf/export")
def export_valorisation_pf(request: Request):
    """Export Excel des prix unitaires PF actuels (format ré-importable)."""
    require_stock_matieres_admin(request)
    with get_db() as conn:
        items = _pf_valo_query(conn)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise HTTPException(500, f"openpyxl indisponible : {e}") from None
    wb = Workbook()
    ws = wb.active
    ws.title = "Valorisation"
    headers = [
        "Référence", "Désignation", "Unité (MySIFA)", "Type",
        "Quantité stock", "Prix unitaire HT", "Valorisation",
        "Date dernière cmd", "Statut", "Source prix",
        "Dernière maj", "Modifié par",
    ]
    head_fill = PatternFill("solid", fgColor="1E293B")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for i, it in enumerate(items, start=2):
        ws.cell(row=i, column=1, value=it["reference"])
        ws.cell(row=i, column=2, value=it["designation"])
        ws.cell(row=i, column=3, value=it["unite"])
        ws.cell(row=i, column=4, value=it["type_label"])
        ws.cell(row=i, column=5, value=it["quantite"])
        ws.cell(row=i, column=6, value=it["prix_unitaire_ht"])
        ws.cell(row=i, column=7, value=it["valorisation"])
        ws.cell(row=i, column=8, value=it.get("date_derniere_cmd") or "")
        ws.cell(row=i, column=9, value=it.get("statut") or "")
        ws.cell(row=i, column=10, value=it.get("source_prix") or "")
        ws.cell(row=i, column=11, value=it.get("updated_at") or "")
        ws.cell(row=i, column=12, value=it.get("updated_by_name") or "")
    widths = [16, 60, 14, 12, 14, 16, 16, 16, 22, 22, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else f"A{chr(64 + i - 26)}"].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"valorisation-pf-{date_str}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )




# ─────────────────────────────────────────────────────────────────────────
# Référentiel mp_laizes (admin /settings)
# ─────────────────────────────────────────────────────────────────────────


def _require_laizes_admin(request: Request) -> dict:
    user = get_current_user(request)
    if user.get("role") != "superadmin":
        raise HTTPException(403, "Accès super admin requis.")
    return user


@router.get("/api/admin/mp_laizes")
def list_mp_laizes(request: Request, all: int = 0):
    # Lecture autorisée pour tous les utilisateurs ayant accès au stock
    require_stock(request)
    include_inactive = bool(all)
    with get_db() as conn:
        where = "" if include_inactive else "WHERE actif = 1"
        rows = conn.execute(
            f"SELECT id, valeur_mm, label, ordre, actif FROM mp_laizes {where} "
            "ORDER BY ordre ASC, valeur_mm ASC"
        ).fetchall()
    return [dict(r) for r in rows]


@router.post("/api/admin/mp_laizes")
async def create_mp_laize(request: Request):
    _require_laizes_admin(request)
    body = await request.json()
    try:
        valeur = float(body.get("valeur_mm") or 0)
    except (TypeError, ValueError):
        raise HTTPException(400, "Valeur (mm) invalide.") from None
    if valeur <= 0 or valeur > 5000:
        raise HTTPException(400, "Valeur (mm) hors limites (1 à 5000).")
    label = (body.get("label") or "").strip()
    if not label:
        # Auto : "333 mm"
        if abs(valeur - int(valeur)) < 1e-9:
            label = f"{int(valeur)} mm"
        else:
            label = f"{valeur:g} mm"
    try:
        ordre = int(body.get("ordre") or 0)
    except (TypeError, ValueError):
        ordre = 0
    with get_db() as conn:
        try:
            cur = conn.execute(
                "INSERT INTO mp_laizes (valeur_mm, label, ordre, actif) VALUES (?,?,?,1)",
                (valeur, label, ordre),
            )
            conn.commit()
            row = conn.execute(
                "SELECT id, valeur_mm, label, ordre, actif FROM mp_laizes WHERE id=?",
                (cur.lastrowid,),
            ).fetchone()
        except sqlite3.IntegrityError:
            raise HTTPException(400, "Cette laize existe déjà.") from None
    return dict(row)


@router.put("/api/admin/mp_laizes/{laize_id}")
async def update_mp_laize(laize_id: int, request: Request):
    _require_laizes_admin(request)
    body = await request.json()
    sets: list[str] = []
    args: list[Any] = []
    if "valeur_mm" in body:
        try:
            v = float(body["valeur_mm"])
        except (TypeError, ValueError):
            raise HTTPException(400, "Valeur (mm) invalide.") from None
        if v <= 0 or v > 5000:
            raise HTTPException(400, "Valeur (mm) hors limites.")
        sets.append("valeur_mm=?"); args.append(v)
    if "label" in body:
        lab = (body.get("label") or "").strip()
        if not lab:
            raise HTTPException(400, "Label obligatoire.")
        sets.append("label=?"); args.append(lab)
    if "ordre" in body:
        try:
            sets.append("ordre=?"); args.append(int(body["ordre"]))
        except (TypeError, ValueError):
            pass
    if "actif" in body:
        sets.append("actif=?"); args.append(1 if body["actif"] else 0)
    if not sets:
        raise HTTPException(400, "Aucun champ à mettre à jour.")
    args.append(laize_id)
    with get_db() as conn:
        try:
            conn.execute(f"UPDATE mp_laizes SET {', '.join(sets)} WHERE id=?", args)
            conn.commit()
            row = conn.execute(
                "SELECT id, valeur_mm, label, ordre, actif FROM mp_laizes WHERE id=?",
                (laize_id,),
            ).fetchone()
        except sqlite3.IntegrityError:
            raise HTTPException(400, "Cette laize existe déjà.") from None
    if not row:
        raise HTTPException(404, "Laize introuvable.")
    return dict(row)


@router.delete("/api/admin/mp_laizes/{laize_id}")
def delete_mp_laize(laize_id: int, request: Request):
    _require_laizes_admin(request)
    with get_db() as conn:
        # Si cette laize est référencée, on refuse la suppression (on permet la désactivation)
        used_mat = conn.execute(
            "SELECT COUNT(*) AS n FROM mp_matiere_laizes WHERE laize_id=?", (laize_id,)
        ).fetchone()
        used_mvt = conn.execute(
            "SELECT COUNT(*) AS n FROM mp_mouvements WHERE laize_id=?", (laize_id,)
        ).fetchone()
        if (used_mat["n"] or 0) > 0 or (used_mvt["n"] or 0) > 0:
            raise HTTPException(
                400,
                "Laize utilisée par des matières ou mouvements — désactivez-la plutôt que de la supprimer.",
            )
        conn.execute("DELETE FROM mp_laizes WHERE id=?", (laize_id,))
        conn.commit()
    return {"ok": True}


# ─────────────────────────────────────────────────────────────────────────
# Helpers matières laizées
# ─────────────────────────────────────────────────────────────────────────


def _mp_laizes_for_matiere(conn, matiere_id: int) -> list[dict]:
    rows = conn.execute(
        """
        SELECT l.id, l.valeur_mm, l.label, l.ordre, l.actif,
               COALESCE(s.quantite, 0) AS quantite,
               ml.prix_eur_m2 AS laize_prix_eur_m2
        FROM mp_matiere_laizes ml
        JOIN mp_laizes l ON l.id = ml.laize_id
        LEFT JOIN mp_stock_laize s ON s.matiere_id = ml.matiere_id AND s.laize_id = ml.laize_id
        WHERE ml.matiere_id = ?
        ORDER BY l.ordre ASC, l.valeur_mm ASC
        """,
        (matiere_id,),
    ).fetchall()
    out = []
    for r in rows:
        try:
            lprix = float(r["laize_prix_eur_m2"]) if r["laize_prix_eur_m2"] is not None else None
        except (TypeError, ValueError):
            lprix = None
        out.append({
            "laize_id": r["id"],
            "valeur_mm": float(r["valeur_mm"] or 0),
            "label": r["label"],
            "actif": int(r["actif"] or 0),
            "quantite": float(r["quantite"] or 0),
            "prix_eur_m2": lprix,
        })
    return out


def _mp_is_complete(mp_row, has_laizes: bool, all_laizes_priced: bool = False) -> bool:
    """Une matière laizée est 'complète' si elle a >=1 laize + metres_lineaires_par_bobine > 0
    et soit prix_eur_m2 (matière) > 0 (mode unique), soit toutes les laizes ont un prix > 0
    (mode prix par laize)."""
    cat = (mp_row["categorie"] or "").lower()
    if cat not in _MP_CATEGORIES_LAIZEES:
        return True
    if not has_laizes:
        return False
    try:
        prix = float(mp_row["prix_eur_m2"] or 0)
        metres = float(mp_row["metres_lineaires_par_bobine"] or 0)
    except (TypeError, ValueError):
        return False
    if metres <= 0:
        return False
    keys = mp_row.keys() if hasattr(mp_row, "keys") else []
    prix_par_laize = False
    if "prix_par_laize" in keys:
        try:
            prix_par_laize = bool(int(mp_row["prix_par_laize"] or 0))
        except (TypeError, ValueError):
            prix_par_laize = False
    if prix_par_laize:
        return all_laizes_priced
    return prix > 0


@router.get("/api/stock/matieres/incompletes")
def count_matieres_incompletes(request: Request):
    require_stock(request)
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT mp.id, mp.categorie, mp.reference, mp.designation,
                   mp.prix_eur_m2, mp.metres_lineaires_par_bobine,
                   COALESCE(mp.prix_par_laize, 0) AS prix_par_laize,
                   (SELECT COUNT(*) FROM mp_matiere_laizes ml WHERE ml.matiere_id = mp.id) AS nb_laizes,
                   (SELECT COUNT(*) FROM mp_matiere_laizes ml WHERE ml.matiere_id = mp.id AND ml.prix_eur_m2 IS NOT NULL AND ml.prix_eur_m2 > 0) AS nb_laizes_priced
            FROM matieres_premieres mp
            WHERE mp.actif = 1 AND mp.categorie IN ('frontal','glassine','complexe')
            """
        ).fetchall()
    incomplete = []
    for r in rows:
        nb_laizes = int(r["nb_laizes"] or 0)
        nb_priced = int(r["nb_laizes_priced"] or 0)
        has_laizes = nb_laizes > 0
        all_priced = has_laizes and nb_priced == nb_laizes
        if not _mp_is_complete(r, has_laizes, all_laizes_priced=all_priced):
            incomplete.append({
                "id": r["id"],
                "categorie": r["categorie"],
                "reference": r["reference"],
                "designation": r["designation"],
            })
    return {"count": len(incomplete), "items": incomplete}


@router.post("/api/stock/matieres/{matiere_id}/laizes")
async def set_matiere_laizes(matiere_id: int, request: Request):
    """Définit la liste des laizes valables pour une matière laizée + métrage + prix m²."""
    require_stock_matieres_admin(request)
    body = await request.json()
    laize_ids_raw = body.get("laize_ids") or []
    if not isinstance(laize_ids_raw, list):
        raise HTTPException(400, "laize_ids doit être une liste.")
    try:
        laize_ids = [int(x) for x in laize_ids_raw]
    except (TypeError, ValueError):
        raise HTTPException(400, "laize_ids invalides.") from None
    metres = body.get("metres_lineaires_par_bobine")
    prix = body.get("prix_eur_m2")
    try:
        metres = float(metres) if metres not in (None, "") else None
        prix = float(prix) if prix not in (None, "") else None
    except (TypeError, ValueError):
        raise HTTPException(400, "métrage ou prix invalide.") from None
    if metres is not None and metres < 0:
        raise HTTPException(400, "métrage négatif.")
    if prix is not None and prix < 0:
        raise HTTPException(400, "prix négatif.")
    # Mode prix par laize (radio) : 0 = prix unique matière, 1 = prix distinct par laize
    prix_par_laize_raw = body.get("prix_par_laize")
    prix_par_laize: Optional[int] = None
    if prix_par_laize_raw is not None:
        try:
            prix_par_laize = 1 if int(prix_par_laize_raw) else 0
        except (TypeError, ValueError):
            raise HTTPException(400, "prix_par_laize invalide.") from None
    # Prix par laize : dict {laize_id: prix_eur_m2}
    laize_prices_raw = body.get("laize_prices") or {}
    if not isinstance(laize_prices_raw, dict):
        raise HTTPException(400, "laize_prices doit être un objet.")
    laize_prices: dict[int, Optional[float]] = {}
    for k, v in laize_prices_raw.items():
        try:
            lid = int(k)
        except (TypeError, ValueError):
            raise HTTPException(400, "laize_prices: clé invalide.") from None
        if v in (None, ""):
            laize_prices[lid] = None
        else:
            try:
                fv = float(v)
            except (TypeError, ValueError):
                raise HTTPException(400, f"laize_prices[{lid}] invalide.") from None
            if fv < 0:
                raise HTTPException(400, f"laize_prices[{lid}] négatif.")
            laize_prices[lid] = fv
    with get_db() as conn:
        mat = conn.execute(
            "SELECT id, categorie FROM matieres_premieres WHERE id=?", (matiere_id,)
        ).fetchone()
        if not mat:
            raise HTTPException(404, "Matière introuvable.")
        if not _mp_is_laizee(mat["categorie"]):
            raise HTTPException(400, "Cette catégorie n'est pas laizée.")
        if laize_ids:
            placeholders = ",".join("?" * len(laize_ids))
            valid_rows = conn.execute(
                f"SELECT id FROM mp_laizes WHERE id IN ({placeholders})", laize_ids
            ).fetchall()
            valid_ids = {r["id"] for r in valid_rows}
            invalid = [lid for lid in laize_ids if lid not in valid_ids]
            if invalid:
                raise HTTPException(400, f"Laizes inconnues : {invalid}")
        # Suppression des laizes retirées (et de leur stock — protégé si quantité > 0)
        current = conn.execute(
            "SELECT laize_id FROM mp_matiere_laizes WHERE matiere_id=?", (matiere_id,)
        ).fetchall()
        current_ids = {r["laize_id"] for r in current}
        new_ids = set(laize_ids)
        removed = current_ids - new_ids
        for rid in removed:
            stock = conn.execute(
                "SELECT quantite FROM mp_stock_laize WHERE matiere_id=? AND laize_id=?",
                (matiere_id, rid),
            ).fetchone()
            if stock and float(stock["quantite"] or 0) > 0:
                raise HTTPException(
                    400,
                    "Impossible de retirer une laize avec du stock > 0.",
                )
            conn.execute(
                "DELETE FROM mp_matiere_laizes WHERE matiere_id=? AND laize_id=?",
                (matiere_id, rid),
            )
            conn.execute(
                "DELETE FROM mp_stock_laize WHERE matiere_id=? AND laize_id=?",
                (matiere_id, rid),
            )
        # Ajout des nouvelles
        added = new_ids - current_ids
        for aid in added:
            conn.execute(
                "INSERT OR IGNORE INTO mp_matiere_laizes (matiere_id, laize_id) VALUES (?, ?)",
                (matiere_id, aid),
            )
            conn.execute(
                "INSERT OR IGNORE INTO mp_stock_laize (matiere_id, laize_id, quantite) VALUES (?, ?, 0)",
                (matiere_id, aid),
            )
        # Mise à jour métrage / prix / mode prix_par_laize
        sets: list[str] = []
        args: list[Any] = []
        if metres is not None:
            sets.append("metres_lineaires_par_bobine=?"); args.append(metres)
        if prix is not None:
            sets.append("prix_eur_m2=?"); args.append(prix)
        if prix_par_laize is not None:
            sets.append("prix_par_laize=?"); args.append(prix_par_laize)
        if sets:
            args.append(matiere_id)
            conn.execute(
                f"UPDATE matieres_premieres SET {', '.join(sets)} WHERE id=?", args
            )
        # Prix par laize (n'est appliqué que sur les laizes associées à la matière)
        if laize_prices:
            active_ids = {r["laize_id"] for r in conn.execute(
                "SELECT laize_id FROM mp_matiere_laizes WHERE matiere_id=?", (matiere_id,)
            ).fetchall()}
            for lid, fv in laize_prices.items():
                if lid not in active_ids:
                    continue
                conn.execute(
                    "UPDATE mp_matiere_laizes SET prix_eur_m2=? WHERE matiere_id=? AND laize_id=?",
                    (fv, matiere_id, lid),
                )
        conn.commit()
        laizes = _mp_laizes_for_matiere(conn, matiere_id)
    return {"ok": True, "laizes": laizes}


def _mp_fournisseurs_by_laize(conn, matiere_id: int) -> dict[int, list[dict]]:
    rows = conn.execute("""
        SELECT mlf.laize_id, f.id, f.nom, COALESCE(f.has_fsc, 1) AS has_fsc,
               f.licence, f.certificat
          FROM matiere_laize_fournisseurs mlf
          JOIN fournisseurs_fsc f ON f.id = mlf.fournisseur_id
         WHERE mlf.matiere_id = ?
         ORDER BY f.nom COLLATE NOCASE ASC
    """, (matiere_id,)).fetchall()
    out: dict[int, list[dict]] = {}
    for r in rows:
        out.setdefault(int(r["laize_id"]), []).append({
            "id": int(r["id"]),
            "nom": r["nom"],
            "has_fsc": int(r["has_fsc"] or 0),
            "licence": r["licence"],
            "certificat": r["certificat"],
        })
    return out


@router.get("/api/stock/fournisseurs")
def stock_list_fournisseurs(request: Request):
    """Liste plate des fournisseurs (accessible a tous les utilisateurs stock)."""
    require_stock(request)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT id, nom, COALESCE(has_fsc, 1) AS has_fsc, licence, certificat
              FROM fournisseurs_fsc
             ORDER BY nom COLLATE NOCASE ASC
        """).fetchall()
    return [dict(r) for r in rows]


@router.put("/api/stock/matieres/{matiere_id}/laizes/{laize_id}/fournisseurs")
async def set_matiere_laize_fournisseurs(matiere_id: int, laize_id: int, request: Request):
    """Sync des fournisseurs lies a une (matiere, laize)."""
    require_stock_matieres_admin(request)
    body = await request.json()
    ids_raw = body.get("fournisseur_ids") or []
    if not isinstance(ids_raw, list):
        raise HTTPException(400, "fournisseur_ids doit etre une liste")
    try:
        ids = sorted({int(x) for x in ids_raw})
    except (TypeError, ValueError):
        raise HTTPException(400, "fournisseur_ids invalide") from None
    with get_db() as conn:
        link = conn.execute(
            "SELECT 1 FROM mp_matiere_laizes WHERE matiere_id=? AND laize_id=?",
            (matiere_id, laize_id),
        ).fetchone()
        if not link:
            raise HTTPException(404, "Cette laize n'est pas associee a la matiere.")
        if ids:
            placeholders = ",".join("?" * len(ids))
            valid = {r["id"] for r in conn.execute(
                f"SELECT id FROM fournisseurs_fsc WHERE id IN ({placeholders})", ids
            ).fetchall()}
            invalid = [i for i in ids if i not in valid]
            if invalid:
                raise HTTPException(400, f"Fournisseurs inconnus : {invalid}")
        conn.execute(
            "DELETE FROM matiere_laize_fournisseurs WHERE matiere_id=? AND laize_id=?",
            (matiere_id, laize_id),
        )
        for fid in ids:
            conn.execute(
                "INSERT INTO matiere_laize_fournisseurs (matiere_id, laize_id, fournisseur_id) VALUES (?, ?, ?)",
                (matiere_id, laize_id, fid),
            )
        conn.commit()
        fournisseurs_by_laize = _mp_fournisseurs_by_laize(conn, matiere_id)
    return {"ok": True, "fournisseurs": fournisseurs_by_laize.get(laize_id, [])}


@router.get("/api/stock/matieres/{matiere_id}/laizes")
def get_matiere_laizes(matiere_id: int, request: Request):
    require_stock(request)
    with get_db() as conn:
        mat = conn.execute(
            """SELECT id, categorie, reference, designation,
                      COALESCE(prix_eur_m2, 0) AS prix_eur_m2,
                      COALESCE(prix_par_laize, 0) AS prix_par_laize,
                      COALESCE(metres_lineaires_par_bobine, 0) AS metres_lineaires_par_bobine
               FROM matieres_premieres WHERE id=?""",
            (matiere_id,),
        ).fetchone()
        if not mat:
            raise HTTPException(404, "Matière introuvable.")
        laizes = _mp_laizes_for_matiere(conn, matiere_id)
        fourn = _mp_fournisseurs_by_laize(conn, matiere_id)
    for l in laizes:
        lid = l.get("laize_id")
        if lid is not None:
            l["fournisseurs"] = fourn.get(int(lid), [])
    return {
        "matiere_id": mat["id"],
        "categorie": mat["categorie"],
        "reference": mat["reference"],
        "designation": mat["designation"],
        "prix_eur_m2": float(mat["prix_eur_m2"] or 0),
        "prix_par_laize": bool(int(mat["prix_par_laize"] or 0)) if "prix_par_laize" in mat.keys() else False,
        "metres_lineaires_par_bobine": float(mat["metres_lineaires_par_bobine"] or 0),
        "laizes": laizes,
    }
