"""MySifa — Router Gestion des Paies
Accès : superadmin + direction + administration + comptabilite
"""
from __future__ import annotations

import io
import json
from datetime import datetime
from typing import Any, Dict, Optional

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse

from app.core.database import get_db
from app.services.auth_service import get_current_user

ROLES_PAIE = {"superadmin", "direction", "administration", "comptabilite"}

router = APIRouter()

# ─── Helpers ──────────────────────────────────────────────────────────────────

MOIS_FR = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
           "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]

# Ordre des lignes dans l'export xlsx (identique au modèle de la comptable)
EXPORT_FIELDS = [
    ("nom",                       "Nom"),
    ("prenom",                    "Prénom"),
    ("matricule",                 "Matricule"),
    ("compteur_hs_m1",            "Compteur HS M-1"),
    ("contrat_type",              "Contrat (TYPE)"),
    ("date_debut",                "Date début"),
    ("date_fin",                  "Date fin"),
    ("nb_heures_base",            "Nb d'heures de base"),
    ("nb_heures_payer",           "Nb d'heures à payer"),
    ("taux_horaire",              "Taux horaire"),
    ("salaire_mensuel",           "Salaire mensuel"),
    ("augmentation_salaire",      "Augmentation de Salaire"),
    ("commissions_ventes",        "Commissions sur ventes"),
    ("mutuelle",                  "MUTUELLE"),
    ("avantage_voiture",          "Avantages en natures voiture"),
    ("heures_nuit",               "Heure de nuit"),
    ("heures_nuit_ferie",         "dont Heures DE NUIT férié"),
    ("heures_nuit_dimanche",      "dont Heures DE NUIT dimanche"),
    ("heures_nuit_dimanche_ferie","dont Heure de nuit dimanche férié"),
    ("heures_sup_25",             "Heures sup 25 %"),
    ("heures_sup_50",             "Heures sup 50 %"),
    ("heures_sup_nuit",           "Heures Supplémentaires DE NUIT"),
    ("panier",                    "Panier (6,47€ par jours)"),
    ("heures_ferie",              "Nb d'heures jour férié ( +150%)"),
    ("prime_anciennete",          "Prime ancienneté"),
    ("prime_objectifs",           "Prime d'objectifs"),
    ("prime_inflation",           "Prime inflation"),
    ("prime_exceptionnelle",      "Prime exceptionnelle"),
    ("solde_tout_compte",         "Solde tout compte (oui non)"),
    ("prime_equipe",              "Prime équipe"),
    ("absence_heures",            "Absence en heures"),
    ("absence_maladie_heures",    "Absence maladie en heures"),
    ("absence_maladie_jours",     "Absence maladie en jours"),
    ("absence_deces_mariage",     "Absence Deces familial - Mariage"),
    ("absence_cp_heures",         "Absence congés payés en heures"),
    ("absence_cp_jours",          "Absence congés payés en jours"),
    ("absence_rtt",               "Absence RTT"),
    ("absence_css_heures",        "Absence Congés sans solde heures"),
    ("absence_css_jours",         "Absence Congés sans solde jours"),
    ("absence_non_justifie_h",    "heures d'absence non justifiés"),
    ("absence_non_justifie_j",    "Jours d'absence non justifiés"),
    ("absence_justifiee_np_h",    "Absence justifiée non payée Heures"),
    ("absence_justifiee_np_j",    "Absence justifiée non payée Jours"),
    ("absence_at_heures",         "Absence AT en heures"),
    ("absence_at_jours",          "Absence AT en jours"),
    ("mi_temps_therapeutique",    "Mi-temps thérapeutique"),
    ("absence_chomage_partiel",   "Absence Chomage partiel"),
    ("absence_conge_parentale",   "Absence congés parentale"),
    ("date_conges_payes",         "DATE CONGES PAYES"),
    ("frais_pro",                 "Frais professionnels"),
    ("frais_transport",           "Frais remboursement transport"),
    ("pret_sifa",                 "Prêt SIFA"),
    ("atd",                       "ATD"),
    ("acompte_exceptionnel",      "Acompte exceptionnel"),
    ("information",               "information"),
]

FIXED_FIELDS = {
    "matricule", "contrat_type", "date_debut", "date_fin",
    "nb_heures_base", "taux_horaire", "salaire_mensuel",
    "prime_anciennete", "mutuelle", "avantage_voiture",
}


def _require_paie(request: Request) -> dict:
    user = get_current_user(request)
    if user.get("role") not in ROLES_PAIE:
        raise HTTPException(status_code=403, detail="Accès réservé")
    return user


def _row_to_dict(row) -> dict:
    return dict(row) if row else {}


# ─── Employees list ────────────────────────────────────────────────────────────

@router.get("/api/paie/employes")
def list_employes(request: Request):
    _require_paie(request)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT u.id, u.nom, u.email, u.actif,
                   pe.id        AS pe_id,
                   pe.matricule, pe.contrat_type, pe.date_debut, pe.date_fin,
                   pe.nb_heures_base, pe.taux_horaire, pe.salaire_mensuel,
                   pe.prime_anciennete, pe.mutuelle, pe.avantage_voiture
            FROM users u
            LEFT JOIN paie_employes pe ON pe.user_id = u.id
            WHERE pe.user_id IS NOT NULL
            ORDER BY u.nom COLLATE NOCASE
        """).fetchall()

    result = []
    for r in rows:
        parts = (r["nom"] or "").strip().split(" ", 1)
        result.append({
            "user_id":        r["id"],
            "nom":            parts[0] if parts else "",
            "prenom":         parts[1] if len(parts) > 1 else "",
            "nom_complet":    r["nom"],
            "email":          r["email"],
            "actif":          bool(r["actif"]),
            "pe_id":          r["pe_id"],
            "matricule":      r["matricule"],
            "contrat_type":   r["contrat_type"] or "CDI",
            "date_debut":     r["date_debut"],
            "date_fin":       r["date_fin"],
            "nb_heures_base": r["nb_heures_base"],
            "taux_horaire":   r["taux_horaire"],
            "salaire_mensuel":r["salaire_mensuel"],
            "prime_anciennete": r["prime_anciennete"],
            "mutuelle":       "Oui" if r["mutuelle"] else "Non",
            "avantage_voiture": r["avantage_voiture"],
        })
    return {"employes": result}


# ─── Fixed employee data ───────────────────────────────────────────────────────

@router.put("/api/paie/employes/{user_id}/fixed")
async def update_employe_fixed(user_id: int, request: Request):
    user = _require_paie(request)
    body = await request.json()
    now = datetime.now().isoformat()
    with get_db() as conn:
        existing = conn.execute("SELECT id FROM paie_employes WHERE user_id=?", (user_id,)).fetchone()
        if existing:
            conn.execute("""
                UPDATE paie_employes SET
                    matricule=?, contrat_type=?, date_debut=?, date_fin=?,
                    nb_heures_base=?, taux_horaire=?, salaire_mensuel=?,
                    prime_anciennete=?, mutuelle=?, avantage_voiture=?,
                    updated_at=?, updated_by=?
                WHERE user_id=?
            """, (
                body.get("matricule"), body.get("contrat_type", "CDI"),
                body.get("date_debut"), body.get("date_fin"),
                body.get("nb_heures_base"), body.get("taux_horaire"),
                body.get("salaire_mensuel"), body.get("prime_anciennete"),
                1 if str(body.get("mutuelle","")).lower() in ("oui","1","true") else 0,
                body.get("avantage_voiture"),
                now, user["email"], user_id,
            ))
        else:
            conn.execute("""
                INSERT INTO paie_employes
                    (user_id, matricule, contrat_type, date_debut, date_fin,
                     nb_heures_base, taux_horaire, salaire_mensuel,
                     prime_anciennete, mutuelle, avantage_voiture, updated_at, updated_by)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                user_id,
                body.get("matricule"), body.get("contrat_type", "CDI"),
                body.get("date_debut"), body.get("date_fin"),
                body.get("nb_heures_base"), body.get("taux_horaire"),
                body.get("salaire_mensuel"), body.get("prime_anciennete"),
                1 if str(body.get("mutuelle","")).lower() in ("oui","1","true") else 0,
                body.get("avantage_voiture"),
                now, user["email"],
            ))
        conn.commit()
    return {"ok": True}


# ─── Monthly variables ─────────────────────────────────────────────────────────

@router.get("/api/paie/variables/{annee}/{mois}")
def get_variables(annee: int, mois: int, request: Request):
    _require_paie(request)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT pv.user_id, pv.data, pv.updated_at, pv.updated_by
            FROM paie_variables pv
            WHERE pv.annee=? AND pv.mois=?
        """, (annee, mois)).fetchall()

    result = {}
    for r in rows:
        try:
            data = json.loads(r["data"]) if r["data"] else {}
        except Exception:
            data = {}
        result[str(r["user_id"])] = {
            "data": data,
            "updated_at": r["updated_at"],
            "updated_by": r["updated_by"],
        }
    return {"annee": annee, "mois": mois, "variables": result}


@router.put("/api/paie/variables/{annee}/{mois}/{user_id}")
async def save_variables(annee: int, mois: int, user_id: int, request: Request):
    user = _require_paie(request)
    body = await request.json()
    now = datetime.now().isoformat()
    data_json = json.dumps(body.get("data", {}), ensure_ascii=False)
    with get_db() as conn:
        conn.execute("""
            INSERT INTO paie_variables (user_id, annee, mois, data, updated_at, updated_by)
            VALUES (?,?,?,?,?,?)
            ON CONFLICT(user_id, annee, mois) DO UPDATE SET
                data=excluded.data, updated_at=excluded.updated_at, updated_by=excluded.updated_by
        """, (user_id, annee, mois, data_json, now, user["email"]))
        conn.commit()
    return {"ok": True}


# ─── History ───────────────────────────────────────────────────────────────────

@router.get("/api/paie/historique")
def get_historique(request: Request):
    _require_paie(request)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT annee, mois, COUNT(DISTINCT user_id) AS nb_employes, MAX(updated_at) AS last_update
            FROM paie_variables
            GROUP BY annee, mois
            ORDER BY annee DESC, mois DESC
        """).fetchall()
    return {"periodes": [
        {"annee": r["annee"], "mois": r["mois"],
         "mois_label": MOIS_FR[r["mois"]] if 1 <= r["mois"] <= 12 else str(r["mois"]),
         "nb_employes": r["nb_employes"], "last_update": r["last_update"]}
        for r in rows
    ]}


# ─── XLSX Export ───────────────────────────────────────────────────────────────

@router.get("/api/paie/export/{annee}/{mois}")
def export_xlsx(annee: int, mois: int, request: Request):
    _require_paie(request)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non installé")

    with get_db() as conn:
        employes = conn.execute("""
            SELECT u.id, u.nom,
                   pe.matricule, pe.contrat_type, pe.date_debut, pe.date_fin,
                   pe.nb_heures_base, pe.taux_horaire, pe.salaire_mensuel,
                   pe.prime_anciennete, pe.mutuelle, pe.avantage_voiture
            FROM users u
            LEFT JOIN paie_employes pe ON pe.user_id = u.id
            WHERE u.actif = 1
            ORDER BY u.nom COLLATE NOCASE
        """).fetchall()

        var_rows = conn.execute(
            "SELECT user_id, data FROM paie_variables WHERE annee=? AND mois=?",
            (annee, mois)
        ).fetchall()

    var_map = {}
    for r in var_rows:
        try:
            var_map[r["user_id"]] = json.loads(r["data"]) if r["data"] else {}
        except Exception:
            var_map[r["user_id"]] = {}

    wb = Workbook()
    ws = wb.active
    mois_label = MOIS_FR[mois] if 1 <= mois <= 12 else str(mois)
    ws.title = f"{mois_label} {annee}"

    # ── Styles ──
    BG_HEADER  = "0A0E17"
    BG_LABEL   = "111827"
    BG_ACCENT  = "0E7490"
    FG_WHITE   = "F1F5F9"
    FG_MUTED   = "94A3B8"
    FG_ACCENT  = "22D3EE"

    hdr_font   = Font(bold=True, color=FG_WHITE, name="Arial", size=9)
    label_font = Font(bold=True, color=FG_ACCENT, name="Arial", size=8)
    data_font  = Font(color=FG_WHITE, name="Arial", size=8)
    title_font = Font(bold=True, color=FG_WHITE, name="Arial", size=11)

    hdr_fill   = PatternFill("solid", fgColor=BG_ACCENT)
    label_fill = PatternFill("solid", fgColor=BG_LABEL)
    dark_fill  = PatternFill("solid", fgColor=BG_HEADER)

    thin_border = Border(
        left=Side(style="thin", color="1E293B"),
        right=Side(style="thin", color="1E293B"),
        top=Side(style="thin", color="1E293B"),
        bottom=Side(style="thin", color="1E293B"),
    )

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Titre ──
    ws.merge_cells(f"A1:{get_column_letter(len(employes) + 1)}1")
    title_cell = ws["A1"]
    title_cell.value = f"Saisie des Paies — {mois_label} {annee}"
    title_cell.font  = title_font
    title_cell.fill  = PatternFill("solid", fgColor=BG_ACCENT)
    title_cell.alignment = center
    ws.row_dimensions[1].height = 24

    # ── Row 2: header Nom ──
    ws["A2"].value = "Champ"
    ws["A2"].font  = hdr_font
    ws["A2"].fill  = hdr_fill
    ws["A2"].alignment = center
    ws["A2"].border = thin_border
    ws.column_dimensions["A"].width = 36

    for col_idx, emp in enumerate(employes, start=2):
        parts = (emp["nom"] or "").strip().split(" ", 1)
        nom    = parts[0].upper()
        prenom = parts[1].capitalize() if len(parts) > 1 else ""
        cell = ws.cell(row=2, column=col_idx, value=f"{nom}\n{prenom}")
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # ── Data rows ──
    def _get_val(emp_dict, var_dict, key):
        if key == "nom":
            parts = (emp_dict["nom"] or "").strip().split(" ", 1)
            return parts[0].upper() if parts else ""
        if key == "prenom":
            parts = (emp_dict["nom"] or "").strip().split(" ", 1)
            return parts[1].capitalize() if len(parts) > 1 else ""
        if key == "matricule":     return emp_dict["matricule"]
        if key == "contrat_type":  return emp_dict["contrat_type"] or "CDI"
        if key == "date_debut":    return emp_dict["date_debut"]
        if key == "date_fin":      return emp_dict["date_fin"]
        if key == "nb_heures_base": return emp_dict["nb_heures_base"]
        if key == "taux_horaire":  return emp_dict["taux_horaire"]
        if key == "salaire_mensuel": return emp_dict["salaire_mensuel"]
        if key == "prime_anciennete": return emp_dict["prime_anciennete"]
        if key == "mutuelle":      return "Oui" if emp_dict["mutuelle"] else "Non"
        if key == "avantage_voiture": return emp_dict["avantage_voiture"]
        return var_dict.get(key)

    # Alternating row colors
    BG_ALT1 = "0D1422"
    BG_ALT2 = "111827"

    for row_idx, (field_key, field_label) in enumerate(EXPORT_FIELDS, start=3):
        # Alternate fill
        bg = BG_ALT1 if row_idx % 2 == 1 else BG_ALT2

        label_cell = ws.cell(row=row_idx, column=1, value=field_label)
        label_cell.font      = label_font
        label_cell.fill      = PatternFill("solid", fgColor=bg)
        label_cell.alignment = left
        label_cell.border    = thin_border
        ws.row_dimensions[row_idx].height = 16

        for col_idx, emp in enumerate(employes, start=2):
            uid = emp["id"]
            var = var_map.get(uid, {})
            val = _get_val(dict(emp), var, field_key)

            cell = ws.cell(row=row_idx, column=col_idx, value=val if val is not None else None)
            cell.font      = data_font
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = center
            cell.border    = thin_border

    # Freeze pane : figer colonnes A et ligne 1-2
    ws.freeze_panes = "B3"

    # ── Output ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"paie_{annee}_{mois:02d}_{mois_label}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
