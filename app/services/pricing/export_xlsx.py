"""Export Excel — produits coûts matières (openpyxl)."""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

_EUR_FMT = '#,##0.0000" €"'
_PCT_FMT = "0.00"


def _role_label(role: str) -> str:
    labels = {
        "frontal": "Frontal",
        "adhesif": "Adhésif",
        "silicone": "Silicone",
        "glassine": "Glassine",
    }
    if role in labels:
        return labels[role]
    if role.startswith("extra_"):
        return f"Extra {role.split('_')[-1]}"
    return role


def build_products_workbook(
    products: list[dict[str, Any]],
    materials_by_id: dict[int, dict[str, Any]],
) -> bytes:
    """
    products: liste de dicts avec keys code, name, cost (ProductCostOut-like dict),
    frontal_id, adhesif_id, silicone_id, glassine_id, component material refs.
    materials_by_id: matériels enrichis (name, appellation, unit_price, computed...).
    """
    wb = Workbook()
    ws_p = wb.active
    ws_p.title = "Produits"

    prod_headers = [
        "Code",
        "Nom",
        "Frontal",
        "Adhésif",
        "Silicone",
        "Glassine",
        "Coût €/m²",
        "Marge €/m²",
        "Prix vente €/m²",
    ]
    ws_p.append(prod_headers)
    for col in range(1, len(prod_headers) + 1):
        ws_p.cell(1, col).font = Font(bold=True)

    def mat_cell(mid: Optional[int]) -> str:
        if not mid:
            return ""
        m = materials_by_id.get(int(mid))
        if not m:
            return f"#{mid}"
        return (m.get("appellation_code") or m.get("name") or "").strip()

    for p in products:
        cost = p.get("cost") or {}
        ws_p.append(
            [
                p.get("code"),
                p.get("name"),
                mat_cell(p.get("frontal_id")),
                mat_cell(p.get("adhesif_id")),
                mat_cell(p.get("silicone_id")),
                mat_cell(p.get("glassine_id")),
                float(cost.get("total_eur_per_m2") or 0),
                float(cost.get("margin_eur_m2") or 0),
                float(cost.get("sell_price_eur_m2") or 0),
            ]
        )

    for row in range(2, ws_p.max_row + 1):
        for col in (7, 8, 9):
            ws_p.cell(row, col).number_format = _EUR_FMT

    ws_p.freeze_panes = "A2"
    for i, w in enumerate([14, 28, 16, 16, 14, 14, 14, 12, 14], start=1):
        ws_p.column_dimensions[get_column_letter(i)].width = w

    ws_m = wb.create_sheet("Matières utilisées")
    mat_headers = [
        "ID",
        "Catégorie",
        "Nom",
        "Appellation",
        "Prix unitaire",
        "Devise",
        "Base",
        "€/m² calculé",
    ]
    ws_m.append(mat_headers)
    for col in range(1, len(mat_headers) + 1):
        ws_m.cell(1, col).font = Font(bold=True)

    used_ids: set[int] = set()
    for p in products:
        for key in ("frontal_id", "adhesif_id", "silicone_id", "glassine_id"):
            if p.get(key):
                used_ids.add(int(p[key]))
        for mid in p.get("extra_material_ids") or []:
            used_ids.add(int(mid))

    for mid in sorted(used_ids):
        m = materials_by_id.get(mid)
        if not m:
            continue
        computed = m.get("computed") or {}
        ws_m.append(
            [
                mid,
                m.get("category_code"),
                m.get("name"),
                m.get("appellation_code"),
                float(m.get("unit_price") or 0),
                m.get("price_currency"),
                m.get("price_basis"),
                float(computed.get("price_eur_per_m2") or 0),
            ]
        )

    for row in range(2, ws_m.max_row + 1):
        ws_m.cell(row, 5).number_format = _EUR_FMT
        ws_m.cell(row, 8).number_format = _EUR_FMT

    ws_m.freeze_panes = "A2"
    for i, w in enumerate([8, 12, 24, 16, 14, 8, 10, 14], start=1):
        ws_m.column_dimensions[get_column_letter(i)].width = w

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
