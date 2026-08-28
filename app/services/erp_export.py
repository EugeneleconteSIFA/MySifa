"""Export xlsx d'une vue ERP.

Ce module ne décide rien : il reçoit une vue déjà construite par
`erp_mirror.lister()` — donc filtrée, triée et nettoyée de ses sentinelles
exactement comme la grille — et l'écrit dans un classeur.

Deux règles tiennent tout le fichier :

1. **Ce qu'on exporte est ce qu'on voit.** Mêmes colonnes, même ordre, mêmes
   libellés, mêmes énumérations traduites. Un export qui ramènerait les
   quarante colonnes de la table pendant que l'écran en montre douze obligerait
   à refaire le tri dans Excel — c'est précisément le travail qu'on supprime.
2. **Les nombres sortent en nombres.** Une quantité écrite « 12 500 » dans une
   cellule texte ne se somme pas. Les dates sortent en dates, les montants en
   montants, avec leur format — sinon le fichier n'est bon qu'à être relu à
   l'œil.

La feuille « Critères » porte le filtre qui a produit le fichier. Un export
qu'on retrouve trois semaines plus tard sans savoir ce qu'il contient ne vaut
rien ; celui-là se relit tout seul.
"""

import io
from datetime import date, datetime

# ── Formats de nombre, dans les codes de MySifa ──────────────────────────────
FMT = {
    "date": "DD/MM/YYYY",
    "datetime": "DD/MM/YYYY HH:MM",
    "nombre": "#,##0",
    "qte": "#,##0",
    "prix": "#,##0.0000",
    "montant": "#,##0.00",
    "pct": "0.00",
    "id": "0",              # un numéro de pièce n'a pas de séparateur de milliers
}

ALIGNE_DROITE = {"nombre", "qte", "prix", "montant", "pct", "id"}

RATT_LIB = {"oui": "Rattaché", "partiel": "Partiel",
            "douteux": "À vérifier", "non": ""}

# Couleurs du thème MySifa (variante claire : un classeur s'imprime).
BG_ENTETE = "0E7490"
FG_ENTETE = "FFFFFF"
BG_BANDE = "F1F5F9"


def _date(v):
    """« 2026-08-26 09:12 » → objet date/datetime, ou le texte tel quel."""
    s = str(v or "").strip()
    if len(s) >= 16:
        try:
            return datetime.strptime(s[:16], "%Y-%m-%d %H:%M")
        except ValueError:
            pass
        try:
            return datetime.strptime(s[:16], "%Y-%m-%dT%H:%M")
        except ValueError:
            pass
    if len(s) >= 10:
        try:
            return date(int(s[0:4]), int(s[5:7]), int(s[8:10]))
        except ValueError:
            return s
    return s or None


def _nb(v):
    try:
        f = float(v)
    except (TypeError, ValueError):
        return v
    return int(f) if f == int(f) else f


def valeur_cellule(col, v, enums):
    """(valeur, format) pour une cellule. `None` = case laissée vide.

    Une valeur absente reste une case VIDE, pas un tiret : dans un tableur, le
    tiret est du texte et casse la somme de la colonne.
    """
    t = col.get("type") or "texte"
    if t == "ratt":
        etat = (v or {}).get("etat") if isinstance(v, dict) else None
        return (RATT_LIB.get(etat or "non") or None), None
    if v is None or v == "":
        return None, None
    if t in ("date", "datetime"):
        d = _date(v)
        return d, (FMT[t] if isinstance(d, (date, datetime)) else None)
    if t == "enum":
        table = (enums or {}).get(col.get("enum")) or {}
        code = str(v)
        if code == "255":
            return None, None          # sentinelle WinDev : octet non renseigné
        return table.get(code, code), None
    if t == "bool":
        return ("Oui" if str(v) in ("1", "True", "true") else None), None
    if t in FMT:
        n = _nb(v)
        return n, (FMT[t] if isinstance(n, (int, float)) else None)
    return v, None


def _largeur(col):
    """La largeur d'écran, en pixels, traduite en largeur de colonne Excel."""
    px = col.get("largeur") or 120
    return max(9, min(52, round(px / 7.2, 1)))


def construire(titre, colonnes, lignes, enums=None, criteres=None,
               note=None, tronque=False):
    """Rend le classeur en octets.

    `colonnes` est déjà dans l'ordre d'affichage ; `lignes` sort de
    `erp_mirror.lister()`. `criteres` est une liste de couples (libellé,
    valeur) écrite sur la seconde feuille.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    # Excel refuse 31 caractères et cinq ponctuations dans un nom d'onglet.
    ws.title = "".join(c for c in titre if c not in "[]:*?/\\")[:31] or "Vue"

    gras = Font(bold=True, color=FG_ENTETE, size=10)
    fond = PatternFill("solid", fgColor=BG_ENTETE)
    centre = Alignment(horizontal="center", vertical="center", wrap_text=False)
    droite = Alignment(horizontal="right", vertical="center")

    for i, c in enumerate(colonnes, start=1):
        cell = ws.cell(row=1, column=i, value=c.get("label") or c.get("nom"))
        cell.font = gras
        cell.fill = fond
        cell.alignment = centre
        ws.column_dimensions[get_column_letter(i)].width = _largeur(c)
    ws.row_dimensions[1].height = 22

    for r, ligne in enumerate(lignes, start=2):
        for i, c in enumerate(colonnes, start=1):
            v, fmt = valeur_cellule(c, ligne.get(c["nom"]), enums)
            if v is None:
                continue
            cell = ws.cell(row=r, column=i, value=v)
            if fmt:
                cell.number_format = fmt
            if (c.get("type") or "") in ALIGNE_DROITE:
                cell.alignment = droite

    # L'en-tête reste à l'écran, et le filtre natif d'Excel est déjà posé : le
    # fichier s'ouvre prêt à être trié, pas prêt à être mis en forme.
    ws.freeze_panes = "A2"
    if colonnes:
        ws.auto_filter.ref = "A1:%s%d" % (
            get_column_letter(len(colonnes)), max(2, len(lignes) + 1))

    fc = wb.create_sheet("Critères")
    fc.column_dimensions["A"].width = 28
    fc.column_dimensions["B"].width = 74
    lignes_c = list(criteres or [])
    if tronque:
        lignes_c.append(("Attention", note or "Export tronqué."))
    elif note:
        lignes_c.append(("Note", note))
    for r, (lib, val) in enumerate(lignes_c, start=1):
        a = fc.cell(row=r, column=1, value=lib)
        a.font = Font(bold=True, size=10)
        a.fill = PatternFill("solid", fgColor=BG_BANDE)
        a.alignment = Alignment(vertical="top")
        b = fc.cell(row=r, column=2, value=val)
        b.alignment = Alignment(vertical="top", wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
